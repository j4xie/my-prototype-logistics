# -*- coding: utf-8 -*-
"""Update test Excel with SKU assembly + full E2E verified flow"""
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

path = os.path.join(os.path.dirname(__file__), '..', '..', 'tests',
                    '\u516d\u6247\u95e8\u7b2c\u4e00\u671f-\u624b\u52a8\u6d4b\u8bd5\u8ba1\u5212.xlsx')
wb = openpyxl.load_workbook(path)

# === Sheet 1 ===
ws1 = wb['\u5b8c\u6574\u6d4b\u8bd5\u7528\u4f8b']
r = ws1.max_row + 1

# Phase G: SKU
ws1.cell(row=r, column=1, value='Phase G: SKU \u7ec4\u88c5\u6d41\u7a0b')
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

sku_cases = [
    ['T45', 'factory_admin1/123456', 'Web',
     '\u4ea7\u54c1\u7ba1\u7406 > \u70b9\u300cSKU\u7ec4\u88c5\u300d\u6309\u94ae',
     '\u8d26\u53f7: factory_admin1',
     '\u5f39\u51fa\u7ec4\u88c5\u5bf9\u8bdd\u6846\n\u52a0\u8f7d\u4ea7\u54c1\u6a21\u677f\u5217\u8868+\u5ba2\u6237\u5217\u8868', '', ''],
    ['T46', 'factory_admin1', 'Web',
     '\u9009\u6a21\u677f(\u58a8\u9c7c\u5708) > \u9009\u5ba2\u6237(\u6c47\u5ead) > \u70b9\u300c\u521b\u5efaSKU\u300d',
     '\u6a21\u677f: \u58a8\u9c7c\u5708\n\u5ba2\u6237: \u6c47\u5ead\u5546\u8d38',
     '\u63d0\u793a"SKU\u521b\u5efa\u6210\u529f: CP-\u6c47\u5ead\u5546\u8d38-\u58a8\u9c7c\u5708"\n\u914d\u65b9\u81ea\u52a8\u4ece\u6a21\u677f\u590d\u5236', '', '\u9700\u6a21\u677f\u5df2\u914dBOM'],
    ['T47', 'factory_admin1', 'Web',
     '\u4ea7\u54c1\u5217\u8868 > \u641c\u7d22"CP-\u6c47\u5ead"',
     '',
     '\u663e\u793aSKU: CP-\u6c47\u5ead\u5546\u8d38-\u58a8\u9c7c\u5708\ntemplateId\u6709\u503c, customerId\u6709\u503c', '', '\u786e\u8ba4SKU\u5c5e\u6027\u5b8c\u6574'],
]

for c in sku_cases:
    for col, val in enumerate(c, 1):
        ws1.cell(row=r, column=col, value=val)
    r += 1

# Phase H: E2E
ws1.cell(row=r, column=1, value='Phase H: \u5b8c\u6574E2E\u94fe\u8def (SKU\u2192\u6392\u4ea7\u2192\u8c03\u62e8\u2192\u7b7e\u6536\u2192\u62a5\u5de5)')
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

e2e_cases = [
    ['T48', 'factory_admin1', 'Web',
     '\u751f\u4ea7\u8ba1\u5212 > \u65b0\u5efa > \u4ea7\u54c1\u9009T46\u7684SKU > \u8ba1\u5212\u91cf200',
     '\u4ea7\u54c1: CP-\u6c47\u5ead-\u58a8\u9c7c\u5708\n\u8ba1\u5212\u91cf: 200kg',
     '\u8ba1\u5212\u521b\u5efa\u6210\u529f, \u72b6\u6001\u5f85\u6267\u884c', '', '\u7528SKU\u6392\u4ea7\u800c\u975e\u6a21\u677f'],
    ['T49', 'factory_admin1', 'Web',
     '\u8ba1\u5212\u5217\u8868 > T48\u8ba1\u5212 > \u70b9\u300c\u751f\u6210\u8c03\u62e8\u5355\u300d',
     '',
     '\u63d0\u793a"\u8c03\u62e8\u5355\u5df2\u751f\u6210,\u5171N\u9879\u7269\u6599"\nBOM\u5c55\u5f00\u8ba1\u7b97\u539f\u8f85\u6599\u9700\u6c42', '', 'BOM\u8f6c\u6362\u7387\u6b63\u786e'],
    ['T50', 'warehouse_mgr1/123456', 'RN',
     '\u9996\u9875 > \u9886\u6599\u8c03\u62e8 > \u627e\u5230T49\u8c03\u62e8\u5355',
     'warehouse_mgr1 / 123456',
     '\u663e\u793a\u8c03\u62e8\u5355\u72b6\u6001:\u5df2\u7533\u8bf7\n\u7269\u6599\u660e\u7ec6\u5b8c\u6574', '', ''],
    ['T51', 'warehouse_mgr1', 'RN',
     '\u8c03\u62e8\u8be6\u60c5 > \u5ba1\u6279\u901a\u8fc7(approve) > \u53d1\u8d27(ship)',
     '',
     'approve\u540e\u5df2\u6279\u51c6\nship\u540e\u5df2\u53d1\u8fd0\n\u539f\u6599\u5e93\u5b58FEFO\u6263\u51cf', '', '\u9a8c\u8bc1material_batches'],
    ['T52', 'workshop_sup1/123456', 'RN',
     '\u6279\u6b21\u7ba1\u7406 > \u7269\u6599\u8c03\u62e8 > \u7b7e\u6536(receive) > \u786e\u8ba4(confirm)',
     'workshop_sup1 / 123456',
     'receive\u540e\u5df2\u7b7e\u6536\nconfirm\u540e\u5df2\u786e\u8ba4\n\u5de5\u5382\u4ed3\u65b0\u589e\u539f\u6599\u6279\u6b21', '', ''],
    ['T53', 'workshop_sup1', 'RN',
     '\u6279\u6b21\u8be6\u60c5 > \u539f\u6599\u9886\u6599 > \u5f55\u5165\u5b9e\u9645\u7528\u91cf',
     '\u6bcf\u4e2a\u7269\u6599\u8f93\u5165\u5b9e\u9645\u91cf',
     '\u5f55\u5165\u6210\u529f\n\u8fbe\u6210\u7387\u663e\u793a\u6b63\u786e', '', ''],
    ['T54', 'workshop_sup1', 'RN',
     '\u5de5\u5e8f\u64cd\u4f5c > \u7b7e\u5230 > \u9009\u5458\u5de5 > \u62a5\u91cf > \u7b7e\u9000',
     '\u5458\u5de5\u5217\u8868\u9009\u62e9+\u62a5\u91cf50',
     '\u7b7e\u5230/\u62a5\u5de5/\u7b7e\u9000\u5168\u6d41\u7a0b\u901a', '', '\u5458\u5de5\u9009\u62e9\u53cc\u6a21\u5f0f'],
]

for c in e2e_cases:
    for col, val in enumerate(c, 1):
        ws1.cell(row=r, column=col, value=val)
    r += 1

# === Sheet 2 ===
ws2 = wb['\u81ea\u52a8\u94fe\u8def\u9a8c\u8bc1']
r2 = ws2.max_row + 1

auto = [
    ['C11', 'T46: \u7ec4\u88c5SKU',
     'SkuAssemblyService.assemblesku() \u514b\u9686\u6a21\u677f + copyRecipe\u590d\u5236\u914d\u65b9',
     'DB: product_types\u65b0\u589e(templateId\u6709\u503c) + material_product_conversions\u65b0\u589e', ''],
    ['C12', 'T49: \u751f\u6210\u8c03\u62e8\u5355',
     'ProductionWorkflowOrchestrator: expandBOM\u7528SKU\u7684productTypeId\u67e5\u8be2\u914d\u65b9',
     'API\u8fd4\u56de InternalTransfer, items\u542b\u6b63\u786e\u7269\u6599\u548c\u6570\u91cf', ''],
]

for c in auto:
    for col, val in enumerate(c, 1):
        ws2.cell(row=r2, column=col, value=val)
    r2 += 1

wb.save(path)
print('Done! Added T45-T54 + C11-C12')
