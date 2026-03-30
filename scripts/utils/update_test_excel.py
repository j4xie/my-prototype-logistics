# -*- coding: utf-8 -*-
"""Add requisition flow + employee picker test cases to 六扇门第一期-手动测试计划.xlsx"""
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

path = os.path.join(os.path.dirname(__file__), '..', '..', 'tests', '六扇门第一期-手动测试计划.xlsx')
wb = openpyxl.load_workbook(path)

# === Sheet 1 ===
ws1 = wb['完整测试用例']
r = ws1.max_row + 1

# Phase E header
ws1.cell(row=r, column=1, value='Phase E: 领料调拨流程')
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

cases = [
    ['T35', 'factory_admin1/123456', 'Web',
     '生产管理 > 生产计划 > PENDING计划 > 点「生成调拨单」',
     '选择已有计划(须有BOM)',
     '提示"调拨单已生成,共N项物料"\n自动创建InternalTransfer(REQUESTED)', '', '需material_product_conversions数据'],
    ['T36', 'warehouse_mgr1/123456', 'RN',
     '首页 > 点「领料调拨」卡片',
     'warehouse_mgr1 / 123456',
     '进入调拨列表,显示T35的调拨单\n状态:已申请', '', ''],
    ['T37', 'warehouse_mgr1', 'RN',
     '调拨列表 > 点击调拨单 > 查看详情',
     '',
     '显示物料清单(名称/数量/单位)\n关联计划号\n操作按钮', '', ''],
    ['T38', 'warehouse_mgr1', 'RN',
     '调拨详情 > approve > ship',
     '依次审批和发货',
     'approve后已批准\nship后已发运\n原料库存FEFO扣减', '', '验证material_batches.usedQuantity'],
    ['T39', 'workshop_sup1/123456', 'RN',
     '批次管理 > 物料调拨 > receive > confirm',
     '签收并确认',
     'receive后已签收\nconfirm后已确认\n工厂仓新增原料批次', '', ''],
    ['T40', 'workshop_sup1', 'RN',
     '批次详情 > 点「原料领料」',
     '选择关联批次',
     '显示BOM物料清单\n计划量/实际量/达成率\n每项可点「录入实际用量」', '', ''],
    ['T41', 'workshop_sup1', 'RN',
     '原料消耗 > 录入实际用量 > 确认',
     '输入计划量的98%',
     '提示录入成功\n刷新显示达成率\n后端创建MaterialConsumption', '', ''],
]

for c in cases:
    for col, val in enumerate(c, 1):
        ws1.cell(row=r, column=col, value=val)
    r += 1

# Phase F header
ws1.cell(row=r, column=1, value='Phase F: 报工员工选择优化')
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

emp_cases = [
    ['T42', 'workshop_sup1', 'RN',
     '工序操作 > 选工序 > 报产量区域',
     '选有签到员工的工序',
     '显示已签到员工列表(蓝色标签)\n扫码工牌备选\n主管自己报按钮', '', '员工需先签到'],
    ['T43', 'workshop_sup1', 'RN',
     '点员工标签 > 输数量 > 提交',
     '选员工+数量50',
     '员工已选中显示\n提交成功,targetWorkerId正确', '', ''],
    ['T44', 'workshop_sup1', 'RN',
     '工序列表 > 右上角扫码 > 扫已签到员工',
     '扫码',
     '自动跳转该员工工序\n员工自动选中', '', '扫码定位功能'],
]

for c in emp_cases:
    for col, val in enumerate(c, 1):
        ws1.cell(row=r, column=col, value=val)
    r += 1

# === Sheet 2 ===
ws2 = wb['自动链路验证']
r2 = ws2.max_row + 1

auto = [
    ['C07', 'T35: 生成调拨单',
     'BomExpansionService.expandBOM() + createTransfer() + requestTransfer()',
     'API返回InternalTransfer含items', ''],
    ['C08', 'T38: ship发货',
     'deductSourceInventory() FEFO扣减RAW_MATERIAL',
     'DB: material_batches.usedQuantity增加', ''],
    ['C09', 'T39: confirm确认',
     'createTargetInventory()创建工厂仓MaterialBatch',
     'DB: material_batches新增目标工厂记录', ''],
    ['C10', 'T41: 录入用量',
     'recordMaterialConsumption()创建消耗记录',
     'DB: material_consumptions新增记录', ''],
]

for c in auto:
    for col, val in enumerate(c, 1):
        ws2.cell(row=r2, column=col, value=val)
    r2 += 1

out = path.replace('.xlsx', '-v2.xlsx')
wb.save(out)
print(f'Done! Added T35-T44 + C07-C10 to {out}')
