#!/usr/bin/env python3
"""
Auto-adaptation script: Extract system data from uploaded Excel files.
Reads xlsx files and generates SQL INSERT statements for:
  - smart_bi_sales_data
  - smart_bi_finance_data (COST, REVENUE, AR, BUDGET)
  - smart_bi_department_data

Usage:
  python generate_system_data.py <xlsx_file> [--factory-id F001] [--output seed_data.sql]
  python generate_system_data.py tests/test-data/Test-mock-food-normal-s42.xlsx
"""

import sys
import os
import argparse
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


def safe_num(val, default=0):
    """Convert to float safely, return default if not numeric."""
    if val is None:
        return default
    try:
        f = float(val)
        if f != f:  # NaN check
            return default
        return f
    except (ValueError, TypeError):
        return default


def safe_str(val, max_len=200):
    """Convert to SQL-safe string."""
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''").strip()
    if len(s) > max_len:
        s = s[:max_len]
    return f"'{s}'"


def safe_date(val):
    """Convert to SQL date string."""
    if val is None:
        return "NULL"
    if isinstance(val, datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    if isinstance(val, date):
        return f"'{val.isoformat()}'"
    s = str(val).strip()
    if not s:
        return "NULL"
    return f"'{s}'"


def detect_sheets(wb):
    """Detect which sheets are available and return a capability map."""
    sheets = set(wb.sheetnames)
    caps = {
        'has_sales_detail': '销售明细' in sheets,
        'has_ar_aging': '应收账款账龄' in sheets,
        'has_monthly_kpi': '月度经营分析' in sheets,
        'has_revenue_summary': '收入及净利简表' in sheets,
        'has_budget_execution': '费用预算执行' in sheets,
        'has_balance_sheet': '资产负债表' in sheets,
        'has_cash_flow': '现金流量表' in sheets,
        'has_inventory': '库存台账' in sheets,
        'has_rebate_detail': '24年返利明细' in sheets,
        'profit_tables': [s for s in sheets if '利润表' in s],
    }
    return caps


def extract_sales_data(wb, factory_id):
    """Extract from 销售明细 sheet → smart_bi_sales_data rows."""
    if '销售明细' not in wb.sheetnames:
        print("  [SKIP] 销售明细 sheet not found - cannot generate sales data")
        return []

    ws = wb['销售明细']
    rows = []

    # Find header row (look for '日期' or '订单号')
    header_row = None
    for r in range(1, min(5, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v and '日期' in str(v) for v in vals if v):
            header_row = r
            break

    if header_row is None:
        print("  [SKIP] Cannot find header row in 销售明细")
        return []

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h = str(h).strip()
        if '日期' in h:
            col_map['date'] = i
        elif '订单' in h:
            col_map['order_no'] = i
        elif '客户' in h:
            col_map['customer'] = i
        elif '产品' == h or '产品名' in h:
            col_map['product'] = i
        elif '规格' in h:
            col_map['spec'] = i
        elif '数量' in h:
            col_map['quantity'] = i
        elif '单价' in h:
            col_map['unit_price'] = i
        elif '金额' in h:
            col_map['amount'] = i
        elif '区域' in h:
            col_map['region'] = i
        elif '业务员' in h or '销售' in h:
            col_map['salesperson'] = i

    print(f"  [OK] 销售明细: found columns {list(col_map.keys())}")

    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(ws.max_column)]
        # Skip empty rows
        if not any(v is not None for v in vals):
            continue
        # Skip summary/total rows
        first_val = vals[0] if vals else None
        if first_val and isinstance(first_val, str) and ('合计' in first_val or '总计' in first_val):
            continue

        order_date = vals[col_map.get('date', 0)] if 'date' in col_map else None
        if order_date is None:
            continue

        customer = vals[col_map.get('customer', 2)] if 'customer' in col_map else None
        product = vals[col_map.get('product', 3)] if 'product' in col_map else None
        quantity = safe_num(vals[col_map.get('quantity', 5)] if 'quantity' in col_map else 0)
        unit_price = safe_num(vals[col_map.get('unit_price', 6)] if 'unit_price' in col_map else 0)
        amount = safe_num(vals[col_map.get('amount', 7)] if 'amount' in col_map else 0)
        region = vals[col_map.get('region', 8)] if 'region' in col_map else None
        salesperson = vals[col_map.get('salesperson', 9)] if 'salesperson' in col_map else None

        # Derive cost and profit (estimate: food industry avg COGS ~65-70%)
        cost = round(amount * 0.67, 2) if amount > 0 else 0
        profit = round(amount - cost, 2)
        gross_margin = round(profit / amount, 4) if amount > 0 else 0

        rows.append({
            'factory_id': factory_id,
            'order_date': safe_date(order_date),
            'salesperson_name': safe_str(salesperson, 100),
            'region': safe_str(region, 100),
            'customer_name': safe_str(customer, 200),
            'product_name': safe_str(product, 200),
            'product_category': safe_str(product, 100),  # Same as product for now
            'quantity': quantity,
            'amount': amount,
            'unit_price': unit_price,
            'cost': cost,
            'profit': profit,
            'gross_margin': gross_margin,
        })

    print(f"  [OK] Extracted {len(rows)} sales records")
    return rows


def extract_finance_cost_data(wb, factory_id):
    """Extract COST records from 利润表 sheets → smart_bi_finance_data."""
    profit_sheets = [s for s in wb.sheetnames if '利润表' in s]
    if not profit_sheets:
        print("  [SKIP] No 利润表 sheets found - cannot generate COST data")
        return []

    rows = []
    # Use the main center profit table for cost breakdown
    main_sheet = None
    for name in profit_sheets:
        if '中心利润表' in name and '销售' not in name:
            main_sheet = name
            break
    if not main_sheet:
        main_sheet = profit_sheets[0]

    ws = wb[main_sheet]
    print(f"  [OK] Reading COST data from: {main_sheet}")

    # Parse the P&L structure - find month columns and cost rows
    # Typical structure: Row labels in col A, months as column groups
    # Each month has: 预算数, 本月实际 (budget, actual)

    # Find header row with month references
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(30, ws.max_column + 1))]
        str_vals = [str(v) for v in vals if v]
        if any('月' in v for v in str_vals) or any('预算' in v for v in str_vals):
            header_row = r
            break

    if header_row is None:
        # Try to find '行次' which is common in Chinese P&L
        for r in range(1, min(10, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, min(10, ws.max_column + 1))]
            if any(v and '行次' in str(v) for v in vals if v):
                header_row = r
                break

    if header_row is None:
        print(f"  [SKIP] Cannot find header in {main_sheet}")
        return []

    # Detect month-column pairs (budget_col, actual_col) for each month
    month_cols = []
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]

    # Look for patterns: "1月预算数", "1月本月实际" or column groups
    # Also check row above header for month labels
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is None:
            continue
        s = str(val).strip()
        if '预算' in s:
            # Next column should be actual
            actual_col = c + 1
            if actual_col <= ws.max_column:
                actual_val = ws.cell(header_row, actual_col).value
                if actual_val and '实际' in str(actual_val):
                    # Extract month number from the label or from row above
                    month_num = None
                    # Try to get month from merged cell above
                    above_val = ws.cell(header_row - 1, c).value if header_row > 1 else None
                    if above_val:
                        for m in range(1, 13):
                            if f'{m}月' in str(above_val):
                                month_num = m
                                break
                    # Try from the label itself
                    if month_num is None:
                        for m in range(1, 13):
                            if f'{m}月' in s:
                                month_num = m
                                break
                    if month_num is None:
                        month_num = len(month_cols) + 1  # Fallback: sequential
                    month_cols.append((month_num, c, actual_col))

    if not month_cols:
        print(f"  [SKIP] Cannot detect month columns in {main_sheet}")
        return []

    print(f"  [OK] Detected {len(month_cols)} month columns")

    # Find key cost rows by label
    cost_categories = {
        '营业成本': 'COGS',
        '销售费用': '销售费用',
        '管理费用': '管理费用',
        '研发费用': '研发费用',
        '财务费用': '财务费用',
        '营业收入': 'REVENUE',
    }

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()

        category = None
        record_type = 'COST'
        for key, cat in cost_categories.items():
            if label == key or label.startswith(key):
                category = cat
                if key == '营业收入':
                    record_type = 'REVENUE'
                break

        if category is None:
            continue

        for month_num, budget_col, actual_col in month_cols:
            budget_val = safe_num(ws.cell(r, budget_col).value)
            actual_val = safe_num(ws.cell(r, actual_col).value)

            if budget_val == 0 and actual_val == 0:
                continue

            record_date = f"'2025-{month_num:02d}-01'"
            variance = round(actual_val - budget_val, 2)

            if record_type == 'REVENUE':
                rows.append({
                    'factory_id': factory_id,
                    'record_date': record_date,
                    'record_type': 'REVENUE',
                    'category': safe_str(category, 100),
                    'budget_amount': budget_val,
                    'actual_amount': actual_val,
                    'variance_amount': variance,
                    'total_cost': 0,
                })
            else:
                rows.append({
                    'factory_id': factory_id,
                    'record_date': record_date,
                    'record_type': 'COST',
                    'category': safe_str(category, 100),
                    'budget_amount': budget_val,
                    'actual_amount': actual_val,
                    'variance_amount': variance,
                    'total_cost': actual_val,
                })

    print(f"  [OK] Extracted {len(rows)} COST/REVENUE records from P&L")
    return rows


def extract_finance_ar_data(wb, factory_id):
    """Extract AR records from 应收账款账龄 sheet."""
    if '应收账款账龄' not in wb.sheetnames:
        print("  [SKIP] 应收账款账龄 sheet not found - cannot generate AR data")
        return []

    ws = wb['应收账款账龄']
    rows = []

    # Find header
    header_row = None
    for r in range(1, min(5, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v and '客户' in str(v) for v in vals if v):
            header_row = r
            break

    if header_row is None:
        print("  [SKIP] Cannot find header in 应收账款账龄")
        return []

    # Multi-row header: aging buckets may be in the row BELOW the main header
    # Scan header_row and header_row+1 for all column names
    col_map = {}
    for scan_row in [header_row, header_row + 1]:
        for i in range(ws.max_column):
            h = ws.cell(scan_row, i + 1).value
            if h is None:
                continue
            h = str(h).strip()
            if '客户' in h:
                col_map['customer'] = i
            elif '应收' in h and ('余额' in h or '合计' in h):
                col_map['balance'] = i
            elif h.startswith('应收'):
                col_map['balance'] = i
            elif '0-30' in h:
                col_map['d0_30'] = i
            elif '31-60' in h:
                col_map['d31_60'] = i
            elif '61-90' in h:
                col_map['d61_90'] = i
            elif '91-180' in h:
                col_map['d91_180'] = i
            elif '180天以上' in h or (h.startswith('180') and '以上' in h):
                col_map['d180plus'] = i
            elif '信用' in h:
                col_map['credit_limit'] = i

    # Data starts after the last header row
    data_start = header_row + 1
    if 'd0_30' in col_map:
        data_start = header_row + 2  # Multi-row header

    print(f"  [OK] 应收账款账龄: found columns {list(col_map.keys())}, data starts row {data_start}")

    for r in range(data_start, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(ws.max_column)]
        customer = vals[col_map.get('customer', 0)] if 'customer' in col_map else None
        if customer is None or (isinstance(customer, str) and ('合计' in customer or '占比' in customer)):
            continue

        balance = safe_num(vals[col_map.get('balance', 1)] if 'balance' in col_map else 0)
        d0_30 = safe_num(vals[col_map.get('d0_30', 2)] if 'd0_30' in col_map else 0)
        d31_60 = safe_num(vals[col_map.get('d31_60', 3)] if 'd31_60' in col_map else 0)
        d61_90 = safe_num(vals[col_map.get('d61_90', 4)] if 'd61_90' in col_map else 0)
        d91_180 = safe_num(vals[col_map.get('d91_180', 5)] if 'd91_180' in col_map else 0)
        d180plus = safe_num(vals[col_map.get('d180plus', 6)] if 'd180plus' in col_map else 0)

        if balance <= 0:
            continue

        # Create one AR record per aging bucket with weighted avg days
        buckets = [
            (d0_30, 15, '0-30天'),
            (d31_60, 45, '31-60天'),
            (d61_90, 75, '61-90天'),
            (d91_180, 135, '91-180天'),
            (d180plus, 270, '180天以上'),
        ]

        # Create a single combined AR record per customer
        # Weighted average aging days
        total_weighted = sum(amt * days for amt, days, _ in buckets if amt > 0)
        avg_aging = int(total_weighted / balance) if balance > 0 else 0
        collection = round(balance * 0.3, 2)  # Estimate 30% collection rate

        rows.append({
            'factory_id': factory_id,
            'record_date': "'2025-12-31'",  # Latest snapshot
            'record_type': 'AR',
            'customer_name': safe_str(customer, 200),
            'receivable_amount': balance,
            'collection_amount': collection,
            'aging_days': avg_aging,
        })

    print(f"  [OK] Extracted {len(rows)} AR records")
    return rows


def extract_finance_budget_data(wb, factory_id):
    """Extract BUDGET records from 费用预算执行 sheet."""
    if '费用预算执行' not in wb.sheetnames:
        print("  [SKIP] 费用预算执行 sheet not found - cannot generate BUDGET data")
        return []

    ws = wb['费用预算执行']
    rows = []

    # Structure: Row 2 has month labels (1月, 2月...), Row 3 has (预算, 实际, 差异) repeated
    # Col 1 = department, Col 2 = expense category
    # Data starts at row 4

    # Find the sub-header row with 预算/实际/差异
    sub_header_row = None
    month_label_row = None
    for r in range(1, min(8, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(40, ws.max_column + 1))]
        str_vals = [str(v) for v in vals if v]
        if any('预算' in v for v in str_vals) and any('实际' in v for v in str_vals):
            sub_header_row = r
            month_label_row = r - 1
            break

    if sub_header_row is None:
        print("  [SKIP] Cannot find header in 费用预算执行")
        return []

    # Detect month-column groups: scan the month label row for "N月" patterns
    # then map to (budget, actual, variance) columns
    month_groups = []
    for c in range(3, ws.max_column + 1):  # Start from col 3 (cols 1-2 are dept/category)
        month_label = ws.cell(month_label_row, c).value if month_label_row >= 1 else None
        sub_label = ws.cell(sub_header_row, c).value

        if sub_label and '预算' in str(sub_label):
            month_num = len(month_groups) + 1
            if month_label:
                for m in range(1, 13):
                    if f'{m}月' in str(month_label):
                        month_num = m
                        break
            # budget=c, actual=c+1, variance=c+2
            month_groups.append((month_num, c, c + 1, c + 2))

    if not month_groups:
        print("  [SKIP] Cannot detect month columns in 费用预算执行")
        return []

    print(f"  [OK] 费用预算执行: detected {len(month_groups)} month groups, data starts row {sub_header_row + 1}")

    # Parse rows - Col 1 = department (may be None for continuation), Col 2 = expense category
    current_dept = None
    dept_keywords = ['销售部', '市场部', '研发部', '生产部', '采购部', '人力资源部', '财务部', '行政部',
                     '品控部', '物流部', '技术部', '管理部']

    for r in range(sub_header_row + 1, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = ws.cell(r, 2).value

        # Update current department from col1
        if col1 is not None:
            col1_str = str(col1).strip()
            for dk in dept_keywords:
                if dk in col1_str:
                    current_dept = dk
                    break
            # Skip if col1 is a subtotal/total
            if '小计' in col1_str or '合计' in col1_str or '总计' in col1_str:
                continue

        if col2 is None or not current_dept:
            continue

        category = str(col2).strip()
        if '小计' in category or '合计' in category:
            continue

        for month_num, budget_col, actual_col, variance_col in month_groups:
            budget_val = safe_num(ws.cell(r, budget_col).value)
            actual_val = safe_num(ws.cell(r, actual_col).value)
            variance_val = safe_num(ws.cell(r, variance_col).value)

            if budget_val == 0 and actual_val == 0:
                continue

            rows.append({
                'factory_id': factory_id,
                'record_date': f"'2025-{month_num:02d}-01'",
                'record_type': 'BUDGET',
                'department': safe_str(current_dept, 100),
                'category': safe_str(category, 100),
                'budget_amount': budget_val,
                'actual_amount': actual_val,
                'variance_amount': variance_val,
            })

    print(f"  [OK] Extracted {len(rows)} BUDGET records")
    return rows


def extract_department_data(wb, factory_id):
    """Extract department performance from 收入及净利简表 + 月度经营分析."""
    rows = []

    # Method 1: From 月度经营分析 (regional breakdown)
    # Structure: rows 12-17 have regional data in columns 8-12:
    #   Col 8 = 区域, Col 9 = 收入(万元), Col 10 = 净利(万元), Col 11 = 净利率, Col 12 = 达成率
    if '月度经营分析' in wb.sheetnames:
        ws = wb['月度经营分析']
        print("  [OK] Reading department data from 月度经营分析")

        # Scan all rows for region names in any column
        for r in range(1, ws.max_row + 1):
            # Check columns 1 and 8 for region names
            for check_col in [8, 1]:
                val = ws.cell(r, check_col).value
                if val is None:
                    continue
                val_str = str(val).strip()
                if '分部' in val_str or '区域' in val_str or '省区' in val_str:
                    dept_name = val_str
                    # Revenue/profit are in adjacent columns
                    if check_col == 8:
                        revenue = safe_num(ws.cell(r, 9).value) * 10000  # 万元 → 元
                        net_profit = safe_num(ws.cell(r, 10).value) * 10000
                        margin = safe_num(ws.cell(r, 11).value)
                        achievement = safe_num(ws.cell(r, 12).value)
                    else:
                        revenue = safe_num(ws.cell(r, 2).value) * 10000
                        net_profit = safe_num(ws.cell(r, 3).value) * 10000
                        margin = safe_num(ws.cell(r, 4).value)
                        achievement = safe_num(ws.cell(r, 5).value)

                    if revenue <= 0:
                        continue

                    cost = round(revenue - net_profit, 2)
                    # Estimate headcount: food industry ~35万/person/year
                    headcount = max(5, int(revenue / 350000))
                    # Derive target from achievement rate
                    sales_target = round(revenue / achievement, 2) if achievement > 0 else round(revenue * 0.98, 2)

                    # Avoid duplicates
                    if any(rr['department'] == safe_str(dept_name, 100) for rr in rows):
                        continue

                    rows.append({
                        'factory_id': factory_id,
                        'record_date': "'2025-12-31'",
                        'department': safe_str(dept_name, 100),
                        'headcount': headcount,
                        'sales_amount': revenue,
                        'sales_target': sales_target,
                        'cost_amount': cost,
                        'per_capita_sales': round(revenue / headcount, 2) if headcount > 0 else 0,
                        'per_capita_cost': round(cost / headcount, 2) if headcount > 0 else 0,
                    })

        if rows:
            print(f"  [OK] Extracted {len(rows)} department records from 月度经营分析")
            return rows

    # Method 2: Fall back to 收入及净利简表
    if '收入及净利简表' in wb.sheetnames:
        ws = wb['收入及净利简表']
        print("  [OK] Falling back to 收入及净利简表 for department data")

        # Find region rows and aggregate their annual data
        regions = {}
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val is None:
                continue
            val = str(val).strip()
            if '分部' in val or '区域' in val or '省区' in val:
                # Sum all months for this region
                total_actual = 0
                total_budget = 0
                for c in range(2, ws.max_column + 1):
                    cell_val = ws.cell(r, c).value
                    if cell_val is not None:
                        total_actual += safe_num(cell_val)
                if val not in regions:
                    regions[val] = {'actual': 0, 'budget': 0}
                regions[val]['actual'] += total_actual

        for dept_name, data in regions.items():
            revenue = data['actual']
            if revenue <= 0:
                continue
            cost = round(revenue * 0.93, 2)  # Estimate 93% cost ratio
            headcount = max(5, int(revenue / 350000))

            rows.append({
                'factory_id': factory_id,
                'record_date': "'2025-12-31'",
                'department': safe_str(dept_name, 100),
                'headcount': headcount,
                'sales_amount': revenue,
                'sales_target': round(revenue * 0.98, 2),
                'cost_amount': cost,
                'per_capita_sales': round(revenue / headcount, 2) if headcount > 0 else 0,
                'per_capita_cost': round(cost / headcount, 2) if headcount > 0 else 0,
            })

        print(f"  [OK] Extracted {len(rows)} department records from 收入及净利简表")

    if not rows:
        print("  [SKIP] No department data sources found")

    return rows


def generate_sql(sales_rows, finance_rows, department_rows, factory_id):
    """Generate SQL INSERT statements."""
    lines = []
    lines.append("-- Auto-generated system data from Excel upload")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- Factory: {factory_id}")
    lines.append("")

    # Clean old data first
    lines.append("-- ========== CLEANUP ==========")
    lines.append(f"DELETE FROM smart_bi_sales_data WHERE factory_id = '{factory_id}';")
    lines.append(f"DELETE FROM smart_bi_finance_data WHERE factory_id = '{factory_id}';")
    lines.append(f"DELETE FROM smart_bi_department_data WHERE factory_id = '{factory_id}';")
    lines.append("")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ts = f"'{now}'"  # Timestamp for created_at/updated_at

    # Sales data
    if sales_rows:
        lines.append(f"-- ========== SALES DATA ({len(sales_rows)} rows) ==========")
        for row in sales_rows:
            lines.append(
                f"INSERT INTO smart_bi_sales_data "
                f"(factory_id, order_date, salesperson_name, region, customer_name, "
                f"product_name, product_category, quantity, amount, unit_price, cost, profit, gross_margin, "
                f"created_at, updated_at) "
                f"VALUES ('{row['factory_id']}', {row['order_date']}, {row['salesperson_name']}, "
                f"{row['region']}, {row['customer_name']}, {row['product_name']}, "
                f"{row['product_category']}, {row['quantity']}, {row['amount']}, "
                f"{row['unit_price']}, {row['cost']}, {row['profit']}, {row['gross_margin']}, "
                f"{ts}, {ts});"
            )
        lines.append("")

    # Finance data
    if finance_rows:
        lines.append(f"-- ========== FINANCE DATA ({len(finance_rows)} rows) ==========")
        for row in finance_rows:
            cols = ['factory_id', 'record_date', 'record_type']
            vals = [f"'{row['factory_id']}'", row['record_date'], f"'{row['record_type']}'"]

            if row.get('department') and row['department'] != 'NULL':
                cols.append('department')
                vals.append(row['department'])
            if row.get('category') and row['category'] != 'NULL':
                cols.append('category')
                vals.append(row['category'])
            if row.get('customer_name') and row['customer_name'] != 'NULL':
                cols.append('customer_name')
                vals.append(row['customer_name'])
            if row.get('total_cost', 0) != 0:
                cols.append('total_cost')
                vals.append(str(row['total_cost']))
            if row.get('receivable_amount', 0) != 0:
                cols.append('receivable_amount')
                vals.append(str(row['receivable_amount']))
            if row.get('collection_amount', 0) != 0:
                cols.append('collection_amount')
                vals.append(str(row['collection_amount']))
            if row.get('aging_days', 0) != 0:
                cols.append('aging_days')
                vals.append(str(row['aging_days']))
            if row.get('budget_amount', 0) != 0:
                cols.append('budget_amount')
                vals.append(str(row['budget_amount']))
            if row.get('actual_amount', 0) != 0:
                cols.append('actual_amount')
                vals.append(str(row['actual_amount']))
            if row.get('variance_amount', 0) != 0:
                cols.append('variance_amount')
                vals.append(str(row['variance_amount']))

            cols.extend(['created_at', 'updated_at'])
            vals.extend([ts, ts])

            lines.append(f"INSERT INTO smart_bi_finance_data ({', '.join(cols)}) VALUES ({', '.join(vals)});")
        lines.append("")

    # Department data
    if department_rows:
        lines.append(f"-- ========== DEPARTMENT DATA ({len(department_rows)} rows) ==========")
        for row in department_rows:
            lines.append(
                f"INSERT INTO smart_bi_department_data "
                f"(factory_id, record_date, department, headcount, sales_amount, sales_target, "
                f"cost_amount, per_capita_sales, per_capita_cost, created_at, updated_at) "
                f"VALUES ('{row['factory_id']}', {row['record_date']}, {row['department']}, "
                f"{row['headcount']}, {row['sales_amount']}, {row['sales_target']}, "
                f"{row['cost_amount']}, {row['per_capita_sales']}, {row['per_capita_cost']}, "
                f"{ts}, {ts});"
            )
        lines.append("")

    # Summary
    lines.append(f"-- ========== SUMMARY ==========")
    lines.append(f"-- Sales records: {len(sales_rows)}")
    lines.append(f"-- Finance records: {len(finance_rows)}")
    lines.append(f"-- Department records: {len(department_rows)}")
    lines.append(f"-- Total: {len(sales_rows) + len(finance_rows) + len(department_rows)}")

    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='Generate SmartBI system data from Excel')
    parser.add_argument('xlsx_file', help='Path to Excel file')
    parser.add_argument('--factory-id', default='F001', help='Factory ID (default: F001)')
    parser.add_argument('--output', default=None, help='Output SQL file path')
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_file)
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    output_path = args.output or f"database/seed_system_data_{args.factory_id}.sql"

    print(f"\n{'='*60}")
    print(f"SmartBI System Data Generator")
    print(f"{'='*60}")
    print(f"Input:   {xlsx_path}")
    print(f"Factory: {args.factory_id}")
    print(f"Output:  {output_path}")
    print(f"{'='*60}\n")

    # Load workbook
    print("Loading workbook...")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # Detect capabilities
    caps = detect_sheets(wb)
    print(f"\nDetected sheets:")
    for key, val in caps.items():
        if isinstance(val, list):
            print(f"  {key}: {val}")
        else:
            print(f"  {key}: {'YES' if val else 'NO'}")

    print(f"\n--- Extracting Sales Data ---")
    sales_rows = extract_sales_data(wb, args.factory_id)

    print(f"\n--- Extracting Finance Data (COST/REVENUE) ---")
    cost_rows = extract_finance_cost_data(wb, args.factory_id)

    print(f"\n--- Extracting Finance Data (AR) ---")
    ar_rows = extract_finance_ar_data(wb, args.factory_id)

    print(f"\n--- Extracting Finance Data (BUDGET) ---")
    budget_rows = extract_finance_budget_data(wb, args.factory_id)

    finance_rows = cost_rows + ar_rows + budget_rows

    print(f"\n--- Extracting Department Data ---")
    department_rows = extract_department_data(wb, args.factory_id)

    # Generate SQL
    print(f"\n--- Generating SQL ---")
    sql = generate_sql(sales_rows, finance_rows, department_rows, args.factory_id)

    # Write output
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(sql)

    print(f"\n{'='*60}")
    print(f"DONE! Generated {output_path}")
    print(f"  Sales:      {len(sales_rows)} rows")
    print(f"  Finance:    {len(finance_rows)} rows (COST:{len(cost_rows)} AR:{len(ar_rows)} BUDGET:{len(budget_rows)})")
    print(f"  Department: {len(department_rows)} rows")
    print(f"  Total:      {len(sales_rows) + len(finance_rows) + len(department_rows)} rows")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
