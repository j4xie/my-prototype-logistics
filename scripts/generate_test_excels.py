#!/usr/bin/env python3
"""
generate_test_excels.py
生成3个不同行业的测试Excel文件, 数据真实、有涨跌趋势。
依赖: pip install openpyxl
输出: scripts/test-data/ 目录
"""

import os
import random
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    raise

OUTPUT_DIR = Path(__file__).parent / "test-data"
OUTPUT_DIR.mkdir(exist_ok=True)

random.seed(42)  # 固定随机种子，保证每次生成结果一致


# ──────────────────────────────────────────────
# 样式助手
# ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F6EBA", end_color="2F6EBA", fill_type="solid")
ALT_FILL    = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=13)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, ncols: int, alt: bool = False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        if alt:
            cell.fill = ALT_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_w=10, max_w=25):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val) * 1.8)  # 中文字符宽度补偿
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len, min_w), max_w)


def _write_sheet(ws, headers: list, rows: list, title: str = ""):
    """通用写入：标题行 + 表头 + 数据行"""
    start = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        start = 2

    # 表头
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    # 数据
    for i, row_data in enumerate(rows):
        r = start + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

    _auto_width(ws)


# ──────────────────────────────────────────────
# 数据生成工具
# ──────────────────────────────────────────────
MONTHS = [f"2025年{m}月" for m in range(1, 13)]


def _trend(base: float, count: int, drift: float = 0.03, noise: float = 0.05) -> list:
    """生成有趋势+噪声的序列 (正增长为主)"""
    vals = []
    v = base
    for _ in range(count):
        v = v * (1 + drift + random.uniform(-noise, noise))
        vals.append(round(v, 2))
    return vals


def _round2(v):
    return round(float(v), 2)


# ──────────────────────────────────────────────
# Excel 1: 张记餐饮-2025经营报表.xlsx
# ──────────────────────────────────────────────
def generate_catering(path: Path):
    wb = Workbook()

    # ── Sheet1: 月度收入明细 ──
    ws1 = wb.active
    ws1.title = "月度收入明细"

    takeout   = _trend(38000, 12, drift=0.025, noise=0.06)   # 外卖收入(波动较大)
    dining    = _trend(52000, 12, drift=0.018, noise=0.04)   # 堂食收入
    catering  = [_round2(t + d) for t, d in zip(takeout, dining)]

    # 营业收入 = 堂食 + 外卖 + 小额其他
    other     = [round(random.uniform(2000, 6000), 2) for _ in range(12)]
    revenue   = [_round2(t + d + o) for t, d, o in zip(takeout, dining, other)]

    rows1 = []
    for i in range(12):
        rows1.append([
            MONTHS[i],
            revenue[i],
            takeout[i],
            dining[i],
            catering[i],
        ])
    _write_sheet(ws1,
        headers=["月份", "营业收入(元)", "外卖收入(元)", "堂食收入(元)", "总收入(元)"],
        rows=rows1,
        title="张记餐饮 2025年月度收入明细"
    )

    # ── Sheet2: 成本分析 ──
    ws2 = wb.create_sheet("成本分析")
    raw_mat   = _trend(28000, 12, drift=0.020, noise=0.05)
    labor     = _trend(22000, 12, drift=0.008, noise=0.02)
    rent      = [18500.0] * 12   # 固定租金
    utilities = _trend(4200,  12, drift=0.010, noise=0.08)
    total_cost = [_round2(a + b + c + d) for a, b, c, d in zip(raw_mat, labor, rent, utilities)]

    rows2 = []
    for i in range(12):
        rows2.append([
            MONTHS[i],
            raw_mat[i],
            labor[i],
            rent[i],
            utilities[i],
            total_cost[i],
        ])
    _write_sheet(ws2,
        headers=["月份", "原材料(元)", "人工(元)", "房租(元)", "水电(元)", "总成本(元)"],
        rows=rows2,
        title="张记餐饮 2025年成本分析"
    )

    # ── Sheet3: 菜品销售排行 ──
    ws3 = wb.create_sheet("菜品销售排行")
    dishes = [
        ("宫保鸡丁", 38),
        ("麻婆豆腐", 28),
        ("红烧肉",   48),
        ("蛋炒饭",   18),
        ("酸辣汤",   22),
        ("鱼香肉丝", 32),
        ("糖醋排骨", 48),
        ("清蒸鲈鱼", 68),
        ("凉拌黄瓜", 12),
        ("口水鸡",   28),
        ("扬州炒饭", 22),
        ("阳春面",   16),
        ("小笼包",   28),
        ("皮蛋豆腐", 18),
        ("西湖牛肉羹", 26),
        ("芒果布丁", 18),
        ("红豆沙",   14),
        ("拍黄瓜",   12),
        ("白斩鸡",   42),
        ("蟹黄豆腐", 58),
    ]
    rows3 = []
    for name, price in dishes:
        qty = int(random.uniform(200, 2800))
        sales = _round2(qty * price * (1 + random.uniform(-0.05, 0.05)))
        cost_rate = random.uniform(0.38, 0.55)
        gross_margin = _round2((1 - cost_rate) * 100)
        rows3.append([name, qty, price, sales, f"{gross_margin}%"])

    # 按销售额排序
    rows3.sort(key=lambda x: x[3], reverse=True)

    _write_sheet(ws3,
        headers=["菜品名", "销量(份)", "单价(元)", "销售额(元)", "毛利率"],
        rows=rows3,
        title="张记餐饮 2025年菜品销售排行(全年汇总)"
    )

    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────
# Excel 2: 绿源食品-2025生产报表.xlsx
# ──────────────────────────────────────────────
def generate_food_processing(path: Path):
    wb = Workbook()

    # ── Sheet1: 生产产量 ──
    ws1 = wb.active
    ws1.title = "生产产量"

    prod_a = _trend(18500, 12, drift=0.022, noise=0.04)   # 产品A (主力产品)
    prod_b = _trend(11200, 12, drift=0.030, noise=0.06)   # 产品B (新品增长快)
    qual_rate = [_round2(min(99.8, 97.0 + random.uniform(0, 2.5))) for _ in range(12)]
    waste_rate = [_round2(100 - q) for q in qual_rate]

    rows1 = []
    for i in range(12):
        rows1.append([
            MONTHS[i],
            int(prod_a[i]),
            int(prod_b[i]),
            f"{qual_rate[i]}%",
            f"{waste_rate[i]}%",
        ])
    _write_sheet(ws1,
        headers=["月份", "产品A产量(箱)", "产品B产量(箱)", "合格率", "废品率"],
        rows=rows1,
        title="绿源食品 2025年生产产量统计"
    )

    # ── Sheet2: 原材料采购 ──
    ws2 = wb.create_sheet("原材料采购")
    suppliers = ["苏州优质农业", "无锡新鲜蔬果", "南京粮油批发", "常州调味品行"]
    rows2 = []
    for i, month in enumerate(MONTHS):
        sup = suppliers[i % len(suppliers)]
        qty = round(random.uniform(8000, 22000), 1)
        unit_price = round(random.uniform(3.2, 8.8), 2)
        amount = _round2(qty * unit_price)
        rows2.append([month, sup, qty, amount, unit_price])

    _write_sheet(ws2,
        headers=["月份", "供应商", "采购量(kg)", "采购金额(元)", "单价(元/kg)"],
        rows=rows2,
        title="绿源食品 2025年原材料采购记录"
    )

    # ── Sheet3: 利润表 ──
    ws3 = wb.create_sheet("利润表")
    rev = _trend(380000, 12, drift=0.025, noise=0.04)
    cost_rate = [random.uniform(0.62, 0.70) for _ in range(12)]
    oper_cost = [_round2(r * c) for r, c in zip(rev, cost_rate)]
    gross     = [_round2(r - c) for r, c in zip(rev, oper_cost)]

    # 期间费用: 管理+销售+财务
    period_expense = [_round2(r * random.uniform(0.14, 0.20)) for r in rev]
    net_profit     = [_round2(g - pe) for g, pe in zip(gross, period_expense)]

    rows3 = []
    for i in range(12):
        rows3.append([
            MONTHS[i],
            rev[i],
            oper_cost[i],
            gross[i],
            period_expense[i],
            net_profit[i],
        ])
    _write_sheet(ws3,
        headers=["月份", "营业收入(元)", "营业成本(元)", "毛利(元)", "期间费用(元)", "净利润(元)"],
        rows=rows3,
        title="绿源食品 2025年利润表"
    )

    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────
# Excel 3: 鲜味零售-2025销售数据.xlsx
# ──────────────────────────────────────────────
def generate_retail(path: Path):
    wb = Workbook()

    # ── Sheet1: 门店销售 ──
    ws1 = wb.active
    ws1.title = "门店销售"

    store_a = _trend(125000, 12, drift=0.020, noise=0.05)   # 旗舰店
    store_b = _trend(82000,  12, drift=0.028, noise=0.06)   # 次旗舰店
    store_c = _trend(54000,  12, drift=0.035, noise=0.07)   # 新店 (增速快)
    total   = [_round2(a + b + c) for a, b, c in zip(store_a, store_b, store_c)]

    rows1 = []
    for i in range(12):
        rows1.append([
            MONTHS[i],
            store_a[i],
            store_b[i],
            store_c[i],
            total[i],
        ])
    _write_sheet(ws1,
        headers=["月份", "旗舰店(元)", "次旗舰店(元)", "新店(元)", "合计(元)"],
        rows=rows1,
        title="鲜味零售 2025年门店销售汇总"
    )

    # ── Sheet2: 会员分析 ──
    ws2 = wb.create_sheet("会员分析")
    member_levels = [
        ("普通会员",  18500, 380, 158),
        ("银卡会员",   5200, 820, 285),
        ("金卡会员",   1850, 1680, 420),
        ("铂金会员",    420, 3250, 580),
        ("钻石会员",     85, 8800, 820),
        ("超级VIP",      12, 28000, 1200),
        ("企业卡",      220, 12500, 680),
        ("学生卡",     3600, 260, 95),
        ("老年优惠",   2100, 320, 110),
        ("家庭卡",      980, 1850, 460),
    ]
    rows2 = []
    for lvl, count, avg_spend, avg_order in member_levels:
        total_spend = _round2(count * avg_spend * (1 + random.uniform(-0.08, 0.08)))
        repurchase = _round2(random.uniform(0.28, 0.85) * 100)
        rows2.append([lvl, count, total_spend, avg_order, f"{repurchase}%"])

    _write_sheet(ws2,
        headers=["等级", "人数", "消费总额(元)", "平均消费(元)", "复购率"],
        rows=rows2,
        title="鲜味零售 2025年会员分析"
    )

    # ── Sheet3: 库存周转 ──
    ws3 = wb.create_sheet("库存周转")
    begin_inv = 320000.0
    rows3 = []
    for i in range(12):
        purchase_in   = round(random.uniform(180000, 420000), 2)
        sales_out     = _round2(total[i] * random.uniform(0.75, 0.88))   # 出库≈营收×成本率
        end_inv       = _round2(begin_inv + purchase_in - sales_out)
        end_inv       = max(end_inv, 80000)  # 不低于安全库存
        # 周转天数 = (期初+期末)/2 / (销售出库/30)
        avg_inv       = (begin_inv + end_inv) / 2
        daily_sales   = sales_out / 30
        turnover_days = round(avg_inv / daily_sales, 1) if daily_sales > 0 else 0

        rows3.append([
            MONTHS[i],
            begin_inv,
            purchase_in,
            sales_out,
            end_inv,
            turnover_days,
        ])
        begin_inv = end_inv   # 滚动计算

    _write_sheet(ws3,
        headers=["月份", "期初库存(元)", "采购入库(元)", "销售出库(元)", "期末库存(元)", "周转天数"],
        rows=rows3,
        title="鲜味零售 2025年库存周转分析"
    )

    wb.save(path)
    print(f"[OK] {path}")


# ──────────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────────
if __name__ == "__main__":
    p1 = OUTPUT_DIR / "张记餐饮-2025经营报表.xlsx"
    p2 = OUTPUT_DIR / "绿源食品-2025生产报表.xlsx"
    p3 = OUTPUT_DIR / "鲜味零售-2025销售数据.xlsx"

    print("=== 生成测试Excel文件 ===")
    generate_catering(p1)
    generate_food_processing(p2)
    generate_retail(p3)

    print("\n=== 生成完成 ===")
    for p in [p1, p2, p3]:
        size = p.stat().st_size / 1024
        print(f"  {p.name}  ({size:.1f} KB)")

    print(f"\n输出目录: {OUTPUT_DIR.resolve()}")
