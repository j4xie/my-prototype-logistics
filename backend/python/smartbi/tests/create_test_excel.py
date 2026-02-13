"""
创建复杂测试Excel文件 - 5个Sheet
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import random

def create_test_excel(output_path: str):
    """创建包含5个不同类型Sheet的测试Excel"""
    wb = openpyxl.Workbook()

    # 删除默认sheet
    wb.remove(wb.active)

    # ============================================================
    # Sheet 1: 利润表 (复杂多层表头 + 合并单元格)
    # ============================================================
    ws1 = wb.create_sheet("2025年利润表")

    # 标题行
    ws1.merge_cells('A1:H1')
    ws1['A1'] = "宁波某某食品有限公司利润表"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center')

    # 元信息行
    ws1['A2'] = "单位：万元"
    ws1['E2'] = "编制日期：2025年1月"

    # 多层表头
    ws1.merge_cells('A3:A4')
    ws1['A3'] = "项目"
    ws1.merge_cells('B3:B4')
    ws1['B3'] = "行次"

    # 1月
    ws1.merge_cells('C3:D3')
    ws1['C3'] = "2025年1月"
    ws1['C4'] = "预算数"
    ws1['D4'] = "实际数"

    # 2月
    ws1.merge_cells('E3:F3')
    ws1['E3'] = "2025年2月"
    ws1['E4'] = "预算数"
    ws1['F4'] = "实际数"

    # 累计
    ws1.merge_cells('G3:H3')
    ws1['G3'] = "累计"
    ws1['G4'] = "预算数"
    ws1['H4'] = "实际数"

    # 数据行
    profit_items = [
        ("一、营业收入", 1, 500, 520, 600, 580, 1100, 1100),
        ("  主营业务收入", 2, 450, 470, 540, 530, 990, 1000),
        ("  其他业务收入", 3, 50, 50, 60, 50, 110, 100),
        ("二、营业成本", 4, 300, 310, 360, 350, 660, 660),
        ("  主营业务成本", 5, 270, 280, 320, 310, 590, 590),
        ("  其他业务成本", 6, 30, 30, 40, 40, 70, 70),
        ("三、营业毛利", 7, 200, 210, 240, 230, 440, 440),
        ("四、期间费用", 8, 80, 85, 90, 88, 170, 173),
        ("  销售费用", 9, 30, 32, 35, 33, 65, 65),
        ("  管理费用", 10, 40, 43, 45, 45, 85, 88),
        ("  财务费用", 11, 10, 10, 10, 10, 20, 20),
        ("五、营业利润", 12, 120, 125, 150, 142, 270, 267),
        ("六、利润总额", 13, 120, 125, 150, 142, 270, 267),
        ("七、净利润", 14, 102, 106, 127.5, 120.7, 229.5, 226.7),
    ]

    for i, item in enumerate(profit_items):
        row = 5 + i
        ws1[f'A{row}'] = item[0]
        ws1[f'B{row}'] = item[1]
        ws1[f'C{row}'] = item[2]
        ws1[f'D{row}'] = item[3]
        ws1[f'E{row}'] = item[4]
        ws1[f'F{row}'] = item[5]
        ws1[f'G{row}'] = item[6]
        ws1[f'H{row}'] = item[7]

    # 备注行
    last_row = 5 + len(profit_items)
    ws1[f'A{last_row}'] = "编制说明："
    ws1.merge_cells(f'B{last_row}:H{last_row}')
    ws1[f'B{last_row}'] = "本表按权责发生制编制，营业收入按实际发货确认，成本按配比原则计算。"

    # ============================================================
    # Sheet 2: 销售明细表 (大量数据行)
    # ============================================================
    ws2 = wb.create_sheet("销售明细")

    ws2['A1'] = "销售明细表"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A2'] = "统计期间：2025年1月1日-2025年1月31日"

    # 表头
    headers = ["序号", "日期", "客户名称", "产品名称", "规格", "数量", "单价", "金额", "业务员", "备注"]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=3, column=col, value=header)
        ws2.cell(row=3, column=col).font = Font(bold=True)

    # 模拟数据
    customers = ["上海沃尔玛", "杭州联华", "宁波大润发", "温州永辉", "嘉兴世纪联华"]
    products = ["火锅底料", "麻辣烫底料", "番茄底料", "菌菇底料", "酸菜鱼底料"]
    specs = ["500g/袋", "1kg/袋", "250g/盒", "500g/盒"]
    salesmen = ["张三", "李四", "王五", "赵六"]

    for i in range(1, 51):  # 50行数据
        row = 3 + i
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=f"2025-01-{(i % 28) + 1:02d}")
        ws2.cell(row=row, column=3, value=random.choice(customers))
        ws2.cell(row=row, column=4, value=random.choice(products))
        ws2.cell(row=row, column=5, value=random.choice(specs))
        qty = random.randint(100, 1000)
        price = round(random.uniform(15, 45), 2)
        ws2.cell(row=row, column=6, value=qty)
        ws2.cell(row=row, column=7, value=price)
        ws2.cell(row=row, column=8, value=round(qty * price, 2))
        ws2.cell(row=row, column=9, value=random.choice(salesmen))
        if i % 10 == 0:
            ws2.cell(row=row, column=10, value="促销订单")

    # 合计行
    total_row = 54
    ws2.cell(row=total_row, column=1, value="合计")
    ws2.cell(row=total_row, column=6, value="=SUM(F4:F53)")
    ws2.cell(row=total_row, column=8, value="=SUM(H4:H53)")

    # ============================================================
    # Sheet 3: 部门预算对比 (横向表头)
    # ============================================================
    ws3 = wb.create_sheet("部门预算对比")

    ws3.merge_cells('A1:G1')
    ws3['A1'] = "2025年各部门预算完成情况"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = Alignment(horizontal='center')

    # 表头
    ws3['A3'] = "部门"
    ws3['B3'] = "年度预算"
    ws3['C3'] = "1月完成"
    ws3['D3'] = "2月完成"
    ws3['E3'] = "累计完成"
    ws3['F3'] = "完成率"
    ws3['G3'] = "排名"

    departments = [
        ("销售一部", 1200, 110, 105),
        ("销售二部", 1000, 85, 90),
        ("销售三部", 800, 70, 72),
        ("电商部", 600, 55, 58),
        ("大客户部", 1500, 130, 125),
    ]

    for i, (dept, budget, m1, m2) in enumerate(departments):
        row = 4 + i
        total = m1 + m2
        rate = total / budget * 100
        ws3.cell(row=row, column=1, value=dept)
        ws3.cell(row=row, column=2, value=budget)
        ws3.cell(row=row, column=3, value=m1)
        ws3.cell(row=row, column=4, value=m2)
        ws3.cell(row=row, column=5, value=total)
        ws3.cell(row=row, column=6, value=f"{rate:.1f}%")
        ws3.cell(row=row, column=7, value=i + 1)

    # ============================================================
    # Sheet 4: 产品成本分析 (包含公式和特殊格式)
    # ============================================================
    ws4 = wb.create_sheet("产品成本分析")

    ws4['A1'] = "产品成本构成分析表"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4['A2'] = "分析期间：2025年Q1"
    ws4['E2'] = "单位：元/kg"

    # 复杂表头
    ws4.merge_cells('A3:A4')
    ws4['A3'] = "产品"
    ws4.merge_cells('B3:D3')
    ws4['B3'] = "直接成本"
    ws4['B4'] = "原料"
    ws4['C4'] = "人工"
    ws4['D4'] = "能耗"
    ws4.merge_cells('E3:F3')
    ws4['E3'] = "间接成本"
    ws4['E4'] = "制造费用"
    ws4['F4'] = "折旧"
    ws4['G3'] = "总成本"
    ws4['G4'] = ""
    ws4['H3'] = "毛利率"
    ws4['H4'] = ""

    products_cost = [
        ("火锅底料A", 8.5, 2.0, 0.8, 1.2, 0.5, 25, 0.48),
        ("火锅底料B", 10.2, 2.2, 0.9, 1.3, 0.6, 30, 0.49),
        ("麻辣底料", 9.0, 1.8, 0.7, 1.1, 0.4, 22, 0.41),
        ("番茄底料", 7.5, 1.5, 0.6, 1.0, 0.4, 18, 0.39),
        ("菌菇底料", 12.0, 2.5, 1.0, 1.5, 0.7, 35, 0.49),
    ]

    for i, item in enumerate(products_cost):
        row = 5 + i
        ws4.cell(row=row, column=1, value=item[0])
        ws4.cell(row=row, column=2, value=item[1])
        ws4.cell(row=row, column=3, value=item[2])
        ws4.cell(row=row, column=4, value=item[3])
        ws4.cell(row=row, column=5, value=item[4])
        ws4.cell(row=row, column=6, value=item[5])
        # 总成本 = 直接成本 + 间接成本
        total_cost = item[1] + item[2] + item[3] + item[4] + item[5]
        ws4.cell(row=row, column=7, value=round(total_cost, 2))
        ws4.cell(row=row, column=8, value=f"{item[7]*100:.0f}%")

    # 说明
    ws4['A11'] = "备注："
    ws4.merge_cells('B11:H11')
    ws4['B11'] = "1. 原料成本按实际采购价加权平均计算；2. 人工成本包含社保公积金；3. 毛利率=(售价-总成本)/售价"

    # ============================================================
    # Sheet 5: 空Sheet + 只有标题 (测试边界情况)
    # ============================================================
    ws5 = wb.create_sheet("待补充数据")

    ws5['A1'] = "2025年客户满意度调查结果"
    ws5['A1'].font = Font(bold=True, size=14)
    ws5['A2'] = "（数据收集中，预计3月完成）"
    ws5['A4'] = "调查维度"
    ws5['B4'] = "非常满意"
    ws5['C4'] = "满意"
    ws5['D4'] = "一般"
    ws5['E4'] = "不满意"
    ws5['F4'] = "平均分"

    # 只有表头，没有数据

    # 保存文件
    wb.save(output_path)
    print(f"测试Excel已创建: {output_path}")
    return output_path


if __name__ == "__main__":
    create_test_excel("test_complex_5sheets.xlsx")
