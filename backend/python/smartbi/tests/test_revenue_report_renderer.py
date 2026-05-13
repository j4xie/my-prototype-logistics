"""Tests for services.excel_renderers.qhj_revenue_v1 + registry + labels.

Spec: docs/qa-specs/2026-05-12-qhj-revenue-report-design.md §7
Plan: docs/superpowers/plans/2026-05-12-qhj-revenue-report.md Task F1

Tests build a workbook from a sample TemplateResult.data dict, then re-open
it via openpyxl.load_workbook() and assert structure: sheet name, key cells,
multi-store stacking, NULL-as-"—" rendering, percentage cells as float (not str).
"""
import io

import pytest
from openpyxl import load_workbook

from smartbi.services.excel_renderers import RENDERERS
from smartbi.services.excel_renderers._labels import LABELS


@pytest.fixture
def sample_data():
    """Mirrors compute_qhj_revenue_report().data shape (per spec §6.2)."""
    return {
        "block1_yoy": [
            {"store_id": 1, "store_name": "青花椒南方百联店",
             "total": 10000, "dine_in": 7000, "takeout": 3000,
             "prev_total": None, "prev_dine_in": None, "prev_takeout": None,
             "total_ratio": None, "dine_in_ratio": None, "takeout_ratio": None},
        ],
        "block2_mom": [
            {"store_id": 1, "store_name": "青花椒南方百联店",
             "total": 10000, "dine_in": 7000, "takeout": 3000,
             "prev_total": 8000, "prev_dine_in": 5000, "prev_takeout": 3000,
             "total_ratio": 25.0, "dine_in_ratio": 40.0, "takeout_ratio": 0.0},
        ],
        "block3_meal_split": [
            {"store_id": 1, "store_name": "青花椒南方百联店",
             "dine_in_revenue": 7000, "takeout_revenue": 3000,
             "dine_in_bills": 50, "takeout_bills": 20,
             "revenue_ratio": 0.7, "bill_ratio": 0.714},
        ],
        "block4_diner_dist": [{
            "store_id": 1, "store_name": "青花椒南方百联店",
            "date_range": "2025-10-01 ~ 2025-10-07",
            "distribution": [
                {"diner_count": 1, "bill_count": 5, "bill_ratio": 0.1,
                 "total_items": 12, "avg_items_per_bill": 2.4,
                 "revenue": 500, "revenue_per_diner": 100,
                 "revenue_per_item": 42, "revenue_ratio": 0.05},
                {"diner_count": 2, "bill_count": 30, "bill_ratio": 0.6,
                 "total_items": 90, "avg_items_per_bill": 3.0,
                 "revenue": 4500, "revenue_per_diner": 75,
                 "revenue_per_item": 50, "revenue_ratio": 0.45},
            ],
        }],
        "meta": {"date_from": "2025-10-01", "date_to": "2025-10-07",
                 "yoy_available": False, "yoy_note": "需要 2024 同期数据"},
    }


def test_registry_has_qhj_renderer():
    assert "qhj_revenue_v1" in RENDERERS
    assert callable(RENDERERS["qhj_revenue_v1"])


def test_render_returns_bytesio_with_xlsx(sample_data):
    render = RENDERERS["qhj_revenue_v1"]
    buf = render(sample_data, labels=LABELS["zh-CN"])
    assert isinstance(buf, io.BytesIO)
    # xlsx is a zip; PK\x03\x04 is the local-file-header magic.
    assert buf.getvalue()[:2] == b"PK"


def test_workbook_has_revenue_report_sheet(sample_data):
    render = RENDERERS["qhj_revenue_v1"]
    buf = render(sample_data, labels=LABELS["zh-CN"])
    wb = load_workbook(buf)
    assert "收入管理报表" in wb.sheetnames


def test_block1_writes_store_row(sample_data):
    render = RENDERERS["qhj_revenue_v1"]
    buf = render(sample_data, labels=LABELS["zh-CN"])
    wb = load_workbook(buf)
    ws = wb.active
    # Scan all cells for the store name.
    found = False
    for row in ws.iter_rows(values_only=True):
        if any(c == "青花椒南方百联店" for c in row if c):
            found = True
            break
    assert found


def test_block1_yoy_disabled_renders_dash_for_null_columns(sample_data):
    """Spec §11.1 — NULL ratio/prev cols rendered as '—' (no_data label)."""
    render = RENDERERS["qhj_revenue_v1"]
    buf = render(sample_data, labels=LABELS["zh-CN"])
    wb = load_workbook(buf)
    ws = wb.active
    # The block1 row exists somewhere; at least one cell must be the no_data label.
    no_data = LABELS["zh-CN"]["no_data"]
    saw_dash = False
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c == no_data:
                saw_dash = True
                break
        if saw_dash:
            break
    assert saw_dash, "Block 1 NULL columns must render no_data label"


def test_block4_multi_store_stacks_blocks(sample_data):
    """Multi-store Block 4 stacks one sub-block per store."""
    sample_data["block4_diner_dist"].append({
        "store_id": 2, "store_name": "青花椒徐汇店",
        "date_range": "2025-10-01 ~ 2025-10-07",
        "distribution": [
            {"diner_count": 1, "bill_count": 2, "bill_ratio": 0.5,
             "total_items": 6, "avg_items_per_bill": 3.0,
             "revenue": 200, "revenue_per_diner": 100,
             "revenue_per_item": 33, "revenue_ratio": 0.5},
        ],
    })
    render = RENDERERS["qhj_revenue_v1"]
    buf = render(sample_data, labels=LABELS["zh-CN"])
    wb = load_workbook(buf)
    ws = wb.active
    # Both store names appear (with date_range suffix).
    text = "\n".join(
        str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None
    )
    assert "青花椒南方百联店" in text
    assert "青花椒徐汇店" in text


def test_block3_ratio_stored_as_numeric_for_excel_format(sample_data):
    """Ratio cells must be number (Excel renders via 0.00% format), not string."""
    render = RENDERERS["qhj_revenue_v1"]
    buf = render(sample_data, labels=LABELS["zh-CN"])
    wb = load_workbook(buf)
    ws = wb.active
    # Find any cell whose value is 0.7 (block3 revenue_ratio).
    found_numeric = False
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and abs(cell.value - 0.7) < 1e-6:
                found_numeric = True
                # Should have percentage format applied.
                assert "%" in (cell.number_format or "")
                break
        if found_numeric:
            break
    assert found_numeric


def test_labels_dict_has_required_keys():
    labels_zh = LABELS["zh-CN"]
    required = {
        "title", "block1_title", "block2_title", "block3_title", "block4_title",
        "store_name", "no_data", "no_yoy_data",
        "diner_count", "bill_count", "revenue", "total_row",
    }
    missing = required - set(labels_zh.keys())
    assert not missing, f"missing labels: {missing}"
