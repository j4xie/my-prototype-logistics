# -*- coding: utf-8 -*-
"""
Create Complex Test Excel for E2E Testing

生成包含多种复杂场景的测试 Excel 文件：
1. 标准利润表 (合并单元格表头)
2. 销售明细 (大量数据行)
3. 预算对比 (多层表头)
4. 时间序列 (日期+数值)
5. 混合类型列 (数字+文本混合)
6. 稀疏数据 (大量空值)
7. 超宽表 (100+列)
8. 嵌套分组 (多级分类)
9. 中英文混合
10. 特殊字符和公式
"""
import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_complex_test_excel(output_path: str):
    """Create comprehensive test Excel with various scenarios"""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. 标准利润表 (合并单元格表头)
    create_profit_statement(wb)

    # 2. 销售明细 (大量数据行)
    create_sales_detail(wb)

    # 3. 预算对比 (多层表头)
    create_budget_comparison(wb)

    # 4. 时间序列数据
    create_time_series(wb)

    # 5. 混合类型列
    create_mixed_types(wb)

    # 6. 稀疏数据
    create_sparse_data(wb)

    # 7. 超宽表 (100+列)
    create_wide_table(wb)

    # 8. 嵌套分组
    create_nested_groups(wb)

    # 9. 中英文混合
    create_bilingual_data(wb)

    # 10. 特殊字符和公式
    create_special_chars(wb)

    # Save
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_profit_statement(wb):
    """1. 标准利润表 - 合并单元格表头"""
    ws = wb.create_sheet("利润表_合并表头")

    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = "2025年第一季度利润表"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 单位
    ws['A2'] = "单位：万元"

    # 合并表头
    ws.merge_cells('A3:A4')
    ws['A3'] = "项目"

    ws.merge_cells('B3:C3')
    ws['B3'] = "本期"
    ws['B4'] = "预算"
    ws['C4'] = "实际"

    ws.merge_cells('D3:E3')
    ws['D3'] = "上期"
    ws['D4'] = "预算"
    ws['E4'] = "实际"

    ws['F3'] = "同比"
    ws['F4'] = "增长率"

    # 数据
    items = [
        ("营业收入", 1000, 1050, 900, 920, 0.14),
        ("  主营业务收入", 800, 840, 720, 735, 0.14),
        ("  其他业务收入", 200, 210, 180, 185, 0.14),
        ("营业成本", 600, 580, 540, 525, 0.10),
        ("毛利润", 400, 470, 360, 395, 0.19),
        ("销售费用", 50, 48, 45, 44, 0.09),
        ("管理费用", 80, 75, 72, 70, 0.07),
        ("财务费用", 20, 18, 18, 17, 0.06),
        ("营业利润", 250, 329, 225, 264, 0.25),
        ("净利润", 200, 263, 180, 211, 0.25),
    ]

    for i, (item, b1, a1, b2, a2, rate) in enumerate(items, start=5):
        ws[f'A{i}'] = item
        ws[f'B{i}'] = b1
        ws[f'C{i}'] = a1
        ws[f'D{i}'] = b2
        ws[f'E{i}'] = a2
        ws[f'F{i}'] = rate


def create_sales_detail(wb):
    """2. 销售明细 - 大量数据行"""
    ws = wb.create_sheet("销售明细_500行")

    # 表头
    headers = ["订单号", "日期", "客户名称", "产品", "数量", "单价", "金额", "区域", "销售员"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 生成500行数据
    products = ["产品A", "产品B", "产品C", "产品D", "产品E"]
    regions = ["华东", "华北", "华南", "西南", "西北"]
    sales_reps = ["张三", "李四", "王五", "赵六", "钱七"]

    base_date = datetime(2025, 1, 1)

    for i in range(500):
        row = i + 2
        qty = random.randint(10, 500)
        price = random.uniform(50, 500)

        ws.cell(row=row, column=1, value=f"ORD-2025-{i+1:05d}")
        ws.cell(row=row, column=2, value=base_date + timedelta(days=random.randint(0, 90)))
        ws.cell(row=row, column=3, value=f"客户{random.randint(1, 100):03d}")
        ws.cell(row=row, column=4, value=random.choice(products))
        ws.cell(row=row, column=5, value=qty)
        ws.cell(row=row, column=6, value=round(price, 2))
        ws.cell(row=row, column=7, value=round(qty * price, 2))
        ws.cell(row=row, column=8, value=random.choice(regions))
        ws.cell(row=row, column=9, value=random.choice(sales_reps))


def create_budget_comparison(wb):
    """3. 预算对比 - 多层表头"""
    ws = wb.create_sheet("预算对比_多层表头")

    # 三层表头
    # Row 1: 大类
    ws.merge_cells('A1:A3')
    ws['A1'] = "部门"

    ws.merge_cells('B1:G1')
    ws['B1'] = "2025年Q1"

    ws.merge_cells('H1:M1')
    ws['H1'] = "2025年Q2"

    # Row 2: 中类
    for start_col, label in [('B', '收入'), ('E', '成本'), ('H', '收入'), ('K', '成本')]:
        col_idx = ord(start_col) - ord('A') + 1
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+2)
        ws.cell(row=2, column=col_idx, value=label)

    # Row 3: 小类
    sub_headers = ["预算", "实际", "差异"] * 4
    for i, h in enumerate(sub_headers, 2):
        ws.cell(row=3, column=i, value=h)

    # 数据
    departments = ["销售部", "市场部", "研发部", "运营部", "人力资源部"]
    for i, dept in enumerate(departments, 4):
        ws.cell(row=i, column=1, value=dept)
        for j in range(12):
            base = random.randint(100, 1000)
            if j % 3 == 2:  # 差异列
                ws.cell(row=i, column=j+2, value=random.randint(-50, 50))
            else:
                ws.cell(row=i, column=j+2, value=base)


def create_time_series(wb):
    """4. 时间序列数据"""
    ws = wb.create_sheet("时间序列_日数据")

    # 表头
    headers = ["日期", "销售额", "订单数", "客单价", "转化率", "UV", "PV"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 365天数据
    base_date = datetime(2025, 1, 1)
    base_sales = 10000

    for i in range(365):
        row = i + 2
        date = base_date + timedelta(days=i)

        # 模拟周末效应和季节性
        weekday_factor = 1.2 if date.weekday() < 5 else 0.8
        month_factor = 1 + 0.1 * (date.month - 1)  # 逐月增长
        noise = random.uniform(0.9, 1.1)

        sales = base_sales * weekday_factor * month_factor * noise
        orders = int(sales / random.uniform(80, 120))
        uv = random.randint(5000, 15000)
        pv = uv * random.uniform(2, 5)

        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=round(sales, 2))
        ws.cell(row=row, column=3, value=orders)
        ws.cell(row=row, column=4, value=round(sales / max(orders, 1), 2))
        ws.cell(row=row, column=5, value=round(orders / uv * 100, 2))
        ws.cell(row=row, column=6, value=uv)
        ws.cell(row=row, column=7, value=int(pv))


def create_mixed_types(wb):
    """5. 混合类型列 - 数字和文本混合"""
    ws = wb.create_sheet("混合类型_复杂")

    headers = ["ID", "金额", "状态码", "备注", "混合列"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for i in range(100):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)

        # 金额列有些是数字，有些是文本
        if random.random() > 0.2:
            ws.cell(row=row, column=2, value=random.randint(100, 10000))
        else:
            ws.cell(row=row, column=2, value="N/A")

        # 状态码有些是数字，有些是文本
        status = random.choice([1, 2, 3, "成功", "失败", "待处理", None])
        ws.cell(row=row, column=3, value=status)

        # 备注
        ws.cell(row=row, column=4, value=f"备注{i}" if random.random() > 0.5 else None)

        # 混合列：完全随机类型
        mixed_values = [
            123, 45.67, "文本", True, None,
            "100", "45.6%", datetime.now(),
            "2025-01-01", "-", "—"
        ]
        ws.cell(row=row, column=5, value=random.choice(mixed_values))


def create_sparse_data(wb):
    """6. 稀疏数据 - 大量空值"""
    ws = wb.create_sheet("稀疏数据_70%空")

    headers = [f"列{i}" for i in range(1, 21)]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row in range(2, 102):
        for col in range(1, 21):
            # 70% 概率为空
            if random.random() > 0.7:
                ws.cell(row=row, column=col, value=random.randint(1, 1000))


def create_wide_table(wb):
    """7. 超宽表 - 100+列"""
    ws = wb.create_sheet("超宽表_120列")

    # 120 列表头
    for col in range(1, 121):
        header = f"指标_{col:03d}"
        ws.cell(row=1, column=col, value=header)

    # 50 行数据
    for row in range(2, 52):
        for col in range(1, 121):
            ws.cell(row=row, column=col, value=random.uniform(0, 100))


def create_nested_groups(wb):
    """8. 嵌套分组 - 多级分类"""
    ws = wb.create_sheet("嵌套分组_3级")

    headers = ["一级分类", "二级分类", "三级分类", "金额", "数量", "占比"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 三级嵌套数据
    categories = {
        "电子产品": {
            "手机": ["iPhone", "华为", "小米", "OPPO"],
            "电脑": ["联想", "戴尔", "华硕", "惠普"],
            "配件": ["充电器", "耳机", "保护壳", "数据线"]
        },
        "服装": {
            "男装": ["T恤", "衬衫", "外套", "裤子"],
            "女装": ["连衣裙", "上衣", "裙子", "裤装"],
            "童装": ["婴儿装", "幼儿装", "儿童装"]
        },
        "食品": {
            "零食": ["饼干", "糖果", "坚果", "薯片"],
            "饮料": ["矿泉水", "果汁", "茶饮", "咖啡"],
            "生鲜": ["水果", "蔬菜", "肉类", "海鲜"]
        }
    }

    row = 2
    for cat1, sub1 in categories.items():
        for cat2, items in sub1.items():
            for cat3 in items:
                ws.cell(row=row, column=1, value=cat1)
                ws.cell(row=row, column=2, value=cat2)
                ws.cell(row=row, column=3, value=cat3)
                ws.cell(row=row, column=4, value=random.randint(1000, 100000))
                ws.cell(row=row, column=5, value=random.randint(10, 1000))
                ws.cell(row=row, column=6, value=round(random.uniform(0.01, 0.1), 4))
                row += 1


def create_bilingual_data(wb):
    """9. 中英文混合"""
    ws = wb.create_sheet("中英文_Bilingual")

    headers = ["Product Name 产品名称", "Category 类别", "Price 价格", "Stock 库存", "Status 状态"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    products = [
        ("iPhone 15 Pro 苹果手机", "Electronics 电子", 7999, 150, "In Stock 有货"),
        ("Nike Air Max 耐克运动鞋", "Footwear 鞋类", 1299, 80, "Low Stock 库存紧张"),
        ("Lenovo ThinkPad 联想笔记本", "Computers 电脑", 6499, 45, "In Stock 有货"),
        ("Sony WH-1000XM5 索尼耳机", "Audio 音频", 2499, 0, "Out of Stock 缺货"),
        ("Dyson V15 戴森吸尘器", "Home 家居", 4990, 25, "In Stock 有货"),
    ]

    for i, (name, cat, price, stock, status) in enumerate(products, 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=cat)
        ws.cell(row=i, column=3, value=price)
        ws.cell(row=i, column=4, value=stock)
        ws.cell(row=i, column=5, value=status)


def create_special_chars(wb):
    """10. 特殊字符和公式"""
    ws = wb.create_sheet("特殊字符_Formula")

    headers = ["名称", "值A", "值B", "公式结果", "特殊字符"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 带公式的数据
    for i in range(2, 22):
        ws.cell(row=i, column=1, value=f"项目{i-1}")
        ws.cell(row=i, column=2, value=random.randint(100, 1000))
        ws.cell(row=i, column=3, value=random.randint(100, 1000))
        ws.cell(row=i, column=4, value=f"=B{i}+C{i}")  # Excel 公式

        # 特殊字符
        special_chars = [
            "包含<HTML>标签",
            "引号\"测试\'",
            "换行\n测试",
            "制表符\t测试",
            "百分比50%",
            "货币$100",
            "日期2025/01/01",
            "邮箱test@example.com",
            "网址https://example.com",
            "表情😀🎉",
            "日文テスト",
            "韩文테스트",
            "零宽字符\u200B测试",
            "NULL",
            "null",
            "#N/A",
            "#REF!",
            "-",
            "—",
            "...",
        ]
        ws.cell(row=i, column=5, value=special_chars[(i-2) % len(special_chars)])


if __name__ == "__main__":
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "test_complex_scenarios.xlsx")
    create_complex_test_excel(output_path)
    print(f"\nTest file created with 10 sheets covering various scenarios:")
    print("  1. 利润表_合并表头 - 合并单元格")
    print("  2. 销售明细_500行 - 大数据量")
    print("  3. 预算对比_多层表头 - 多层表头")
    print("  4. 时间序列_日数据 - 365天数据")
    print("  5. 混合类型_复杂 - 数字+文本混合")
    print("  6. 稀疏数据_70%空 - 大量空值")
    print("  7. 超宽表_120列 - 超宽表")
    print("  8. 嵌套分组_3级 - 多级分类")
    print("  9. 中英文_Bilingual - 中英文混合")
    print("  10. 特殊字符_Formula - 特殊字符和公式")
