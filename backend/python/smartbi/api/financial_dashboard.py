"""Financial Dashboard API — Chart generation, batch, analysis, PPT/PDF/Excel export."""
import logging
import asyncio
from typing import Optional, Dict, List, Any
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
import pandas as pd
import io
import json

from smartbi.services.financial_dashboard import FinancialDashboardService
from smartbi.services.financial.base import _sanitize_for_json

logger = logging.getLogger(__name__)
router = APIRouter()

# Service singleton
dashboard_service = FinancialDashboardService()


class FinancialDashboardRequest(BaseModel):
    upload_id: Optional[int] = None
    raw_data: Optional[List[Dict[str, Any]]] = Field(None, max_items=50000)  # Direct data input, max 50k rows
    chart_type: str = "all"  # specific type, "all", or "auto"
    year: int = 2026
    period_type: str = "year"  # month|quarter|year|month_range
    start_month: int = Field(1, ge=1, le=12)
    end_month: int = Field(12, ge=1, le=12)
    factory_id: str = "F001"
    filters: Optional[Dict[str, List[str]]] = None  # Fix 72: Global slicer filters {col: [values]}


class AnalyzeRequest(BaseModel):
    chart_type: str
    analysis_context: str
    chart_result: Optional[Dict] = None


class PPTExportRequest(BaseModel):
    upload_id: Optional[int] = None
    year: int = 2026
    period_type: str = "year"
    start_month: int = 1
    end_month: int = 12
    chart_images: Dict[str, str] = Field(default_factory=dict)  # {chartType: base64PNG}
    analysis_results: Dict[str, str] = Field(default_factory=dict)
    template: str = "default"
    company_name: str = "白垩纪科技"
    kpi_summary: Optional[Dict] = None


async def _load_data(request: FinancialDashboardRequest) -> pd.DataFrame:
    """Load data from upload_id or raw_data."""
    if request.raw_data:
        return pd.DataFrame(request.raw_data)

    if request.upload_id:
        # Load from smart_bi_dynamic_data table (with factory_id isolation)
        try:
            from smartbi.config import get_settings
            from sqlalchemy import create_engine, text
            settings = get_settings()
            engine = create_engine(settings.postgres_url)
            with engine.connect() as conn:
                row = conn.execute(
                    text("SELECT parsed_data FROM smart_bi_dynamic_data WHERE id = :id AND factory_id = :factory_id"),
                    {"id": request.upload_id, "factory_id": request.factory_id}
                ).fetchone()
                if row and row[0]:
                    import json
                    data = json.loads(row[0]) if isinstance(row[0], str) else row[0]
                    if isinstance(data, list):
                        return pd.DataFrame(data)
                    elif isinstance(data, dict):
                        # {sheetName: [rows]} format — merge all sheets
                        all_rows = []
                        for key, val in data.items():
                            if isinstance(val, list) and len(val) > 0:
                                all_rows.extend(val)
                        if all_rows:
                            return pd.DataFrame(all_rows)
        except Exception as e:
            logger.error(f"Failed to load upload {request.upload_id}: {e}")
            raise HTTPException(status_code=400, detail="数据加载失败，请确认上传ID正确")

    raise HTTPException(status_code=400, detail="Either upload_id or raw_data is required")


def _apply_filters(df: pd.DataFrame, filters: Optional[Dict[str, List[str]]]) -> pd.DataFrame:
    """Fix 72: Apply global slicer filters to DataFrame."""
    if not filters:
        return df
    for col, values in filters.items():
        if col in df.columns and values:
            df = df[df[col].isin(values)]
    return df


def _extract_dimensions(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Fix 72: Extract available filter dimensions from DataFrame."""
    dims = []
    for col in df.columns:
        if df[col].dtype == 'object':
            unique_vals = df[col].dropna().unique().tolist()
            if 2 <= len(unique_vals) <= 20:
                dims.append({"name": col, "values": sorted(unique_vals)})
    return dims


@router.post("/generate")
async def generate_chart(request: FinancialDashboardRequest):
    """Generate a single financial chart or all charts."""
    try:
        df = await _load_data(request)
        df = _apply_filters(df, request.filters)  # Fix 72
        result = dashboard_service.generate_chart(
            chart_type=request.chart_type,
            raw_data=df,
            year=request.year,
            period_type=request.period_type,
            start_month=request.start_month,
            end_month=request.end_month,
        )
        return _sanitize_for_json(result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Chart generation failed: {e}", exc_info=True)
        return {"success": False, "error": "图表生成失败，请检查数据格式后重试"}


@router.post("/batch")
async def batch_generate(request: FinancialDashboardRequest):
    """Generate all available charts (batch mode)."""
    try:
        request.chart_type = "all"
        df = await _load_data(request)
        # Fix 72: Extract dimensions before filtering
        dimensions = _extract_dimensions(df)
        # Fix 72: Apply slicer filters
        df = _apply_filters(df, request.filters)
        result = dashboard_service.generate_dashboard(
            raw_data=df,
            year=request.year,
            period_type=request.period_type,
            start_month=request.start_month,
            end_month=request.end_month,
        )
        # Fix 72: Attach available dimensions to response
        if isinstance(result, dict):
            result["availableDimensions"] = dimensions
        return _sanitize_for_json(result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Batch generation failed: {e}", exc_info=True)
        return {"success": False, "error": "批量生成失败，请检查数据格式后重试"}


@router.post("/analyze")
async def analyze_chart(request: AnalyzeRequest):
    """AI analysis for a single chart."""
    try:
        prompt = (
            f"请对以下财务数据进行专业分析：\n\n"
            f"图表类型：{request.chart_type}\n"
            f"数据摘要：{request.analysis_context}\n\n"
            f"请提供：\n"
            f"1. 核心发现（3-5条）\n"
            f"2. 风险预警（如有）\n"
            f"3. 改进建议（2-3条）\n\n"
            f"要求：结合食品加工行业特点，数据驱动，具体可操作。"
        )

        from smartbi.config import get_settings
        from openai import OpenAI
        settings = get_settings()
        client = OpenAI(api_key=settings.llm_api_key, base_url=settings.llm_base_url)

        response = client.chat.completions.create(
            model=settings.llm_insight_model,
            messages=[
                {"role": "system", "content": "你是服务于食品加工企业CFO的资深财务分析师。请用简洁的中文文本回复，使用要点列表格式。"},
                {"role": "user", "content": prompt},
            ],
            max_tokens=2000,
            temperature=0.3,
        )

        analysis_text = response.choices[0].message.content
        return {
            "success": True,
            "chartType": request.chart_type,
            "analysis": analysis_text,
        }
    except Exception as e:
        logger.error(f"Analysis failed: {e}", exc_info=True)
        return {"success": False, "error": "AI 分析暂时不可用，请稍后重试"}


@router.get("/templates")
async def list_templates():
    """List all available chart templates."""
    return {
        "success": True,
        "templates": dashboard_service.list_templates(),
    }


@router.post("/export-ppt")
async def export_ppt(request: PPTExportRequest):
    """Export financial dashboard to PPTX file."""
    try:
        from smartbi.services.ppt_generator import FinancialPPTGenerator
        generator = FinancialPPTGenerator()
        pptx_bytes = generator.generate(
            chart_images=request.chart_images,
            analysis_results=request.analysis_results,
            company_name=request.company_name,
            year=request.year,
            period_type=request.period_type,
            start_month=request.start_month,
            end_month=request.end_month,
            kpi_summary=request.kpi_summary,
            template=request.template,
        )
        from urllib.parse import quote
        filename = f"财务分析报告_{request.year}年{request.start_month}-{request.end_month}月.pptx"
        filename_ascii = f"report_{request.year}_{request.start_month}-{request.end_month}.pptx"
        filename_encoded = quote(filename)
        return StreamingResponse(
            io.BytesIO(pptx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_encoded}",
            },
        )
    except Exception as e:
        logger.error(f"PPT export failed: {e}", exc_info=True)
        return {"success": False, "error": "PPT 导出失败，请稍后重试"}


class ExcelExportRequest(BaseModel):
    charts: List[Dict[str, Any]] = Field(default_factory=list)  # Chart results from batch
    analysis_results: Dict[str, str] = Field(default_factory=dict)
    company_name: str = "白垩纪科技"
    year: int = 2026
    period_type: str = "year"
    start_month: int = 1
    end_month: int = 12


@router.post("/export-excel")
async def export_excel(request: ExcelExportRequest):
    """Export financial dashboard data to Excel with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1B65A8", end_color="1B65A8", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # Sheet 1: KPI Summary
        ws_summary = wb.active
        ws_summary.title = "KPI汇总"
        ws_summary.append(["图表类型", "指标", "数值", "单位", "趋势"])
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for chart in request.charts:
            chart_type = chart.get("chartType", "")
            title = chart.get("title", chart_type)
            for kpi in chart.get("kpis", []):
                ws_summary.append([
                    title,
                    kpi.get("label", ""),
                    kpi.get("value", ""),
                    kpi.get("unit", ""),
                    {"up": "↑", "down": "↓", "flat": "→"}.get(kpi.get("trend", ""), ""),
                ])

        # Auto-width for summary
        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws_summary.column_dimensions[col[0].column_letter].width = min(max_len, 30)

        # Sheet per chart with tableData
        from smartbi.services.ppt_generator import CHART_DISPLAY_NAMES
        for chart in request.charts:
            chart_type = chart.get("chartType", "")
            table_data = chart.get("tableData")
            if not table_data:
                continue

            display_name = CHART_DISPLAY_NAMES.get(chart_type, chart_type)
            # Sheet name max 31 chars
            sheet_name = display_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            headers = table_data.get("headers", [])
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row in table_data.get("rows", []):
                label = row.get("label", "")
                values = row.get("values", [])
                ws.append([label] + values)

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 25)

        # Sheet: AI Analysis
        if request.analysis_results:
            ws_ai = wb.create_sheet(title="AI分析")
            ws_ai.append(["图表", "分析内容"])
            for cell in ws_ai[1]:
                cell.font = header_font
                cell.fill = header_fill
            for chart_type, analysis in request.analysis_results.items():
                display_name = CHART_DISPLAY_NAMES.get(chart_type, chart_type)
                ws_ai.append([display_name, analysis[:5000]])
            ws_ai.column_dimensions["A"].width = 20
            ws_ai.column_dimensions["B"].width = 80

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from urllib.parse import quote
        filename = f"财务分析数据_{request.year}年{request.start_month}-{request.end_month}月.xlsx"
        filename_ascii = f"financial_data_{request.year}_{request.start_month}-{request.end_month}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{quote(filename)}",
            },
        )
    except ImportError:
        return {"success": False, "error": "服务器未安装 openpyxl，请联系管理员"}
    except Exception as e:
        logger.error(f"Excel export failed: {e}", exc_info=True)
        return {"success": False, "error": "Excel 导出失败，请稍后重试"}


class PDFExportRequest(BaseModel):
    chart_images: Dict[str, str] = Field(default_factory=dict)  # {chartType: base64PNG}
    analysis_results: Dict[str, str] = Field(default_factory=dict)
    company_name: str = "白垩纪科技"
    year: int = 2026
    period_type: str = "year"
    start_month: int = 1
    end_month: int = 12
    kpi_summary: Optional[Dict] = None


@router.post("/export-pdf")
async def export_pdf(request: PDFExportRequest):
    """Export financial dashboard to PDF with chart images and analysis."""
    try:
        from smartbi.services.pdf_generator import FinancialPDFGenerator
        generator = FinancialPDFGenerator()
        pdf_bytes = generator.generate(
            chart_images=request.chart_images,
            analysis_results=request.analysis_results,
            company_name=request.company_name,
            year=request.year,
            period_type=request.period_type,
            start_month=request.start_month,
            end_month=request.end_month,
            kpi_summary=request.kpi_summary,
        )

        from urllib.parse import quote
        filename = f"财务分析报告_{request.year}年{request.start_month}-{request.end_month}月.pdf"
        filename_ascii = f"report_{request.year}_{request.start_month}-{request.end_month}.pdf"

        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{quote(filename)}",
            },
        )
    except ImportError as e:
        logger.error(f"PDF export import failed: {e}")
        return {"success": False, "error": "PDF 导出依赖未安装，请联系管理员"}
    except Exception as e:
        logger.error(f"PDF export failed: {e}", exc_info=True)
        return {"success": False, "error": "PDF 导出失败，请稍后重试"}
