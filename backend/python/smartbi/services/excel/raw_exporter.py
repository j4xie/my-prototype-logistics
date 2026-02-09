"""
Raw Exporter - 原始数据导出（100% FIDELITY MODE）

Use this when you need UNPROCESSED, EXACT data from Excel.
All structure detection and interpretation is left to the consumer (e.g., LLM).

设计原则：
1. 不做任何智能处理，100%保留原始数据
2. 不猜测表头，所有行都作为数据导出
3. 保留合并单元格信息、原始行号、列号
4. 输出结构化JSON，供后续LLM处理

Related services:
- DataExporter (data_exporter.py): Intelligent processing, header detection, type inference
  Use when you need CLEAN, TRANSFORMED data ready for analysis
- ExportValidator (export_validator.py): Validates that exported data matches original

输出结构：
{
    "sheet_name": "利润表",
    "total_rows": 20,
    "total_cols": 10,
    "merged_cells": [...],      # 合并单元格信息
    "raw_rows": [               # 原始行数据（带行号）
        {"row_index": 0, "cells": [...]},
        ...
    ]
}
"""
import io
import json
import csv
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class CellInfo:
    """单元格信息"""
    col_index: int          # 列索引 (0-based)
    col_letter: str         # 列字母 (A, B, C...)
    value: Any              # 原始值
    value_type: str         # 值类型: text, number, date, bool, empty
    is_merged: bool = False # 是否是合并单元格的一部分
    merge_master: Optional[str] = None  # 如果是合并单元格，指向主单元格

    def to_dict(self) -> Dict:
        return {
            "col": self.col_index,
            "col_letter": self.col_letter,
            "value": self.value,
            "type": self.value_type,
            "is_merged": self.is_merged,
            "merge_master": self.merge_master
        }


@dataclass
class RowInfo:
    """行信息"""
    row_index: int          # 行索引 (0-based)
    row_number: int         # Excel行号 (1-based)
    cells: List[CellInfo]   # 单元格列表
    non_empty_count: int    # 非空单元格数量
    numeric_count: int      # 数值单元格数量

    def to_dict(self) -> Dict:
        return {
            "row": self.row_index,
            "row_number": self.row_number,
            "non_empty": self.non_empty_count,
            "numeric": self.numeric_count,
            "cells": [c.to_dict() for c in self.cells]
        }

    def to_simple_dict(self) -> Dict:
        """简化版：只返回值列表"""
        return {
            "row": self.row_index,
            "values": [c.value for c in self.cells]
        }


@dataclass
class MergedCellInfo:
    """合并单元格信息"""
    range_str: str          # 范围字符串 (如 "A1:C1")
    min_row: int            # 起始行 (1-based)
    max_row: int            # 结束行 (1-based)
    min_col: int            # 起始列 (1-based)
    max_col: int            # 结束列 (1-based)
    value: Any              # 合并单元格的值

    def to_dict(self) -> Dict:
        return {
            "range": self.range_str,
            "min_row": self.min_row,
            "max_row": self.max_row,
            "min_col": self.min_col,
            "max_col": self.max_col,
            "value": self.value
        }


@dataclass
class RawSheetData:
    """原始Sheet数据"""
    sheet_name: str
    sheet_index: int
    total_rows: int
    total_cols: int
    merged_cells: List[MergedCellInfo]
    rows: List[RowInfo]

    # 统计信息（帮助LLM理解数据结构）
    stats: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict:
        return {
            "sheet_name": self.sheet_name,
            "sheet_index": self.sheet_index,
            "total_rows": self.total_rows,
            "total_cols": self.total_cols,
            "merged_cells": [m.to_dict() for m in self.merged_cells],
            "stats": self.stats,
            "rows": [r.to_dict() for r in self.rows]
        }

    def to_simple_dict(self) -> Dict:
        """简化版JSON（更紧凑）"""
        return {
            "sheet_name": self.sheet_name,
            "sheet_index": self.sheet_index,
            "dimensions": f"{self.total_rows}x{self.total_cols}",
            "merged_cells": [m.to_dict() for m in self.merged_cells],
            "stats": self.stats,
            "rows": [r.to_simple_dict() for r in self.rows]
        }


class RawExporter:
    """
    原始数据导出器 - 保真模式

    不做任何智能处理，把Excel数据原样导出
    """

    def __init__(self):
        pass

    def export_sheet(
        self,
        file_bytes: bytes,
        sheet_index: int = 0,
        max_rows: Optional[int] = None
    ) -> RawSheetData:
        """
        导出单个Sheet的原始数据

        Args:
            file_bytes: Excel文件字节
            sheet_index: Sheet索引 (0-based)
            max_rows: 最大行数限制（None表示全部）

        Returns:
            RawSheetData 原始数据结构
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        if sheet_index >= len(wb.sheetnames):
            sheet_index = 0

        ws = wb[wb.sheetnames[sheet_index]]
        sheet_name = ws.title
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0

        # 1. 提取合并单元格信息
        merged_cells = self._extract_merged_cells(ws)
        merge_map = self._build_merge_map(merged_cells)

        # 2. 提取所有行数据
        rows = []
        row_limit = min(total_rows, max_rows) if max_rows else total_rows

        for row_idx in range(row_limit):
            row_number = row_idx + 1  # Excel行号 (1-based)
            cells = []
            non_empty_count = 0
            numeric_count = 0

            for col_idx in range(total_cols):
                col_number = col_idx + 1
                col_letter = openpyxl.utils.get_column_letter(col_number)

                cell = ws.cell(row=row_number, column=col_number)
                value = cell.value
                value_type = self._detect_value_type(value)

                # 检查是否是合并单元格
                cell_key = (row_number, col_number)
                is_merged = cell_key in merge_map
                merge_master = merge_map.get(cell_key)

                # 统计
                if value is not None and str(value).strip():
                    non_empty_count += 1
                if value_type == "number":
                    numeric_count += 1

                # 处理日期类型的序列化
                if isinstance(value, datetime):
                    value = value.isoformat()

                cells.append(CellInfo(
                    col_index=col_idx,
                    col_letter=col_letter,
                    value=value,
                    value_type=value_type,
                    is_merged=is_merged,
                    merge_master=merge_master
                ))

            rows.append(RowInfo(
                row_index=row_idx,
                row_number=row_number,
                cells=cells,
                non_empty_count=non_empty_count,
                numeric_count=numeric_count
            ))

        wb.close()

        # 3. 计算统计信息
        stats = self._compute_stats(rows, total_cols)

        return RawSheetData(
            sheet_name=sheet_name,
            sheet_index=sheet_index,
            total_rows=total_rows,
            total_cols=total_cols,
            merged_cells=merged_cells,
            rows=rows,
            stats=stats
        )

    def export_all_sheets(
        self,
        file_bytes: bytes,
        max_rows_per_sheet: Optional[int] = None
    ) -> List[RawSheetData]:
        """导出所有Sheet的原始数据"""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()

        results = []
        for i in range(sheet_count):
            data = self.export_sheet(file_bytes, sheet_index=i, max_rows=max_rows_per_sheet)
            results.append(data)

        return results

    def to_json(self, data: RawSheetData, simple: bool = False, indent: int = 2) -> str:
        """导出为JSON"""
        if simple:
            return json.dumps(data.to_simple_dict(), ensure_ascii=False, indent=indent)
        return json.dumps(data.to_dict(), ensure_ascii=False, indent=indent)

    def to_markdown(
        self,
        data: RawSheetData,
        max_rows: int = 50,
        truncate: bool = True,
        max_cell_width: int = 20
    ) -> str:
        """
        导出为Markdown表格（原始版）

        特点：
        - 不猜测表头，第一行就是第一行
        - 显示行号
        - 显示合并单元格信息

        Args:
            data: RawSheetData 原始数据
            max_rows: 最大显示行数（默认50）
            truncate: 是否截断过长的单元格值（默认True）
            max_cell_width: 单元格最大宽度（默认20，仅当truncate=True时生效）
        """
        lines = []

        # 标题
        lines.append(f"# Sheet: {data.sheet_name}")
        lines.append(f"")
        lines.append(f"- 总行数: {data.total_rows}")
        lines.append(f"- 总列数: {data.total_cols}")
        lines.append(f"- 合并单元格: {len(data.merged_cells)} 个")
        lines.append(f"")

        # 合并单元格信息
        if data.merged_cells:
            lines.append("## 合并单元格")
            for m in data.merged_cells[:10]:  # 最多显示10个
                lines.append(f"- `{m.range_str}`: {m.value}")
            if len(data.merged_cells) > 10:
                lines.append(f"- ... 还有 {len(data.merged_cells) - 10} 个")
            lines.append("")

        # 统计信息
        lines.append("## 行统计")
        lines.append("| 行号 | 非空数 | 数值数 | 预估类型 |")
        lines.append("|------|--------|--------|----------|")
        for row in data.rows[:15]:  # 前15行的统计
            row_type = self._guess_row_type(row, data.total_cols)
            lines.append(f"| {row.row_number} | {row.non_empty_count} | {row.numeric_count} | {row_type} |")
        lines.append("")

        # 原始数据表格
        lines.append("## 原始数据")
        lines.append("")

        # 表头：使用Excel列字母 (A, B, C...)
        # 对于宽表（>30列），智能采样：前5列 + 均匀采样 + 后5列
        MAX_LLM_COLS = 30
        display_rows = data.rows[:max_rows]
        if display_rows:
            total_cols = len(display_rows[0].cells)
            if total_cols <= MAX_LLM_COLS:
                sample_indices = list(range(total_cols))
            else:
                # 智能采样：前5 + 均匀采样中间20 + 后5
                first_cols = list(range(5))
                last_cols = list(range(total_cols - 5, total_cols))
                middle_count = MAX_LLM_COLS - 10
                step = max(1, (total_cols - 10) // middle_count)
                middle_indices = [5 + i * step for i in range(middle_count) if 5 + i * step < total_cols - 5]
                sample_indices = first_cols + middle_indices + last_cols
                # 去重并排序
                sample_indices = sorted(list(set(sample_indices)))

            col_count = len(sample_indices)
            # 添加列采样说明
            if total_cols > MAX_LLM_COLS:
                lines.append(f"*注意: 原表有 {total_cols} 列，已智能采样 {col_count} 列供分析*")
                lines.append("")
            header_cols = [display_rows[0].cells[i].col_letter for i in sample_indices]
            lines.append("| 行号 | " + " | ".join(header_cols) + " |")
            lines.append("|------|" + "|".join(["---"] * col_count) + "|")

            for row in display_rows:
                values = []
                for idx in sample_indices:
                    cell = row.cells[idx] if idx < len(row.cells) else None
                    if cell is None:
                        values.append("")
                        continue
                    v = str(cell.value) if cell.value is not None else ""
                    # 截断过长的值（如果启用）
                    if truncate and len(v) > max_cell_width:
                        v = v[:max_cell_width - 3] + "..."
                    # 转义管道符和换行符
                    v = v.replace("|", "\\|").replace("\n", " ")
                    values.append(v)
                lines.append(f"| {row.row_number} | " + " | ".join(values) + " |")

        if len(data.rows) > max_rows:
            lines.append(f"")
            lines.append(f"*（显示前 {max_rows} 行，共 {len(data.rows)} 行）*")

        return "\n".join(lines)

    def to_csv(
        self,
        data: RawSheetData,
        include_row_number: bool = True,
        include_metadata: bool = False
    ) -> str:
        """
        导出为CSV（原始版）

        特点：
        - 第一列是行号（可选）
        - 可包含元数据注释（可选）
        - 不加任何处理，原样输出

        Args:
            data: RawSheetData 原始数据
            include_row_number: 是否包含行号列（默认True）
            include_metadata: 是否在开头添加元数据注释（默认False）
        """
        lines = []

        # 添加元数据注释
        if include_metadata:
            lines.append(f"# sheet: {data.sheet_name}")
            lines.append(f"# rows: {data.total_rows}")
            lines.append(f"# cols: {data.total_cols}")
            lines.append(f"# merged_cells: {len(data.merged_cells)}")
            lines.append("")

        output = io.StringIO()
        writer = csv.writer(output)

        for row in data.rows:
            values = [cell.value if cell.value is not None else "" for cell in row.cells]
            if include_row_number:
                values = [row.row_number] + values
            writer.writerow(values)

        csv_content = output.getvalue()

        if include_metadata:
            return "\n".join(lines) + csv_content
        return csv_content

    def _extract_merged_cells(self, ws) -> List[MergedCellInfo]:
        """提取合并单元格信息"""
        merged = []
        for merged_range in ws.merged_cells.ranges:
            cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            value = cell.value
            if isinstance(value, datetime):
                value = value.isoformat()

            merged.append(MergedCellInfo(
                range_str=str(merged_range),
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
                value=value
            ))
        return merged

    def _build_merge_map(self, merged_cells: List[MergedCellInfo]) -> Dict[tuple, str]:
        """构建合并单元格映射：(row, col) -> master_cell_range"""
        merge_map = {}
        for m in merged_cells:
            master = f"{openpyxl.utils.get_column_letter(m.min_col)}{m.min_row}"
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    if (r, c) != (m.min_row, m.min_col):
                        merge_map[(r, c)] = master
        return merge_map

    def _detect_value_type(self, value: Any) -> str:
        """检测值类型"""
        if value is None:
            return "empty"
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, datetime):
            return "date"
        # 尝试解析字符串中的数字
        if isinstance(value, str):
            v = value.strip()
            if not v:
                return "empty"
            try:
                float(v.replace(",", "").replace("¥", "").replace("$", "").replace("%", ""))
                return "number"
            except:
                pass
        return "text"

    def _compute_stats(self, rows: List[RowInfo], total_cols: int) -> Dict[str, Any]:
        """计算统计信息，帮助LLM理解数据结构"""
        if not rows:
            return {}

        # 找出非空行
        non_empty_rows = [r for r in rows if r.non_empty_count > 0]

        # 计算每行的填充率
        fill_rates = [r.non_empty_count / total_cols for r in rows if total_cols > 0]

        # 找出"数据行"的起始位置（填充率>50% 且 数值占比>30%）
        potential_data_start = None
        for row in rows:
            fill_rate = row.non_empty_count / total_cols if total_cols > 0 else 0
            numeric_rate = row.numeric_count / row.non_empty_count if row.non_empty_count > 0 else 0
            if fill_rate > 0.5 and numeric_rate > 0.3:
                potential_data_start = row.row_index
                break

        return {
            "non_empty_row_count": len(non_empty_rows),
            "avg_fill_rate": sum(fill_rates) / len(fill_rates) if fill_rates else 0,
            "potential_data_start_row": potential_data_start,
            "first_10_rows_summary": [
                {
                    "row": r.row_index,
                    "non_empty": r.non_empty_count,
                    "numeric": r.numeric_count,
                    "fill_rate": round(r.non_empty_count / total_cols, 2) if total_cols > 0 else 0
                }
                for r in rows[:10]
            ]
        }

    def _guess_row_type(self, row: RowInfo, total_cols: int) -> str:
        """猜测行类型（仅供参考，不作为最终判断）"""
        if row.non_empty_count == 0:
            return "空行"

        fill_rate = row.non_empty_count / total_cols if total_cols > 0 else 0
        numeric_rate = row.numeric_count / row.non_empty_count if row.non_empty_count > 0 else 0

        if fill_rate < 0.2:
            return "稀疏行(标题?)"
        if numeric_rate > 0.5:
            return "数据行"
        if fill_rate > 0.5 and numeric_rate < 0.3:
            return "文本行(表头?)"

        return "混合行"
