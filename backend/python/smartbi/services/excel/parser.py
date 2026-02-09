from __future__ import annotations
"""
Excel Parser Service

Handles Excel file parsing with support for:
- Multi-level headers
- Data direction detection (row vs column oriented)
- Data type detection (DATE, NUMERIC, CATEGORICAL, ID, TEXT)
- Field mapping with synonym dictionary
- Various Excel formats (.xlsx, .xls, .csv)

Python equivalent of Java's ExcelDynamicParserServiceImpl
"""
import logging
from io import BytesIO
from typing import Any, Optional, Union, List, Dict
from enum import Enum
from datetime import datetime

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

from ..data.feature_analyzer import (
    DataFeatureAnalyzer,
    DataFeatureResult,
    DataType,
    count_time_pattern_headers,
    is_time_pattern
)
from ..field.mapping import (
    FieldMappingService,
    FieldMappingResult,
    FieldMappingDictionary
)
from ..structure.table_classifier import TableClassifier, TableType

logger = logging.getLogger(__name__)


class DataDirection(str, Enum):
    """Data orientation in the spreadsheet"""
    ROW_ORIENTED = "row"  # Headers in first row, data in subsequent rows
    COLUMN_ORIENTED = "column"  # Headers in first column, data in subsequent columns
    UNKNOWN = "unknown"


class SheetInfo:
    """Information about a sheet in an Excel file"""

    def __init__(
        self,
        index: int,
        name: str,
        row_count: int = 0,
        column_count: int = 0,
        is_empty: bool = True,
        preview_headers: Optional[List[str]] = None
    ):
        self.index = index
        self.name = name
        self.row_count = row_count
        self.column_count = column_count
        self.is_empty = is_empty
        self.preview_headers = preview_headers or []

    def to_dict(self) -> Dict[str, Any]:
        return {
            "index": self.index,
            "name": self.name,
            "rowCount": self.row_count,
            "columnCount": self.column_count,
            "empty": self.is_empty,
            "previewHeaders": self.preview_headers
        }


class ExcelParser:
    """
    Excel file parser with full dynamic analysis support.

    Features:
    - Multi-level header detection and flattening
    - Data direction detection (row vs column oriented)
    - Data type analysis per column
    - Field mapping with synonym dictionary
    - Sheet enumeration with preview
    """

    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls", ".csv"]
        self.feature_analyzer = DataFeatureAnalyzer()
        self.field_mapping_service = FieldMappingService()
        self.table_classifier = TableClassifier()

    def parse(
        self,
        file_bytes: bytes,
        sheet_name: Optional[Union[str, int]] = 0,
        header_rows: int = 1,
        skip_rows: int = 0,
        transpose: bool = False
    ) -> Dict[str, Any]:
        """
        Parse Excel file and return structured data

        Args:
            file_bytes: Raw file bytes
            sheet_name: Sheet name or index (default: first sheet)
            header_rows: Number of header rows (1 or 2 for multi-level)
            skip_rows: Number of rows to skip at the top
            transpose: Whether to transpose the data

        Returns:
            Dictionary with headers, rows, and metadata
        """
        try:
            buffer = BytesIO(file_bytes)

            # Determine header parameter
            if header_rows == 2:
                header = [0, 1]
            elif header_rows > 2:
                header = list(range(header_rows))
            else:
                header = 0

            # Read Excel file
            df = pd.read_excel(
                buffer,
                sheet_name=sheet_name,
                header=header,
                skiprows=skip_rows
            )

            # Handle multi-level columns
            if isinstance(df.columns, pd.MultiIndex):
                df = self._flatten_multi_index_columns(df)

            # Transpose if needed
            if transpose:
                df = df.T.reset_index()
                # Use first row as headers after transpose
                df.columns = df.iloc[0].tolist()
                df = df[1:]

            # Detect data direction
            direction = self._detect_data_direction(df)

            # Clean data
            df = self._clean_dataframe(df)

            # Convert to structured format
            headers = df.columns.tolist()
            rows = df.values.tolist()

            # Get sheet names
            buffer.seek(0)
            xl = pd.ExcelFile(buffer)
            sheet_names = xl.sheet_names

            return {
                "success": True,
                "headers": headers,
                "rows": rows,
                "rowCount": len(rows),
                "columnCount": len(headers),
                "direction": direction.value,
                "sheetNames": sheet_names,
                "currentSheet": sheet_name if isinstance(sheet_name, str) else sheet_names[sheet_name] if sheet_name < len(sheet_names) else sheet_names[0]
            }

        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": [],
                "rowCount": 0,
                "columnCount": 0
            }

    def parse_csv(
        self,
        file_bytes: bytes,
        encoding: str = "utf-8",
        delimiter: str = ","
    ) -> Dict[str, Any]:
        """Parse CSV file"""
        try:
            buffer = BytesIO(file_bytes)
            df = pd.read_csv(buffer, encoding=encoding, delimiter=delimiter)
            df = self._clean_dataframe(df)

            return {
                "success": True,
                "headers": df.columns.tolist(),
                "rows": df.values.tolist(),
                "rowCount": len(df),
                "columnCount": len(df.columns),
                "direction": self._detect_data_direction(df).value
            }
        except Exception as e:
            logger.error(f"Failed to parse CSV file: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": []
            }

    def get_sheet_names(self, file_bytes: bytes) -> List[str]:
        """Get all sheet names from Excel file"""
        try:
            buffer = BytesIO(file_bytes)
            xl = pd.ExcelFile(buffer)
            return xl.sheet_names
        except Exception as e:
            logger.error(f"Failed to get sheet names: {e}")
            return []

    def list_sheets_detailed(self, file_bytes: bytes) -> List[SheetInfo]:
        """
        Get detailed information about all sheets in an Excel file.

        Equivalent to Java's listSheets() method.

        Returns:
            List of SheetInfo with row count, column count, and preview headers
        """
        try:
            buffer = BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(buffer, read_only=True, data_only=True)

            sheets = []
            for index, sheet_name in enumerate(workbook.sheetnames):
                try:
                    worksheet = workbook[sheet_name]

                    # Count rows (max 1000 for performance)
                    row_count = 0
                    max_column = 0
                    preview_headers = []

                    # Iterate through rows to count and get dimensions
                    for row_idx, row in enumerate(worksheet.iter_rows(max_row=1000), start=1):
                        # Check if row has any data
                        has_data = any(cell.value is not None for cell in row)
                        if has_data:
                            row_count = row_idx

                            # Track max column count from first 100 rows
                            if row_idx <= 100:
                                col_count = sum(1 for cell in row if cell.value is not None)
                                max_column = max(max_column, col_count)

                            # Get preview headers from first row (max 10 columns)
                            if row_idx == 1:
                                for cell in row[:10]:
                                    if cell.value is not None:
                                        preview_headers.append(str(cell.value))

                    # Determine if sheet is empty
                    is_empty = row_count == 0

                    # If we counted rows, use worksheet dimensions as fallback for column count
                    if not is_empty and max_column == 0:
                        max_column = worksheet.max_column or 0

                    sheet_info = SheetInfo(
                        index=index,
                        name=sheet_name,
                        row_count=row_count,
                        column_count=max_column,
                        is_empty=is_empty,
                        preview_headers=preview_headers
                    )
                    sheets.append(sheet_info)

                except Exception as sheet_error:
                    logger.warning(f"Failed to analyze sheet '{sheet_name}': {sheet_error}")
                    # Add minimal info for failed sheet
                    sheets.append(SheetInfo(
                        index=index,
                        name=sheet_name,
                        row_count=0,
                        column_count=0,
                        is_empty=True,
                        preview_headers=[]
                    ))

            workbook.close()
            return sheets

        except Exception as e:
            logger.error(f"Failed to list sheets detailed: {e}", exc_info=True)
            return []

    def detect_multi_header(self, file_bytes: bytes, sheet_name: Optional[Union[str, int]] = 0) -> Dict[str, Any]:
        """
        Detect if the Excel file has multi-level headers using openpyxl for merged cell detection.

        Returns:
            Dictionary with detection results including:
            - hasMultiHeader: true if merged cells span multiple rows/columns
            - originalHeaders: raw header values before merging
            - recommendedHeaderRows: suggested number of header rows
        """
        try:
            buffer = BytesIO(file_bytes)

            # Use openpyxl for merged cell detection
            workbook = openpyxl.load_workbook(buffer, read_only=False, data_only=True)

            # Get the target sheet
            if isinstance(sheet_name, int):
                if sheet_name < len(workbook.sheetnames):
                    ws = workbook[workbook.sheetnames[sheet_name]]
                else:
                    ws = workbook.active
            else:
                ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

            # Analyze merged cells in header region (first 5 rows)
            merged_ranges = list(ws.merged_cells.ranges)
            header_merged_cells = []
            max_merge_row = 1
            has_multi_header = False

            for merged_range in merged_ranges:
                # Check if merge is in header region (first 5 rows)
                if merged_range.min_row <= 5:
                    header_merged_cells.append({
                        "range": str(merged_range),
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col
                    })

                    # Check if merge spans multiple rows
                    if merged_range.max_row > merged_range.min_row:
                        has_multi_header = True
                        max_merge_row = max(max_merge_row, merged_range.max_row)

                    # Check if merge spans multiple columns in header rows
                    if merged_range.min_row <= 2 and merged_range.max_col > merged_range.min_col:
                        has_multi_header = True

            # Get original headers (raw values from first rows)
            original_headers = []
            preview_rows = []

            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_values = []
                for col_idx in range(1, min(ws.max_column + 1, 51)):  # Max 50 columns
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is not None:
                        row_values.append(str(value))
                    else:
                        row_values.append(None)

                if row_idx == 1:
                    original_headers = row_values
                preview_rows.append(row_values)

            # Determine recommended header rows
            if has_multi_header:
                recommended_header_rows = min(max_merge_row, 3)  # Cap at 3 rows
            else:
                # Fall back to pandas-based detection for non-merged multi-headers
                buffer.seek(0)
                df_raw = pd.read_excel(buffer, sheet_name=sheet_name, header=None, nrows=5)

                if len(df_raw) >= 2:
                    first_row = df_raw.iloc[0]
                    second_row = df_raw.iloc[1]

                    first_row_empty_ratio = first_row.isna().sum() / len(first_row) if len(first_row) > 0 else 0
                    second_row_types = second_row.apply(lambda x: 'str' if isinstance(x, str) else 'num')

                    if first_row_empty_ratio > 0.3 and (second_row_types == 'str').sum() > len(second_row) * 0.5:
                        has_multi_header = True
                        recommended_header_rows = 2
                    else:
                        recommended_header_rows = 1
                else:
                    recommended_header_rows = 1

            workbook.close()

            return {
                "success": True,
                "hasMultiHeader": has_multi_header,
                "isMultiHeader": has_multi_header,  # Alias for backward compatibility
                "recommendedHeaderRows": recommended_header_rows,
                "originalHeaders": original_headers,
                "mergedCells": header_merged_cells,
                "previewRows": preview_rows
            }

        except Exception as e:
            logger.error(f"Failed to detect multi-header: {e}", exc_info=True)
            return {
                "success": False,
                "hasMultiHeader": False,
                "isMultiHeader": False,
                "recommendedHeaderRows": 1,
                "originalHeaders": [],
                "mergedCells": [],
                "previewRows": [],
                "error": str(e)
            }

    def _flatten_multi_index_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Flatten multi-level column index to single level"""
        if isinstance(df.columns, pd.MultiIndex):
            # Join multi-level column names with separator
            df.columns = [
                '_'.join(str(c) for c in col if str(c) != 'nan' and str(c) != '')
                for col in df.columns.values
            ]
        return df

    def parse_with_merged_headers(
        self,
        file_bytes: bytes,
        sheet_name: Optional[Union[str, int]] = 0,
        header_row_count: int = 3,
        data_start_row: int = 3
    ) -> Dict[str, Any]:
        """
        Parse Excel with complex merged multi-level headers.

        This method uses openpyxl to read raw cell values and properly
        propagate merged cell values to build correct column headers.

        Designed for sheets like "收入及净利简表" which have:
        - Row 0: Title (merged across all columns)
        - Row 1: Category groupings (e.g., months, merged across sub-columns)
        - Row 2: Sub-categories (e.g., 预算收入, 实际收入)
        - Row 3+: Data

        Args:
            file_bytes: Raw file bytes
            sheet_name: Sheet name or index
            header_row_count: Number of header rows (default 3)
            data_start_row: 0-indexed row where data starts (default 3)

        Returns:
            Dictionary with headers, rows, and metadata
        """
        try:
            buffer = BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(buffer, data_only=True)

            # Get the target sheet
            if isinstance(sheet_name, int):
                if sheet_name < len(workbook.sheetnames):
                    ws = workbook[workbook.sheetnames[sheet_name]]
                    actual_sheet_name = workbook.sheetnames[sheet_name]
                else:
                    ws = workbook.active
                    actual_sheet_name = ws.title
            else:
                ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
                actual_sheet_name = ws.title

            max_col = ws.max_column or 1
            max_row = ws.max_row or 1

            # Step 1: Build a map of merged cell values
            merged_values = {}
            for merged_range in ws.merged_cells.ranges:
                # Get the value from the top-left cell
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                value = top_left_cell.value

                # Fill all cells in the merged range with this value
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_values[(row, col)] = value

            # Step 2: Read header rows and build column names
            header_rows_data = []
            for row_idx in range(1, header_row_count + 1):  # 1-indexed in openpyxl
                row_values = []
                for col_idx in range(1, max_col + 1):
                    # Check if this cell is part of a merged range
                    if (row_idx, col_idx) in merged_values:
                        value = merged_values[(row_idx, col_idx)]
                    else:
                        value = ws.cell(row=row_idx, column=col_idx).value
                    row_values.append(value)
                header_rows_data.append(row_values)

            # Step 3: Forward-fill None values in header rows (for non-merged gaps)
            for row_idx, row in enumerate(header_rows_data):
                last_value = None
                for col_idx in range(len(row)):
                    if row[col_idx] is not None:
                        last_value = row[col_idx]
                    elif last_value is not None and row_idx > 0:  # Don't fill first row
                        row[col_idx] = last_value

            # Step 4: Build final column names by combining header rows
            # Skip the title row (row 0) if it has only 1-2 unique values
            unique_first_row = set(v for v in header_rows_data[0] if v is not None)
            start_header_row = 1 if len(unique_first_row) <= 2 else 0

            headers = []
            for col_idx in range(max_col):
                parts = []
                for row_idx in range(start_header_row, header_row_count):
                    if row_idx < len(header_rows_data) and col_idx < len(header_rows_data[row_idx]):
                        value = header_rows_data[row_idx][col_idx]
                        if value is not None:
                            str_value = str(value).strip()
                            # Avoid duplicating the same value
                            if str_value and (not parts or parts[-1] != str_value):
                                parts.append(str_value)

                if parts:
                    headers.append('_'.join(parts))
                else:
                    headers.append(f'Column_{col_idx + 1}')

            # Step 5: Read data rows
            rows = []
            for row_idx in range(data_start_row + 1, max_row + 1):  # +1 for 1-indexed
                row_values = []
                has_data = False
                for col_idx in range(1, max_col + 1):
                    # Check merged values first
                    if (row_idx, col_idx) in merged_values:
                        value = merged_values[(row_idx, col_idx)]
                    else:
                        value = ws.cell(row=row_idx, column=col_idx).value

                    if value is not None:
                        has_data = True

                    row_values.append(value)

                # Only add rows that have some data
                if has_data:
                    rows.append(row_values)

            workbook.close()

            # Get all sheet names
            buffer.seek(0)
            xl = pd.ExcelFile(buffer)
            sheet_names = xl.sheet_names

            logger.info(f"Parsed with merged headers: {len(headers)} columns, {len(rows)} rows")

            return {
                "success": True,
                "headers": headers,
                "rows": rows,
                "rowCount": len(rows),
                "columnCount": len(headers),
                "direction": "row",
                "sheetNames": sheet_names,
                "currentSheet": actual_sheet_name,
                "_parse_method": "merged_headers"
            }

        except Exception as e:
            logger.error(f"Failed to parse with merged headers: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": [],
                "rowCount": 0,
                "columnCount": 0
            }

    def _detect_data_direction(self, df: pd.DataFrame) -> DataDirection:
        """
        Detect whether data is row-oriented or column-oriented

        Heuristics:
        - Row-oriented: First row is headers, subsequent rows are data records
        - Column-oriented: First column is headers, subsequent columns are data series
        """
        if df.empty:
            return DataDirection.ROW_ORIENTED

        # Check first column vs first row for date/time patterns
        first_col = df.iloc[:, 0] if len(df.columns) > 0 else pd.Series()
        first_row = df.iloc[0] if len(df) > 0 else pd.Series()

        # Count numeric values
        first_col_numeric = pd.to_numeric(first_col, errors='coerce').notna().sum()
        first_row_numeric = pd.to_numeric(first_row, errors='coerce').notna().sum()

        # If first column has more numeric values, likely column-oriented (time series)
        if first_col_numeric > len(first_col) * 0.7 and first_row_numeric < len(first_row) * 0.3:
            return DataDirection.COLUMN_ORIENTED

        return DataDirection.ROW_ORIENTED

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean dataframe by handling NaN and converting types"""
        # Replace NaN with None for JSON serialization
        df = df.replace({np.nan: None})

        # Convert datetime columns to ISO format strings
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].apply(
                    lambda x: x.isoformat() if pd.notna(x) else None
                )

        return df

    def parse_with_analysis(
        self,
        file_bytes: bytes,
        sheet_name: Optional[Union[str, int]] = 0,
        header_rows: int = 1,
        skip_rows: int = 0,
        transpose: bool = False,
        factory_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Parse Excel file with full data feature analysis and field mapping.

        Enhanced version of parse() that includes:
        - Data feature analysis for each column
        - Field mapping to standard fields
        - Missing required fields detection
        - Enhanced data orientation detection

        Args:
            file_bytes: Raw file bytes
            sheet_name: Sheet name or index (default: first sheet)
            header_rows: Number of header rows (1 or 2 for multi-level)
            skip_rows: Number of rows to skip at the top
            transpose: Whether to transpose the data
            factory_id: Optional factory ID for factory-specific mappings

        Returns:
            Dictionary with headers, rows, metadata, dataFeatures, fieldMappings, etc.
        """
        try:
            buffer = BytesIO(file_bytes)

            # Determine header parameter
            if header_rows == 2:
                header = [0, 1]
            elif header_rows > 2:
                header = list(range(header_rows))
            else:
                header = 0

            # Read Excel file
            df = pd.read_excel(
                buffer,
                sheet_name=sheet_name,
                header=header,
                skiprows=skip_rows
            )

            # Handle multi-level columns
            if isinstance(df.columns, pd.MultiIndex):
                df = self._flatten_multi_index_columns(df)

            # Transpose if needed
            if transpose:
                df = df.T.reset_index()
                # Use first row as headers after transpose
                df.columns = df.iloc[0].tolist()
                df = df[1:]

            # Detect data direction using enhanced method
            direction = self._detect_data_orientation_enhanced(df)

            # Clean data
            df_cleaned = self._clean_dataframe(df.copy())

            # Convert to structured format
            headers = df_cleaned.columns.tolist()
            rows = df_cleaned.values.tolist()

            # Get sheet names
            buffer.seek(0)
            xl = pd.ExcelFile(buffer)
            sheet_names = xl.sheet_names

            # Analyze data features for each column
            data_features: List[DataFeatureResult] = []
            for idx, col_name in enumerate(headers):
                values = self._extract_column_values(df, col_name)
                feature = self.feature_analyzer.analyze_column(
                    column_name=str(col_name),
                    values=values,
                    column_index=idx
                )
                data_features.append(feature)

            # Map fields to standard fields
            field_mappings = self.field_mapping_service.map_fields(
                headers=[str(h) for h in headers],
                features=data_features,
                factory_id=factory_id
            )

            # Determine missing required/recommended fields
            mapped_standard_fields = {
                fm.standard_field for fm in field_mappings
                if fm.standard_field is not None
            }
            missing_required_fields = self.field_mapping_service.dictionary.get_missing_recommended_fields(
                mapped_standard_fields
            )

            current_sheet = (
                sheet_name if isinstance(sheet_name, str)
                else sheet_names[sheet_name] if sheet_name < len(sheet_names)
                else sheet_names[0]
            )

            # Classify table type using TableClassifier
            table_type = TableType.DATA_TABLE.value
            try:
                classification = self.table_classifier.classify(current_sheet, df_cleaned)
                table_type = classification.table_type.value
                logger.info(f"Table classified as: {table_type} (confidence: {classification.confidence:.2f})")
            except Exception as e:
                logger.warning(f"Table classification failed: {e}")

            return {
                "success": True,
                "headers": headers,
                "rows": rows,
                "rowCount": len(rows),
                "columnCount": len(headers),
                "direction": direction.value,
                "sheetNames": sheet_names,
                "currentSheet": current_sheet,
                "tableType": table_type,
                "dataFeatures": [f.to_dict() for f in data_features],
                "fieldMappings": [fm.to_dict() for fm in field_mappings],
                "missingRequiredFields": missing_required_fields,
                "status": "COMPLETE"  # Dynamic adaptation - always complete
            }

        except Exception as e:
            logger.error(f"Failed to parse Excel file with analysis: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": [],
                "rowCount": 0,
                "columnCount": 0,
                "tableType": TableType.UNKNOWN.value,
                "dataFeatures": [],
                "fieldMappings": [],
                "missingRequiredFields": [],
                "status": "COMPLETE"
            }

    def _extract_column_values(self, df: pd.DataFrame, column_name: str) -> List[Any]:
        """
        Extract values from a DataFrame column.

        Args:
            df: The DataFrame
            column_name: Name of the column to extract

        Returns:
            List of column values
        """
        if column_name not in df.columns:
            return []

        column_data = df[column_name]
        values = []

        for value in column_data:
            if pd.isna(value):
                values.append(None)
            elif isinstance(value, (pd.Timestamp,)):
                # Convert datetime to string for consistency
                values.append(value.isoformat() if pd.notna(value) else None)
            else:
                values.append(value)

        return values

    def _detect_data_orientation_enhanced(self, df: pd.DataFrame) -> DataDirection:
        """
        Enhanced data direction detection using time pattern analysis.

        Based on Java implementation logic:
        1. Count time pattern headers using is_time_pattern()
        2. Check if first column is a label column (text, high uniqueness)
        3. Return COLUMN_ORIENTED if many time patterns in headers AND first column is labels

        Args:
            df: The DataFrame to analyze

        Returns:
            DataDirection enum value
        """
        if df.empty:
            return DataDirection.ROW_ORIENTED

        headers = [str(h) for h in df.columns.tolist()]

        # Count time patterns in headers
        time_pattern_count = count_time_pattern_headers(headers)
        time_pattern_ratio = time_pattern_count / len(headers) if headers else 0

        logger.debug(
            f"Time pattern analysis: count={time_pattern_count}, "
            f"total_headers={len(headers)}, ratio={time_pattern_ratio:.2%}"
        )

        # Check if first column is a label column
        first_col_is_label = False
        if len(df.columns) > 0 and len(df) > 0:
            first_col = df.iloc[:, 0]

            # Check if first column is mostly text
            text_count = sum(
                1 for v in first_col
                if v is not None and isinstance(v, str) and not self._is_numeric_string(str(v))
            )
            text_ratio = text_count / len(first_col) if len(first_col) > 0 else 0

            # Check uniqueness of first column
            non_null_values = [v for v in first_col if v is not None and str(v).strip()]
            unique_count = len(set(str(v) for v in non_null_values))
            unique_ratio = unique_count / len(non_null_values) if non_null_values else 0

            # First column is a label column if:
            # - High text ratio (> 50%)
            # - High uniqueness (> 50%)
            first_col_is_label = text_ratio > 0.5 and unique_ratio > 0.5

            logger.debug(
                f"First column analysis: text_ratio={text_ratio:.2%}, "
                f"unique_ratio={unique_ratio:.2%}, is_label={first_col_is_label}"
            )

        # Decision logic (from Java implementation):
        # If many headers are time patterns (> 30%) AND first column is labels
        # -> COLUMN_ORIENTED (time series data)
        if time_pattern_ratio > 0.3 and first_col_is_label:
            logger.info(
                f"Detected COLUMN_ORIENTED data: time_headers={time_pattern_count}, "
                f"first_col_is_label={first_col_is_label}"
            )
            return DataDirection.COLUMN_ORIENTED

        # Fall back to original detection method
        return self._detect_data_direction(df)

    def _is_numeric_string(self, value: str) -> bool:
        """
        Check if a string represents a numeric value.

        Args:
            value: String to check

        Returns:
            True if the string is numeric
        """
        if not value or not value.strip():
            return False

        # Remove common formatting
        cleaned = value.strip().replace(',', '').replace(' ', '')

        # Check for currency symbols and percentage
        if cleaned.startswith(('¥', '$', '€', '£')):
            cleaned = cleaned[1:]
        if cleaned.endswith('%'):
            cleaned = cleaned[:-1]

        try:
            float(cleaned)
            return True
        except ValueError:
            return False

    def smart_parse(
        self,
        file_bytes: bytes,
        sheet_name: Optional[Union[str, int]] = 0,
        factory_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Smart parsing with automatic structure detection.

        This method automatically detects and handles:
        - Multi-level headers (2+ header rows)
        - Title/metadata rows that need to be skipped
        - Merged cells in headers
        - Column-oriented vs row-oriented data

        It tries multiple parsing strategies and picks the best result
        based on data quality metrics.

        Args:
            file_bytes: Raw file bytes
            sheet_name: Sheet name or index (default: first sheet)
            factory_id: Optional factory ID for factory-specific mappings

        Returns:
            Best parsing result with headers, rows, and metadata
        """
        logger.info(f"Smart parse starting for sheet: {sheet_name}")

        # Step 1: Detect structure using multi-header detection
        detection = self.detect_multi_header(file_bytes, sheet_name)
        detected_header_rows = detection.get("recommendedHeaderRows", 1)
        has_multi_header = detection.get("hasMultiHeader", False)
        merged_cells = detection.get("mergedCells", [])
        preview_rows = detection.get("previewRows", [])

        logger.info(f"Detection result: header_rows={detected_header_rows}, "
                   f"has_multi_header={has_multi_header}, merged_cells={len(merged_cells)}")

        # Step 2: Check if this sheet needs special merged header handling
        # Criteria: many merged cells in header region AND multi-header detected
        needs_merged_header_parsing = self._needs_merged_header_parsing(
            merged_cells, has_multi_header, detected_header_rows, preview_rows
        )

        if needs_merged_header_parsing:
            logger.info("Detected complex merged headers, trying parse_with_merged_headers")
            try:
                # Try the specialized merged header parser
                merged_result = self.parse_with_merged_headers(
                    file_bytes,
                    sheet_name=sheet_name,
                    header_row_count=detected_header_rows,
                    data_start_row=detected_header_rows
                )
                if merged_result.get("success"):
                    merged_score = self._score_parse_result(merged_result)
                    logger.info(f"Merged header parse score: {merged_score}")

                    # If score is good enough, use this result
                    if merged_score >= 50:
                        merged_result["_parse_strategy"] = {"method": "merged_headers", "header_rows": detected_header_rows}
                        merged_result["_parse_score"] = merged_score

                        # Analyze data features
                        headers = merged_result.get("headers", [])
                        rows = merged_result.get("rows", [])
                        if headers and rows:
                            data_features = []
                            for idx, col_name in enumerate(headers):
                                values = [row[idx] if idx < len(row) else None for row in rows]
                                feature = self.feature_analyzer.analyze_column(
                                    column_name=str(col_name),
                                    values=values,
                                    column_index=idx
                                )
                                data_features.append(feature)

                            merged_result["dataFeatures"] = [f.to_dict() for f in data_features]

                            # Map fields
                            field_mappings = self.field_mapping_service.map_fields(
                                headers=[str(h) for h in headers],
                                features=data_features,
                                factory_id=factory_id
                            )
                            merged_result["fieldMappings"] = [fm.to_dict() for fm in field_mappings]

                            # Check if this gives us measures (numeric columns)
                            measure_count = sum(1 for f in data_features if f.data_type == DataType.NUMERIC)
                            if measure_count >= 1:
                                logger.info(f"Merged header parse successful: {measure_count} measures found")
                                return merged_result
                            else:
                                logger.info(f"Merged header parse found no measures, trying other strategies")
            except Exception as e:
                logger.warning(f"parse_with_merged_headers failed: {e}")

        # Step 3: Analyze preview rows to find data start
        data_start_row = self._find_data_start_row(preview_rows)
        logger.info(f"Detected data start row: {data_start_row}")

        # Step 4: Generate parsing strategies to try
        strategies = self._generate_parsing_strategies(
            detected_header_rows, data_start_row, has_multi_header, merged_cells
        )

        # Step 5: Try each strategy and score results
        best_result = None
        best_score = -1

        for strategy in strategies:
            skip_rows = strategy["skip_rows"]
            header_rows = strategy["header_rows"]

            logger.debug(f"Trying strategy: skip_rows={skip_rows}, header_rows={header_rows}")

            try:
                result = self.parse(
                    file_bytes,
                    sheet_name=sheet_name,
                    header_rows=header_rows,
                    skip_rows=skip_rows,
                    transpose=False
                )

                if not result.get("success"):
                    continue

                # Score this result
                score = self._score_parse_result(result)
                logger.debug(f"Strategy score: {score} (skip={skip_rows}, header={header_rows})")

                if score > best_score:
                    best_score = score
                    best_result = result
                    best_result["_parse_strategy"] = strategy
                    best_result["_parse_score"] = score

            except Exception as e:
                logger.warning(f"Strategy failed: skip={skip_rows}, header={header_rows}, error={e}")
                continue

        if best_result is None:
            logger.error("All parsing strategies failed")
            return {
                "success": False,
                "error": "Failed to parse sheet with any strategy",
                "headers": [],
                "rows": [],
                "rowCount": 0,
                "columnCount": 0
            }

        # Step 5: Analyze data features for the best result
        if factory_id or True:  # Always analyze
            headers = best_result.get("headers", [])
            rows = best_result.get("rows", [])

            if headers and rows:
                data_features = []
                for idx, col_name in enumerate(headers):
                    values = [row[idx] if idx < len(row) else None for row in rows]
                    feature = self.feature_analyzer.analyze_column(
                        column_name=str(col_name),
                        values=values,
                        column_index=idx
                    )
                    data_features.append(feature)

                best_result["dataFeatures"] = [f.to_dict() for f in data_features]

                # Map fields
                field_mappings = self.field_mapping_service.map_fields(
                    headers=[str(h) for h in headers],
                    features=data_features,
                    factory_id=factory_id
                )
                best_result["fieldMappings"] = [fm.to_dict() for fm in field_mappings]

                # Count dimensions and measures
                dimension_count = sum(1 for f in data_features if f.data_type in ("CATEGORICAL", "ID", "TEXT"))
                measure_count = sum(1 for f in data_features if f.data_type == "NUMERIC")
                best_result["_dimension_count"] = dimension_count
                best_result["_measure_count"] = measure_count

        logger.info(f"Smart parse complete: score={best_score}, "
                   f"headers={len(best_result.get('headers', []))}, "
                   f"rows={best_result.get('rowCount', 0)}")

        return best_result

    def _find_data_start_row(self, preview_rows: List[List[Any]]) -> int:
        """
        Find the first row that contains actual data (numeric values).

        Args:
            preview_rows: Preview of first N rows from the sheet

        Returns:
            0-indexed row number where data starts
        """
        if not preview_rows:
            return 1

        for row_idx, row in enumerate(preview_rows):
            if not row:
                continue

            # Count numeric values in this row
            numeric_count = 0
            non_empty_count = 0

            for cell in row:
                if cell is not None and str(cell).strip():
                    non_empty_count += 1
                    if self._is_numeric_string(str(cell)):
                        numeric_count += 1

            # Data row: has many numeric values
            if non_empty_count > 0 and numeric_count / non_empty_count > 0.3:
                logger.debug(f"Row {row_idx} looks like data: {numeric_count}/{non_empty_count} numeric")
                return row_idx

        # Default: assume data starts after first row
        return 1

    def _needs_merged_header_parsing(
        self,
        merged_cells: List[Dict],
        has_multi_header: bool,
        detected_header_rows: int,
        preview_rows: List[List[Any]]
    ) -> bool:
        """
        Determine if the sheet needs special merged header parsing.

        This is needed when:
        1. There are merged cells that span columns (creating header groups)
        2. Multiple header rows are detected (3+)
        3. The header structure has repeating patterns (like month groups)

        Args:
            merged_cells: List of merged cell information
            has_multi_header: Whether multi-header was detected
            detected_header_rows: Number of detected header rows
            preview_rows: Preview of first N rows

        Returns:
            True if special merged header parsing is needed
        """
        if not has_multi_header:
            return False

        if detected_header_rows < 2:
            return False

        # Check for column-spanning merged cells in header region
        col_spanning_merges = 0
        for m in merged_cells:
            min_row = m.get("min_row", 1)
            max_col = m.get("max_col", 1)
            min_col = m.get("min_col", 1)

            # Only count merges in header region
            if min_row <= detected_header_rows:
                # Count if merge spans 2+ columns
                if max_col - min_col >= 1:
                    col_spanning_merges += 1

        # Need special handling if many column-spanning merges
        if col_spanning_merges >= 3:
            logger.debug(f"Detected {col_spanning_merges} column-spanning merges in header region")
            return True

        # Check for repeating patterns in headers (like month names)
        if preview_rows and len(preview_rows) >= 2:
            # Check second row for repeating patterns
            second_row = preview_rows[1] if len(preview_rows) > 1 else []
            if second_row:
                # Look for month patterns (1月, 2月, etc.)
                import re
                month_pattern = re.compile(r'(\d{1,2}月|一月|二月|三月|四月|五月|六月|七月|八月|九月|十月|十一月|十二月|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', re.IGNORECASE)
                month_count = sum(1 for v in second_row if v and month_pattern.search(str(v)))

                if month_count >= 3:
                    logger.debug(f"Detected {month_count} month patterns in header row")
                    return True

            # Check for repeating sub-headers (like 预算数, 实际数 repeated)
            if len(preview_rows) > 2:
                third_row = preview_rows[2]
                if third_row:
                    # Count unique values vs total values
                    non_null_values = [str(v) for v in third_row if v is not None]
                    if non_null_values:
                        unique_count = len(set(non_null_values))
                        total_count = len(non_null_values)
                        # If many repeated values, likely sub-headers
                        if total_count >= 6 and unique_count <= total_count / 3:
                            logger.debug(f"Detected repeating sub-headers: {unique_count} unique in {total_count} values")
                            return True

        return False

    def _generate_parsing_strategies(
        self,
        detected_header_rows: int,
        data_start_row: int,
        has_multi_header: bool,
        merged_cells: List[Dict]
    ) -> List[Dict[str, int]]:
        """
        Generate a list of parsing strategies to try.

        Each strategy specifies skip_rows and header_rows parameters.

        Returns:
            List of strategy dictionaries
        """
        strategies = []

        # Strategy 1: Use detected values
        skip = max(0, data_start_row - detected_header_rows)
        strategies.append({
            "skip_rows": skip,
            "header_rows": detected_header_rows,
            "reason": "detected"
        })

        # Strategy 2: Single header, no skip
        strategies.append({
            "skip_rows": 0,
            "header_rows": 1,
            "reason": "default"
        })

        # Strategy 3: If data starts late, try skipping title rows
        if data_start_row >= 3:
            # Skip title rows, use 2 header rows
            strategies.append({
                "skip_rows": data_start_row - 2,
                "header_rows": 2,
                "reason": "skip_titles_2_headers"
            })
            # Skip title rows, use 1 header row
            strategies.append({
                "skip_rows": data_start_row - 1,
                "header_rows": 1,
                "reason": "skip_titles_1_header"
            })

        # Strategy 4: If multi-header detected, try 2 header rows
        if has_multi_header and detected_header_rows >= 2:
            strategies.append({
                "skip_rows": 0,
                "header_rows": 2,
                "reason": "multi_header"
            })

        # Strategy 5: For merged cells spanning rows, adjust skip
        if merged_cells:
            max_merge_row = max((m.get("max_row", 1) for m in merged_cells), default=1)
            if max_merge_row >= 2:
                strategies.append({
                    "skip_rows": max_merge_row,
                    "header_rows": 1,
                    "reason": "after_merge"
                })
                strategies.append({
                    "skip_rows": max_merge_row - 1,
                    "header_rows": 1,
                    "reason": "after_merge_minus_1"
                })

        # Strategy 6: Common Chinese financial report patterns
        # Title (1 row) + Subtitle/Unit (1-2 rows) + Headers (1-2 rows)
        for skip in [1, 2, 3, 4]:
            for headers in [1, 2]:
                strategy = {"skip_rows": skip, "header_rows": headers, "reason": f"pattern_{skip}_{headers}"}
                if strategy not in strategies:
                    strategies.append(strategy)

        # Remove duplicates while preserving order
        seen = set()
        unique_strategies = []
        for s in strategies:
            key = (s["skip_rows"], s["header_rows"])
            if key not in seen:
                seen.add(key)
                unique_strategies.append(s)

        logger.debug(f"Generated {len(unique_strategies)} parsing strategies")
        return unique_strategies

    def _score_parse_result(self, result: Dict[str, Any]) -> float:
        """
        Score a parsing result based on data quality metrics.

        Higher score = better parsing result.

        Scoring criteria:
        - Valid headers (non-empty, non-numeric strings)
        - Number of data rows
        - Presence of dimension columns (categorical)
        - Presence of measure columns (numeric)
        - Data density (fewer empty cells)

        Args:
            result: Parsing result from parse()

        Returns:
            Quality score (0-100)
        """
        if not result.get("success"):
            return 0

        headers = result.get("headers", [])
        rows = result.get("rows", [])

        if not headers or not rows:
            return 0

        score = 0

        # 1. Header quality (0-30 points)
        valid_headers = 0
        for h in headers:
            h_str = str(h).strip()
            if h_str and h_str.lower() not in ("nan", "none", "null", "unnamed"):
                # Check if header looks like a column name (not a numeric value)
                if not self._is_numeric_string(h_str):
                    valid_headers += 1

        header_ratio = valid_headers / len(headers) if headers else 0
        score += header_ratio * 30

        # 2. Row count (0-20 points)
        row_count = len(rows)
        if row_count >= 100:
            score += 20
        elif row_count >= 10:
            score += 15
        elif row_count >= 1:
            score += 10

        # 3. Dimension detection (0-25 points)
        # A good parse should have at least one categorical/text column
        dimension_count = 0
        measure_count = 0

        for col_idx, header in enumerate(headers):
            col_values = [row[col_idx] if col_idx < len(row) else None for row in rows[:50]]
            non_null = [v for v in col_values if v is not None]

            if not non_null:
                continue

            # Check if mostly numeric
            numeric_count = sum(1 for v in non_null if self._is_numeric_string(str(v)))
            numeric_ratio = numeric_count / len(non_null)

            if numeric_ratio > 0.8:
                measure_count += 1
            else:
                # Check uniqueness for dimension
                unique_count = len(set(str(v) for v in non_null))
                if unique_count > 1:
                    dimension_count += 1

        # Good parse has both dimensions and measures
        if dimension_count >= 1 and measure_count >= 1:
            score += 25
        elif dimension_count >= 1 or measure_count >= 1:
            score += 15
        elif dimension_count + measure_count > 0:
            score += 5

        # 4. Data density (0-15 points)
        if rows:
            total_cells = len(headers) * len(rows[:50])
            non_empty_cells = 0
            for row in rows[:50]:
                for i, cell in enumerate(row):
                    if i < len(headers) and cell is not None and str(cell).strip():
                        non_empty_cells += 1

            density = non_empty_cells / total_cells if total_cells > 0 else 0
            score += density * 15

        # 5. First row sanity check (-10 points if first row looks like title)
        if rows:
            first_row = rows[0]
            non_empty_first = sum(1 for c in first_row if c is not None and str(c).strip())
            if non_empty_first <= 2 and len(headers) > 5:
                # First "data" row has very few values - might be parsing titles as data
                score -= 10

        return max(0, score)
