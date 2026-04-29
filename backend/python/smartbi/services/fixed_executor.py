from __future__ import annotations
"""
Fixed Executor Service

Deterministic data extraction engine that processes Excel files
based on JSON configuration from structure detection and semantic mapping.

Part of the Zero-Code SmartBI architecture.

Key principle: This engine does NOT generate code dynamically.
It uses pre-written, tested logic to process data based on configuration.
"""
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

import openpyxl
import pandas as pd
import numpy as np

from services.structure_detector import StructureDetectionResult, ColumnInfo
from services.semantic_mapper import SemanticMappingResult, FieldMapping
from services.context_extractor import ContextExtractor, ContextInfo

logger = logging.getLogger(__name__)


@dataclass
class ExtractedData:
    """Result of data extraction"""
    success: bool = True
    error: Optional[str] = None

    # Extracted data
    headers: List[str] = field(default_factory=list)  # Standard field names
    original_headers: List[str] = field(default_factory=list)  # Original column names
    rows: List[Dict[str, Any]] = field(default_factory=list)  # Data rows as dicts
    row_count: int = 0
    column_count: int = 0

    # Metadata
    data_types: Dict[str, str] = field(default_factory=dict)  # column -> type
    statistics: Dict[str, Dict[str, Any]] = field(default_factory=dict)  # column -> stats

    # Context (Three-Layer Model - Layer 3)
    context: Optional[ContextInfo] = None

    # Processing info
    skipped_rows: int = 0
    processing_notes: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "error": self.error,
            "headers": self.headers,
            "original_headers": self.original_headers,
            "rows": self.rows,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "data_types": self.data_types,
            "statistics": self.statistics,
            "context": self.context.to_dict() if self.context else None,
            "skipped_rows": self.skipped_rows,
            "processing_notes": self.processing_notes
        }

    def to_dataframe(self) -> pd.DataFrame:
        """Convert to pandas DataFrame"""
        if not self.rows:
            return pd.DataFrame()
        return pd.DataFrame(self.rows)


class FixedExecutor:
    """
    Fixed data extraction engine.

    This class uses deterministic, pre-written logic to:
    1. Read data from Excel based on structure config
    2. Apply field mappings to standardize column names
    3. Clean and validate data
    4. Calculate basic statistics

    NO dynamic code generation - all logic is pre-defined and tested.
    """

    def __init__(self):
        self._value_cleaners = {
            "numeric": self._clean_numeric,
            "percentage": self._clean_percentage,
            "currency": self._clean_currency,
            "text": self._clean_text,
            "date": self._clean_date
        }
        self._context_extractor = ContextExtractor()

    def execute(
        self,
        file_bytes: bytes,
        structure_config: StructureDetectionResult,
        mapping_config: SemanticMappingResult,
        options: Optional[Dict[str, Any]] = None
    ) -> ExtractedData:
        """
        Execute data extraction based on configurations.

        Args:
            file_bytes: Raw Excel file bytes
            structure_config: Structure detection result
            mapping_config: Semantic mapping result
            options: Optional extraction options

        Returns:
            ExtractedData with standardized data
        """
        options = options or {}
        result = ExtractedData()

        try:
            # Load workbook
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            # Get target sheet
            sheet_name = structure_config.sheet_name
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Build column mapping
            column_map = self._build_column_map(structure_config, mapping_config)

            # Extract data starting from data_start_row
            data_start_row = structure_config.data_start_row + 1  # openpyxl is 1-indexed
            max_rows = options.get("max_rows", 10000)

            # Get original headers from the last header row
            header_row_num = structure_config.data_start_row  # 0-indexed, so this is the header row
            original_headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row_num + 1, column=col_idx)  # +1 for openpyxl indexing
                header = str(cell.value) if cell.value else f"Column_{col_idx}"
                original_headers.append(header)

            result.original_headers = original_headers

            # Map to standard headers
            standard_headers = []
            for orig in original_headers:
                standard = column_map.get(orig, orig)
                standard_headers.append(standard)
            result.headers = standard_headers

            # Build data type map
            data_type_map = self._build_data_type_map(structure_config, mapping_config, original_headers)
            result.data_types = data_type_map

            # Extract data rows
            rows = []
            skipped = 0

            for row_idx in range(data_start_row + 1, min(data_start_row + max_rows + 1, ws.max_row + 1)):
                row_data = {}
                has_data = False

                for col_idx, (orig_header, std_header) in enumerate(zip(original_headers, standard_headers), start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    if value is not None:
                        has_data = True

                    # Clean value based on data type
                    data_type = data_type_map.get(std_header, "text")
                    cleaned_value = self._clean_value(value, data_type)

                    row_data[std_header] = cleaned_value

                # Skip empty rows if configured
                if not has_data and options.get("skip_empty_rows", True):
                    skipped += 1
                    continue

                rows.append(row_data)

            result.rows = rows
            result.row_count = len(rows)
            result.column_count = len(standard_headers)
            result.skipped_rows = skipped

            # Calculate statistics
            if options.get("calculate_stats", True) and rows:
                result.statistics = self._calculate_statistics(rows, standard_headers, data_type_map)

            # Extract context (Three-Layer Model - Layer 3)
            if options.get("extract_context", True):
                data_end_row = data_start_row + len(rows)
                result.context = self._context_extractor.extract(
                    ws=ws,
                    data_end_row=data_end_row,
                    total_rows=ws.max_row or data_end_row,
                    header_rows=structure_config.header_row_count
                )

            wb.close()

            # Add processing notes
            if mapping_config.unmapped_fields:
                result.processing_notes.append(
                    f"Unmapped fields: {', '.join(mapping_config.unmapped_fields)}"
                )

            if structure_config.note:
                result.processing_notes.append(structure_config.note)

            return result

        except Exception as e:
            logger.error(f"Data extraction failed: {e}", exc_info=True)
            return ExtractedData(success=False, error=str(e))

    def execute_with_pandas(
        self,
        file_bytes: Optional[bytes],
        structure_config: StructureDetectionResult,
        mapping_config: SemanticMappingResult,
        options: Optional[Dict[str, Any]] = None,
        preparsed_df: Optional["pd.DataFrame"] = None,
    ) -> ExtractedData:
        """
        Execute data extraction using pandas for better performance.

        Args:
            file_bytes: Raw Excel file bytes
            structure_config: Structure detection result
            mapping_config: Semantic mapping result
            options: Optional extraction options

        Returns:
            ExtractedData with standardized data
        """
        options = options or {}
        result = ExtractedData()

        try:
            header_rows = structure_config.header_row_count
            data_start_row = structure_config.data_start_row

            # CSV fast-path: when auto_parse used csv_passthrough, file is CSV not xlsx.
            # pd.read_excel would fail; dispatch to pd.read_csv.
            # CRITICAL: pass nrows=max_rows so we don't load a 55MB / 470K-row
            # CSV fully into memory and then ship it to Java (caused repeated
            # prod OOM at 2026-04-15 23:36 — -Xmx1280m heap can't hold 470K
            # Map<String,Object> entries). Default 10000 matches the xlsx path.
            csv_max_rows = options.get("max_rows", 10000)
            # FIX (Apr 15 2026, BUG #4): honour csv_skiprows so CSV files with leading
            # metadata rows (大众点评/美团/客如云 export format) are parsed correctly.
            # auto_parse passes csv_skiprows = max(0, (effective_header_override or 1) - 1).
            csv_skiprows = options.get("csv_skiprows", 0)
            if structure_config.method == "csv_passthrough":
                # Step 1d (Apr 20 2026): caller may pass already-parsed DataFrame
                # so we don't re-parse from 263MB+ content bytes. This is what
                # lets auto_parse_excel `del content` safely after its own
                # read_csv — cuts ~3× 263MB memory duplication on the BG
                # worker self-call path. Validate the df has the expected
                # columns; otherwise fall back to re-parse.
                if preparsed_df is not None and not preparsed_df.empty:
                    df = preparsed_df
                else:
                    if file_bytes is None:
                        return ExtractedData(
                            success=False,
                            error="execute_with_pandas: file_bytes is None and no preparsed_df provided",
                        )
                    try:
                        df = pd.read_csv(io.BytesIO(file_bytes), nrows=csv_max_rows, skiprows=csv_skiprows)
                    except UnicodeDecodeError:
                        df = pd.read_csv(io.BytesIO(file_bytes), encoding="gbk", nrows=csv_max_rows, skiprows=csv_skiprows)
            else:
                # 对于复杂多层表头 (>2行)，使用智能合并而不是 pandas 默认拼接
                if header_rows > 2 or structure_config.merged_cells:
                    # OOM guard: smart_merge 调 openpyxl.load_workbook(data_only=True) 全量加载
                    # 所有 cell 对象 (~500-2000 bytes/cell), 宽文件在 swap-full 服务器上 OOM。
                    # pandas 1.5.3 内部用 read_only=True 按行流式读, 内存安全。
                    # 超出 300K cells → 降级到 pandas 路径 + nrows cap。
                    _should_use_smart = True
                    if file_bytes:
                        try:
                            import openpyxl as _opxl_pre
                            _wb_pre = _opxl_pre.load_workbook(
                                io.BytesIO(file_bytes), read_only=True, data_only=True
                            )
                            _sname_pre = structure_config.sheet_name or _wb_pre.sheetnames[0]
                            _ws_pre = (_wb_pre[_sname_pre]
                                       if _sname_pre in _wb_pre.sheetnames
                                       else _wb_pre[_wb_pre.sheetnames[0]])
                            _pre_rows = _ws_pre.max_row or 0
                            _pre_cols = _ws_pre.max_column or 1
                            _wb_pre.close()
                            _SMART_BUDGET = 300_000
                            logger.info(
                                f"[smart-merge-probe] rows={_pre_rows} cols={_pre_cols} "
                                f"cells={_pre_rows*_pre_cols:,} budget={_SMART_BUDGET:,} "
                                f"file={len(file_bytes)//1024}KB"
                            )
                            _dim_unreliable = _pre_rows <= 1  # xlsx 无 dimension 标签
                            if _pre_rows > 1 and _pre_rows * _pre_cols > _SMART_BUDGET:
                                # 确切维度 × cells 超预算
                                _cap = max(1000, _SMART_BUDGET // max(1, _pre_cols))
                                logger.warning(
                                    f"[smart-merge-oom-guard] {_pre_rows}r × {_pre_cols}c "
                                    f"= {_pre_rows*_pre_cols:,} cells > budget {_SMART_BUDGET:,}. "
                                    f"Downgrade to pandas read_only path, cap={_cap} rows."
                                )
                                options = dict(options)
                                options["nrows"] = _cap
                                _should_use_smart = False
                            elif _dim_unreliable and len(file_bytes) > 300_000:
                                # xlsx 无 dimension tag + 大文件 → openpyxl 全量加载估算不出行数,
                                # 但实际数据量可能很大。用保守 cap 避免 OOM。
                                _cap_safe = 2000
                                logger.warning(
                                    f"[smart-merge-oom-guard] {len(file_bytes)//1024}KB xlsx "
                                    f"+ no dimension tag (max_row={_pre_rows}). "
                                    f"Downgrade to pandas path, cap={_cap_safe} rows."
                                )
                                options = dict(options)
                                options["nrows"] = _cap_safe
                                _should_use_smart = False
                        except Exception as _pre_e:
                            logger.warning(f"[smart-merge-oom-guard] probe failed: {_pre_e}")
                    if _should_use_smart:
                        return self._execute_with_smart_header_merge(
                            file_bytes, structure_config, mapping_config, options
                        )
                    # else: fall through to pandas read_only path with capped nrows

                # 简单表头情况，使用 pandas 默认处理
                if header_rows == 2:
                    header = [0, 1]
                elif header_rows == 1:
                    header = 0
                else:
                    header = data_start_row - 1 if data_start_row > 0 else 0

                # Read with pandas
                # Bug #25b (2026-04-18): honour options["nrows"] to crop to a
                # user-selected region. Default is read-all (historical behaviour).
                _nrows_opt = options.get("nrows")
                # 2026-04-29: cell-budget safety cap (parity with CSV path in
                # excel.py L1053). xlsx 路径之前没有 cap, 1.37MB×4000行×112列
                # 这种文件会让 pd.read_excel 全量加载 → 加 openpyxl XML 解析峰值
                # → 14GB swap-full 服务器上必 OOM。Probe 行数 + 列数 via openpyxl
                # read_only 模式 (cheap), 按 15M cells 截断, 跟 CSV 一致。
                if not _nrows_opt or _nrows_opt <= 0:
                    try:
                        import openpyxl as _opxl_probe
                        _wb_probe = _opxl_probe.load_workbook(
                            io.BytesIO(file_bytes), read_only=True, data_only=True
                        )
                        _sheet_name = structure_config.sheet_name or _wb_probe.sheetnames[0]
                        _ws_probe = _wb_probe[_sheet_name] if _sheet_name in _wb_probe.sheetnames else _wb_probe[_wb_probe.sheetnames[0]]
                        _probe_rows = _ws_probe.max_row or 0
                        _probe_cols = _ws_probe.max_column or 1
                        _wb_probe.close()
                        # xlsx openpyxl Cell 对象开销约 500-2000 bytes/cell (Python 对象开销).
                        # CSV 路径 excel.py 用 15M cells, 但 CSV 用 C 解析器, 开销低 50x.
                        # 300K cells × ~1KB/cell ≈ 300MB 峰值 — 在 swap-full 服务器上安全。
                        # 真实例: 1.37MB xlsx, 4033 rows × 112 cols = 451K cells → OOM。
                        _CELL_BUDGET = 300_000
                        _safety_cap = max(1000, _CELL_BUDGET // max(1, _probe_cols))
                        if _probe_rows > _safety_cap:
                            logger.warning(
                                f"[xlsx-safety-cap] {_probe_rows} rows × {_probe_cols} cols "
                                f"= {_probe_rows*_probe_cols:,} cells exceeds budget "
                                f"({_CELL_BUDGET:,}). Capping to {_safety_cap} rows "
                                f"({100*_safety_cap//_probe_rows}% of data). "
                                f"TODO: streaming persist for full coverage."
                            )
                            _nrows_opt = _safety_cap
                        else:
                            logger.info(
                                f"[xlsx-probe] {_probe_rows} rows × {_probe_cols} cols "
                                f"= {_probe_rows*_probe_cols:,} cells — within budget, full load."
                            )
                    except Exception as _probe_e:
                        logger.warning(f"[xlsx-safety-cap] probe failed: {_probe_e}, reading full file")

                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    sheet_name=structure_config.sheet_name or 0,
                    header=header,
                    skiprows=options.get("skip_rows", 0),
                    nrows=_nrows_opt if _nrows_opt and _nrows_opt > 0 else None,
                )

            # Flatten multi-level columns if needed
            if isinstance(df.columns, pd.MultiIndex):
                import re as _re_unnamed
                _unnamed_re = _re_unnamed.compile(r'^Unnamed:?\s*\d+', _re_unnamed.IGNORECASE)

                def _level_is_noise(cell: str) -> bool:
                    """Pandas injects 'Unnamed: X_level_N' when header row cell is empty
                    or part of a merged-cell span. Treat as noise."""
                    s = str(cell).strip()
                    return (s == 'nan' or s == '' or bool(_unnamed_re.match(s)))

                # Pass 1: flatten keeping only meaningful level parts
                flat_cols = []
                for i, col in enumerate(df.columns.values):
                    parts = [str(c) for c in col if not _level_is_noise(c)]
                    flat_cols.append('_'.join(parts) if parts else f'Column_{i}')

                # Bug #25 true fix (Apr 17 2026): if header detection picked wrong rows,
                # most flat columns will be Column_N placeholders. In that case re-scan
                # the sheet via openpyxl to find the real header row (first row that's
                # >=50% non-empty AND >=80% text cells), then re-read with that header.
                placeholder_count = sum(1 for c in flat_cols if str(c).startswith('Column_'))
                if placeholder_count > len(flat_cols) * 0.5 and structure_config.method != "csv_passthrough":
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
                        sheet = wb[structure_config.sheet_name] if structure_config.sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
                        detected_header_row = None
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True, max_row=20)):
                            cells = list(row)
                            non_empty = [v for v in cells if v is not None and str(v).strip() != '']
                            if len(non_empty) < max(3, len(cells) * 0.5):
                                continue
                            text_vals = [
                                v for v in non_empty
                                if isinstance(v, str) and not str(v).replace('.', '').replace('-', '').replace(',', '').strip().isdigit()
                            ]
                            if len(text_vals) >= len(non_empty) * 0.8:
                                detected_header_row = row_idx
                                logger.info(
                                    f"Bug #25 auto-header-detect: row_idx={row_idx} headers={[str(v)[:20] for v in non_empty[:5]]}"
                                )
                                break
                        wb.close()
                        if detected_header_row is not None:
                            df = pd.read_excel(
                                io.BytesIO(file_bytes),
                                sheet_name=structure_config.sheet_name or 0,
                                header=detected_header_row,
                                skiprows=None,
                            )
                            logger.info(f"Bug #25 re-read success: {len(df.columns)} cols, {len(df)} rows")
                    except Exception as _e:
                        logger.warning(f"Bug #25 auto-header-detect failed (keeping placeholders): {_e}")
                        df.columns = flat_cols
                else:
                    df.columns = flat_cols

            # Store original headers
            result.original_headers = df.columns.tolist()

            # Build column mapping and rename
            column_map = self._build_column_map(structure_config, mapping_config)

            # FIX (Apr 15 2026, BUG #3): if the rename would produce duplicate target names
            # (e.g. 开单时间/分单时间/结单时间 all → "period") OR if csv_passthrough method,
            # PRESERVE original Chinese column names. Restaurant section handlers expect literal
            # business names (开单时间/营业日期/区域 etc), the generic English aliases break them.
            from collections import Counter
            target_names = [column_map.get(orig, orig) for orig in df.columns]
            target_counts = Counter(target_names)
            has_collision = any(c > 1 for c in target_counts.values())
            is_csv_passthrough = structure_config.method == "csv_passthrough"

            if has_collision or is_csv_passthrough:
                logger.info(
                    f"Column rename SKIPPED — preserving original headers ({'collision' if has_collision else 'csv_passthrough'}): {df.columns.tolist()[:5]}..."
                )
                # Don't rename, but still ensure no duplicate original names (rare edge case)
                seen: Dict[str, int] = {}
                new_cols = []
                for c in df.columns:
                    if c in seen:
                        seen[c] += 1
                        new_cols.append(f"{c}_{seen[c]}")
                    else:
                        seen[c] = 1
                        new_cols.append(c)
                df.columns = new_cols
            else:
                renamed_columns = {orig: column_map.get(orig, orig) for orig in df.columns}
                df = df.rename(columns=renamed_columns)
                seen: Dict[str, int] = {}
                new_cols = []
                for c in df.columns:
                    if c in seen:
                        seen[c] += 1
                        new_cols.append(f"{c}_{seen[c]}")
                    else:
                        seen[c] = 1
                        new_cols.append(c)
                df.columns = new_cols

            result.headers = df.columns.tolist()

            # Clean data — iterate by position to avoid duplicate-name issues
            data_type_map = self._build_data_type_map(structure_config, mapping_config, result.original_headers)
            result.data_types = data_type_map

            # Step 1e (Apr 20 2026): skip per-cell clean_value loop for
            # csv_passthrough — pandas CSV values are already JSON-scalar
            # (str/int/float/nan). The lambda apply was doing N*M function
            # calls creating N*M new objects, doubling memory churn for wide
            # CSVs. For xlsx still needed (type coercion, timestamps).
            if structure_config.method != "csv_passthrough":
                for col_idx in range(len(df.columns)):
                    col_name = df.columns[col_idx]
                    data_type = data_type_map.get(col_name, "text")
                    df.iloc[:, col_idx] = df.iloc[:, col_idx].apply(lambda x: self._clean_value(x, data_type))

            # Remove empty rows if configured
            if options.get("skip_empty_rows", True):
                original_count = len(df)
                df = df.dropna(how='all')
                result.skipped_rows = original_count - len(df)

            # Convert to records and ensure all values are JSON-safe scalars
            df = df.replace({np.nan: None})
            raw_rows = df.to_dict(orient='records')
            if structure_config.method == "csv_passthrough":
                # Step 1e: CSV values from pandas are already scalar. Skip the
                # ensure_scalar dict rebuild (halves peak memory for wide CSVs).
                result.rows = raw_rows
            else:
                result.rows = [
                    {k: self._ensure_scalar(v) for k, v in row.items()}
                    for row in raw_rows
                ]
            result.row_count = len(result.rows)
            result.column_count = len(result.headers)

            # Calculate statistics
            if options.get("calculate_stats", True) and result.rows:
                result.statistics = self._calculate_statistics(
                    result.rows, result.headers, data_type_map
                )

            # Extract context (Three-Layer Model - Layer 3)
            # Skip for CSV: context extractor + _get_sheet_index both use openpyxl,
            # which crashes on CSV bytes with "File is not a zip file".
            if options.get("extract_context", True) and structure_config.method != "csv_passthrough":
                result.context = self._context_extractor.extract_from_bytes(
                    file_bytes=file_bytes,
                    sheet_index=0 if not structure_config.sheet_name else
                        self._get_sheet_index(file_bytes, structure_config.sheet_name),
                    data_end_row=structure_config.data_start_row + result.row_count
                )

            return result

        except Exception as e:
            logger.error(f"Pandas extraction failed: {e}", exc_info=True)
            return ExtractedData(success=False, error=str(e))

    def _execute_with_smart_header_merge(
        self,
        file_bytes: bytes,
        structure_config: StructureDetectionResult,
        mapping_config: SemanticMappingResult,
        options: Optional[Dict[str, Any]] = None
    ) -> ExtractedData:
        """
        使用智能表头合并处理复杂多层表头。

        处理逻辑:
        1. 用 openpyxl 读取原始数据
        2. 分析合并单元格，构建表头层级关系
        3. 智能合并表头：只保留有意义的层级（如 "1月_预算数"）
        4. 跳过标题行、单位行等非数据列名行
        """
        options = options or {}
        result = ExtractedData()

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            # 获取目标 sheet
            if structure_config.sheet_name and structure_config.sheet_name in wb.sheetnames:
                ws = wb[structure_config.sheet_name]
            else:
                ws = wb.active

            data_start_row = structure_config.data_start_row
            header_row_count = structure_config.header_row_count
            merged_cells = structure_config.merged_cells

            # 构建合并单元格映射 (用于获取合并区域的值)
            merge_map = self._build_merge_map(ws, merged_cells, header_row_count)

            # 智能合并表头（返回表头和实际数据开始行）
            merged_headers, actual_data_start = self._smart_merge_headers(
                ws, header_row_count, data_start_row, merge_map
            )

            result.original_headers = merged_headers
            logger.info(f"智能合并后的表头: {merged_headers[:5]}...")
            logger.info(f"数据从第 {actual_data_start} 行开始 (1-indexed)")

            # 读取数据行 (actual_data_start 已经是 1-indexed)
            max_rows = options.get("max_rows", 10000)
            rows = []
            skipped = 0

            for row_idx in range(actual_data_start, min(actual_data_start + max_rows, ws.max_row + 1)):
                row_data = {}
                has_data = False

                for col_idx, header in enumerate(merged_headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    if value is not None:
                        has_data = True

                    row_data[header] = value

                if not has_data and options.get("skip_empty_rows", True):
                    skipped += 1
                    continue

                rows.append(row_data)

            wb.close()

            # 应用字段映射
            column_map = self._build_column_map(structure_config, mapping_config)
            mapped_rows = []
            mapped_headers = []

            for orig_header in merged_headers:
                mapped = column_map.get(orig_header, orig_header)
                mapped_headers.append(mapped)

            # Deduplicate mapped headers (e.g. both 日期 and 月份 → period)
            seen_mapped: Dict[str, int] = {}
            deduped_mapped = []
            for h in mapped_headers:
                if h in seen_mapped:
                    seen_mapped[h] += 1
                    deduped_mapped.append(f"{h}_{seen_mapped[h]}")
                else:
                    seen_mapped[h] = 1
                    deduped_mapped.append(h)
            mapped_headers = deduped_mapped

            for row in rows:
                mapped_row = {}
                for idx, orig_header in enumerate(merged_headers):
                    deduped_header = mapped_headers[idx]
                    value = row.get(orig_header)
                    # 清理数据
                    data_type = "numeric" if self._looks_numeric(value) else "text"
                    mapped_row[deduped_header] = self._clean_value(value, data_type)
                mapped_rows.append(mapped_row)

            result.headers = mapped_headers
            result.rows = mapped_rows
            result.row_count = len(mapped_rows)
            result.column_count = len(mapped_headers)
            result.skipped_rows = skipped
            result.processing_notes.append(
                f"智能表头合并: 原始表头行数={header_row_count}, 实际数据起始行={actual_data_start}"
            )

            # 计算统计
            if options.get("calculate_stats", True) and result.rows:
                data_type_map = {h: "numeric" for h in mapped_headers}
                result.statistics = self._calculate_statistics(
                    result.rows, result.headers, data_type_map
                )

            # Extract context (Three-Layer Model - Layer 3)
            if options.get("extract_context", True):
                data_end_row = actual_data_start + len(mapped_rows) - 1
                result.context = self._context_extractor.extract(
                    ws=ws,
                    data_end_row=data_end_row,
                    total_rows=ws.max_row or data_end_row,
                    header_rows=header_row_count
                )

            return result

        except Exception as e:
            logger.error(f"Smart header merge extraction failed: {e}", exc_info=True)
            return ExtractedData(success=False, error=str(e))

    def _build_merge_map(
        self,
        ws,
        merged_cells: List,
        max_row: int
    ) -> Dict[tuple, str]:
        """
        构建合并单元格映射。

        返回: {(row, col): merged_value} 的字典
        """
        merge_map = {}

        for merge_info in merged_cells:
            # 获取合并区域的值（左上角单元格的值）
            min_row = merge_info.min_row
            max_row_merge = merge_info.max_row
            min_col = merge_info.min_col
            max_col = merge_info.max_col

            cell_value = ws.cell(row=min_row, column=min_col).value
            value_str = str(cell_value) if cell_value else ""

            # 将合并区域内所有单元格都映射到这个值
            for r in range(min_row, max_row_merge + 1):
                for c in range(min_col, max_col + 1):
                    merge_map[(r, c)] = value_str

        return merge_map

    def _smart_merge_headers(
        self,
        ws,
        header_row_count: int,
        data_start_row: int,
        merge_map: Dict[tuple, str]
    ) -> tuple:
        """
        智能合并多层表头。

        策略:
        1. 跳过标题行（通常是第1行，跨全宽的合并单元格）
        2. 跳过单位行（包含"单位"字样）
        3. 跳过数据行（数字占比 > 30%，且有足够数据）
        4. 合并有意义的层级（如 "1月" + "预算数" -> "1月_预算数"）
        5. 避免重复（如果子列名已经包含父列名的信息）

        返回:
            tuple: (merged_headers, actual_data_start_row)
        """
        max_col = ws.max_column or 1
        merged_headers = []

        # 识别哪些行是"有意义的表头行"（排除标题行、单位行、数据行）
        meaningful_rows = []
        actual_data_start = None  # 跟踪实际数据开始行（1-indexed）

        # 检查结构检测器给的表头范围内
        for row_idx in range(1, header_row_count + 1):
            row_values = []
            numeric_count = 0
            total_count = 0

            for col_idx in range(1, max_col + 1):
                # 优先使用合并单元格的值
                if (row_idx, col_idx) in merge_map:
                    val = merge_map[(row_idx, col_idx)]
                else:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = str(cell.value) if cell.value else ""

                row_values.append(val)

                # 统计数字比例
                if val and val.strip():
                    total_count += 1
                    try:
                        float(val.replace(',', '').replace('¥', '').replace('%', ''))
                        numeric_count += 1
                    except ValueError:
                        pass

            # 判断是否是标题行
            non_empty = [v for v in row_values if v and v.strip()]
            unique_values = set(non_empty)

            # 标题行特征：
            # 1. 唯一值很少 (<=2) - 表示可能是合并单元格
            # 2. 或者非空值很少 (<=3) 且包含关键词
            # 3. 包含中文报表关键词
            TITLE_KEYWORDS = ["利润表", "资产负债表", "现金流量表", "报表", "汇总表", "明细表", "统计表"]
            has_title_keyword = any(kw in v for v in non_empty for kw in TITLE_KEYWORDS)
            # 如果只有1-2个唯一值，很可能是标题行（合并单元格）
            # Title row if: few unique values AND (contains keyword OR very few non-empty)
            is_title_row = len(unique_values) <= 2 and (has_title_keyword or len(non_empty) <= 3)

            # Debug logging (temporary, use INFO to see in logs)
            if row_idx <= 5:
                logger.info(f"Row {row_idx} CHECK: unique={len(unique_values)}, has_keyword={has_title_keyword}, non_empty={len(non_empty)}, is_title={is_title_row}, first='{non_empty[0][:30] if non_empty else ''}'...")

            is_unit_row = any("单位" in v or "编制" in v or "Unit" in v.lower() for v in non_empty)
            # Date row: few non-empty values with datetime pattern
            is_date_row = len(unique_values) <= 3 and any("00:00:00" in v or (len(v) == 10 and "-" in v) for v in non_empty)

            # 判断是否是数据行（数字占比 > 30%，且有足够多的非空值）
            numeric_ratio = numeric_count / total_count if total_count > 0 else 0
            # 数据行特征：数字比例高，且第一列通常是文本（项目名）
            first_cell = row_values[0] if row_values else ""
            first_is_category = first_cell and not self._looks_numeric(first_cell) and len(first_cell) > 1
            is_data_row = numeric_ratio > 0.3 and total_count > 5 and first_is_category

            if is_data_row:
                # 这是数据行，不是表头！更新实际数据开始行
                if actual_data_start is None:
                    actual_data_start = row_idx
                logger.info(f"Row {row_idx} detected as DATA row (numeric_ratio={numeric_ratio:.2f}, first_cell='{first_cell[:20]}...')")
                continue

            if is_title_row:
                logger.info(f"Row {row_idx} is TITLE row (non_empty={len(non_empty)}, unique={len(unique_values)}), skipping: '{non_empty[0][:30] if non_empty else ''}'")
                continue

            if is_unit_row:
                logger.info(f"Row {row_idx} is UNIT row, skipping")
                continue

            if is_date_row:
                logger.info(f"Row {row_idx} is DATE row (non_empty={len(non_empty)}), skipping")
                continue

            meaningful_rows.append((row_idx, row_values))
            logger.info(f"Row {row_idx} is MEANINGFUL header row (non_empty={len(non_empty)}, unique={len(unique_values)})")

        # 如果检测到数据行在表头范围内，使用检测到的实际数据开始行
        # 否则使用结构检测器提供的 data_start_row
        if actual_data_start is None:
            actual_data_start = data_start_row + 1  # 转为 1-indexed

        logger.info(f"Actual data start row (1-indexed): {actual_data_start}")

        # 如果没有有意义的行，使用实际数据行之前的最后一行
        if not meaningful_rows:
            row_idx = actual_data_start - 1
            if row_idx < 1:
                row_idx = 1
            row_values = []
            for col_idx in range(1, max_col + 1):
                if (row_idx, col_idx) in merge_map:
                    row_values.append(merge_map[(row_idx, col_idx)])
                else:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_values.append(str(cell.value) if cell.value else "")
            meaningful_rows = [(row_idx, row_values)]
            logger.info(f"No meaningful header rows found, using row {row_idx} as header")

        # 合并有意义的行
        for col_idx in range(max_col):
            parts = []
            seen_parts = set()

            for row_idx, row_values in meaningful_rows:
                if col_idx < len(row_values):
                    value = row_values[col_idx].strip()
                    if value and value not in seen_parts:
                        # 避免添加纯数字（可能是数据值泄漏）
                        try:
                            float(value.replace(',', ''))
                            # 是数字，跳过
                            continue
                        except ValueError:
                            pass

                        # 避免添加日期时间戳
                        if "00:00:00" in value:
                            # 提取日期部分
                            value = value.split(" ")[0] if " " in value else value

                        parts.append(value)
                        seen_parts.add(value)

            # 生成最终列名
            if parts:
                # 合并所有层级，保留完整语义 (e.g., "2024年_上半年_收入")
                final_name = "_".join(parts)
            else:
                final_name = f"Column_{col_idx + 1}"

            merged_headers.append(final_name)

        # Deduplicate column names: append _2, _3, ... for duplicates
        seen_names: Dict[str, int] = {}
        deduped = []
        for name in merged_headers:
            if name in seen_names:
                seen_names[name] += 1
                deduped.append(f"{name}_{seen_names[name]}")
            else:
                seen_names[name] = 1
                deduped.append(name)

        return deduped, actual_data_start

    def _looks_numeric(self, value: Any) -> bool:
        """判断值是否看起来像数字"""
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        try:
            float(str(value).replace(',', '').replace('¥', '').replace('%', ''))
            return True
        except (ValueError, TypeError):
            return False

    def _build_column_map(
        self,
        structure_config: StructureDetectionResult,
        mapping_config: SemanticMappingResult
    ) -> Dict[str, str]:
        """Build original -> standard column name mapping"""
        column_map = {}

        for fm in mapping_config.field_mappings:
            if fm.standard:
                column_map[fm.original] = fm.standard
            else:
                # Keep original name if no mapping
                column_map[fm.original] = fm.original

        return column_map

    def _build_data_type_map(
        self,
        structure_config: StructureDetectionResult,
        mapping_config: SemanticMappingResult,
        original_headers: List[str]
    ) -> Dict[str, str]:
        """Build column -> data type mapping"""
        type_map = {}

        # From structure detection
        for col_info in structure_config.columns:
            standard_name = None
            for fm in mapping_config.field_mappings:
                if fm.original == col_info.name:
                    standard_name = fm.standard or fm.original
                    break
            if standard_name:
                type_map[standard_name] = col_info.data_type

        # From semantic mapping (field categories)
        for fm in mapping_config.field_mappings:
            name = fm.standard or fm.original
            if name not in type_map and fm.category:
                # Map category to data type
                category_to_type = {
                    "amount": "numeric",
                    "rate": "percentage",
                    "category": "text",
                    "time": "date"
                }
                type_map[name] = category_to_type.get(fm.category, "text")

        return type_map

    def _ensure_scalar(self, value: Any) -> Any:
        """Convert any non-scalar value to a JSON-safe scalar.

        Handles pandas Series, numpy arrays, Timestamps, and other
        non-primitive types that can leak through pd.read_excel() or
        MultiIndex flattening.
        """
        if value is None:
            return None
        # pandas Series → take first element
        if isinstance(value, pd.Series):
            return self._ensure_scalar(value.iloc[0]) if len(value) > 0 else None
        # pandas Timestamp
        if isinstance(value, pd.Timestamp):
            return value.isoformat() if not pd.isna(value) else None
        # numpy ndarray → take first element (check before hasattr 'item' since ndarray has .item() too)
        if isinstance(value, np.ndarray):
            return self._ensure_scalar(value.flat[0]) if value.size > 0 else None
        # numpy scalar (int64, float64, bool_) — has .item() method
        if hasattr(value, 'item') and not isinstance(value, (str, bytes)):
            try:
                return value.item()
            except (ValueError, OverflowError):
                return None
        # list/tuple → take first element
        if isinstance(value, (list, tuple)):
            return self._ensure_scalar(value[0]) if len(value) > 0 else None
        # dict → stringify
        if isinstance(value, dict):
            return str(value)
        # Check for NaN/NaT
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        return value

    def _clean_value(self, value: Any, data_type: str) -> Any:
        """Clean value based on data type"""
        if value is None:
            return None

        # Ensure we have a scalar before type-specific cleaning
        value = self._ensure_scalar(value)
        if value is None:
            return None

        cleaner = self._value_cleaners.get(data_type, self._clean_text)
        try:
            return cleaner(value)
        except Exception:
            return value

    def _clean_numeric(self, value: Any) -> Optional[float]:
        """Clean numeric value"""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        # String cleaning
        s = str(value).strip()
        if not s or s.lower() in ('nan', 'none', '-', '—', 'n/a'):
            return None

        # Remove formatting
        s = s.replace(',', '').replace(' ', '')
        s = s.replace('¥', '').replace('$', '').replace('€', '').replace('£', '')

        # Handle parentheses for negative
        if s.startswith('(') and s.endswith(')'):
            s = '-' + s[1:-1]

        try:
            return float(s)
        except ValueError:
            return None

    def _clean_percentage(self, value: Any) -> Optional[float]:
        """Clean percentage value (returns decimal, e.g., 0.5 for 50%)"""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            # Already decimal
            if -1 <= value <= 1:
                return float(value)
            # Likely already percentage (e.g., 50 for 50%)
            return float(value) / 100

        s = str(value).strip()
        if not s or s.lower() in ('nan', 'none', '-', '—', 'n/a'):
            return None

        # Remove % sign
        s = s.replace('%', '').replace(',', '').strip()

        try:
            num = float(s)
            # If value > 1 or < -1, assume it's percentage points
            if num > 1 or num < -1:
                return num / 100
            return num
        except ValueError:
            return None

    def _clean_currency(self, value: Any) -> Optional[float]:
        """Clean currency value"""
        # Same as numeric but preserves sign
        return self._clean_numeric(value)

    def _clean_text(self, value: Any) -> Optional[str]:
        """Clean text value"""
        if value is None:
            return None

        s = str(value).strip()
        if not s or s.lower() in ('nan', 'none'):
            return None

        return s

    def _clean_date(self, value: Any) -> Optional[str]:
        """Clean date value (returns ISO format string)"""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.isoformat()

        if isinstance(value, (int, float)):
            # Excel serial date
            try:
                from openpyxl.utils.datetime import from_excel
                return from_excel(value).isoformat()
            except Exception:
                pass

        s = str(value).strip()
        if not s or s.lower() in ('nan', 'none', '-', '—'):
            return None

        # Try common date formats
        formats = [
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%d-%m-%Y', '%d/%m/%Y',
            '%Y年%m月%d日', '%Y年%m月',
            '%m/%d/%Y', '%m-%d-%Y'
        ]

        for fmt in formats:
            try:
                return datetime.strptime(s, fmt).isoformat()
            except ValueError:
                continue

        return s

    def _calculate_statistics(
        self,
        rows: List[Dict[str, Any]],
        headers: List[str],
        data_type_map: Dict[str, str]
    ) -> Dict[str, Dict[str, Any]]:
        """Calculate statistics for each column"""
        stats = {}

        for header in headers:
            values = [row.get(header) for row in rows]
            non_null = [v for v in values if v is not None]

            col_stats = {
                "count": len(non_null),
                "null_count": len(values) - len(non_null),
                "null_ratio": (len(values) - len(non_null)) / len(values) if values else 0
            }

            data_type = data_type_map.get(header, "text")

            if data_type in ("numeric", "percentage", "currency"):
                numeric_values = [v for v in non_null if isinstance(v, (int, float))]
                if numeric_values:
                    col_stats.update({
                        "min": min(numeric_values),
                        "max": max(numeric_values),
                        "sum": sum(numeric_values),
                        "mean": sum(numeric_values) / len(numeric_values),
                        "numeric_count": len(numeric_values)
                    })

            elif data_type == "text":
                if non_null:
                    unique_values = set(str(v) for v in non_null)
                    col_stats.update({
                        "unique_count": len(unique_values),
                        "unique_ratio": len(unique_values) / len(non_null)
                    })

            stats[header] = col_stats

        return stats

    def _get_sheet_index(self, file_bytes: bytes, sheet_name: str) -> int:
        """Get sheet index by name"""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            if sheet_name in wb.sheetnames:
                return wb.sheetnames.index(sheet_name)
            wb.close()
        except Exception:
            pass
        return 0


class DataTransformer:
    """
    Additional data transformation utilities.

    Pre-defined transformations that can be applied to extracted data.
    """

    @staticmethod
    def pivot_time_series(
        data: ExtractedData,
        category_column: str,
        time_columns: List[str],
        value_type: str = "actual"
    ) -> pd.DataFrame:
        """
        Pivot time series data from wide to long format.

        Args:
            data: Extracted data
            category_column: Column containing category names
            time_columns: Columns containing time-series values
            value_type: Name for the value column

        Returns:
            Long-format DataFrame
        """
        df = data.to_dataframe()

        if category_column not in df.columns:
            raise ValueError(f"Category column '{category_column}' not found")

        # Melt the dataframe
        id_vars = [col for col in df.columns if col not in time_columns]
        melted = df.melt(
            id_vars=id_vars,
            value_vars=time_columns,
            var_name='period',
            value_name=value_type
        )

        return melted

    @staticmethod
    def aggregate_by_category(
        data: ExtractedData,
        category_column: str,
        value_columns: List[str],
        agg_func: str = "sum"
    ) -> pd.DataFrame:
        """
        Aggregate data by category.

        Args:
            data: Extracted data
            category_column: Column to group by
            value_columns: Columns to aggregate
            agg_func: Aggregation function (sum, mean, count, etc.)

        Returns:
            Aggregated DataFrame
        """
        df = data.to_dataframe()

        if category_column not in df.columns:
            raise ValueError(f"Category column '{category_column}' not found")

        agg_dict = {col: agg_func for col in value_columns if col in df.columns}
        return df.groupby(category_column).agg(agg_dict).reset_index()

    # calculate_derived_columns removed in Round 8 audit (dead code, df.eval security risk)
