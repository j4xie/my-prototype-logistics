"""QHJ 收入管理报表 v1 — pure-code openpyxl renderer.

Spec: docs/qa-specs/2026-05-12-qhj-revenue-report-design.md §7.1, §7.2
Plan: docs/superpowers/plans/2026-05-12-qhj-revenue-report.md Task F1

Builds the workbook from scratch (no template-file load) so any
number of stores is supported. Visual fidelity to the customer's reference
xlsx (`docs/qa-specs/_assets/qhj_revenue_v1_template.xlsx` when present) is
maintained through the merge-cells / fonts / number-format conventions below.

Returns a BytesIO. Caller streams to client (Phase G API endpoint) or to a
Redis cache; never writes to local disk.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ─── Style constants ────────────────────────────────────────────────────

_HEADER_FONT = Font(name="微软雅黑", size=10, bold=True)
_BODY_FONT = Font(name="微软雅黑", size=10)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_CENTER = Alignment(horizontal="center", vertical="center")

_MONEY_FMT = "#,##0.00"
_RATIO_FMT = "0.00%"
_INT_FMT = "#,##0"
_DECIMAL1_FMT = "#,##0.0"


# ─── Public entry ───────────────────────────────────────────────────────

def render(report_data: dict, labels: dict) -> BytesIO:
    """report_data must match compute_qhj_revenue_report().data shape (spec §6.2)."""
    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"]

    cursor = 3  # leave 2 blank rows up top
    cursor = _write_compare_block(
        ws, cursor, report_data["block1_yoy"], labels,
        block_title=labels["block1_title"],
        prev_label=labels["prev_yoy"],
        ratio_label=labels["ratio_yoy"],
    )
    cursor += 2
    cursor = _write_compare_block(
        ws, cursor, report_data["block2_mom"], labels,
        block_title=labels["block2_title"],
        prev_label=labels["prev_mom"],
        ratio_label=labels["ratio_mom"],
    )
    cursor += 2
    cursor = _write_block3(ws, cursor, report_data["block3_meal_split"], labels)
    cursor += 2
    cursor = _write_block4(ws, cursor, report_data["block4_diner_dist"], labels)

    # Column widths chosen to match customer template visual.
    for col_letter, width in (
        ("A", 4), ("B", 28), ("C", 14), ("D", 14), ("E", 14), ("F", 10),
        ("G", 14), ("H", 14), ("I", 14), ("J", 10),
        ("K", 14), ("L", 14), ("M", 14), ("N", 10),
    ):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Helpers ────────────────────────────────────────────────────────────

def _set_header(cell, text):
    cell.value = text
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    cell.fill = _HEADER_FILL
    cell.border = _BORDER


def _set_body_numeric(cell, value, fmt: str, no_data_label: str):
    """Sets numeric value + format; NULL/None renders as no_data_label."""
    cell.font = _BODY_FONT
    cell.border = _BORDER
    if value is None:
        cell.value = no_data_label
        return
    if fmt == _RATIO_FMT:
        # Spec: backend returns 0-100 (percent) for ratios; convert to 0-1 so
        # Excel "0.00%" cell format renders as expected. For Block 3/4 the
        # value already comes pre-divided (0-1); detect by magnitude.
        cell.value = float(value) / 100 if abs(float(value)) > 1 else float(value)
    else:
        cell.value = value
    cell.number_format = fmt


def _set_body_text(cell, text):
    cell.value = text
    cell.font = _BODY_FONT
    cell.border = _BORDER


# ─── Blocks 1 & 2 (compare blocks: yoy / mom) ───────────────────────────

def _write_compare_block(
    ws, start_row: int, rows: list[dict], labels: dict,
    *, block_title: str, prev_label: str, ratio_label: str,
) -> int:
    """Write a yoy- or mom-style block: 1 title row + 1 label row + N data rows."""
    no_data = labels["no_data"]

    # Row 1: block title + merged column-group labels.
    ws.cell(start_row, 2, block_title).font = _HEADER_FONT
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=3, end_column=5)
    _set_header(ws.cell(start_row, 3), labels["total_summary"])
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=6, end_column=8)
    _set_header(ws.cell(start_row, 6), labels["dine_in"])
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=9, end_column=11)
    _set_header(ws.cell(start_row, 9), labels["takeout"])

    # Row 2: detail headers.
    hdr_row = start_row + 1
    headers = [
        labels["store_name"],
        labels["current"], prev_label, ratio_label,
        labels["current"], prev_label, ratio_label,
        labels["current"], prev_label, ratio_label,
    ]
    for col, label in enumerate(headers, start=2):
        _set_header(ws.cell(hdr_row, col), label)

    # Rows 3..N: data.
    for i, row in enumerate(rows):
        r = hdr_row + 1 + i
        _set_body_text(ws.cell(r, 2), row.get("store_name", ""))
        cells = [
            (row.get("total"),         _MONEY_FMT),
            (row.get("prev_total"),    _MONEY_FMT),
            (row.get("total_ratio"),   _RATIO_FMT),
            (row.get("dine_in"),       _MONEY_FMT),
            (row.get("prev_dine_in"),  _MONEY_FMT),
            (row.get("dine_in_ratio"), _RATIO_FMT),
            (row.get("takeout"),       _MONEY_FMT),
            (row.get("prev_takeout"),  _MONEY_FMT),
            (row.get("takeout_ratio"), _RATIO_FMT),
        ]
        for col, (val, fmt) in enumerate(cells, start=3):
            _set_body_numeric(ws.cell(r, col), val, fmt, no_data)

    return hdr_row + 1 + len(rows)


# ─── Block 3 (meal split) ───────────────────────────────────────────────

def _write_block3(ws, start_row: int, rows: list[dict], labels: dict) -> int:
    no_data = labels["no_data"]

    ws.cell(start_row, 2, labels["block3_title"]).font = _HEADER_FONT

    hdr_row = start_row + 1
    headers = [
        labels["store_name"],
        f"{labels['actual_revenue']}{labels['dine_in']}",
        f"{labels['actual_revenue']}{labels['takeout']}",
        labels["revenue_ratio"],
        f"{labels['bill_count_label']}{labels['dine_in']}",
        f"{labels['bill_count_label']}{labels['takeout']}",
        labels["bill_ratio_label"],
    ]
    for col, label in enumerate(headers, start=2):
        _set_header(ws.cell(hdr_row, col), label)

    for i, row in enumerate(rows):
        r = hdr_row + 1 + i
        _set_body_text(ws.cell(r, 2), row.get("store_name", ""))
        cells = [
            (row.get("dine_in_revenue"),  _MONEY_FMT),
            (row.get("takeout_revenue"),  _MONEY_FMT),
            (row.get("revenue_ratio"),    _RATIO_FMT),
            (row.get("dine_in_bills"),    _INT_FMT),
            (row.get("takeout_bills"),    _INT_FMT),
            (row.get("bill_ratio"),       _RATIO_FMT),
        ]
        for col, (val, fmt) in enumerate(cells, start=3):
            _set_body_numeric(ws.cell(r, col), val, fmt, no_data)

    return hdr_row + 1 + len(rows)


# ─── Block 4 (per-store diner-count distribution; stacked) ──────────────

def _write_block4(ws, start_row: int, store_blocks: list[dict], labels: dict) -> int:
    no_data = labels["no_data"]
    cursor = start_row

    for block in store_blocks:
        title = f"{block['store_name']}  {block['date_range']}"
        ws.cell(cursor, 2, title).font = _HEADER_FONT

        hdr_row = cursor + 1
        headers = [
            labels["diner_count"],
            labels["bill_count"],
            labels["bill_ratio"],
            labels["total_items"],
            labels["avg_items"],
            labels["revenue"],
            labels["revenue_per_diner"],
            labels["revenue_per_item"],
            labels["block4_revenue_ratio"],
        ]
        for col, label in enumerate(headers, start=2):
            _set_header(ws.cell(hdr_row, col), label)

        distribution = block.get("distribution", [])
        for i, d in enumerate(distribution):
            r = hdr_row + 1 + i
            cells = [
                (d.get("diner_count"),         _INT_FMT),
                (d.get("bill_count"),          _INT_FMT),
                (d.get("bill_ratio"),          _RATIO_FMT),
                (d.get("total_items"),         _INT_FMT),
                (d.get("avg_items_per_bill"),  _DECIMAL1_FMT),
                (d.get("revenue"),             _MONEY_FMT),
                (d.get("revenue_per_diner"),   _INT_FMT),
                (d.get("revenue_per_item"),    _INT_FMT),
                (d.get("revenue_ratio"),       _RATIO_FMT),
            ]
            for col, (val, fmt) in enumerate(cells, start=2):
                _set_body_numeric(ws.cell(r, col), val, fmt, no_data)

        # Total row (sums bill_count and revenue per spec §7.2 reference).
        total_row = hdr_row + 1 + len(distribution)
        ws.cell(total_row, 2, labels["total_row"]).font = _HEADER_FONT
        if distribution:
            _set_body_numeric(
                ws.cell(total_row, 3),
                sum(int(d.get("bill_count") or 0) for d in distribution),
                _INT_FMT, no_data,
            )
            _set_body_numeric(
                ws.cell(total_row, 7),
                sum(float(d.get("revenue") or 0) for d in distribution),
                _MONEY_FMT, no_data,
            )
        cursor = total_row + 2

    return cursor
