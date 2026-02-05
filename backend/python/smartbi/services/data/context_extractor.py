from __future__ import annotations
"""
Context Extractor Service

Extracts contextual information from Excel sheets:
- Footer notes (备注)
- Compilation notes (编制说明)
- Term definitions
- Metadata from header area

Part of the Three-Layer Data Model architecture:
Layer 1: Metadata (title, unit, period)
Layer 2: Data (business records)
Layer 3: Context (notes, explanations, definitions)
"""
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class ContextInfo:
    """Extracted context information from Excel sheet"""

    # Notes (备注)
    notes: List[str] = field(default_factory=list)

    # Compilation explanations (编制说明)
    explanations: List[str] = field(default_factory=list)

    # Term definitions extracted from notes/explanations
    definitions: Dict[str, str] = field(default_factory=dict)

    # Source row numbers (for reference)
    source_rows: List[int] = field(default_factory=list)

    # Raw text for vector embedding
    raw_text: str = ""

    # Metadata extracted from header area
    title: Optional[str] = None
    unit: Optional[str] = None
    period: Optional[str] = None
    department: Optional[str] = None
    compiled_by: Optional[str] = None
    compiled_date: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for API response"""
        return {
            "notes": self.notes,
            "explanations": self.explanations,
            "definitions": self.definitions,
            "source_rows": self.source_rows,
            "raw_text": self.raw_text,
            "metadata": {
                "title": self.title,
                "unit": self.unit,
                "period": self.period,
                "department": self.department,
                "compiled_by": self.compiled_by,
                "compiled_date": self.compiled_date
            }
        }

    def has_content(self) -> bool:
        """Check if any context was extracted"""
        return bool(
            self.notes or
            self.explanations or
            self.definitions or
            self.title or
            self.unit
        )

    def to_prompt_text(self) -> str:
        """Format context for LLM prompt injection"""
        parts = []

        if self.title:
            parts.append(f"报表名称: {self.title}")
        if self.unit:
            parts.append(f"数据单位: {self.unit}")
        if self.period:
            parts.append(f"统计周期: {self.period}")
        if self.department:
            parts.append(f"部门: {self.department}")

        if self.notes:
            parts.append("\n备注:")
            for note in self.notes:
                parts.append(f"  - {note}")

        if self.explanations:
            parts.append("\n编制说明:")
            for exp in self.explanations:
                parts.append(f"  {exp}")

        if self.definitions:
            parts.append("\n术语定义:")
            for term, definition in self.definitions.items():
                parts.append(f"  - {term}: {definition}")

        return "\n".join(parts)


class ContextExtractor:
    """
    Extracts contextual information from Excel worksheets.

    Processes:
    1. Header area: title, unit, period, department
    2. Footer area: notes, explanations, definitions
    """

    # Patterns for detecting note/explanation rows
    NOTE_PATTERNS = [
        re.compile(r'^备注[:：]?\s*', re.IGNORECASE),
        re.compile(r'^注[:：]?\s*', re.IGNORECASE),
        re.compile(r'^说明[:：]?\s*', re.IGNORECASE),
        re.compile(r'^Note[:：]?\s*', re.IGNORECASE),
        re.compile(r'^\*\s*'),  # Asterisk prefix
        re.compile(r'^※\s*'),  # Special note marker
    ]

    EXPLANATION_PATTERNS = [
        re.compile(r'^编制说明[:：]?\s*', re.IGNORECASE),
        re.compile(r'^编表说明[:：]?\s*', re.IGNORECASE),
        re.compile(r'^填表说明[:：]?\s*', re.IGNORECASE),
        re.compile(r'^数据说明[:：]?\s*', re.IGNORECASE),
        re.compile(r'^口径说明[:：]?\s*', re.IGNORECASE),
        re.compile(r'^统计口径[:：]?\s*', re.IGNORECASE),
    ]

    # Numbered item pattern (for multi-line explanations)
    NUMBERED_ITEM_PATTERN = re.compile(r'^[（\(]?\s*(\d+)\s*[）\)\.、]\s*(.+)')

    # Unit patterns
    UNIT_PATTERNS = [
        re.compile(r'单位[:：]?\s*(.+)', re.IGNORECASE),
        re.compile(r'Unit[:：]?\s*(.+)', re.IGNORECASE),
        re.compile(r'[（\(]单位[:：]?\s*([^）\)]+)[）\)]'),
        re.compile(r'[（\(]([万亿元%]+)[）\)]'),
    ]

    # Period patterns
    PERIOD_PATTERNS = [
        re.compile(r'(\d{4})年(\d{1,2})月[-—至到](\d{1,2})月'),
        re.compile(r'(\d{4})年(\d{1,2})月'),
        re.compile(r'(\d{4})年度?'),
        re.compile(r'(\d{4})[年\-/](\d{1,2})[月\-/]?'),
        re.compile(r'第([一二三四1234])季度'),
        re.compile(r'Q([1-4])'),
    ]

    # Definition patterns (term: definition)
    DEFINITION_PATTERNS = [
        re.compile(r'["""]?([^""":：]+)["""]?\s*[:：]\s*(.+)'),
        re.compile(r'([^，,]+)\s*(?:是指|指的是|为|即)\s*(.+)'),
    ]

    def __init__(self):
        pass

    def extract(
        self,
        ws: Worksheet,
        data_end_row: int,
        total_rows: int,
        header_rows: int = 1
    ) -> ContextInfo:
        """
        Extract context information from worksheet.

        Args:
            ws: openpyxl worksheet
            data_end_row: Row where data ends (1-indexed)
            total_rows: Total rows in sheet
            header_rows: Number of header rows

        Returns:
            ContextInfo with extracted notes, explanations, definitions
        """
        context = ContextInfo()

        try:
            # Extract metadata from header area
            self._extract_header_metadata(ws, header_rows, context)

            # Extract context from footer area (after data)
            self._extract_footer_context(ws, data_end_row, total_rows, context)

            # Build raw text for embedding
            context.raw_text = self._build_raw_text(context)

            logger.info(
                f"Context extracted: notes={len(context.notes)}, "
                f"explanations={len(context.explanations)}, "
                f"definitions={len(context.definitions)}"
            )

        except Exception as e:
            logger.error(f"Context extraction failed: {e}", exc_info=True)

        return context

    def extract_from_bytes(
        self,
        file_bytes: bytes,
        sheet_index: int = 0,
        data_end_row: Optional[int] = None
    ) -> ContextInfo:
        """
        Extract context from raw file bytes.

        Args:
            file_bytes: Excel file bytes
            sheet_index: Sheet index (0-based)
            data_end_row: Optional data end row (auto-detected if not provided)

        Returns:
            ContextInfo with extracted context
        """
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        if sheet_index < len(wb.sheetnames):
            ws = wb[wb.sheetnames[sheet_index]]
        else:
            ws = wb.active

        total_rows = ws.max_row or 1

        # Auto-detect data end if not provided
        if data_end_row is None:
            data_end_row = self._detect_data_end_row(ws, total_rows)

        context = self.extract(ws, data_end_row, total_rows)
        wb.close()

        return context

    def _extract_header_metadata(
        self,
        ws: Worksheet,
        header_rows: int,
        context: ContextInfo
    ) -> None:
        """Extract metadata from header area"""
        max_col = ws.max_column or 1

        # Scan first few rows for metadata
        scan_rows = min(header_rows + 3, 10)

        for row_idx in range(1, scan_rows + 1):
            row_text = self._get_row_text(ws, row_idx, max_col)
            if not row_text:
                continue

            # Check for title (usually first non-empty row with merged cells or long text)
            if context.title is None and row_idx <= 2:
                # Check if this looks like a title (long text, no unit pattern)
                if len(row_text) > 5 and not any(p.search(row_text) for p in self.UNIT_PATTERNS):
                    # Check if row is mostly this one value (merged cell indicator)
                    non_empty_count = sum(
                        1 for c in range(1, max_col + 1)
                        if ws.cell(row=row_idx, column=c).value
                    )
                    if non_empty_count <= 3:
                        context.title = row_text.strip()
                        continue

            # Check for unit
            if context.unit is None:
                for pattern in self.UNIT_PATTERNS:
                    match = pattern.search(row_text)
                    if match:
                        context.unit = match.group(1).strip()
                        break

            # Check for period
            if context.period is None:
                for pattern in self.PERIOD_PATTERNS:
                    match = pattern.search(row_text)
                    if match:
                        context.period = row_text.strip()
                        break

            # Check for department
            if context.department is None and row_idx <= 3:
                if any(kw in row_text for kw in ['部', '中心', '处', '科', '组', '部门']):
                    context.department = row_text.strip()

    def _extract_footer_context(
        self,
        ws: Worksheet,
        data_end_row: int,
        total_rows: int,
        context: ContextInfo
    ) -> None:
        """Extract notes and explanations from footer area"""
        max_col = ws.max_column or 1

        # Start scanning from after data ends
        start_row = data_end_row + 1
        if start_row > total_rows:
            return

        in_explanation_block = False
        explanation_buffer = []

        for row_idx in range(start_row, total_rows + 1):
            row_text = self._get_row_text(ws, row_idx, max_col)
            if not row_text:
                # Empty row might end an explanation block
                if explanation_buffer:
                    context.explanations.extend(explanation_buffer)
                    explanation_buffer = []
                    in_explanation_block = False
                continue

            context.source_rows.append(row_idx)

            # Check if this starts an explanation block
            for pattern in self.EXPLANATION_PATTERNS:
                if pattern.match(row_text):
                    in_explanation_block = True
                    # Remove the prefix
                    content = pattern.sub('', row_text).strip()
                    if content:
                        explanation_buffer.append(content)
                    break
            else:
                # Check for note patterns
                is_note = False
                for pattern in self.NOTE_PATTERNS:
                    if pattern.match(row_text):
                        is_note = True
                        content = pattern.sub('', row_text).strip()
                        if content:
                            context.notes.append(content)
                            # Try to extract definitions from notes
                            self._extract_definitions_from_text(content, context)
                        break

                if not is_note:
                    # Check for numbered items (continuation of explanation)
                    match = self.NUMBERED_ITEM_PATTERN.match(row_text)
                    if match or in_explanation_block:
                        if match:
                            explanation_buffer.append(row_text.strip())
                        elif in_explanation_block:
                            # Continuation of previous explanation
                            explanation_buffer.append(row_text.strip())

                        # Try to extract definitions
                        self._extract_definitions_from_text(row_text, context)
                    elif self._looks_like_note(row_text):
                        # Heuristic: short text after data might be a note
                        context.notes.append(row_text.strip())

        # Flush remaining buffer
        if explanation_buffer:
            context.explanations.extend(explanation_buffer)

    def _extract_definitions_from_text(
        self,
        text: str,
        context: ContextInfo
    ) -> None:
        """Extract term definitions from text"""
        for pattern in self.DEFINITION_PATTERNS:
            match = pattern.search(text)
            if match:
                term = match.group(1).strip()
                definition = match.group(2).strip()

                # Filter out false positives
                if (
                    len(term) >= 2 and
                    len(definition) >= 4 and
                    not term.isdigit() and
                    '。' not in term
                ):
                    context.definitions[term] = definition

    def _get_row_text(self, ws: Worksheet, row_idx: int, max_col: int) -> str:
        """Get concatenated text from a row"""
        values = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value).strip()
                if val and val.lower() not in ('nan', 'none'):
                    values.append(val)

        return ' '.join(values)

    def _looks_like_note(self, text: str) -> bool:
        """Heuristic check if text looks like a note"""
        if not text:
            return False

        # Note indicators
        note_indicators = [
            '含', '不含', '包含', '不包括', '除', '按', '指', '即',
            '数据', '统计', '口径', '计算', '来源', '说明',
            '其中', '注意', '提示', '*', '※'
        ]

        text_lower = text.lower()
        return any(ind in text or ind in text_lower for ind in note_indicators)

    def _detect_data_end_row(self, ws: Worksheet, total_rows: int) -> int:
        """
        Auto-detect where data ends in the sheet.

        Strategy: Find the last row that contains numeric data in multiple columns.
        """
        max_col = ws.max_column or 1
        last_data_row = total_rows

        # Scan from bottom up to find last data row
        for row_idx in range(total_rows, 0, -1):
            numeric_count = 0
            non_empty_count = 0

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value

                if val is not None:
                    non_empty_count += 1
                    if isinstance(val, (int, float)):
                        numeric_count += 1

            # If row has multiple numeric values, it's likely data
            if numeric_count >= 2 and non_empty_count >= 3:
                last_data_row = row_idx
                break

            # If we've scanned many rows without finding data, use current position
            if total_rows - row_idx > 20:
                break

        return last_data_row

    def _build_raw_text(self, context: ContextInfo) -> str:
        """Build raw text for vector embedding"""
        parts = []

        if context.title:
            parts.append(context.title)

        if context.notes:
            parts.extend(context.notes)

        if context.explanations:
            parts.extend(context.explanations)

        if context.definitions:
            for term, definition in context.definitions.items():
                parts.append(f"{term}: {definition}")

        return "\n".join(parts)

    async def extract_context(
        self,
        file_bytes: bytes,
        sheet_index: int = 0
    ) -> ContextInfo:
        """
        Async wrapper for extract_from_bytes.

        This method provides an async interface for compatibility with
        async analysis pipelines.

        Args:
            file_bytes: Excel file bytes
            sheet_index: Sheet index (0-based)

        Returns:
            ContextInfo with extracted context
        """
        return self.extract_from_bytes(file_bytes, sheet_index)
