"""
Export Validator - LLM 二次核对导出完整性

流程:
1. 规则解析导出数据
2. LLM 核对原始 vs 导出
3. 生成核对报告

核对项:
- 行数是否匹配
- 列数是否匹配
- 关键数据是否准确
- 备注/说明是否保留
- 元信息是否完整
"""

import json
import logging
import io
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class ValidationIssue:
    """验证问题"""
    level: str  # "error" | "warning" | "info"
    category: str  # "row_count" | "column_count" | "data" | "metadata" | "notes"
    message: str
    details: Optional[Dict[str, Any]] = None


@dataclass
class ValidationResult:
    """验证结果"""
    success: bool
    sheet_name: str
    issues: List[ValidationIssue] = field(default_factory=list)
    summary: Dict[str, Any] = field(default_factory=dict)
    llm_review: Optional[str] = None  # LLM 核对意见

    def has_errors(self) -> bool:
        return any(i.level == "error" for i in self.issues)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "sheet_name": self.sheet_name,
            "has_errors": self.has_errors(),
            "issue_count": len(self.issues),
            "issues": [
                {
                    "level": i.level,
                    "category": i.category,
                    "message": i.message,
                    "details": i.details
                }
                for i in self.issues
            ],
            "summary": self.summary,
            "llm_review": self.llm_review
        }


@dataclass
class BatchValidationResult:
    """批量验证结果"""
    success: bool
    total_sheets: int
    validated_sheets: int
    sheets_with_errors: int
    sheets_with_warnings: int
    results: List[ValidationResult] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "total_sheets": self.total_sheets,
            "validated_sheets": self.validated_sheets,
            "sheets_with_errors": self.sheets_with_errors,
            "sheets_with_warnings": self.sheets_with_warnings,
            "results": [r.to_dict() for r in self.results]
        }


class ExportValidator:
    """
    导出验证器 - 规则验证 + LLM 二次核对
    """

    # LLM 核对提示词
    VALIDATION_PROMPT = """你是一个数据核对专家。请对比以下Excel原始数据和导出数据，检查是否有信息丢失或错误。

## Excel 原始信息
- Sheet名: {sheet_name}
- 总行数: {excel_rows}
- 总列数: {excel_cols}
- 前5行预览:
{excel_preview}

- 最后5行预览:
{excel_tail}

## 导出数据信息
- 导出行数: {export_rows}
- 导出列数: {export_cols}
- 元信息: {metadata}
- 前3行数据:
{export_preview}

- 最后3行数据:
{export_tail}

## 请核对以下项目，返回JSON格式:
{{
  "row_count_ok": true/false,
  "row_count_note": "行数核对说明",
  "column_count_ok": true/false,
  "column_count_note": "列数核对说明",
  "metadata_ok": true/false,
  "metadata_note": "元信息核对说明（标题、单位、期间等）",
  "notes_ok": true/false,
  "notes_note": "备注/说明核对（是否保留了编制说明等）",
  "data_accuracy_ok": true/false,
  "data_accuracy_note": "数据准确性说明",
  "overall_ok": true/false,
  "overall_summary": "整体核对结论",
  "issues": ["问题1", "问题2"]
}}

只返回JSON，不要其他内容。"""

    def __init__(self, llm_client=None):
        """
        Args:
            llm_client: LLM客户端，需实现 async chat(prompt) -> str
        """
        self.llm_client = llm_client

    async def validate_export(
        self,
        file_bytes: bytes,
        exported_data: Dict[str, Any],
        sheet_index: int = 0,
        use_llm: bool = True
    ) -> ValidationResult:
        """
        验证单个Sheet的导出完整性

        Args:
            file_bytes: 原始Excel文件
            exported_data: 导出的数据（包含 metadata, columns, rows）
            sheet_index: Sheet索引
            use_llm: 是否使用LLM二次核对

        Returns:
            ValidationResult
        """
        # 加载原始Excel
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        if sheet_index >= len(wb.sheetnames):
            sheet_index = 0
        ws = wb[wb.sheetnames[sheet_index]]

        result = ValidationResult(
            success=True,
            sheet_name=ws.title,
            summary={
                "excel_rows": ws.max_row,
                "excel_cols": ws.max_column,
                "export_rows": exported_data.get("row_count", 0),
                "export_cols": exported_data.get("column_count", 0),
            }
        )

        # 1. 规则验证
        self._rule_validate(ws, exported_data, result)

        # 2. LLM 二次核对
        if use_llm and self.llm_client:
            await self._llm_validate(ws, exported_data, result)

        # 设置最终状态
        result.success = not result.has_errors()

        wb.close()
        return result

    async def validate_batch_export(
        self,
        file_bytes: bytes,
        exported_sheets: List[Dict[str, Any]],
        use_llm: bool = True
    ) -> BatchValidationResult:
        """
        验证批量导出的完整性

        Args:
            file_bytes: 原始Excel文件
            exported_sheets: 导出的所有sheets数据
            use_llm: 是否使用LLM核对

        Returns:
            BatchValidationResult
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        batch_result = BatchValidationResult(
            success=True,
            total_sheets=len(wb.sheetnames),
            validated_sheets=0,
            sheets_with_errors=0,
            sheets_with_warnings=0
        )

        for idx, sheet_data in enumerate(exported_sheets):
            result = await self.validate_export(
                file_bytes,
                sheet_data,
                sheet_index=idx,
                use_llm=use_llm
            )

            batch_result.results.append(result)
            batch_result.validated_sheets += 1

            if result.has_errors():
                batch_result.sheets_with_errors += 1
            elif any(i.level == "warning" for i in result.issues):
                batch_result.sheets_with_warnings += 1

        batch_result.success = batch_result.sheets_with_errors == 0

        wb.close()
        return batch_result

    def _rule_validate(
        self,
        ws,
        exported_data: Dict[str, Any],
        result: ValidationResult
    ):
        """规则验证"""

        # 1. 行数验证
        excel_rows = ws.max_row
        export_rows = exported_data.get("row_count", 0)

        # 找到Excel最后一行有数据的行
        last_data_row = excel_rows
        for row_idx in range(excel_rows, 0, -1):
            has_data = False
            for col_idx in range(1, min(10, ws.max_column + 1)):
                if ws.cell(row=row_idx, column=col_idx).value:
                    has_data = True
                    break
            if has_data:
                last_data_row = row_idx
                break

        # 估算预期数据行（减去表头）
        header_rows = exported_data.get("metadata", {}).get("header_rows", 1)
        expected_min = last_data_row - header_rows - 10  # 允许10行空行误差
        expected_max = last_data_row - header_rows + 5

        if not (expected_min <= export_rows <= expected_max):
            result.issues.append(ValidationIssue(
                level="warning",
                category="row_count",
                message=f"行数差异: Excel最后数据行={last_data_row}, 导出={export_rows}",
                details={
                    "excel_last_data_row": last_data_row,
                    "export_rows": export_rows,
                    "expected_range": [expected_min, expected_max]
                }
            ))

        # 2. 列数验证
        excel_cols = ws.max_column
        export_cols = exported_data.get("column_count", 0)

        if excel_cols != export_cols:
            result.issues.append(ValidationIssue(
                level="warning",
                category="column_count",
                message=f"列数差异: Excel={excel_cols}, 导出={export_cols}",
                details={"excel_cols": excel_cols, "export_cols": export_cols}
            ))

        # 3. 元信息验证
        metadata = exported_data.get("metadata", {})
        if not metadata.get("title"):
            result.issues.append(ValidationIssue(
                level="info",
                category="metadata",
                message="未提取到标题"
            ))

        # 4. 备注验证
        rows = exported_data.get("rows", [])
        has_note = any(
            "说明" in str(list(r.values())[0]) or "备注" in str(list(r.values())[0])
            for r in rows if r
        )

        # 检查Excel是否有备注
        excel_has_note = False
        for row_idx in range(max(1, last_data_row - 10), last_data_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and ("说明" in str(cell_val) or "备注" in str(cell_val)):
                excel_has_note = True
                break

        if excel_has_note and not has_note:
            result.issues.append(ValidationIssue(
                level="error",
                category="notes",
                message="Excel有备注/说明，但导出中缺失"
            ))

        # 5. 首行数据验证
        if rows:
            first_export = list(rows[0].values())[0] if rows[0] else None

            # 找Excel第一行数据
            data_start = metadata.get("data_start_row", header_rows + 1)
            first_excel = ws.cell(row=data_start, column=1).value

            if first_export and first_excel:
                if str(first_export).strip() != str(first_excel).strip():
                    result.issues.append(ValidationIssue(
                        level="warning",
                        category="data",
                        message=f"首行数据不匹配: Excel='{first_excel}', 导出='{first_export}'"
                    ))

        result.summary["issues_count"] = len(result.issues)
        result.summary["has_notes"] = has_note

    async def _llm_validate(
        self,
        ws,
        exported_data: Dict[str, Any],
        result: ValidationResult
    ):
        """LLM 二次核对"""
        try:
            # 准备Excel预览
            excel_preview = []
            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_vals = []
                for col_idx in range(1, min(8, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_vals.append(str(val)[:30] if val else "")
                excel_preview.append(f"Row {row_idx}: {row_vals}")

            # Excel 末尾预览
            excel_tail = []
            for row_idx in range(max(1, ws.max_row - 4), ws.max_row + 1):
                row_vals = []
                for col_idx in range(1, min(8, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_vals.append(str(val)[:30] if val else "")
                if any(row_vals):
                    excel_tail.append(f"Row {row_idx}: {row_vals}")

            # 导出数据预览
            rows = exported_data.get("rows", [])
            export_preview = []
            for i, row in enumerate(rows[:3]):
                vals = [str(v)[:30] if v else "" for v in list(row.values())[:8]]
                export_preview.append(f"Row {i+1}: {vals}")

            export_tail = []
            for i, row in enumerate(rows[-3:]):
                vals = [str(v)[:30] if v else "" for v in list(row.values())[:8]]
                export_tail.append(f"Row {len(rows)-2+i}: {vals}")

            # 构建提示词
            prompt = self.VALIDATION_PROMPT.format(
                sheet_name=ws.title,
                excel_rows=ws.max_row,
                excel_cols=ws.max_column,
                excel_preview="\n".join(excel_preview),
                excel_tail="\n".join(excel_tail),
                export_rows=exported_data.get("row_count", 0),
                export_cols=exported_data.get("column_count", 0),
                metadata=json.dumps(exported_data.get("metadata", {}), ensure_ascii=False),
                export_preview="\n".join(export_preview),
                export_tail="\n".join(export_tail)
            )

            # 调用 LLM
            response = await self.llm_client.chat(prompt)

            # 解析响应
            llm_result = self._parse_llm_response(response)

            result.llm_review = llm_result.get("overall_summary", "")

            # 添加 LLM 发现的问题
            if not llm_result.get("overall_ok", True):
                for issue in llm_result.get("issues", []):
                    result.issues.append(ValidationIssue(
                        level="warning",
                        category="llm_review",
                        message=f"[LLM] {issue}"
                    ))

            # 具体项目检查
            if not llm_result.get("notes_ok", True):
                result.issues.append(ValidationIssue(
                    level="warning",
                    category="notes",
                    message=f"[LLM] {llm_result.get('notes_note', '备注可能不完整')}"
                ))

        except Exception as e:
            logger.error(f"LLM validation failed: {e}")
            result.issues.append(ValidationIssue(
                level="info",
                category="llm_review",
                message=f"LLM核对跳过: {str(e)}"
            ))

    def _parse_llm_response(self, response: str) -> Dict[str, Any]:
        """解析 LLM 响应"""
        import re

        # 尝试提取 JSON
        json_match = re.search(r'```json\s*([\s\S]*?)\s*```', response)
        if json_match:
            response = json_match.group(1)
        else:
            brace_match = re.search(r'\{[\s\S]*\}', response)
            if brace_match:
                response = brace_match.group(0)

        try:
            return json.loads(response)
        except json.JSONDecodeError:
            return {
                "overall_ok": True,
                "overall_summary": "LLM响应解析失败",
                "issues": []
            }


# ============================================================
# 简化的 LLM 客户端接口
# ============================================================

class SimpleLLMClient:
    """
    简化的 LLM 客户端示例

    实际使用时替换为真实的 API 调用
    """

    def __init__(self, api_key: str = None, base_url: str = None, model: str = "gpt-4"):
        self.api_key = api_key
        self.base_url = base_url or "https://api.openai.com/v1"
        self.model = model

    async def chat(self, prompt: str) -> str:
        """
        发送聊天请求

        实际实现需要调用 OpenAI/Claude API
        """
        # 示例实现 - 替换为真实 API 调用
        import httpx

        if not self.api_key:
            raise ValueError("API key not configured")

        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"{self.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": self.model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1
                },
                timeout=60.0
            )

            response.raise_for_status()
            data = response.json()
            return data["choices"][0]["message"]["content"]


# ============================================================
# LLM 补充修复
# ============================================================

class ExportFixer:
    """
    导出修复器 - LLM 发现差异后补充修复
    """

    FIX_PROMPT = """你是一个数据修复专家。以下是Excel原始数据和导出数据的对比，发现了一些差异需要修复。

## 差异问题
{issues}

## Excel 原始数据（相关部分）
{excel_data}

## 当前导出数据
{export_data}

## 请修复差异，返回JSON格式:
{{
  "fixes_applied": [
    {{
      "type": "add_row",
      "position": "末尾",
      "data": {{"列名1": "值1", "列名2": "值2"}}
    }},
    {{
      "type": "update_metadata",
      "field": "title",
      "value": "正确的标题"
    }},
    {{
      "type": "add_note",
      "content": "备注内容"
    }}
  ],
  "fix_summary": "修复说明"
}}

只返回JSON。"""

    def __init__(self, llm_client):
        self.llm_client = llm_client

    async def fix_export(
        self,
        file_bytes: bytes,
        exported_data: Dict[str, Any],
        issues: List[ValidationIssue],
        sheet_index: int = 0
    ) -> Tuple[Dict[str, Any], List[str]]:
        """
        修复导出数据中的差异

        Args:
            file_bytes: 原始Excel
            exported_data: 导出的数据
            issues: 验证发现的问题
            sheet_index: Sheet索引

        Returns:
            (修复后的数据, 修复记录)
        """
        if not issues or not self.llm_client:
            return exported_data, []

        # 只处理需要修复的问题
        fixable_issues = [i for i in issues if i.level in ["error", "warning"]]
        if not fixable_issues:
            return exported_data, []

        # 获取Excel相关数据
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[sheet_index]]

        # 准备问题描述
        issues_text = "\n".join([
            f"- [{i.level}] {i.category}: {i.message}"
            for i in fixable_issues
        ])

        # 准备Excel数据（重点是差异相关的部分）
        excel_data = self._extract_relevant_excel_data(ws, fixable_issues)

        # 准备当前导出数据摘要
        export_summary = {
            "metadata": exported_data.get("metadata", {}),
            "row_count": exported_data.get("row_count", 0),
            "first_row": exported_data.get("rows", [{}])[0] if exported_data.get("rows") else {},
            "last_rows": exported_data.get("rows", [])[-3:] if exported_data.get("rows") else []
        }

        # 调用 LLM 获取修复建议
        prompt = self.FIX_PROMPT.format(
            issues=issues_text,
            excel_data=json.dumps(excel_data, ensure_ascii=False, indent=2),
            export_data=json.dumps(export_summary, ensure_ascii=False, indent=2)
        )

        try:
            response = await self.llm_client.chat(prompt)
            fix_plan = self._parse_fix_response(response)

            # 应用修复
            fixed_data, fix_log = self._apply_fixes(exported_data, fix_plan, ws)

            wb.close()
            return fixed_data, fix_log

        except Exception as e:
            logger.error(f"LLM fix failed: {e}")
            wb.close()
            return exported_data, [f"修复失败: {str(e)}"]

    def _extract_relevant_excel_data(
        self,
        ws,
        issues: List[ValidationIssue]
    ) -> Dict[str, Any]:
        """提取与问题相关的Excel数据"""
        data = {
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column
        }

        # 如果有备注问题，提取末尾数据
        if any(i.category == "notes" for i in issues):
            tail_rows = []
            for row_idx in range(max(1, ws.max_row - 5), ws.max_row + 1):
                row_data = {}
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        row_data[f"col_{col_idx}"] = str(val)[:100]
                if row_data:
                    tail_rows.append({"row": row_idx, "data": row_data})
            data["tail_rows"] = tail_rows

        # 如果有元信息问题，提取表头
        if any(i.category == "metadata" for i in issues):
            header_rows = []
            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_data = {}
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        row_data[f"col_{col_idx}"] = str(val)[:100]
                if row_data:
                    header_rows.append({"row": row_idx, "data": row_data})
            data["header_rows"] = header_rows

        return data

    def _parse_fix_response(self, response: str) -> Dict[str, Any]:
        """解析 LLM 修复响应"""
        import re

        json_match = re.search(r'```json\s*([\s\S]*?)\s*```', response)
        if json_match:
            response = json_match.group(1)
        else:
            brace_match = re.search(r'\{[\s\S]*\}', response)
            if brace_match:
                response = brace_match.group(0)

        try:
            return json.loads(response)
        except json.JSONDecodeError:
            return {"fixes_applied": [], "fix_summary": "解析失败"}

    def _apply_fixes(
        self,
        exported_data: Dict[str, Any],
        fix_plan: Dict[str, Any],
        ws
    ) -> Tuple[Dict[str, Any], List[str]]:
        """应用修复"""
        import copy
        fixed_data = copy.deepcopy(exported_data)
        fix_log = []

        for fix in fix_plan.get("fixes_applied", []):
            fix_type = fix.get("type")

            if fix_type == "add_row":
                # 添加行
                row_data = fix.get("data", {})
                if row_data:
                    fixed_data["rows"].append(row_data)
                    fixed_data["row_count"] = len(fixed_data["rows"])
                    fix_log.append(f"添加行: {list(row_data.values())[0] if row_data else 'empty'}")

            elif fix_type == "update_metadata":
                # 更新元信息
                field = fix.get("field")
                value = fix.get("value")
                if field and value:
                    fixed_data["metadata"][field] = value
                    fix_log.append(f"更新元信息: {field}={value}")

            elif fix_type == "add_note":
                # 添加备注（作为最后一行）
                content = fix.get("content", "")
                if content:
                    # 获取列名
                    cols = fixed_data.get("columns", [])
                    if cols:
                        note_row = {cols[0].get("name", "col_0"): "编制说明："}
                        if len(cols) > 1:
                            note_row[cols[1].get("name", "col_1")] = content
                        fixed_data["rows"].append(note_row)
                        fixed_data["row_count"] = len(fixed_data["rows"])
                        fix_log.append(f"添加备注: {content[:50]}...")

        if fix_plan.get("fix_summary"):
            fix_log.append(f"[总结] {fix_plan['fix_summary']}")

        return fixed_data, fix_log


# ============================================================
# 完整的导出 + 验证 + 修复流程
# ============================================================

async def export_with_validation_and_fix(
    file_bytes: bytes,
    sheet_index: int = 0,
    llm_client=None,
    auto_fix: bool = True
) -> Tuple[Dict[str, Any], ValidationResult, List[str]]:
    """
    导出 + 验证 + 自动修复

    Args:
        file_bytes: Excel文件
        sheet_index: Sheet索引
        llm_client: LLM客户端
        auto_fix: 是否自动修复差异

    Returns:
        (导出数据, 验证结果, 修复记录)
    """
    from services.data_exporter import DataExporter

    # 1. 导出
    exporter = DataExporter()
    data = await exporter.from_excel(file_bytes, sheet_index=sheet_index)

    exported = {
        "metadata": data.metadata,
        "columns": [c.to_dict() for c in data.columns],
        "rows": data.rows,
        "row_count": data.row_count,
        "column_count": data.column_count
    }

    # 2. 验证
    validator = ExportValidator(llm_client=llm_client)
    validation = await validator.validate_export(
        file_bytes,
        exported,
        sheet_index=sheet_index,
        use_llm=llm_client is not None
    )

    fix_log = []

    # 3. 如果有问题且启用自动修复
    if auto_fix and llm_client and validation.issues:
        fixer = ExportFixer(llm_client)
        exported, fix_log = await fixer.fix_export(
            file_bytes,
            exported,
            validation.issues,
            sheet_index=sheet_index
        )

        # 重新验证
        validation = await validator.validate_export(
            file_bytes,
            exported,
            sheet_index=sheet_index,
            use_llm=False  # 修复后用规则验证即可
        )

    return exported, validation, fix_log


# ============================================================
# 完整的导出 + 验证流程
# ============================================================

async def export_with_validation(
    file_bytes: bytes,
    sheet_index: int = 0,
    llm_client=None
) -> Tuple[Dict[str, Any], ValidationResult]:
    """
    导出并验证单个Sheet

    Args:
        file_bytes: Excel文件
        sheet_index: Sheet索引
        llm_client: LLM客户端（可选）

    Returns:
        (导出数据, 验证结果)
    """
    from services.data_exporter import DataExporter

    # 1. 导出
    exporter = DataExporter()
    data = await exporter.from_excel(file_bytes, sheet_index=sheet_index)

    exported = {
        "metadata": data.metadata,
        "columns": [c.to_dict() for c in data.columns],
        "rows": data.rows,
        "row_count": data.row_count,
        "column_count": data.column_count
    }

    # 2. 验证
    validator = ExportValidator(llm_client=llm_client)
    validation = await validator.validate_export(
        file_bytes,
        exported,
        sheet_index=sheet_index,
        use_llm=llm_client is not None
    )

    return exported, validation


async def batch_export_with_validation(
    file_bytes: bytes,
    llm_client=None,
    max_rows_per_md: int = 500
) -> Tuple[Dict[str, Any], BatchValidationResult]:
    """
    批量导出并验证所有Sheets

    Args:
        file_bytes: Excel文件
        llm_client: LLM客户端（可选）
        max_rows_per_md: Markdown最大行数

    Returns:
        (批量导出结果, 批量验证结果)
    """
    from services.data_exporter import BatchExporter

    # 1. 批量导出
    exporter = BatchExporter()
    export_result = await exporter.export_all_sheets(
        file_bytes,
        max_rows_per_md=max_rows_per_md
    )

    # 2. 准备验证数据
    sheets_data = []
    for sheet in export_result.sheets:
        sheets_data.append({
            "metadata": sheet.data.metadata,
            "columns": [c.to_dict() for c in sheet.data.columns],
            "rows": sheet.data.rows,
            "row_count": sheet.data.row_count,
            "column_count": sheet.data.column_count
        })

    # 3. 批量验证
    validator = ExportValidator(llm_client=llm_client)
    validation = await validator.validate_batch_export(
        file_bytes,
        sheets_data,
        use_llm=llm_client is not None
    )

    return export_result, validation
