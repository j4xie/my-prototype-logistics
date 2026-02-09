from __future__ import annotations
"""
Parallel Processor Service

Handles parallel processing of multiple Excel sheets to improve performance.
Target: Reduce processing time from 205s to ~60s for multi-sheet workbooks.

Part of SmartBI Phase 1 Performance Optimization.
"""
import asyncio
import hashlib
import io
import logging
import time
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class SheetPreview:
    """Quick preview of a sheet for filtering"""
    index: int
    name: str
    row_count: int
    column_count: int
    is_empty: bool
    has_data: bool
    preview_headers: List[str] = field(default_factory=list)
    estimated_size: str = "small"  # small, medium, large


@dataclass
class SheetResult:
    """Result of processing a single sheet"""
    index: int
    name: str
    success: bool
    error: Optional[str] = None

    # Structure detection
    structure: Optional[Dict[str, Any]] = None

    # Field mappings
    field_mappings: Optional[List[Dict[str, Any]]] = None
    unmapped_fields: Optional[List[str]] = None
    table_type: Optional[str] = None

    # Data
    headers: List[str] = field(default_factory=list)
    row_count: int = 0
    column_count: int = 0
    preview_data: Optional[List[Dict[str, Any]]] = None
    data_types: Optional[Dict[str, str]] = None
    statistics: Optional[Dict[str, Dict[str, Any]]] = None

    # Context
    context: Optional[Dict[str, Any]] = None

    # Processing info
    detection_method: Optional[str] = None
    confidence: float = 0.0
    processing_time_ms: int = 0
    notes: List[str] = field(default_factory=list)


@dataclass
class IndexSheetMapping:
    """Mapping from index sheet row to report information"""
    index: int
    sheet_name: str  # Original sheet name
    report_name: str  # Display name from index
    description: str = ""  # 编制说明


@dataclass
class IndexMetadata:
    """Metadata extracted from index/TOC sheet"""
    has_index: bool = False
    index_sheet_index: int = -1
    index_sheet_name: str = ""
    sheet_mappings: List[IndexSheetMapping] = field(default_factory=list)

    def get_report_name(self, sheet_index: int) -> Optional[str]:
        """Get report name for a sheet index"""
        for mapping in self.sheet_mappings:
            if mapping.index == sheet_index:
                return mapping.report_name
        return None

    def get_description(self, sheet_index: int) -> str:
        """Get description for a sheet index"""
        for mapping in self.sheet_mappings:
            if mapping.index == sheet_index:
                return mapping.description
        return ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "hasIndex": self.has_index,
            "indexSheetIndex": self.index_sheet_index,
            "indexSheetName": self.index_sheet_name,
            "sheetMappings": [
                {
                    "index": m.index,
                    "sheetName": m.sheet_name,
                    "reportName": m.report_name,
                    "description": m.description,
                }
                for m in self.sheet_mappings
            ],
        }


@dataclass
class WorkbookResult:
    """Result of processing entire workbook"""
    success: bool
    error: Optional[str] = None

    total_sheets: int = 0
    processed_sheets: int = 0
    failed_sheets: int = 0
    skipped_sheets: int = 0

    sheets: List[SheetResult] = field(default_factory=list)

    # Summary
    total_rows: int = 0
    total_processing_time_ms: int = 0
    parallelization_speedup: float = 1.0

    # Metadata
    filename: Optional[str] = None

    # Index sheet metadata
    index_metadata: Optional[IndexMetadata] = None


class ParallelProcessor:
    """
    Parallel processing for multi-sheet Excel workbooks.

    Features:
    - Async parallel processing of sheets
    - Smart sheet filtering (skip empty/index sheets)
    - Resource-aware batching
    - Progress tracking
    """

    def __init__(
        self,
        max_workers: int = 4,
        max_concurrent_llm_calls: int = 2
    ):
        """
        Initialize parallel processor.

        Args:
            max_workers: Maximum parallel workers for CPU-bound tasks
            max_concurrent_llm_calls: Maximum concurrent LLM API calls
        """
        self.max_workers = max_workers
        self.max_concurrent_llm_calls = max_concurrent_llm_calls
        self._llm_semaphore = asyncio.Semaphore(max_concurrent_llm_calls)
        self._executor = ThreadPoolExecutor(max_workers=max_workers)

    async def process_workbook(
        self,
        file_bytes: bytes,
        factory_id: Optional[str] = None,
        sheet_indices: Optional[List[int]] = None,
        skip_empty: bool = True,
        skip_index_sheets: bool = True,
        extract_index_metadata: bool = True,
        max_rows_per_sheet: int = 10000,
        calculate_stats: bool = True,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> WorkbookResult:
        """
        Process all sheets in a workbook in parallel.

        Args:
            file_bytes: Raw Excel file bytes
            factory_id: Optional factory ID for mappings
            sheet_indices: Optional list of specific sheet indices to process
            skip_empty: Skip empty sheets
            skip_index_sheets: Skip index/table-of-contents sheets
            extract_index_metadata: Extract metadata from index sheet before skipping
            max_rows_per_sheet: Maximum rows per sheet
            calculate_stats: Calculate statistics for each sheet
            progress_callback: Optional callback(current, total, sheet_name)

        Returns:
            WorkbookResult with all processed sheets
        """
        start_time = time.time()
        result = WorkbookResult(success=True)

        try:
            # Step 1: Quick scan all sheets (synchronous, fast)
            sheets_preview = await self._scan_sheets(file_bytes)
            result.total_sheets = len(sheets_preview)

            # Step 1.5: Extract index metadata before filtering
            if extract_index_metadata:
                index_metadata = await self._extract_index_metadata(file_bytes, sheets_preview)
                result.index_metadata = index_metadata
                if index_metadata.has_index:
                    logger.info(
                        f"Extracted index metadata: {len(index_metadata.sheet_mappings)} mappings"
                    )

            # Step 2: Filter sheets
            sheets_to_process = self._filter_sheets(
                sheets_preview,
                sheet_indices=sheet_indices,
                skip_empty=skip_empty,
                skip_index_sheets=skip_index_sheets
            )

            result.skipped_sheets = result.total_sheets - len(sheets_to_process)
            logger.info(
                f"Processing {len(sheets_to_process)}/{result.total_sheets} sheets "
                f"(skipped {result.skipped_sheets})"
            )

            if not sheets_to_process:
                result.error = "No sheets to process"
                result.success = False
                return result

            # Step 3: Process sheets in parallel
            tasks = []
            for preview in sheets_to_process:
                task = self._process_single_sheet(
                    file_bytes,
                    preview,
                    factory_id=factory_id,
                    max_rows=max_rows_per_sheet,
                    calculate_stats=calculate_stats
                )
                tasks.append(task)

            # Execute with progress tracking
            completed = 0
            for coro in asyncio.as_completed(tasks):
                try:
                    sheet_result = await coro
                    result.sheets.append(sheet_result)

                    if sheet_result.success:
                        result.processed_sheets += 1
                        result.total_rows += sheet_result.row_count
                    else:
                        result.failed_sheets += 1

                    completed += 1
                    if progress_callback:
                        progress_callback(
                            completed,
                            len(sheets_to_process),
                            sheet_result.name
                        )

                except Exception as e:
                    logger.error(f"Sheet processing error: {e}")
                    result.failed_sheets += 1

            # Sort results by sheet index
            result.sheets.sort(key=lambda s: s.index)

            # Calculate speedup
            total_time = time.time() - start_time
            result.total_processing_time_ms = int(total_time * 1000)

            sequential_estimate = sum(s.processing_time_ms for s in result.sheets)
            if total_time > 0:
                result.parallelization_speedup = sequential_estimate / (total_time * 1000)

            logger.info(
                f"Workbook processed: {result.processed_sheets}/{result.total_sheets} sheets, "
                f"{result.total_rows} total rows, {result.total_processing_time_ms}ms "
                f"(speedup: {result.parallelization_speedup:.1f}x)"
            )

            return result

        except Exception as e:
            logger.error(f"Workbook processing failed: {e}", exc_info=True)
            result.success = False
            result.error = str(e)
            result.total_processing_time_ms = int((time.time() - start_time) * 1000)
            return result

    async def _scan_sheets(self, file_bytes: bytes) -> List[SheetPreview]:
        """
        Quick scan of all sheets in workbook (synchronous, <1s).

        Returns basic info about each sheet for filtering decisions.
        """
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self._executor,
            self._scan_sheets_sync,
            file_bytes
        )

    def _scan_sheets_sync(self, file_bytes: bytes) -> List[SheetPreview]:
        """Synchronous implementation of sheet scanning"""
        sheets = []

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                read_only=True,
                data_only=True
            )

            for index, sheet_name in enumerate(wb.sheetnames):
                try:
                    ws = wb[sheet_name]

                    # Quick dimension check
                    row_count = ws.max_row or 0
                    col_count = ws.max_column or 0

                    # Check if empty
                    is_empty = row_count == 0 or col_count == 0

                    # Get preview headers (first row, first 10 cols)
                    preview_headers = []
                    if not is_empty:
                        for cell in ws.iter_rows(min_row=1, max_row=1, max_col=10, values_only=True):
                            preview_headers = [
                                str(v)[:50] if v is not None else ""
                                for v in cell
                            ]

                    # Estimate size
                    total_cells = row_count * col_count
                    if total_cells < 1000:
                        size = "small"
                    elif total_cells < 100000:
                        size = "medium"
                    else:
                        size = "large"

                    sheets.append(SheetPreview(
                        index=index,
                        name=sheet_name,
                        row_count=row_count,
                        column_count=col_count,
                        is_empty=is_empty,
                        has_data=row_count > 1,  # More than header row
                        preview_headers=preview_headers,
                        estimated_size=size
                    ))

                except Exception as e:
                    logger.warning(f"Failed to scan sheet '{sheet_name}': {e}")
                    sheets.append(SheetPreview(
                        index=index,
                        name=sheet_name,
                        row_count=0,
                        column_count=0,
                        is_empty=True,
                        has_data=False
                    ))

            wb.close()

        except Exception as e:
            logger.error(f"Workbook scan failed: {e}")

        return sheets

    def _filter_sheets(
        self,
        sheets: List[SheetPreview],
        sheet_indices: Optional[List[int]] = None,
        skip_empty: bool = True,
        skip_index_sheets: bool = True
    ) -> List[SheetPreview]:
        """
        Filter sheets based on criteria.

        Args:
            sheets: All sheets from scan
            sheet_indices: Optional specific indices to include
            skip_empty: Skip sheets with no data
            skip_index_sheets: Skip index/TOC sheets

        Returns:
            Filtered list of sheets to process
        """
        filtered = []

        # Index sheet name patterns
        index_patterns = [
            "目录", "索引", "index", "toc", "contents",
            "menu", "导航", "封面", "cover", "sheet list"
        ]

        for sheet in sheets:
            # Check specific indices
            if sheet_indices is not None and sheet.index not in sheet_indices:
                logger.debug(f"Skipping sheet '{sheet.name}': not in requested indices")
                continue

            # Skip empty sheets
            if skip_empty and sheet.is_empty:
                logger.debug(f"Skipping sheet '{sheet.name}': empty")
                continue

            # Skip index sheets
            if skip_index_sheets:
                name_lower = sheet.name.lower()
                is_index = any(pattern in name_lower for pattern in index_patterns)

                # Also check if it's likely an index based on structure
                if not is_index and sheet.has_data:
                    # Index sheets often have very few columns and contain "sheet" references
                    if sheet.column_count <= 3:
                        headers_text = " ".join(sheet.preview_headers).lower()
                        if "sheet" in headers_text or "表" in headers_text:
                            is_index = True

                if is_index:
                    logger.debug(f"Skipping sheet '{sheet.name}': detected as index sheet")
                    continue

            filtered.append(sheet)

        return filtered

    async def _extract_index_metadata(
        self,
        file_bytes: bytes,
        sheets_preview: List[SheetPreview]
    ) -> IndexMetadata:
        """
        Extract metadata from index/TOC sheet.

        Identifies the index sheet and extracts:
        - Report names for each data sheet
        - 编制说明 (compilation notes/descriptions) for each sheet

        Args:
            file_bytes: Raw Excel file bytes
            sheets_preview: List of sheet previews

        Returns:
            IndexMetadata with sheet mappings and descriptions
        """
        index_metadata = IndexMetadata()

        # Index sheet detection patterns
        index_patterns = [
            "目录", "索引", "index", "toc", "contents",
            "menu", "导航", "封面", "cover", "sheet list"
        ]

        # Find the index sheet
        index_sheet = None
        for sheet in sheets_preview:
            name_lower = sheet.name.lower()
            if any(pattern in name_lower for pattern in index_patterns):
                index_sheet = sheet
                break

            # Also check structure-based detection
            if sheet.has_data and sheet.column_count <= 4:
                headers_text = " ".join(sheet.preview_headers).lower()
                if "sheet" in headers_text or "表" in headers_text or "报表" in headers_text:
                    index_sheet = sheet
                    break

        if not index_sheet:
            return index_metadata

        # Read the index sheet content
        try:
            loop = asyncio.get_event_loop()
            index_data = await loop.run_in_executor(
                self._executor,
                self._read_index_sheet,
                file_bytes,
                index_sheet.index
            )

            if index_data is None:
                return index_metadata

            index_metadata.has_index = True
            index_metadata.index_sheet_index = index_sheet.index
            index_metadata.index_sheet_name = index_sheet.name

            # Process index data to extract mappings
            sheet_mappings = self._parse_index_content(index_data, sheets_preview)
            index_metadata.sheet_mappings = sheet_mappings

        except Exception as e:
            logger.warning(f"Failed to extract index metadata: {e}")

        return index_metadata

    def _read_index_sheet(
        self,
        file_bytes: bytes,
        sheet_index: int
    ) -> Optional[pd.DataFrame]:
        """Read index sheet as DataFrame"""
        try:
            import io
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                sheet_name=sheet_index,
                header=None,  # Read raw data
                nrows=100  # Limit rows
            )
            return df
        except Exception as e:
            logger.warning(f"Failed to read index sheet: {e}")
            return None

    def _parse_index_content(
        self,
        df: pd.DataFrame,
        sheets_preview: List[SheetPreview]
    ) -> List[IndexSheetMapping]:
        """
        Parse index sheet content to extract sheet mappings.

        Looks for patterns like:
        - Row with sheet name/number and report name
        - 编制说明 (description) text

        Args:
            df: Index sheet DataFrame
            sheets_preview: List of available sheets

        Returns:
            List of IndexSheetMapping
        """
        import re
        mappings: List[IndexSheetMapping] = []

        # Get all sheet names for matching
        sheet_names = {s.name.lower(): s for s in sheets_preview}
        sheet_indices = {s.index: s for s in sheets_preview}

        # Convert DataFrame to string matrix for analysis
        rows_data = []
        for _, row in df.iterrows():
            row_strings = [str(v).strip() if pd.notna(v) else "" for v in row]
            rows_data.append(row_strings)

        current_description = ""

        for row_idx, row in enumerate(rows_data):
            row_text = " ".join(row)

            # Check if this row references a sheet
            matched_sheet = None
            report_name = ""

            for cell_idx, cell in enumerate(row):
                cell_lower = cell.lower()

                # Check for direct sheet name match
                if cell_lower in sheet_names:
                    matched_sheet = sheet_names[cell_lower]
                    # Report name might be in another column
                    for other_cell in row:
                        if other_cell and other_cell != cell and len(other_cell) > 1:
                            report_name = other_cell
                            break
                    if not report_name:
                        report_name = cell
                    break

                # Check for partial matches (e.g., "Sheet1" -> "sheet1表")
                for sheet_name, sheet in sheet_names.items():
                    if sheet_name in cell_lower or cell_lower in sheet_name:
                        matched_sheet = sheet
                        # Look for report name in same row
                        for other_idx, other_cell in enumerate(row):
                            if other_idx != cell_idx and other_cell and len(other_cell) > 1:
                                # Skip if it looks like a number or index
                                if not re.match(r'^\d+$', other_cell):
                                    report_name = other_cell
                                    break
                        if not report_name:
                            report_name = cell
                        break

                if matched_sheet:
                    break

            # Check if row contains 编制说明 or similar
            if "编制说明" in row_text or "说明" in row_text:
                # Description might be in the same row or following rows
                for cell in row:
                    if cell and "编制说明" not in cell and "说明" not in cell:
                        if len(cell) > 10:  # Likely a description
                            current_description = cell
                            break
                # Check next few rows for description content
                for next_idx in range(row_idx + 1, min(row_idx + 5, len(rows_data))):
                    next_row = rows_data[next_idx]
                    for cell in next_row:
                        if cell and len(cell) > 20:
                            current_description = cell
                            break

            if matched_sheet:
                # Skip the index sheet itself
                if matched_sheet.index != sheets_preview[0].index or matched_sheet.name.lower() not in [p for p in ["目录", "索引", "index", "toc", "contents"]]:
                    mappings.append(IndexSheetMapping(
                        index=matched_sheet.index,
                        sheet_name=matched_sheet.name,
                        report_name=report_name or matched_sheet.name,
                        description=current_description
                    ))
                    current_description = ""  # Reset after use

        # Also add mappings for sheets not explicitly listed but with names
        # that suggest they're data sheets
        mapped_indices = {m.index for m in mappings}
        for sheet in sheets_preview:
            if sheet.index not in mapped_indices:
                # Check if it's not the index sheet
                name_lower = sheet.name.lower()
                is_index = any(p in name_lower for p in ["目录", "索引", "index", "toc", "contents"])
                if not is_index and not sheet.is_empty:
                    mappings.append(IndexSheetMapping(
                        index=sheet.index,
                        sheet_name=sheet.name,
                        report_name=sheet.name,  # Use original name
                        description=""
                    ))

        # Sort by index
        mappings.sort(key=lambda m: m.index)

        return mappings

    async def _process_single_sheet(
        self,
        file_bytes: bytes,
        preview: SheetPreview,
        factory_id: Optional[str] = None,
        max_rows: int = 10000,
        calculate_stats: bool = True
    ) -> SheetResult:
        """
        Process a single sheet with all detection and extraction.

        Uses semaphore to limit concurrent LLM calls.
        """
        start_time = time.time()

        result = SheetResult(
            index=preview.index,
            name=preview.name,
            success=False
        )

        try:
            # Import services here to avoid circular imports
            from services.structure_detector import StructureDetector
            from services.semantic_mapper import SemanticMapper
            from services.fixed_executor import FixedExecutor

            # Initialize services
            detector = StructureDetector()
            mapper = SemanticMapper()
            executor = FixedExecutor()

            # Step 1: Structure detection (may use LLM)
            async with self._llm_semaphore:
                structure = await detector.detect(
                    file_bytes,
                    sheet_index=preview.index,
                    max_header_rows=10
                )

            if not structure.success:
                result.error = structure.error or "Structure detection failed"
                result.processing_time_ms = int((time.time() - start_time) * 1000)
                return result

            result.structure = structure.to_dict()
            result.detection_method = structure.method

            # Step 2: Field mapping (may use LLM)
            columns = [col.name for col in structure.columns]
            sample_data = (
                structure.preview_rows[structure.data_start_row:][:5]
                if structure.preview_rows else None
            )

            async with self._llm_semaphore:
                mapping = await mapper.map_fields(
                    columns=columns,
                    sample_data=sample_data,
                    factory_id=factory_id,
                    table_context=preview.name
                )

            result.field_mappings = [fm.to_dict() for fm in mapping.field_mappings]
            result.unmapped_fields = mapping.unmapped_fields
            result.table_type = mapping.table_type

            # Step 3: Data extraction (CPU-bound, use executor)
            loop = asyncio.get_event_loop()
            extracted = await loop.run_in_executor(
                self._executor,
                lambda: executor.execute_with_pandas(
                    file_bytes,
                    structure,
                    mapping,
                    options={
                        "max_rows": max_rows,
                        "skip_empty_rows": True,
                        "calculate_stats": calculate_stats
                    }
                )
            )

            if not extracted.success:
                result.error = extracted.error or "Data extraction failed"
                result.processing_time_ms = int((time.time() - start_time) * 1000)
                return result

            # Populate result
            result.success = True
            result.headers = extracted.headers
            result.row_count = extracted.row_count
            result.column_count = extracted.column_count
            result.preview_data = extracted.rows[:100]
            result.data_types = extracted.data_types
            result.statistics = extracted.statistics if calculate_stats else None
            result.confidence = min(structure.confidence, mapping.confidence)
            result.notes = extracted.processing_notes.copy()

            if structure.note:
                result.notes.append(structure.note)
            if mapping.note:
                result.notes.append(mapping.note)

            # Context
            if extracted.context and extracted.context.has_content():
                result.context = extracted.context.to_dict()

        except Exception as e:
            logger.error(f"Sheet '{preview.name}' processing failed: {e}", exc_info=True)
            result.error = str(e)

        result.processing_time_ms = int((time.time() - start_time) * 1000)
        return result

    def close(self):
        """Clean up resources"""
        self._executor.shutdown(wait=False)


class StructureFingerprint:
    """
    Generates fingerprints for Excel structure to enable caching.

    Same structure fingerprint = same detection result expected.
    """

    @staticmethod
    def compute(
        columns: List[str],
        dtypes: Optional[Dict[str, str]] = None,
        row_count: int = 0
    ) -> str:
        """
        Compute structure fingerprint.

        Args:
            columns: Column names
            dtypes: Optional column data types
            row_count: Approximate row count (rounded)

        Returns:
            MD5 hash of structure signature
        """
        # Sort columns for consistent ordering
        sorted_cols = sorted(columns)

        # Round row count to nearest power of 10
        if row_count > 0:
            import math
            magnitude = math.floor(math.log10(row_count))
            rounded_rows = 10 ** magnitude
        else:
            rounded_rows = 0

        # Build signature
        signature_parts = [
            f"cols:{','.join(sorted_cols)}",
            f"rows:{rounded_rows}"
        ]

        if dtypes:
            type_signature = ",".join(
                f"{k}:{v}" for k, v in sorted(dtypes.items())
            )
            signature_parts.append(f"types:{type_signature}")

        signature = "|".join(signature_parts)

        return hashlib.md5(signature.encode()).hexdigest()

    @staticmethod
    def compute_from_df(df: pd.DataFrame) -> str:
        """
        Compute fingerprint directly from DataFrame.

        Args:
            df: pandas DataFrame

        Returns:
            MD5 hash fingerprint
        """
        columns = df.columns.tolist()
        dtypes = {str(col): str(df[col].dtype) for col in df.columns}
        row_count = len(df)

        return StructureFingerprint.compute(columns, dtypes, row_count)


# Convenience function for single-sheet processing
async def auto_parse_sheet(
    file_bytes: bytes,
    sheet_index: int = 0,
    factory_id: Optional[str] = None,
    max_rows: int = 10000
) -> SheetResult:
    """
    Convenience function to auto-parse a single sheet.

    Args:
        file_bytes: Excel file bytes
        sheet_index: Sheet index
        factory_id: Optional factory ID
        max_rows: Maximum rows

    Returns:
        SheetResult
    """
    processor = ParallelProcessor(max_workers=2)

    try:
        result = await processor.process_workbook(
            file_bytes,
            factory_id=factory_id,
            sheet_indices=[sheet_index],
            max_rows_per_sheet=max_rows
        )

        if result.sheets:
            return result.sheets[0]
        else:
            return SheetResult(
                index=sheet_index,
                name="Unknown",
                success=False,
                error="No sheets processed"
            )
    finally:
        processor.close()
