from __future__ import annotations
"""
Excel Parser Service

Handles Excel file parsing with support for:
- Multi-level headers
- Data direction detection (row vs column oriented)
- Data type detection (DATE, NUMERIC, CATEGORICAL, ID, TEXT)
- Field mapping with synonym dictionary
- Various Excel formats (.xlsx, .xls, .csv)

Python equivalent of Java's ExcelDynamicParserServiceImpl
"""
import logging
from io import BytesIO
from typing import Any, Optional, Union, List, Dict
from enum import Enum
from datetime import datetime

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

from .data_feature_analyzer import (
    DataFeatureAnalyzer,
    DataFeatureResult,
    DataType,
    count_time_pattern_headers,
    is_time_pattern
)
from .field_mapping import (
    FieldMappingService,
    FieldMappingResult,
    FieldMappingDictionary
)

logger = logging.getLogger(__name__)


class DataDirection(str, Enum):
    """Data orientation in the spreadsheet"""
    ROW_ORIENTED = "row"  # Headers in first row, data in subsequent rows
    COLUMN_ORIENTED = "column"  # Headers in first column, data in subsequent columns
    UNKNOWN = "unknown"


class SheetInfo:
    """Information about a sheet in an Excel file"""

    def __init__(
        self,
        index: int,
        name: str,
        row_count: int = 0,
        column_count: int = 0,
        is_empty: bool = True,
        preview_headers: Optional[List[str]] = None
    ):
        self.index = index
        self.name = name
        self.row_count = row_count
        self.column_count = column_count
        self.is_empty = is_empty
        self.preview_headers = preview_headers or []

    def to_dict(self) -> Dict[str, Any]:
        return {
            "index": self.index,
            "name": self.name,
            "rowCount": self.row_count,
            "columnCount": self.column_count,
            "empty": self.is_empty,
            "previewHeaders": self.preview_headers
        }


class ExcelParser:
    """
    Excel file parser with full dynamic analysis support.

    Features:
    - Multi-level header detection and flattening
    - Data direction detection (row vs column oriented)
    - Data type analysis per column
    - Field mapping with synonym dictionary
    - Sheet enumeration with preview
    """

    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls", ".csv"]
        self.feature_analyzer = DataFeatureAnalyzer()
        self.field_mapping_service = FieldMappingService()

    def parse(
        self,
        file_bytes: bytes,
        sheet_name: Optional[Union[str, int]] = 0,
        header_rows: int = 1,
        skip_rows: int = 0,
        transpose: bool = False
    ) -> Dict[str, Any]:
        """
        Parse Excel file and return structured data

        Args:
            file_bytes: Raw file bytes
            sheet_name: Sheet name or index (default: first sheet)
            header_rows: Number of header rows (1 or 2 for multi-level)
            skip_rows: Number of rows to skip at the top
            transpose: Whether to transpose the data

        Returns:
            Dictionary with headers, rows, and metadata
        """
        try:
            buffer = BytesIO(file_bytes)

            # Determine header parameter
            if header_rows == 2:
                header = [0, 1]
            elif header_rows > 2:
                header = list(range(header_rows))
            else:
                header = 0

            # Read Excel file
            df = pd.read_excel(
                buffer,
                sheet_name=sheet_name,
                header=header,
                skiprows=skip_rows
            )

            # Handle multi-level columns
            if isinstance(df.columns, pd.MultiIndex):
                df = self._flatten_multi_index_columns(df)

            # Transpose if needed
            if transpose:
                df = df.T.reset_index()
                # Use first row as headers after transpose
                df.columns = df.iloc[0].tolist()
                df = df[1:]

            # Detect data direction
            direction = self._detect_data_direction(df)

            # Clean data
            df = self._clean_dataframe(df)

            # Convert to structured format
            headers = df.columns.tolist()
            rows = df.values.tolist()

            # Get sheet names
            buffer.seek(0)
            xl = pd.ExcelFile(buffer)
            sheet_names = xl.sheet_names

            return {
                "success": True,
                "headers": headers,
                "rows": rows,
                "rowCount": len(rows),
                "columnCount": len(headers),
                "direction": direction.value,
                "sheetNames": sheet_names,
                "currentSheet": sheet_name if isinstance(sheet_name, str) else sheet_names[sheet_name] if sheet_name < len(sheet_names) else sheet_names[0]
            }

        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": [],
                "rowCount": 0,
                "columnCount": 0
            }

    def parse_csv(
        self,
        file_bytes: bytes,
        encoding: str = "utf-8",
        delimiter: str = ","
    ) -> Dict[str, Any]:
        """Parse CSV file"""
        try:
            buffer = BytesIO(file_bytes)
            df = pd.read_csv(buffer, encoding=encoding, delimiter=delimiter)
            df = self._clean_dataframe(df)

            return {
                "success": True,
                "headers": df.columns.tolist(),
                "rows": df.values.tolist(),
                "rowCount": len(df),
                "columnCount": len(df.columns),
                "direction": self._detect_data_direction(df).value
            }
        except Exception as e:
            logger.error(f"Failed to parse CSV file: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": []
            }

    def get_sheet_names(self, file_bytes: bytes) -> List[str]:
        """Get all sheet names from Excel file"""
        try:
            buffer = BytesIO(file_bytes)
            xl = pd.ExcelFile(buffer)
            return xl.sheet_names
        except Exception as e:
            logger.error(f"Failed to get sheet names: {e}")
            return []

    def list_sheets_detailed(self, file_bytes: bytes) -> List[SheetInfo]:
        """
        Get detailed information about all sheets in an Excel file.

        Equivalent to Java's listSheets() method.

        Returns:
            List of SheetInfo with row count, column count, and preview headers
        """
        try:
            buffer = BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(buffer, read_only=True, data_only=True)

            sheets = []
            for index, sheet_name in enumerate(workbook.sheetnames):
                try:
                    worksheet = workbook[sheet_name]

                    # Count rows (max 1000 for performance)
                    row_count = 0
                    max_column = 0
                    preview_headers = []

                    # Iterate through rows to count and get dimensions
                    for row_idx, row in enumerate(worksheet.iter_rows(max_row=1000), start=1):
                        # Check if row has any data
                        has_data = any(cell.value is not None for cell in row)
                        if has_data:
                            row_count = row_idx

                            # Track max column count from first 100 rows
                            if row_idx <= 100:
                                col_count = sum(1 for cell in row if cell.value is not None)
                                max_column = max(max_column, col_count)

                            # Get preview headers from first row (max 10 columns)
                            if row_idx == 1:
                                for cell in row[:10]:
                                    if cell.value is not None:
                                        preview_headers.append(str(cell.value))

                    # Determine if sheet is empty
                    is_empty = row_count == 0

                    # If we counted rows, use worksheet dimensions as fallback for column count
                    if not is_empty and max_column == 0:
                        max_column = worksheet.max_column or 0

                    sheet_info = SheetInfo(
                        index=index,
                        name=sheet_name,
                        row_count=row_count,
                        column_count=max_column,
                        is_empty=is_empty,
                        preview_headers=preview_headers
                    )
                    sheets.append(sheet_info)

                except Exception as sheet_error:
                    logger.warning(f"Failed to analyze sheet '{sheet_name}': {sheet_error}")
                    # Add minimal info for failed sheet
                    sheets.append(SheetInfo(
                        index=index,
                        name=sheet_name,
                        row_count=0,
                        column_count=0,
                        is_empty=True,
                        preview_headers=[]
                    ))

            workbook.close()
            return sheets

        except Exception as e:
            logger.error(f"Failed to list sheets detailed: {e}", exc_info=True)
            return []

    def detect_multi_header(self, file_bytes: bytes, sheet_name: Optional[Union[str, int]] = 0) -> Dict[str, Any]:
        """
        Detect if the Excel file has multi-level headers using openpyxl for merged cell detection.

        Returns:
            Dictionary with detection results including:
            - hasMultiHeader: true if merged cells span multiple rows/columns
            - originalHeaders: raw header values before merging
            - recommendedHeaderRows: suggested number of header rows
        """
        try:
            buffer = BytesIO(file_bytes)

            # Use openpyxl for merged cell detection
            workbook = openpyxl.load_workbook(buffer, read_only=False, data_only=True)

            # Get the target sheet
            if isinstance(sheet_name, int):
                if sheet_name < len(workbook.sheetnames):
                    ws = workbook[workbook.sheetnames[sheet_name]]
                else:
                    ws = workbook.active
            else:
                ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

            # Analyze merged cells in header region (first 5 rows)
            merged_ranges = list(ws.merged_cells.ranges)
            header_merged_cells = []
            max_merge_row = 1
            has_multi_header = False

            for merged_range in merged_ranges:
                # Check if merge is in header region (first 5 rows)
                if merged_range.min_row <= 5:
                    header_merged_cells.append({
                        "range": str(merged_range),
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col
                    })

                    # Check if merge spans multiple rows
                    if merged_range.max_row > merged_range.min_row:
                        has_multi_header = True
                        max_merge_row = max(max_merge_row, merged_range.max_row)

                    # Check if merge spans multiple columns in header rows
                    if merged_range.min_row <= 2 and merged_range.max_col > merged_range.min_col:
                        has_multi_header = True

            # Get original headers (raw values from first rows)
            original_headers = []
            preview_rows = []

            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_values = []
                for col_idx in range(1, min(ws.max_column + 1, 51)):  # Max 50 columns
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is not None:
                        row_values.append(str(value))
                    else:
                        row_values.append(None)

                if row_idx == 1:
                    original_headers = row_values
                preview_rows.append(row_values)

            # Determine recommended header rows
            if has_multi_header:
                recommended_header_rows = min(max_merge_row, 3)  # Cap at 3 rows
            else:
                # Fall back to pandas-based detection for non-merged multi-headers
                buffer.seek(0)
                df_raw = pd.read_excel(buffer, sheet_name=sheet_name, header=None, nrows=5)

                if len(df_raw) >= 2:
                    first_row = df_raw.iloc[0]
                    second_row = df_raw.iloc[1]

                    first_row_empty_ratio = first_row.isna().sum() / len(first_row) if len(first_row) > 0 else 0
                    second_row_types = second_row.apply(lambda x: 'str' if isinstance(x, str) else 'num')

                    if first_row_empty_ratio > 0.3 and (second_row_types == 'str').sum() > len(second_row) * 0.5:
                        has_multi_header = True
                        recommended_header_rows = 2
                    else:
                        recommended_header_rows = 1
                else:
                    recommended_header_rows = 1

            workbook.close()

            return {
                "success": True,
                "hasMultiHeader": has_multi_header,
                "isMultiHeader": has_multi_header,  # Alias for backward compatibility
                "recommendedHeaderRows": recommended_header_rows,
                "originalHeaders": original_headers,
                "mergedCells": header_merged_cells,
                "previewRows": preview_rows
            }

        except Exception as e:
            logger.error(f"Failed to detect multi-header: {e}", exc_info=True)
            return {
                "success": False,
                "hasMultiHeader": False,
                "isMultiHeader": False,
                "recommendedHeaderRows": 1,
                "originalHeaders": [],
                "mergedCells": [],
                "previewRows": [],
                "error": str(e)
            }

    def _flatten_multi_index_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Flatten multi-level column index to single level"""
        if isinstance(df.columns, pd.MultiIndex):
            # Join multi-level column names with separator
            df.columns = [
                '_'.join(str(c) for c in col if str(c) != 'nan' and str(c) != '')
                for col in df.columns.values
            ]
        return df

    def _detect_data_direction(self, df: pd.DataFrame) -> DataDirection:
        """
        Detect whether data is row-oriented or column-oriented

        Heuristics:
        - Row-oriented: First row is headers, subsequent rows are data records
        - Column-oriented: First column is headers, subsequent columns are data series
        """
        if df.empty:
            return DataDirection.ROW_ORIENTED

        # Check first column vs first row for date/time patterns
        first_col = df.iloc[:, 0] if len(df.columns) > 0 else pd.Series()
        first_row = df.iloc[0] if len(df) > 0 else pd.Series()

        # Count numeric values
        first_col_numeric = pd.to_numeric(first_col, errors='coerce').notna().sum()
        first_row_numeric = pd.to_numeric(first_row, errors='coerce').notna().sum()

        # If first column has more numeric values, likely column-oriented (time series)
        if first_col_numeric > len(first_col) * 0.7 and first_row_numeric < len(first_row) * 0.3:
            return DataDirection.COLUMN_ORIENTED

        return DataDirection.ROW_ORIENTED

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean dataframe by handling NaN and converting types"""
        # Replace NaN with None for JSON serialization
        df = df.replace({np.nan: None})

        # Convert datetime columns to ISO format strings
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].apply(
                    lambda x: x.isoformat() if pd.notna(x) else None
                )

        return df

    def parse_with_analysis(
        self,
        file_bytes: bytes,
        sheet_name: Optional[Union[str, int]] = 0,
        header_rows: int = 1,
        skip_rows: int = 0,
        transpose: bool = False,
        factory_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Parse Excel file with full data feature analysis and field mapping.

        Enhanced version of parse() that includes:
        - Data feature analysis for each column
        - Field mapping to standard fields
        - Missing required fields detection
        - Enhanced data orientation detection

        Args:
            file_bytes: Raw file bytes
            sheet_name: Sheet name or index (default: first sheet)
            header_rows: Number of header rows (1 or 2 for multi-level)
            skip_rows: Number of rows to skip at the top
            transpose: Whether to transpose the data
            factory_id: Optional factory ID for factory-specific mappings

        Returns:
            Dictionary with headers, rows, metadata, dataFeatures, fieldMappings, etc.
        """
        try:
            buffer = BytesIO(file_bytes)

            # Determine header parameter
            if header_rows == 2:
                header = [0, 1]
            elif header_rows > 2:
                header = list(range(header_rows))
            else:
                header = 0

            # Read Excel file
            df = pd.read_excel(
                buffer,
                sheet_name=sheet_name,
                header=header,
                skiprows=skip_rows
            )

            # Handle multi-level columns
            if isinstance(df.columns, pd.MultiIndex):
                df = self._flatten_multi_index_columns(df)

            # Transpose if needed
            if transpose:
                df = df.T.reset_index()
                # Use first row as headers after transpose
                df.columns = df.iloc[0].tolist()
                df = df[1:]

            # Detect data direction using enhanced method
            direction = self._detect_data_orientation_enhanced(df)

            # Clean data
            df_cleaned = self._clean_dataframe(df.copy())

            # Convert to structured format
            headers = df_cleaned.columns.tolist()
            rows = df_cleaned.values.tolist()

            # Get sheet names
            buffer.seek(0)
            xl = pd.ExcelFile(buffer)
            sheet_names = xl.sheet_names

            # Analyze data features for each column
            data_features: List[DataFeatureResult] = []
            for idx, col_name in enumerate(headers):
                values = self._extract_column_values(df, col_name)
                feature = self.feature_analyzer.analyze_column(
                    column_name=str(col_name),
                    values=values,
                    column_index=idx
                )
                data_features.append(feature)

            # Map fields to standard fields
            field_mappings = self.field_mapping_service.map_fields(
                headers=[str(h) for h in headers],
                features=data_features,
                factory_id=factory_id
            )

            # Determine missing required/recommended fields
            mapped_standard_fields = {
                fm.standard_field for fm in field_mappings
                if fm.standard_field is not None
            }
            missing_required_fields = self.field_mapping_service.dictionary.get_missing_recommended_fields(
                mapped_standard_fields
            )

            current_sheet = (
                sheet_name if isinstance(sheet_name, str)
                else sheet_names[sheet_name] if sheet_name < len(sheet_names)
                else sheet_names[0]
            )

            return {
                "success": True,
                "headers": headers,
                "rows": rows,
                "rowCount": len(rows),
                "columnCount": len(headers),
                "direction": direction.value,
                "sheetNames": sheet_names,
                "currentSheet": current_sheet,
                "dataFeatures": [f.to_dict() for f in data_features],
                "fieldMappings": [fm.to_dict() for fm in field_mappings],
                "missingRequiredFields": missing_required_fields,
                "status": "COMPLETE"  # Dynamic adaptation - always complete
            }

        except Exception as e:
            logger.error(f"Failed to parse Excel file with analysis: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "headers": [],
                "rows": [],
                "rowCount": 0,
                "columnCount": 0,
                "dataFeatures": [],
                "fieldMappings": [],
                "missingRequiredFields": [],
                "status": "COMPLETE"
            }

    def _extract_column_values(self, df: pd.DataFrame, column_name: str) -> List[Any]:
        """
        Extract values from a DataFrame column.

        Args:
            df: The DataFrame
            column_name: Name of the column to extract

        Returns:
            List of column values
        """
        if column_name not in df.columns:
            return []

        column_data = df[column_name]
        values = []

        for value in column_data:
            if pd.isna(value):
                values.append(None)
            elif isinstance(value, (pd.Timestamp,)):
                # Convert datetime to string for consistency
                values.append(value.isoformat() if pd.notna(value) else None)
            else:
                values.append(value)

        return values

    def _detect_data_orientation_enhanced(self, df: pd.DataFrame) -> DataDirection:
        """
        Enhanced data direction detection using time pattern analysis.

        Based on Java implementation logic:
        1. Count time pattern headers using is_time_pattern()
        2. Check if first column is a label column (text, high uniqueness)
        3. Return COLUMN_ORIENTED if many time patterns in headers AND first column is labels

        Args:
            df: The DataFrame to analyze

        Returns:
            DataDirection enum value
        """
        if df.empty:
            return DataDirection.ROW_ORIENTED

        headers = [str(h) for h in df.columns.tolist()]

        # Count time patterns in headers
        time_pattern_count = count_time_pattern_headers(headers)
        time_pattern_ratio = time_pattern_count / len(headers) if headers else 0

        logger.debug(
            f"Time pattern analysis: count={time_pattern_count}, "
            f"total_headers={len(headers)}, ratio={time_pattern_ratio:.2%}"
        )

        # Check if first column is a label column
        first_col_is_label = False
        if len(df.columns) > 0 and len(df) > 0:
            first_col = df.iloc[:, 0]

            # Check if first column is mostly text
            text_count = sum(
                1 for v in first_col
                if v is not None and isinstance(v, str) and not self._is_numeric_string(str(v))
            )
            text_ratio = text_count / len(first_col) if len(first_col) > 0 else 0

            # Check uniqueness of first column
            non_null_values = [v for v in first_col if v is not None and str(v).strip()]
            unique_count = len(set(str(v) for v in non_null_values))
            unique_ratio = unique_count / len(non_null_values) if non_null_values else 0

            # First column is a label column if:
            # - High text ratio (> 50%)
            # - High uniqueness (> 50%)
            first_col_is_label = text_ratio > 0.5 and unique_ratio > 0.5

            logger.debug(
                f"First column analysis: text_ratio={text_ratio:.2%}, "
                f"unique_ratio={unique_ratio:.2%}, is_label={first_col_is_label}"
            )

        # Decision logic (from Java implementation):
        # If many headers are time patterns (> 30%) AND first column is labels
        # -> COLUMN_ORIENTED (time series data)
        if time_pattern_ratio > 0.3 and first_col_is_label:
            logger.info(
                f"Detected COLUMN_ORIENTED data: time_headers={time_pattern_count}, "
                f"first_col_is_label={first_col_is_label}"
            )
            return DataDirection.COLUMN_ORIENTED

        # Fall back to original detection method
        return self._detect_data_direction(df)

    def _is_numeric_string(self, value: str) -> bool:
        """
        Check if a string represents a numeric value.

        Args:
            value: String to check

        Returns:
            True if the string is numeric
        """
        if not value or not value.strip():
            return False

        # Remove common formatting
        cleaned = value.strip().replace(',', '').replace(' ', '')

        # Check for currency symbols and percentage
        if cleaned.startswith(('¥', '$', '€', '£')):
            cleaned = cleaned[1:]
        if cleaned.endswith('%'):
            cleaned = cleaned[:-1]

        try:
            float(cleaned)
            return True
        except ValueError:
            return False
