from __future__ import annotations
"""
Structure Detector Service

Uses LLM (with optional vision capabilities) to detect Excel structure:
- Number of header rows
- Data start row
- Merged cells
- Column types

Part of the Zero-Code SmartBI architecture.
"""
import base64
import hashlib
import io
import json
import logging
from dataclasses import dataclass, field, asdict
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class RowInfo:
    """Information about a single row in the header region"""
    index: int
    type: str  # title, subtitle, period, category, column_names, data
    content: str
    is_empty: bool = False
    merged_count: int = 0


@dataclass
class ColumnInfo:
    """Information about a column"""
    index: int
    name: str
    data_type: str  # text, numeric, date, percentage, currency
    sample_values: List[Any] = field(default_factory=list)


@dataclass
class MergedCellInfo:
    """Information about a merged cell region"""
    range: str
    value: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass
class StructureDetectionResult:
    """Result of structure detection"""
    version: str = "1.0"
    success: bool = True
    confidence: float = 0.0
    method: str = "unknown"  # rule, llm_fast, llm_vl, multi_model

    # Sheet info
    sheet_name: str = ""
    total_rows: int = 0
    total_cols: int = 0

    # Header structure
    header_row_count: int = 1
    data_start_row: int = 1  # 0-indexed
    header_rows: List[RowInfo] = field(default_factory=list)

    # Merged cells
    merged_cells: List[MergedCellInfo] = field(default_factory=list)

    # Column information
    columns: List[ColumnInfo] = field(default_factory=list)

    # Preview data
    preview_rows: List[List[Any]] = field(default_factory=list)

    # Error information
    error: Optional[str] = None
    note: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization"""
        result = {
            "version": self.version,
            "success": self.success,
            "confidence": self.confidence,
            "method": self.method,
            "sheet_info": {
                "name": self.sheet_name,
                "total_rows": self.total_rows,
                "total_cols": self.total_cols
            },
            "header": {
                "row_count": self.header_row_count,
                "data_start_row": self.data_start_row,
                "rows": [asdict(r) for r in self.header_rows]
            },
            "merged_cells": [asdict(m) for m in self.merged_cells],
            "columns": [asdict(c) for c in self.columns],
            "preview_rows": self.preview_rows
        }
        if self.error:
            result["error"] = self.error
        if self.note:
            result["note"] = self.note
        return result


class StructureDetector:
    """
    Detects Excel structure using a layered approach:

    Layer 1: Rule-based detection (fast, free)
    Layer 2: LLM-fast (qwen-turbo, cheap)
    Layer 3: LLM-VL (qwen-vl-max, accurate)
    Layer 4: Multi-model voting (for low confidence cases)
    """

    def __init__(self):
        self._llm_client = None
        self._settings = None

    @property
    def settings(self):
        if self._settings is None:
            from config import get_settings
            self._settings = get_settings()
        return self._settings

    async def detect(
        self,
        file_bytes: bytes,
        sheet_index: int = 0,
        max_header_rows: int = 10,
        force_method: Optional[str] = None
    ) -> StructureDetectionResult:
        """
        Detect Excel structure automatically.

        Args:
            file_bytes: Raw Excel file bytes
            sheet_index: Sheet index to analyze (0-based)
            max_header_rows: Maximum rows to scan for header detection
            force_method: Force a specific detection method (rule, llm_fast, llm_vl)

        Returns:
            StructureDetectionResult with detected structure
        """
        try:
            # Load workbook
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            if sheet_index >= len(wb.sheetnames):
                return StructureDetectionResult(
                    success=False,
                    error=f"Sheet index {sheet_index} out of range (max: {len(wb.sheetnames) - 1})"
                )

            ws = wb[wb.sheetnames[sheet_index]]

            # Get basic info
            sheet_name = ws.title
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0

            # Extract raw data for analysis
            raw_rows = self._extract_raw_rows(ws, max_header_rows + 10)
            merged_cells = self._extract_merged_cells(ws, max_header_rows)

            wb.close()

            # Check for complex header patterns - if complex, skip rules and use LLM directly
            is_complex = self._is_complex_header(raw_rows, merged_cells, total_cols)
            if is_complex and force_method is None:
                logger.info("Complex header detected, using LLM directly")
                result = await self._detect_with_llm_for_complex(
                    raw_rows, merged_cells, sheet_name, total_rows, total_cols
                )
                if result.confidence >= 0.6:  # Lower threshold for complex cases
                    logger.info(f"Complex structure detected via LLM: confidence={result.confidence:.2f}")
                    return result

            # Layer 1: Rule-based detection (for simple cases)
            if force_method is None or force_method == "rule":
                result = self._detect_with_rules(
                    raw_rows, merged_cells, sheet_name, total_rows, total_cols
                )
                if result.confidence >= self.settings.structure_detection_confidence_threshold:
                    logger.info(f"Structure detected via rules: confidence={result.confidence:.2f}")
                    return result

                if force_method == "rule":
                    return result

            # Layer 2: LLM-fast detection
            if force_method is None or force_method == "llm_fast":
                result = await self._detect_with_llm_fast(
                    raw_rows, merged_cells, sheet_name, total_rows, total_cols
                )
                if result.confidence >= self.settings.structure_detection_confidence_threshold:
                    logger.info(f"Structure detected via LLM-fast: confidence={result.confidence:.2f}")
                    return result

                if force_method == "llm_fast":
                    return result

            # Layer 3: LLM-VL detection (vision model)
            if force_method is None or force_method == "llm_vl":
                result = await self._detect_with_llm_vl(
                    file_bytes, sheet_index, raw_rows, merged_cells,
                    sheet_name, total_rows, total_cols
                )
                if result.confidence >= self.settings.structure_detection_confidence_threshold:
                    logger.info(f"Structure detected via LLM-VL: confidence={result.confidence:.2f}")
                    return result

                if force_method == "llm_vl":
                    return result

            # Layer 4: Multi-model enhancement
            if self.settings.enable_multi_model_enhancement:
                result = await self._detect_with_multi_model(
                    file_bytes, sheet_index, raw_rows, merged_cells,
                    sheet_name, total_rows, total_cols
                )
                logger.info(f"Structure detected via multi-model: confidence={result.confidence:.2f}")
                return result

            # Fallback: return best guess with low confidence
            logger.warning("All detection methods failed, returning best guess")
            return self._create_fallback_result(
                raw_rows, merged_cells, sheet_name, total_rows, total_cols
            )

        except Exception as e:
            logger.error(f"Structure detection failed: {e}", exc_info=True)
            return StructureDetectionResult(
                success=False,
                error=str(e)
            )

    def _is_complex_header(
        self,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        total_cols: int
    ) -> bool:
        """
        Detect if the header structure is complex and needs LLM analysis.

        Complex patterns:
        1. Multiple merged cell regions in header area
        2. Date values appearing as column groupings (e.g., "2025-01-01", "2025-02-01")
        3. More than 3 potential header rows before numeric data
        4. Merged cells spanning multiple rows
        """
        if not raw_rows:
            return False

        # Check 1: Multiple merged cells with row span
        row_spanning_merges = [m for m in merged_cells if m.max_row > m.min_row]
        col_spanning_merges = [m for m in merged_cells if m.max_col - m.min_col >= 2]

        if len(row_spanning_merges) >= 2 or len(col_spanning_merges) >= 3:
            logger.debug(f"Complex: row_spanning={len(row_spanning_merges)}, col_spanning={len(col_spanning_merges)}")
            return True

        # Check 2: Date values appearing in header rows (monthly grouping pattern)
        import re
        date_pattern = re.compile(r'\d{4}-\d{2}-\d{2}|\d{4}/\d{2}/\d{2}|\d{4}年\d{1,2}月')
        date_header_rows = 0

        for row_idx, row in enumerate(raw_rows[:8]):
            if row_idx == 0:
                continue  # Skip title row
            date_count = sum(1 for v in row if v and date_pattern.search(str(v)))
            if date_count >= 3:  # Multiple dates in one row = monthly grouping
                date_header_rows += 1

        if date_header_rows >= 1:
            logger.debug(f"Complex: date grouping pattern detected in {date_header_rows} rows")
            return True

        # Check 3: Many potential header rows before data
        first_data_row = -1
        for row_idx, row in enumerate(raw_rows[:10]):
            numeric_count = sum(1 for v in row if v and self._is_numeric(str(v)))
            if numeric_count >= len(row) * 0.4 and numeric_count >= 5:
                first_data_row = row_idx
                break

        if first_data_row >= 5:
            logger.debug(f"Complex: {first_data_row} rows before data")
            return True

        return False

    def _is_numeric(self, value: str) -> bool:
        """Check if value is numeric"""
        try:
            float(value.replace(',', '').replace('¥', '').replace('$', '').replace('%', ''))
            return True
        except (ValueError, TypeError):
            return False

    def _extract_raw_rows(
        self, ws, max_rows: int = 20
    ) -> List[List[Any]]:
        """Extract raw row data from worksheet"""
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=0):
            row_values = []
            for cell_value in row:
                if cell_value is not None:
                    row_values.append(str(cell_value))
                else:
                    row_values.append(None)
            rows.append(row_values)
        return rows

    def _extract_merged_cells(
        self, ws, max_rows: int = 10
    ) -> List[MergedCellInfo]:
        """Extract merged cell information from header region"""
        merged = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= max_rows:
                # Get value from top-left cell
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                value = str(cell.value) if cell.value is not None else ""

                merged.append(MergedCellInfo(
                    range=str(merged_range),
                    value=value,
                    min_row=merged_range.min_row,
                    max_row=merged_range.max_row,
                    min_col=merged_range.min_col,
                    max_col=merged_range.max_col
                ))
        return merged

    def _detect_with_rules(
        self,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        sheet_name: str,
        total_rows: int,
        total_cols: int
    ) -> StructureDetectionResult:
        """
        Rule-based structure detection (Layer 1).

        Rules:
        1. Title row: Single merged cell spanning most columns
        2. Subtitle row: Contains units like "单位:", "Unit:"
        3. Period row: Contains date patterns like "2025年", "1月-12月"
        4. Category row: Merged cells creating column groups
        5. Column names row: Many unique values, no merges spanning columns

        注意: 规则检测只能识别结构，不能智能合并表头。
        对于复杂多层表头（>2行或有跨行合并），置信度会降低以触发 LLM 层。
        """
        result = StructureDetectionResult(
            method="rule",
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols
        )

        if not raw_rows:
            result.confidence = 0.3
            result.header_row_count = 1
            result.data_start_row = 1
            return result

        # Analyze each row in the header region
        header_rows = []
        data_start_row = 0
        confidence_factors = []

        for row_idx, row in enumerate(raw_rows[:10]):
            non_null_count = sum(1 for v in row if v is not None)
            unique_values = set(v for v in row if v is not None)

            # Check for merged cells in this row
            row_merges = [m for m in merged_cells if m.min_row == row_idx + 1]
            wide_merges = [m for m in row_merges if m.max_col - m.min_col >= total_cols * 0.5]

            row_type = "unknown"

            # Rule: Title row - single wide merge
            if len(wide_merges) == 1 and non_null_count <= 3:
                row_type = "title"
                confidence_factors.append(0.9)

            # Rule: Subtitle row - contains "单位" or similar
            elif any(v and ("单位" in str(v) or "Unit" in str(v).lower()) for v in row if v):
                row_type = "subtitle"
                confidence_factors.append(0.85)

            # Rule: Period row - contains year/month patterns
            elif any(v and self._is_period_pattern(str(v)) for v in row if v):
                row_type = "period"
                confidence_factors.append(0.85)

            # Rule: Category row - has merged cells but not spanning full width
            elif row_merges and not wide_merges and non_null_count <= total_cols * 0.5:
                row_type = "category"
                confidence_factors.append(0.8)

            # Rule: Data row - numeric values present (check this BEFORE column_names)
            # A row with >40% numeric values is data, not a header
            if self._has_numeric_values(row):
                row_type = "data"
                if data_start_row == 0:
                    data_start_row = row_idx
                break

            # Rule: Column names row - many unique values, high fill rate
            # BUT only if it's not mostly numeric (already checked above)
            elif non_null_count >= total_cols * 0.5 and len(unique_values) >= non_null_count * 0.7:
                row_type = "column_names"
                confidence_factors.append(0.85)
                # After column names comes data
                data_start_row = row_idx + 1

            if row_type != "unknown" and row_type != "data":
                content = " | ".join(str(v) for v in row if v is not None)[:100]
                header_rows.append(RowInfo(
                    index=row_idx,
                    type=row_type,
                    content=content,
                    is_empty=non_null_count == 0,
                    merged_count=len(row_merges)
                ))

        # If no explicit column names found, use first row
        if data_start_row == 0:
            data_start_row = 1
            confidence_factors.append(0.5)

        result.header_row_count = data_start_row
        result.data_start_row = data_start_row
        result.header_rows = header_rows
        result.merged_cells = merged_cells
        result.preview_rows = raw_rows[:min(5, len(raw_rows))]

        # Extract column information from the last header row
        if data_start_row > 0 and data_start_row <= len(raw_rows):
            column_names_row = raw_rows[data_start_row - 1]
            result.columns = self._extract_columns(column_names_row, raw_rows[data_start_row:data_start_row + 5])

        # Calculate confidence
        if confidence_factors:
            result.confidence = sum(confidence_factors) / len(confidence_factors)
        else:
            result.confidence = 0.5

        return result

    def _is_period_pattern(self, value: str) -> bool:
        """Check if value matches a period pattern"""
        import re
        patterns = [
            r'\d{4}年',  # 2025年
            r'\d{1,2}月',  # 1月, 12月
            r'Q[1-4]',  # Q1-Q4
            r'\d{4}-\d{2}',  # 2025-01
            r'\d{4}/\d{2}',  # 2025/01
        ]
        for pattern in patterns:
            if re.search(pattern, value):
                return True
        return False

    def _has_numeric_values(self, row: List[Any]) -> bool:
        """Check if row contains numeric values"""
        numeric_count = 0
        for v in row:
            if v is not None:
                try:
                    cleaned = str(v).replace(',', '').replace('¥', '').replace('$', '').replace('%', '')
                    float(cleaned)
                    numeric_count += 1
                except (ValueError, TypeError):
                    pass
        return numeric_count >= len(row) * 0.3

    def _extract_columns(
        self, header_row: List[Any], data_rows: List[List[Any]]
    ) -> List[ColumnInfo]:
        """Extract column information from header and sample data"""
        columns = []
        for col_idx, header in enumerate(header_row):
            name = str(header) if header else f"Column_{col_idx + 1}"

            # Sample values from data rows
            sample_values = []
            for data_row in data_rows:
                if col_idx < len(data_row):
                    sample_values.append(data_row[col_idx])

            # Detect data type
            data_type = self._detect_column_type(sample_values)

            columns.append(ColumnInfo(
                index=col_idx,
                name=name,
                data_type=data_type,
                sample_values=sample_values[:3]
            ))

        return columns

    def _detect_column_type(self, values: List[Any]) -> str:
        """Detect column data type from sample values"""
        if not values:
            return "text"

        numeric_count = 0
        percentage_count = 0
        currency_count = 0

        for v in values:
            if v is None:
                continue
            v_str = str(v)

            if '%' in v_str or v_str.endswith('%'):
                percentage_count += 1
            elif v_str.startswith(('¥', '$', '€', '£')):
                currency_count += 1
            else:
                try:
                    float(v_str.replace(',', ''))
                    numeric_count += 1
                except ValueError:
                    pass

        total = len([v for v in values if v is not None])
        if total == 0:
            return "text"

        if percentage_count > total * 0.5:
            return "percentage"
        if currency_count > total * 0.5:
            return "currency"
        if numeric_count > total * 0.5:
            return "numeric"

        return "text"

    async def _detect_with_llm_for_complex(
        self,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        sheet_name: str,
        total_rows: int,
        total_cols: int
    ) -> StructureDetectionResult:
        """
        LLM detection specifically for complex multi-level headers.
        Uses detailed prompting to understand header structure.
        """
        result = StructureDetectionResult(
            method="llm_complex",
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols,
            preview_rows=raw_rows[:8],
            merged_cells=merged_cells
        )

        try:
            # Build detailed row representation
            rows_text = "Excel Data Preview (first 10 rows, first 15 columns):\n"
            rows_text += "=" * 60 + "\n"

            for idx, row in enumerate(raw_rows[:10]):
                row_content = [str(v)[:20] if v else "(empty)" for v in row[:15]]
                # Count numeric values
                numeric_count = sum(1 for v in row[:15] if v and self._is_numeric(str(v)))
                total_non_empty = sum(1 for v in row[:15] if v)
                numeric_pct = int(numeric_count / total_non_empty * 100) if total_non_empty > 0 else 0

                rows_text += f"Row {idx}: [{numeric_pct}% numeric] {' | '.join(row_content)}\n"

            # Merged cells info
            merge_text = "\nMerged Cells in Header Region:\n"
            if merged_cells:
                for m in merged_cells[:15]:
                    merge_text += f"  {m.range}: '{m.value[:30] if m.value else ''}' (rows {m.min_row}-{m.max_row}, cols {m.min_col}-{m.max_col})\n"
            else:
                merge_text += "  (none)\n"

            prompt = f"""Analyze this Excel sheet structure. Sheet name: "{sheet_name}"

{rows_text}
{merge_text}

This appears to be a complex table with multiple header rows. Common patterns include:
- Row 1: Title (often merged across all columns)
- Row 2-3: Date/period info, unit info
- Row 4: Column group names (e.g., months like "1月", "2月" merged across sub-columns)
- Row 5: Sub-column names (e.g., "预算数", "本月实际" repeated under each month)
- Row 6+: Actual data rows (mostly numeric values)

Key indicators of data rows:
- High percentage of numeric values (>40%)
- First column often contains category names like "一、营业收入", "销售收入"
- Consistent data pattern across rows

Analyze and determine:
1. How many rows are HEADER rows (before actual data)?
2. Which row index (0-based) is the FIRST DATA ROW?
3. What are the actual column names from the last header row?

IMPORTANT: Data rows have many numeric values. The first row with >40% numeric values is likely the first data row, NOT a header.

Return JSON only:
{{
  "header_row_count": <number of header rows>,
  "data_start_row": <0-based index of first data row>,
  "column_names_row": <0-based index of the row containing actual column names>,
  "confidence": <0.0-1.0>,
  "reasoning": "<brief explanation>"
}}"""

            response = await self._call_llm(prompt, model="default")
            if response:
                parsed = self._parse_llm_response(response)
                if parsed:
                    header_count = parsed.get("header_row_count", 5)
                    data_start = parsed.get("data_start_row", 5)
                    column_names_row = parsed.get("column_names_row", data_start - 1)

                    # Validate: data_start should be where numeric data begins
                    # Double-check by looking for numeric ratio
                    for idx, row in enumerate(raw_rows[:10]):
                        numeric_count = sum(1 for v in row if v and self._is_numeric(str(v)))
                        total_non_empty = sum(1 for v in row if v)
                        if total_non_empty > 5 and numeric_count / total_non_empty > 0.4:
                            if idx < data_start:
                                logger.warning(f"LLM said data_start={data_start}, but row {idx} has {numeric_count}/{total_non_empty} numeric values. Adjusting.")
                                data_start = idx
                                header_count = idx
                            break

                    result.header_row_count = header_count
                    result.data_start_row = data_start
                    result.confidence = parsed.get("confidence", 0.75)
                    result.note = parsed.get("reasoning", "")

                    # Extract columns from the correct row
                    if 0 <= column_names_row < len(raw_rows):
                        column_row = raw_rows[column_names_row]
                        result.columns = self._extract_columns(
                            column_row,
                            raw_rows[data_start:data_start + 5] if data_start < len(raw_rows) else []
                        )

                    logger.info(f"LLM complex detection: header_rows={header_count}, data_start={data_start}, confidence={result.confidence:.2f}")
                    return result

        except Exception as e:
            logger.warning(f"LLM complex detection failed: {e}")

        # Fallback
        result.confidence = 0.4
        result.note = "LLM complex detection failed"
        return result

    async def _detect_with_llm_fast(
        self,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        sheet_name: str,
        total_rows: int,
        total_cols: int
    ) -> StructureDetectionResult:
        """
        LLM-fast structure detection (Layer 2).
        Uses qwen-turbo for quick text-based analysis.
        """
        result = StructureDetectionResult(
            method="llm_fast",
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols,
            preview_rows=raw_rows[:5],
            merged_cells=merged_cells
        )

        try:
            # Prepare prompt with row data
            rows_text = ""
            for idx, row in enumerate(raw_rows[:10]):
                row_content = [str(v) if v else "" for v in row[:15]]  # Limit columns
                rows_text += f"Row {idx}: {' | '.join(row_content)}\n"

            merge_text = ""
            if merged_cells:
                merge_text = "Merged cells:\n"
                for m in merged_cells[:10]:
                    merge_text += f"  {m.range}: '{m.value}'\n"

            prompt = f"""Analyze this Excel sheet structure. The sheet name is "{sheet_name}".

{rows_text}
{merge_text}

Determine:
1. How many rows are header rows (before actual data starts)?
2. What type is each header row? (title/subtitle/period/category/column_names)
3. Which row is the first data row (0-indexed)?

Return JSON only:
{{
  "header_row_count": <number>,
  "data_start_row": <number, 0-indexed>,
  "row_types": [
    {{"index": 0, "type": "title"}},
    {{"index": 1, "type": "subtitle"}},
    ...
  ],
  "confidence": <0.0-1.0>
}}"""

            response = await self._call_llm(prompt, model="fast")
            if response:
                parsed = self._parse_llm_response(response)
                if parsed:
                    result.header_row_count = parsed.get("header_row_count", 1)
                    result.data_start_row = parsed.get("data_start_row", 1)
                    result.confidence = parsed.get("confidence", 0.7)

                    # Convert row types
                    for rt in parsed.get("row_types", []):
                        idx = rt.get("index", 0)
                        if idx < len(raw_rows):
                            row = raw_rows[idx]
                            content = " | ".join(str(v) for v in row if v)[:100]
                            result.header_rows.append(RowInfo(
                                index=idx,
                                type=rt.get("type", "unknown"),
                                content=content
                            ))

                    # Extract columns
                    if result.data_start_row > 0 and result.data_start_row <= len(raw_rows):
                        header_row = raw_rows[result.data_start_row - 1]
                        result.columns = self._extract_columns(
                            header_row, raw_rows[result.data_start_row:result.data_start_row + 5]
                        )

                    return result

        except Exception as e:
            logger.warning(f"LLM-fast detection failed: {e}")

        result.confidence = 0.4
        result.note = "LLM-fast detection failed, using fallback"
        return result

    async def _detect_with_llm_vl(
        self,
        file_bytes: bytes,
        sheet_index: int,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        sheet_name: str,
        total_rows: int,
        total_cols: int
    ) -> StructureDetectionResult:
        """
        LLM-VL structure detection (Layer 3).
        Uses vision model to analyze Excel screenshot.
        """
        result = StructureDetectionResult(
            method="llm_vl",
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols,
            preview_rows=raw_rows[:5],
            merged_cells=merged_cells
        )

        try:
            # Generate Excel preview image
            image_bytes = self._generate_excel_preview_image(file_bytes, sheet_index)

            if image_bytes:
                # Encode image for API
                image_base64 = base64.b64encode(image_bytes).decode('utf-8')

                prompt = """Analyze this Excel sheet structure from the screenshot.

Identify:
1. The total number of header rows (before data starts)
2. The type of each header row (title, subtitle, period, category, column_names)
3. Where the actual data begins (0-indexed row number)
4. Any merged cells in the header region

Return JSON only:
{
  "header_row_count": <number>,
  "data_start_row": <number, 0-indexed>,
  "row_types": [{"index": 0, "type": "title"}, ...],
  "confidence": <0.0-1.0>
}"""

                response = await self._call_llm_vl(prompt, image_base64)
                if response:
                    parsed = self._parse_llm_response(response)
                    if parsed:
                        result.header_row_count = parsed.get("header_row_count", 1)
                        result.data_start_row = parsed.get("data_start_row", 1)
                        result.confidence = parsed.get("confidence", 0.8)

                        # Convert row types
                        for rt in parsed.get("row_types", []):
                            idx = rt.get("index", 0)
                            if idx < len(raw_rows):
                                row = raw_rows[idx]
                                content = " | ".join(str(v) for v in row if v)[:100]
                                result.header_rows.append(RowInfo(
                                    index=idx,
                                    type=rt.get("type", "unknown"),
                                    content=content
                                ))

                        # Extract columns
                        if result.data_start_row > 0 and result.data_start_row <= len(raw_rows):
                            header_row = raw_rows[result.data_start_row - 1]
                            result.columns = self._extract_columns(
                                header_row, raw_rows[result.data_start_row:result.data_start_row + 5]
                            )

                        return result

        except Exception as e:
            logger.warning(f"LLM-VL detection failed: {e}")

        # Fallback to text-based analysis without vision
        result.confidence = 0.5
        result.note = "LLM-VL detection failed, using partial results"
        return result

    async def _detect_with_multi_model(
        self,
        file_bytes: bytes,
        sheet_index: int,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        sheet_name: str,
        total_rows: int,
        total_cols: int
    ) -> StructureDetectionResult:
        """
        Multi-model structure detection (Layer 4).
        Uses multiple models and voting for complex cases.
        """
        result = StructureDetectionResult(
            method="multi_model",
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols,
            preview_rows=raw_rows[:5],
            merged_cells=merged_cells
        )

        try:
            # Strategy A: Get results from multiple models
            results = []

            # Result from rule-based
            rule_result = self._detect_with_rules(
                raw_rows, merged_cells, sheet_name, total_rows, total_cols
            )
            results.append(("rule", rule_result))

            # Result from fast LLM
            fast_result = await self._detect_with_llm_fast(
                raw_rows, merged_cells, sheet_name, total_rows, total_cols
            )
            results.append(("llm_fast", fast_result))

            # Vote on header_row_count
            header_counts = [r[1].header_row_count for r in results]
            voted_header_count = max(set(header_counts), key=header_counts.count)

            # Vote on data_start_row
            data_starts = [r[1].data_start_row for r in results]
            voted_data_start = max(set(data_starts), key=data_starts.count)

            # Calculate agreement
            header_agreement = header_counts.count(voted_header_count) / len(header_counts)
            data_agreement = data_starts.count(voted_data_start) / len(data_starts)

            result.header_row_count = voted_header_count
            result.data_start_row = voted_data_start
            result.confidence = (header_agreement + data_agreement) / 2

            # If still low confidence, use self-correction
            if result.confidence < 0.6:
                for round_num in range(self.settings.max_self_correction_rounds):
                    corrected = await self._self_correct(
                        raw_rows, merged_cells, result, round_num + 1
                    )
                    if corrected.confidence >= 0.6:
                        return corrected

                # Return best guess with note
                result.note = "Low confidence auto-inference, results marked"

            # Extract columns based on voted result
            if result.data_start_row > 0 and result.data_start_row <= len(raw_rows):
                header_row = raw_rows[result.data_start_row - 1]
                result.columns = self._extract_columns(
                    header_row, raw_rows[result.data_start_row:result.data_start_row + 5]
                )

            return result

        except Exception as e:
            logger.error(f"Multi-model detection failed: {e}")
            return self._create_fallback_result(
                raw_rows, merged_cells, sheet_name, total_rows, total_cols
            )

    async def _self_correct(
        self,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        current_result: StructureDetectionResult,
        round_num: int
    ) -> StructureDetectionResult:
        """Self-correction loop for improving detection"""
        try:
            rows_text = ""
            for idx, row in enumerate(raw_rows[:10]):
                row_content = [str(v) if v else "" for v in row[:15]]
                rows_text += f"Row {idx}: {' | '.join(row_content)}\n"

            prompt = f"""Previous analysis of this Excel determined:
- Header rows: {current_result.header_row_count}
- Data starts at row: {current_result.data_start_row}

Review and verify. Here's the data:
{rows_text}

Is this correct? If not, provide corrected values.
Return JSON:
{{
  "header_row_count": <number>,
  "data_start_row": <number>,
  "confidence": <0.0-1.0>,
  "reasoning": "<brief explanation>"
}}"""

            response = await self._call_llm(prompt, model="reasoning")
            if response:
                parsed = self._parse_llm_response(response)
                if parsed:
                    new_result = StructureDetectionResult(
                        method=f"multi_model_round_{round_num}",
                        sheet_name=current_result.sheet_name,
                        total_rows=current_result.total_rows,
                        total_cols=current_result.total_cols,
                        header_row_count=parsed.get("header_row_count", current_result.header_row_count),
                        data_start_row=parsed.get("data_start_row", current_result.data_start_row),
                        confidence=parsed.get("confidence", 0.6),
                        preview_rows=current_result.preview_rows,
                        merged_cells=current_result.merged_cells
                    )
                    return new_result

        except Exception as e:
            logger.warning(f"Self-correction round {round_num} failed: {e}")

        return current_result

    def _create_fallback_result(
        self,
        raw_rows: List[List[Any]],
        merged_cells: List[MergedCellInfo],
        sheet_name: str,
        total_rows: int,
        total_cols: int
    ) -> StructureDetectionResult:
        """Create fallback result when all detection methods fail"""
        result = StructureDetectionResult(
            method="fallback",
            sheet_name=sheet_name,
            total_rows=total_rows,
            total_cols=total_cols,
            header_row_count=1,
            data_start_row=1,
            confidence=0.5,
            preview_rows=raw_rows[:5],
            merged_cells=merged_cells,
            note="Fallback detection - using default single header row"
        )

        if raw_rows:
            result.columns = self._extract_columns(raw_rows[0], raw_rows[1:6])

        return result

    def _generate_excel_preview_image(
        self, file_bytes: bytes, sheet_index: int
    ) -> Optional[bytes]:
        """Generate a preview image of the Excel sheet"""
        # This is a placeholder - in production, you would use
        # libraries like xlrd2, openpyxl + PIL, or external services
        # to render the Excel as an image

        # For now, return None to skip vision-based detection
        # and rely on text-based analysis
        logger.debug("Excel preview image generation not implemented, skipping vision analysis")
        return None

    async def _call_llm(self, prompt: str, model: str = "default") -> Optional[str]:
        """Call LLM API"""
        try:
            from openai import AsyncOpenAI

            model_name = {
                "default": self.settings.llm_model,
                "fast": self.settings.llm_fast_model,
                "reasoning": self.settings.llm_reasoning_model
            }.get(model, self.settings.llm_model)

            client = AsyncOpenAI(
                api_key=self.settings.llm_api_key,
                base_url=self.settings.llm_base_url
            )

            response = await client.chat.completions.create(
                model=model_name,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=1000
            )

            return response.choices[0].message.content

        except Exception as e:
            logger.error(f"LLM call failed: {e}")
            return None

    async def _call_llm_vl(self, prompt: str, image_base64: str) -> Optional[str]:
        """Call Vision-Language LLM API"""
        try:
            from openai import AsyncOpenAI

            client = AsyncOpenAI(
                api_key=self.settings.llm_api_key,
                base_url=self.settings.llm_vl_base_url
            )

            response = await client.chat.completions.create(
                model=self.settings.llm_vl_model,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_base64}"}}
                    ]
                }],
                temperature=0.1,
                max_tokens=1000
            )

            return response.choices[0].message.content

        except Exception as e:
            logger.error(f"LLM-VL call failed: {e}")
            return None

    def _parse_llm_response(self, response: str) -> Optional[Dict]:
        """Parse JSON from LLM response"""
        try:
            # Try to extract JSON from response
            import re

            # Look for JSON block
            json_match = re.search(r'\{[\s\S]*\}', response)
            if json_match:
                return json.loads(json_match.group())

            return None
        except json.JSONDecodeError as e:
            logger.warning(f"Failed to parse LLM response as JSON: {e}")
            return None
