#!/usr/bin/env python3
"""
SmartBI 压测数据生成器 — 用于测试系统处理大数据量的能力和上限

生成 L1-L2 两个级别的压力测试文件，分为工厂端和餐饮端：
- L1: 50,000 行 (~6MB)   — 上传端测试（10MB 内，含 merge/样式/百分比）
- L2: 200,000 行 (~24MB) — API 直接测试（超 10MB 上传限制，测 Python 解析性能）

L1 用于完整 E2E 流程（上传→SSE→解析→图表→AI），有 3 级合并表头和百分比格式。
L2 仅用于 Python API 直接调用测试（跳过 Java 上传端点），使用 write_only 模式。

注意：Java max-file-size=10MB，L2 永远无法通过上传端点。
L2 测试方式：直接用 openpyxl 读取后调用 Python /api/smartbi/chart/batch-build 等端点。

Usage:
    python tests/generate_stress_test.py --level L1            # 50K 行 (上传端)
    python tests/generate_stress_test.py --level L2            # 200K 行 (API 直接)
    python tests/generate_stress_test.py --all                 # L1+L2 × factory+restaurant
"""
import argparse
import random
import time
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 压测级别定义
# ============================================================

LEVELS = {
    "L1": {"rows": 50_000,    "label": "上传端测试(10MB内，含merge/样式)", "mode": "upload"},
    "L2": {"rows": 200_000,   "label": "API直接测试(超10MB，纯性能)", "mode": "api-direct"},
}

# ============================================================
# 工厂端数据模板
# ============================================================

FACTORY_PRODUCTS = [
    "火锅底料-麻辣", "火锅底料-番茄", "火锅底料-清汤", "火锅底料-菌汤",
    "复合调味料-红烧", "复合调味料-糖醋", "复合调味料-麻婆",
    "酱料-豆瓣酱", "酱料-辣椒酱", "酱料-甜面酱",
    "速食-自热火锅", "速食-酸辣粉", "速食-螺蛳粉",
    "预制菜-水煮鱼", "预制菜-酸菜鱼", "预制菜-毛血旺",
]

FACTORY_REGIONS = [
    "华东分部", "华南分部", "华北分部", "西南分部", "华中分部", "东北分部",
]

FACTORY_WAREHOUSES = [
    "上海总仓", "广州仓", "北京仓", "成都仓", "武汉仓", "沈阳仓",
    "杭州仓", "深圳仓", "天津仓", "重庆仓", "长沙仓", "大连仓",
]

FACTORY_CUSTOMERS = [
    "永辉超市", "大润发", "沃尔玛", "家乐福", "盒马鲜生", "叮咚买菜",
    "美团优选", "多多买菜", "朴朴超市", "山姆会员店", "Costco", "麦德龙",
    "7-Eleven", "全家便利", "罗森", "红旗连锁", "步步高", "华润万家",
    "物美超市", "联华超市", "中百仓储", "新华都", "人人乐", "北京华联",
]

FACTORY_EXPENSE_TYPES = [
    "原材料采购", "包装材料", "人工工资", "设备折旧", "水电燃气",
    "运输费用", "仓储费用", "质检费用", "研发费用", "营销费用",
    "管理费用", "租赁费用", "维修费用", "保险费用", "税金",
]

# ============================================================
# 餐饮端数据模板
# ============================================================

RESTAURANT_STORES = [
    "旗舰店-南京西路", "旗舰店-天河城", "旗舰店-三里屯", "旗舰店-春熙路",
    "标准店-人民广场", "标准店-体育西", "标准店-望京", "标准店-锦里",
    "社区店-浦东", "社区店-番禺", "社区店-朝阳", "社区店-武侯",
    "外卖店-虹口", "外卖店-海珠", "外卖店-丰台", "外卖店-金牛",
    "商场店-静安嘉里", "商场店-太古汇", "商场店-朝阳大悦城", "商场店-IFS",
]

RESTAURANT_MENU = [
    "招牌酸菜鱼", "藤椒鱼", "番茄鱼", "金汤酸菜鱼", "麻辣水煮鱼",
    "酸汤肥牛", "毛血旺", "沸腾虾", "干锅花菜", "口水鸡",
    "蒜泥白肉", "麻婆豆腐", "宫保鸡丁", "鱼香肉丝", "回锅肉",
    "米饭", "蛋炒饭", "酸梅汁", "豆浆", "冰粉",
    "凉拌木耳", "拍黄瓜", "泡椒凤爪", "红糖糍粑", "炸酱面",
]

RESTAURANT_CHANNELS = ["堂食", "美团外卖", "饿了么", "抖音团购", "小红书团购", "自营小程序"]

PAYMENT_METHODS = ["微信", "支付宝", "现金", "银行卡", "美团余额", "饿了么余额", "储值卡", "信用卡"]

# ============================================================
# 工厂端 sheet 生成器
# ============================================================

def _style_header(ws, row, col, value, merge_end_col=None, merge_end_row=None):
    """Helper: styled header cell with optional merge."""
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=merge_end_row or row, end_column=merge_end_col)


def gen_factory_sales_detail(wb, n_rows, seed, use_write_only=False):
    """工厂销售明细 — 订单级别交易流水 + 3级merge header + 百分比格式。"""
    random.seed(seed)
    ws = wb.create_sheet(title="销售明细")

    # L1 (non-write_only): add 3-level merge headers
    data_start_row = 1
    if not use_write_only:
        border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        # Level 1: major groups
        _style_header(ws, 1, 1, "订单信息", merge_end_col=4)
        _style_header(ws, 1, 5, "产品信息", merge_end_col=8)
        _style_header(ws, 1, 9, "金额明细", merge_end_col=19)
        _style_header(ws, 1, 20, "物流信息", merge_end_col=25)
        _style_header(ws, 1, 26, "审批信息", merge_end_col=28)
        # Level 2: sub-groups
        _style_header(ws, 2, 1, "基本", merge_end_col=2)
        _style_header(ws, 2, 3, "客户", merge_end_col=4)
        _style_header(ws, 2, 5, "名称", merge_end_col=6)
        _style_header(ws, 2, 7, "规格", merge_end_col=8)
        _style_header(ws, 2, 9, "原始金额", merge_end_col=11)
        _style_header(ws, 2, 12, "折扣与税", merge_end_col=16)
        _style_header(ws, 2, 17, "利润分析", merge_end_col=19)
        _style_header(ws, 2, 20, "发货", merge_end_col=23)
        _style_header(ws, 2, 24, "回款", merge_end_col=25)
        _style_header(ws, 2, 26, "人员", merge_end_col=27)
        _style_header(ws, 2, 28, "备注")
        # Level 3: individual column headers
        headers = [
            "订单编号", "订单日期", "客户名称", "客户区域", "产品名称", "产品类别",
            "规格", "单位", "数量", "单价(元)", "金额(元)", "折扣率", "折后金额(元)",
            "税率", "税额(元)", "含税金额(元)", "成本(元)", "毛利(元)", "毛利率",
            "仓库", "发货日期", "物流公司", "运单号", "回款状态", "回款日期",
            "销售员", "审批人", "备注",
        ]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        data_start_row = 4
    else:
        headers = [
            "订单编号", "订单日期", "客户名称", "客户区域", "产品名称", "产品类别",
            "规格", "单位", "数量", "单价(元)", "金额(元)", "折扣率", "折后金额(元)",
            "税率", "税额(元)", "含税金额(元)", "成本(元)", "毛利(元)", "毛利率",
            "仓库", "发货日期", "物流公司", "运单号", "回款状态", "回款日期",
            "销售员", "审批人", "备注",
        ]
        ws.append(headers)

    # Percentage column indices (1-based): 折扣率=12, 税率=14, 毛利率=19
    pct_cols = {12, 14, 19}

    specs = ["250g", "500g", "1kg", "2.5kg", "5kg", "10kg", "袋装", "瓶装", "罐装", "箱装"]
    logistics = ["顺丰", "京东物流", "德邦", "安能", "中通", "圆通", "韵达", "极兔"]
    sales_reps = ["张经理", "李主管", "王总监", "赵经理", "孙主管", "周专员", "吴经理", "郑主管"]
    approvers = ["陈总", "林副总", "黄总监", "刘经理"]
    pay_statuses = ["已回款", "未回款", "部分回款", "逾期"]
    notes_pool = [None, None, None, "加急", "样品", "赠品", "退货补发", "特价促销", "年度合同", "新客户首单"]
    # Mixed type: occasionally inject text into numeric-dominated备注 column
    mixed_notes = [None, None, None, "加急", 1, 2, "赠品", 999, "退货", True, False]

    base_date = date(2025, 1, 1)
    order_id = 100000
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    for i in range(n_rows):
        order_id += 1
        order_date = base_date + timedelta(days=random.randint(0, 364))
        customer = random.choice(FACTORY_CUSTOMERS)
        region = random.choice(FACTORY_REGIONS)
        product = random.choice(FACTORY_PRODUCTS)
        category = product.split("-")[0]
        spec = random.choice(specs)
        unit = "箱" if spec in ("箱装",) else "袋" if "袋" in spec else "瓶" if "瓶" in spec else "件"
        qty = random.randint(10, 5000)
        unit_price = round(random.uniform(8.5, 120.0), 2)
        amount = round(qty * unit_price, 2)
        discount = round(random.uniform(0.85, 1.0), 4)
        disc_amount = round(amount * discount, 2)
        tax_rate = round(random.choice([0.06, 0.09, 0.13]), 2)
        tax = round(disc_amount * tax_rate, 2)
        total = round(disc_amount + tax, 2)
        cost = round(disc_amount * random.uniform(0.45, 0.75), 2)
        profit = round(disc_amount - cost, 2)
        margin = round(profit / disc_amount, 4) if disc_amount else 0
        warehouse = random.choice(FACTORY_WAREHOUSES)
        ship_date = order_date + timedelta(days=random.randint(1, 7))
        logi = random.choice(logistics)
        tracking = f"SF{random.randint(1000000000, 9999999999)}"
        pay_status = random.choice(pay_statuses)
        pay_date = (ship_date + timedelta(days=random.randint(15, 90))).isoformat() if pay_status != "未回款" else None
        sales_rep = random.choice(sales_reps)
        approver = random.choice(approvers)
        note = random.choice(mixed_notes if not use_write_only else notes_pool)

        row_data = [
            f"SO-{order_id}", order_date.isoformat(), customer, region, product, category,
            spec, unit, qty, unit_price, amount, discount, disc_amount,
            tax_rate, tax, total, cost, profit, margin,
            warehouse, ship_date.isoformat(), logi, tracking, pay_status, pay_date,
            sales_rep, approver, note,
        ]

        if use_write_only:
            ws.append(row_data)
        else:
            row_num = data_start_row + i
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.border = border
                if ci in pct_cols:
                    cell.number_format = "0.00%"

        if (i + 1) % 50000 == 0:
            print(f"    销售明细: {i + 1:,}/{n_rows:,} rows...")


def gen_factory_expense_ledger(wb, n_rows, seed, use_write_only=False):
    """工厂费用台账 — 每行一笔费用记录 + merge header + 百分比格式。"""
    random.seed(seed + 1)
    ws = wb.create_sheet(title="费用台账")

    headers = [
        "凭证号", "记账日期", "费用类型", "科目编码", "科目名称",
        "部门", "项目", "摘要", "借方金额", "贷方金额",
        "预算金额", "预算执行率", "审批人", "状态",
    ]

    data_start_row = 1
    if not use_write_only:
        border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        # Level 1: major groups
        _style_header(ws, 1, 1, "凭证信息", merge_end_col=3)
        _style_header(ws, 1, 4, "科目信息", merge_end_col=5)
        _style_header(ws, 1, 6, "归属", merge_end_col=8)
        _style_header(ws, 1, 9, "金额", merge_end_col=12)
        _style_header(ws, 1, 13, "审批", merge_end_col=14)
        # Level 2: column headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        data_start_row = 3
    else:
        ws.append(headers)

    # Percentage column: 预算执行率=12
    pct_cols = {12}

    departments = ["生产部", "销售部", "研发部", "财务部", "行政部", "品控部", "物流部", "采购部"]
    projects = ["日常运营", "新品开发", "设备升级", "市场推广", "信息化建设", "产线扩建", None]
    statuses = ["已审批", "待审批", "已驳回", "已支付"]
    approvers = ["陈总", "林副总", "黄总监", "刘经理", "财务张"]
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    base_date = date(2025, 1, 1)

    for i in range(n_rows):
        voucher = f"PZ-{2025}{(i % 12) + 1:02d}-{i + 1:06d}"
        rec_date = base_date + timedelta(days=random.randint(0, 364))
        exp_type = random.choice(FACTORY_EXPENSE_TYPES)
        code = f"6{random.randint(1, 999):03d}"
        dept = random.choice(departments)
        proj = random.choice(projects)
        summary = f"{dept}{exp_type}-{rec_date.strftime('%m月')}"
        debit = round(random.uniform(100, 500000), 2)
        credit = 0 if random.random() > 0.1 else round(debit * random.uniform(0.1, 0.5), 2)
        budget = round(debit * random.uniform(1.0, 1.5), 2)
        exec_rate = round(debit / budget, 4) if budget else 0

        row_data = [
            voucher, rec_date.isoformat(), exp_type, code, exp_type,
            dept, proj, summary, debit, credit,
            budget, exec_rate, random.choice(approvers), random.choice(statuses),
        ]

        if use_write_only:
            ws.append(row_data)
        else:
            row_num = data_start_row + i
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.border = border
                if ci in pct_cols:
                    cell.number_format = "0.00%"

        if (i + 1) % 50000 == 0:
            print(f"    费用台账: {i + 1:,}/{n_rows:,} rows...")


def gen_factory_inventory(wb, n_rows, seed, use_write_only=False):
    """工厂库存流水 — 入库出库记录 + merge header。"""
    random.seed(seed + 2)
    ws = wb.create_sheet(title="库存流水")

    headers = [
        "流水号", "日期", "仓库", "产品名称", "批次号", "操作类型",
        "数量", "单位", "单价", "金额", "库存余量", "保质期至",
        "操作人", "备注",
    ]

    data_start_row = 1
    if not use_write_only:
        border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        # Level 1: major groups
        _style_header(ws, 1, 1, "流水标识", merge_end_col=2)
        _style_header(ws, 1, 3, "产品信息", merge_end_col=6)
        _style_header(ws, 1, 7, "数量金额", merge_end_col=11)
        _style_header(ws, 1, 12, "其他", merge_end_col=14)
        # Level 2: column headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        data_start_row = 3
    else:
        ws.append(headers)

    ops = ["入库-采购", "入库-生产", "入库-退货", "出库-销售", "出库-调拨", "出库-报废", "盘点调整"]
    operators = ["仓管张", "仓管李", "仓管王", "主管陈", "质检赵"]
    mixed_notes = [None, None, None, None, "异常记录", "温度超标", 0, 1, True, "霉变", "破损"]
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    base_date = date(2025, 1, 1)
    stock_balance = {p: random.randint(1000, 50000) for p in FACTORY_PRODUCTS}

    for i in range(n_rows):
        rec_date = base_date + timedelta(days=random.randint(0, 364))
        warehouse = random.choice(FACTORY_WAREHOUSES)
        product = random.choice(FACTORY_PRODUCTS)
        batch = f"B{rec_date.strftime('%Y%m%d')}-{random.randint(1, 99):02d}"
        op_type = random.choice(ops)
        qty = random.randint(10, 3000)
        if "出库" in op_type:
            qty = -qty
        unit = "箱"
        price = round(random.uniform(8, 80), 2)
        amount = round(abs(qty) * price, 2)
        stock_balance[product] = max(0, stock_balance[product] + qty)
        shelf_life = (rec_date + timedelta(days=random.randint(90, 730))).isoformat()
        note = random.choice(mixed_notes if not use_write_only else [None, None, "异常记录"])

        row_data = [
            f"INV-{i + 1:08d}", rec_date.isoformat(), warehouse, product, batch, op_type,
            qty, unit, price, amount, stock_balance[product], shelf_life,
            random.choice(operators), note,
        ]

        if use_write_only:
            ws.append(row_data)
        else:
            row_num = data_start_row + i
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.border = border

        if (i + 1) % 50000 == 0:
            print(f"    库存流水: {i + 1:,}/{n_rows:,} rows...")


# ============================================================
# 餐饮端 sheet 生成器
# ============================================================

def gen_restaurant_orders(wb, n_rows, seed, use_write_only=False):
    """餐饮订单流水 — 单品级别交易明细 + merge headers + 百分比格式。"""
    random.seed(seed)
    ws = wb.create_sheet(title="订单流水")

    headers = [
        "订单号", "下单时间", "门店", "渠道", "桌号/骑手",
        "菜品名称", "菜品类别", "数量", "单价", "小计",
        "折扣", "实收", "支付方式", "顾客类型", "会员等级",
        "评分", "评价标签", "配送费", "打包费", "优惠券抵扣",
        "平台抽成率", "平台抽成额", "实际到账", "备注",
    ]

    data_start_row = 1
    if not use_write_only:
        border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        # Level 1: major groups
        _style_header(ws, 1, 1, "订单基本信息", merge_end_col=5)
        _style_header(ws, 1, 6, "菜品信息", merge_end_col=7)
        _style_header(ws, 1, 8, "金额明细", merge_end_col=12)
        _style_header(ws, 1, 13, "顾客信息", merge_end_col=17)
        _style_header(ws, 1, 18, "附加费用", merge_end_col=20)
        _style_header(ws, 1, 21, "平台结算", merge_end_col=23)
        _style_header(ws, 1, 24, "备注")
        # Level 2: sub-groups
        _style_header(ws, 2, 1, "标识", merge_end_col=2)
        _style_header(ws, 2, 3, "渠道", merge_end_col=5)
        _style_header(ws, 2, 6, "品类", merge_end_col=7)
        _style_header(ws, 2, 8, "原始", merge_end_col=10)
        _style_header(ws, 2, 11, "折后", merge_end_col=12)
        _style_header(ws, 2, 13, "属性", merge_end_col=15)
        _style_header(ws, 2, 16, "评价", merge_end_col=17)
        _style_header(ws, 2, 18, "费用", merge_end_col=20)
        _style_header(ws, 2, 21, "抽成", merge_end_col=23)
        _style_header(ws, 2, 24, "")
        # Level 3: column headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        data_start_row = 4
    else:
        ws.append(headers)

    # Percentage columns: 折扣=11, 平台抽成率=21
    pct_cols = {11, 21}

    categories = {item: ("招牌菜" if "鱼" in item or "虾" in item else
                         "主食" if "饭" in item or "面" in item else
                         "饮品" if "汁" in item or "浆" in item or "粉" in item else
                         "凉菜" if "凉" in item or "拍" in item or "泡" in item else
                         "热菜") for item in RESTAURANT_MENU}
    customer_types = ["新客", "回头客", "会员", "团购客"]
    member_levels = [None, "普通会员", "银卡", "金卡", "钻石卡", "黑卡"]
    ratings = [None, 3, 3.5, 4, 4, 4.5, 4.5, 5, 5, 5]
    tags = [None, None, "好评", "味道好", "分量足", "上菜快", "环境好", "服务好", "一般", "偏咸", "等太久"]

    base_date = datetime(2025, 1, 1, 10, 0)
    order_id = 200000
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    mixed_notes = [None, None, "加辣", "少盐", 1, 2, True, "催单", 888, False, "不要葱"]

    for i in range(n_rows):
        order_id += 1
        order_time = base_date + timedelta(
            days=random.randint(0, 364),
            hours=random.randint(0, 13),
            minutes=random.randint(0, 59),
        )
        store = random.choice(RESTAURANT_STORES)
        channel = random.choice(RESTAURANT_CHANNELS)
        table = f"T{random.randint(1, 30)}" if channel == "堂食" else f"骑手{random.randint(1001, 9999)}"
        item = random.choice(RESTAURANT_MENU)
        cat = categories[item]
        qty = random.randint(1, 5)
        price = round(random.uniform(6, 88), 2)
        subtotal = round(qty * price, 2)
        discount = round(random.uniform(0.7, 1.0), 4)
        actual = round(subtotal * discount, 2)
        payment = random.choice(PAYMENT_METHODS)
        cust_type = random.choice(customer_types)
        member = random.choice(member_levels) if cust_type == "会员" else None
        rating = random.choice(ratings)
        tag = random.choice(tags) if rating else None
        delivery_fee = round(random.uniform(0, 8), 2) if channel != "堂食" else 0
        packing_fee = round(random.uniform(0, 3), 2) if channel != "堂食" else 0
        coupon = round(random.uniform(0, 15), 2) if random.random() > 0.7 else 0
        platform_rate = round(random.uniform(0.15, 0.25), 4) if "外卖" in channel or "团购" in channel else 0
        platform_fee = round(actual * platform_rate, 2)
        net = round(actual - platform_fee + delivery_fee, 2)
        note = random.choice(mixed_notes if not use_write_only else ["加辣", "少盐", "不要葱", "打包", "催单", None, None])

        row_data = [
            f"RO-{order_id}", order_time.strftime("%Y-%m-%d %H:%M"), store, channel, table,
            item, cat, qty, price, subtotal,
            discount, actual, payment, cust_type, member,
            rating, tag, delivery_fee, packing_fee, coupon,
            platform_rate, platform_fee, net, note,
        ]

        if use_write_only:
            ws.append(row_data)
        else:
            row_num = data_start_row + i
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.border = border
                if ci in pct_cols:
                    cell.number_format = "0.00%"

        if (i + 1) % 50000 == 0:
            print(f"    订单流水: {i + 1:,}/{n_rows:,} rows...")


def gen_restaurant_daily_ops(wb, n_rows, seed, use_write_only=False):
    """餐饮日运营数据 — 每日每门店运营指标 + merge header + 百分比格式。"""
    random.seed(seed + 1)
    ws = wb.create_sheet(title="日运营数据")

    headers = [
        "日期", "门店", "堂食营业额", "外卖营业额", "团购营业额", "总营业额",
        "堂食订单数", "外卖订单数", "团购订单数", "总订单数",
        "堂食客流", "外卖客流", "总客流", "客单价",
        "食材成本", "人工成本", "租金分摊", "水电费", "其他费用",
        "总成本", "毛利", "毛利率", "净利润", "净利率",
        "好评数", "差评数", "好评率", "翻台率", "上座率",
    ]

    data_start_row = 1
    if not use_write_only:
        border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        # Level 1: major groups
        _style_header(ws, 1, 1, "基本信息", merge_end_col=2)
        _style_header(ws, 1, 3, "营业额", merge_end_col=6)
        _style_header(ws, 1, 7, "订单量", merge_end_col=10)
        _style_header(ws, 1, 11, "客流", merge_end_col=14)
        _style_header(ws, 1, 15, "成本明细", merge_end_col=19)
        _style_header(ws, 1, 20, "利润分析", merge_end_col=24)
        _style_header(ws, 1, 25, "服务质量", merge_end_col=29)
        # Level 2: column headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        data_start_row = 3
    else:
        ws.append(headers)

    # Percentage columns: 毛利率=22, 净利率=24, 好评率=27, 上座率=29
    pct_cols = {22, 24, 27, 29}
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    # Generate enough days×stores to fill n_rows
    n_days = max(365, n_rows // len(RESTAURANT_STORES) + 1)
    base_date = date(2025, 1, 1)
    count = 0

    for d in range(n_days):
        if count >= n_rows:
            break
        rec_date = base_date + timedelta(days=d % 365)
        for store in RESTAURANT_STORES:
            if count >= n_rows:
                break
            # Weekend boost
            is_weekend = rec_date.weekday() >= 5
            mult = 1.3 if is_weekend else 1.0

            dine_rev = round(random.uniform(8000, 45000) * mult, 2)
            dlvr_rev = round(random.uniform(3000, 25000) * mult, 2)
            group_rev = round(random.uniform(1000, 10000) * mult, 2)
            total_rev = round(dine_rev + dlvr_rev + group_rev, 2)
            dine_orders = int(random.uniform(40, 200) * mult)
            dlvr_orders = int(random.uniform(20, 120) * mult)
            group_orders = int(random.uniform(5, 40) * mult)
            total_orders = dine_orders + dlvr_orders + group_orders
            dine_traffic = int(dine_orders * random.uniform(1.5, 3.0))
            dlvr_traffic = dlvr_orders
            total_traffic = dine_traffic + dlvr_traffic
            avg_ticket = round(total_rev / total_orders, 2) if total_orders else 0
            food_cost = round(total_rev * random.uniform(0.30, 0.42), 2)
            labor_cost = round(total_rev * random.uniform(0.20, 0.30), 2)
            rent = round(random.uniform(800, 3000), 2)
            utility = round(random.uniform(200, 800), 2)
            other = round(random.uniform(100, 500), 2)
            total_cost = round(food_cost + labor_cost + rent + utility + other, 2)
            gross_profit = round(total_rev - food_cost, 2)
            gross_margin = round(gross_profit / total_rev, 4) if total_rev else 0
            net_profit = round(total_rev - total_cost, 2)
            net_margin = round(net_profit / total_rev, 4) if total_rev else 0
            good_reviews = int(random.uniform(5, 30) * mult)
            bad_reviews = int(random.uniform(0, 3))
            good_rate = round(good_reviews / (good_reviews + bad_reviews), 4) if (good_reviews + bad_reviews) else 0
            turnover = round(random.uniform(2.0, 6.0) * mult, 1)
            occupancy = round(random.uniform(0.50, 0.95) * mult, 4)
            occupancy = min(occupancy, 1.0)

            row_data = [
                rec_date.isoformat(), store, dine_rev, dlvr_rev, group_rev, total_rev,
                dine_orders, dlvr_orders, group_orders, total_orders,
                dine_traffic, dlvr_traffic, total_traffic, avg_ticket,
                food_cost, labor_cost, rent, utility, other,
                total_cost, gross_profit, gross_margin, net_profit, net_margin,
                good_reviews, bad_reviews, good_rate, turnover, occupancy,
            ]

            if use_write_only:
                ws.append(row_data)
            else:
                row_num = data_start_row + count
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.border = border
                    if ci in pct_cols:
                        cell.number_format = "0.00%"
            count += 1

        if count % 50000 < len(RESTAURANT_STORES):
            print(f"    日运营数据: {count:,}/{n_rows:,} rows...")


def gen_restaurant_hourly_detail(wb, n_rows, seed, use_write_only=False):
    """餐饮时段明细 — 每小时每门店客流和销售 + merge header。"""
    random.seed(seed + 2)
    ws = wb.create_sheet(title="时段明细")

    headers = [
        "日期", "门店", "时段", "堂食客流", "外卖订单",
        "堂食营业额", "外卖营业额", "总营业额",
        "在岗人数", "人效(元/人)", "翻台次数", "等位数",
    ]

    data_start_row = 1
    if not use_write_only:
        border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        # Level 1: major groups
        _style_header(ws, 1, 1, "时间地点", merge_end_col=3)
        _style_header(ws, 1, 4, "客流量", merge_end_col=5)
        _style_header(ws, 1, 6, "营业额", merge_end_col=8)
        _style_header(ws, 1, 9, "人效与翻台", merge_end_col=12)
        # Level 2: column headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        data_start_row = 3
    else:
        ws.append(headers)

    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hours = list(range(10, 23))  # 10:00 - 22:00
    base_date = date(2025, 1, 1)
    count = 0

    for d in range(366):
        if count >= n_rows:
            break
        rec_date = base_date + timedelta(days=d)
        for store in RESTAURANT_STORES:
            for hour in hours:
                if count >= n_rows:
                    break
                # Peak hours: 11-13, 17-20
                is_peak = hour in (11, 12, 13, 17, 18, 19, 20)
                mult = 2.5 if is_peak else 1.0

                dine_traffic = int(random.uniform(2, 25) * mult)
                dlvr_orders = int(random.uniform(1, 15) * mult)
                dine_rev = round(dine_traffic * random.uniform(35, 85), 2)
                dlvr_rev = round(dlvr_orders * random.uniform(25, 65), 2)
                total_rev = round(dine_rev + dlvr_rev, 2)
                staff = int(random.uniform(3, 12) * (1.5 if is_peak else 1.0))
                per_person = round(total_rev / staff, 2) if staff else 0
                turns = round(random.uniform(0, 3) * mult, 1)
                waiting = int(random.uniform(0, 10) * mult) if is_peak else 0

                row_data = [
                    rec_date.isoformat(), store, f"{hour}:00-{hour + 1}:00",
                    dine_traffic, dlvr_orders, dine_rev, dlvr_rev, total_rev,
                    staff, per_person, turns, waiting,
                ]

                if use_write_only:
                    ws.append(row_data)
                else:
                    row_num = data_start_row + count
                    for ci, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=ci, value=val)
                        cell.border = border
                count += 1

            if count % 50000 < len(hours):
                print(f"    时段明细: {count:,}/{n_rows:,} rows...")


# ============================================================
# 主函数
# ============================================================

def generate(level="L1", data_type="factory", seed=42):
    """生成压测 Excel 文件。"""
    if level not in LEVELS:
        raise ValueError(f"Unknown level: {level}. Use: {list(LEVELS.keys())}")

    n_rows = LEVELS[level]["rows"]
    label = LEVELS[level]["label"]
    print(f"\n{'='*60}")
    print(f"压测数据生成: {level} ({label}) — {data_type}")
    print(f"目标行数: {n_rows:,}")
    print(f"{'='*60}")

    start = time.time()

    # Use write_only mode for L2+ to avoid OOM
    use_write_only = n_rows >= 100_000
    wb = Workbook(write_only=use_write_only)

    # Remove default sheet if not write_only
    if not use_write_only and wb.active:
        wb.remove(wb.active)

    # Split rows across sheets (each sheet gets ~33% of total rows)
    rows_per_sheet = n_rows // 3

    if data_type == "factory":
        print(f"\n生成工厂端数据...")
        gen_factory_sales_detail(wb, rows_per_sheet, seed, use_write_only)
        gen_factory_expense_ledger(wb, rows_per_sheet, seed, use_write_only)
        gen_factory_inventory(wb, n_rows - 2 * rows_per_sheet, seed, use_write_only)
    else:
        print(f"\n生成餐饮端数据...")
        gen_restaurant_orders(wb, rows_per_sheet, seed, use_write_only)
        gen_restaurant_daily_ops(wb, rows_per_sheet, seed, use_write_only)
        gen_restaurant_hourly_detail(wb, n_rows - 2 * rows_per_sheet, seed, use_write_only)

    # Save
    output_dir = Path(__file__).parent / "test-data" / "stress"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"Stress-{data_type}-{level}-s{seed}.xlsx"
    filepath = output_dir / filename

    print(f"\n保存文件: {filepath}")
    wb.save(str(filepath))
    elapsed = time.time() - start

    # File size
    size_mb = filepath.stat().st_size / (1024 * 1024)
    mode = LEVELS[level].get("mode", "upload")
    print(f"完成! 耗时: {elapsed:.1f}s, 文件大小: {size_mb:.1f} MB")
    print(f"  测试模式: {mode}")
    if mode == "api-direct":
        print(f"  [!] 此文件超过 10MB 上传限制，仅用于 Python API 直接调用测试")
        print(f"  用法: openpyxl.load_workbook() → pandas → 直接调用 /api/smartbi/chart/batch-build")
    elif mode == "upload":
        print(f"  [OK] 此文件可通过前端上传端点进行完整 E2E 测试")
        if not use_write_only:
            print(f"  [OK] 包含 3 级合并表头 + 百分比格式 + 混合类型列")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for sn in wb.sheetnames:
        print(f"    {sn}")

    wb.close()
    return filepath


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SmartBI 压测数据生成器")
    parser.add_argument("--level", choices=list(LEVELS.keys()), default="L1",
                        help="压测级别: L1=50K, L2=200K, L3=500K, L4=1M")
    parser.add_argument("--type", choices=["factory", "restaurant"], default="factory",
                        help="数据类型: factory=工厂, restaurant=餐饮")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--all", action="store_true", help="生成全部级别×类型")
    args = parser.parse_args()

    if args.all:
        for lvl in LEVELS:
            for dt in ["factory", "restaurant"]:
                generate(lvl, dt, args.seed)
        print(f"\n全部生成完成: {len(LEVELS) * 2} 个文件")
    else:
        generate(args.level, args.type, args.seed)
