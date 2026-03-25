"""Generate 六扇门第一期 detailed manual test plan Excel with exact data inputs."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
hfont = Font(bold=True, size=11, color="FFFFFF")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
role_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

def header(ws, cols, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for i, (name, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = hfont; c.fill = fill; c.alignment = ctr; c.border = bdr
        ws.column_dimensions[chr(64+i) if i <= 26 else "A"].width = w
    ws.freeze_panes = "A2"

def role_row(ws, row, text):
    for col in range(1, 9):
        c = ws.cell(row=row, column=col); c.fill = role_fill; c.border = bdr
        c.font = Font(bold=True, size=11)
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

def rows(ws, data, start):
    for ri, row in enumerate(data, start):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr; c.alignment = wrap

# ==================== Sheet 1 ====================
ws1 = wb.active; ws1.title = "完整测试用例"
header(ws1, [("编号",6),("角色/账号",20),("平台",8),("操作描述",50),("输入数据/具体操作",45),("预期结果",40),("通过?",7),("备注",18)], "2F5496")

r = 2
# --- dispatcher ---
role_row(ws1, r, "Phase A: 调度员排产 — 登录 dispatcher1 / 123456"); r += 1
t = [
["T01","dispatcher1/123456","RN","打开App > 输入账号密码 > 登录","账号: dispatcher1\n密码: 123456","进入紫色主题调度员界面, 底部Tab: 首页/排程/AI/人员/我的","",""],
["T02","dispatcher1","RN","排程Tab > 点右上角\"+\"创建计划","产品: 选\"墨鱼圈\"\n计划量: 100\n计划日期: 明天\n来源: 手动创建\n优先级: 中","创建成功, 列表出现新计划, 状态\"待执行\"","产品必须选已有的"],
["T03","dispatcher1","RN","点击刚创建的计划 > 进入详情","(无需输入)","显示: 墨鱼圈 / 100kg / 待执行 / 进度0%","",""],
["T04","dispatcher1","RN","底部点击绿色\"生成工序任务\"按钮","(无需输入, 点确定)","弹确认\"确定为墨鱼圈生成4个工序任务?\" > 确定 > \"已生成4个工序任务\"","","[新按钮] 4个=拆箱/卤制/包装/质检"],
["T05","dispatcher1","RN","任务分配页 > 查看生成的任务","(无需输入)","看到4个工序任务:\n1.拆箱分拣(箱)\n2.卤制加工(kg)\n3.真空包装(袋)\n4.质检抽样(批)","","验证单位不同"],
]
rows(ws1, t, r); r += len(t)

# --- workshop_sup 签到 ---
role_row(ws1, r, "Phase B: 车间主管签到 — 登录 workshop_sup1 / 123456"); r += 1
t = [
["T06","workshop_sup1/123456","RN","退出登录 > 重新登录车间主管","账号: workshop_sup1\n密码: 123456","进入车间主管界面","",""],
["T07","workshop_sup1","RN","进入NFC签到页面","(无需输入)","看到工序任务下拉列表, 包含刚生成的4个工序","",""],
["T08","workshop_sup1","RN","选择\"卤制加工\"工序 > 点\"扫码签到\" > 扫描QR码","QR码内容: 003\n(workshop_sup1的工号)","\"签到成功: workshop_sup1 已签到 卤制加工\"","","003=workshop_sup1工号"],
["T09","workshop_sup1","RN","等2分钟 > 点\"签退\" > 扫同一QR码","QR码内容: 003","\"签退成功\", 显示工时记录","","验证工时 >= 2分钟"],
]
rows(ws1, t, r); r += len(t)

# --- workshop_sup 三步报工 ---
role_row(ws1, r, "Phase C: 三步报工 [新功能] — 继续使用 workshop_sup1"); r += 1
t = [
["T10","workshop_sup1","RN","工序任务列表 > 右上角扫码图标","(点击图标)","进入\"报工\"页面, 显示3步指示器: 扫人 > 工序 > 报量","","[新页面] ThreeStepReport"],
["T11","workshop_sup1","RN","Step1: 点\"主管自己报工(跳过)\"","(点击跳过链接)","步骤指示器更新, 进入Step2","","跳过扫人"],
["T12","workshop_sup1","RN","Step2: 查看工序卡片","(无需输入)","显示4个工序卡片:\n卤制加工 - 计划100 完成0\n每张卡片有进度条","","验证字体: 工序名20px"],
["T13","workshop_sup1","RN","Step2: 点击\"卤制加工\"卡片","(点击卡片)","进入Step3, 顶部摘要:\n员工: 主管自己\n工序: 卤制加工\n剩余: 100 kg","",""],
["T14","workshop_sup1","RN","Step3: 输入产出 > 提交","产出数量: 50\n(投入量和备注可不填)","\"报工成功\"\n卤制加工\n产出: 50 kg\n按钮: 继续报工 / 返回","",""],
["T15","workshop_sup1","RN","点\"继续报工\" > 再报一次","Step1跳过 > Step2选卤制 > Step3输入: 30","\"报工成功\" 产出: 30 kg","","累加测试: 50+30=80"],
]
rows(ws1, t, r); r += len(t)

# --- workshop_sup 卡片报工 ---
role_row(ws1, r, "Phase D: 卡片直接报工 — 继续使用 workshop_sup1"); r += 1
t = [
["T16","workshop_sup1","RN","返回工序列表 > 点\"卤制加工\"卡片 > 进入详情","(点击卡片)","详情页显示:\n计划量: 100\n已完成: 0 (未审批)\n待审批: 80 (50+30)\n进度条","","数量字体26px"],
["T17","workshop_sup1","RN","查看\"投入产出对比\"区域","(无需输入)","如果有inputQuantity则显示:\n投入量 / 产出量 / 转化率(颜色)","","出成率卡片"],
["T18","workshop_sup1","RN","点\"报工\"按钮 > 输入产出","产出数量: 20\n备注: 最后一批","\"报工已提交，等待审批\"","",""],
["T19","workshop_sup1","RN","返回详情 > 查看报工记录","(无需输入)","3条报工记录:\n1. 50kg 03-25 XX:XX 待审批\n2. 30kg 03-25 XX:XX 待审批\n3. 20kg 03-25 XX:XX 待审批\n时间格式: MM-DD HH:mm","","[新] 时间戳格式"],
["T20","workshop_sup1","RN","确认待审批总量","(无需输入)","待审批 = 100 (50+30+20)","",""],
]
rows(ws1, t, r); r += len(t)

# --- factory_admin 审批 ---
role_row(ws1, r, "Phase E: 管理员审批 — 登录 factory_admin1 / 123456"); r += 1
t = [
["T21","factory_admin1/123456","RN","退出 > 登录管理员","账号: factory_admin1\n密码: 123456","进入管理员界面","",""],
["T22","factory_admin1","RN","工序任务 > 审批图标(clipboard)","(点击图标)","待审批列表: 3条报工记录\n(50kg + 30kg + 20kg)","",""],
["T23","factory_admin1","RN","选第一条(50kg) > 点\"通过\"","(点击通过)","\"审批成功\"","",""],
["T24","factory_admin1","RN","全选剩余 > \"批量通过\"","(勾选30kg+20kg > 批量通过)","\"2条审批成功\"","",""],
["T25","factory_admin1","RN","返回工序详情 > 确认数据","(无需输入)","已完成: 100\n待审批: 0\n进度: 100%\n(后端已自动执行BOM扣料+成本计算)","","自动链路触发"],
]
rows(ws1, t, r); r += len(t)

# --- warehouse_mgr 入库 ---
role_row(ws1, r, "Phase F: 仓储入库 — 登录 warehouse_mgr1 / 123456"); r += 1
t = [
["T26","warehouse_mgr1/123456","RN","退出 > 登录仓储主管","账号: warehouse_mgr1\n密码: 123456","进入仓储界面","",""],
["T27","warehouse_mgr1","RN","原料入库 > 手动入库","原料类型: GPS牛腩\n数量: 500\n单位: kg\n单价: 28\n供应商: 六扇门供应商","\"入库成功\"\n批次号自动生成","移动均价=28"],
["T28","warehouse_mgr1","RN","再次入库(不同价格)","原料类型: GPS牛腩\n数量: 200\n单位: kg\n单价: 32\n供应商: 六扇门供应商","\"入库成功\"\n移动均价自动更新为:\n(500*28+200*32)/700=29.14","验证移动均价"],
["T29","warehouse_mgr1","RN","AI辅助入库","在AI对话框输入:\n\"酱油入库100公斤 单价8元\"","AI识别: 物料=酱油, 数量=100, 单价=8\n自动填表 > 确认 > 入库成功","AI自然语言"],
]
rows(ws1, t, r); r += len(t)

# --- Web-Admin ---
role_row(ws1, r, "Phase G: Web-Admin分析 — 浏览器登录 factory_admin1 / 123456"); r += 1
t = [
["T30","factory_admin1/123456","Web","打开浏览器 > 登录Web-Admin","URL: http://139.196.165.140:8086\n账号: factory_admin1\n密码: 123456","进入管理后台","",""],
["T31","factory_admin1","Web","系统管理 > 工序管理","(查看列表)","显示4个工序:\n拆箱分拣(箱) / 卤制加工(kg) / 真空包装(袋) / 质检抽样(批)\n可编辑单位","验证单位可改"],
["T32","factory_admin1","Web","系统管理 > 产品工序关联 > 选\"墨鱼圈\"","(选择产品)","显示4个已关联工序, 可上移下移排序\n有\"生成工序任务\"按钮","",""],
["T33","factory_admin1","Web","生产管理 > BOM达成率","(查看页面)","显示墨鱼圈批次的:\n每种物料: 计划量 vs 实际消耗 vs 差异%\n总体达成率%","","依赖T25审批后触发"],
["T34","factory_admin1","Web","财务 > SKU毛利率","(查看页面)","显示产品收入/成本/毛利率","",""],
]
rows(ws1, t, r); r += len(t)

# ==================== Sheet 2: 自动链路 ====================
ws2 = wb.create_sheet("自动链路验证")
header(ws2, [("编号",6),("触发动作",30),("自动行为",35),("验证方式",45),("通过?",7)], "548235")
chains = [
["C01","入库T27: GPS牛腩500kg@28","movingAvgPrice = 28.00","Web-Admin > 原料类型GPS牛腩 > 移动均价字段",""],
["C02","入库T28: GPS牛腩200kg@32","movingAvgPrice = (500*28+200*32)/700 = 29.14","Web-Admin > 原料类型GPS牛腩 > 移动均价更新",""],
["C03","审批T23-T24通过100kg","autoConsumeForBatch()自动触发","后端日志: grep \"自动扣料完成\" cretas-prod.log",""],
["C04","自动扣料","BOM展开: GPS牛腩0.5kg*100=50kg 酱油0.05*100=5kg...","原料批次剩余量: GPS牛腩从500→450",""],
["C05","自动扣料","MaterialConsumption记录(sourceType=AUTO_BOM)","Web-Admin > 物料消耗记录 > 查看AUTO_BOM类型记录",""],
["C06","自动扣料","batch.materialCost = SUM(消耗*单价)","Web-Admin > 生产批次详情 > 物料成本字段有值",""],
["C07","查BOM达成率T33","getConsumptionSummary()返回overallAchievementRate","页面显示: 达成率% + 每种物料variance%",""],
["C08","查SKU毛利率T34","SkuGrossMarginTool计算","页面显示: 收入 - 成本 = 毛利率",""],
]
rows(ws2, chains, 2)

# ==================== Sheet 3: 测试数据清单 ====================
ws3 = wb.create_sheet("测试数据清单")
header(ws3, [("类型",12),("ID/账号",20),("名称/值",20),("说明",30)], "7030A0")
data = [
["账号","factory_admin1 / 123456","工厂管理员 (工号001)","审批+分析+Web-Admin"],
["账号","dispatcher1 / 123456","调度员 (工号002)","排产+任务分配"],
["账号","workshop_sup1 / 123456","车间主管 (工号003)","签到+报工"],
["账号","warehouse_mgr1 / 123456","仓储主管 (工号004)","原料入库"],
["账号","quality_insp1 / 123456","品控员 (工号005)","二期品控"],
["产品","PT-F001-003","墨鱼圈","主测试产品"],
["工序","WP001","拆箱分拣 (单位:箱)","第1道工序"],
["工序","WP002","卤制加工 (单位:kg)","第2道工序"],
["工序","WP003","真空包装 (单位:袋)","第3道工序"],
["工序","WP004","质检抽样 (单位:批)","第4道工序"],
["BOM","墨鱼圈 (8项)","GPS牛腩0.5kg + 酱油0.05kg + 花椒0.02kg + ...","已创建"],
["原料","GPS牛腩","28元/kg","主料"],
["原料","酱油","8元/kg","辅料"],
["原料","花椒","35元/kg","辅料"],
["原料","八角","25元/kg","辅料"],
["原料","包装袋(真空)","0.5元/个","包材"],
["客户","CUS-YONGYOU","永佑食品","联系人:张权"],
["客户","CUS-F001-001","永辉超市","已有"],
["客户","CUS-F001-002","盒马鲜生","已有"],
["Web-Admin","http://139.196.165.140:8086","管理后台","浏览器打开"],
]
rows(ws3, data, 2)

path = "C:/Users/Steve/my-prototype-logistics/tests/六扇门第一期-手动测试计划.xlsx"
wb.save(path)
print(f"Done: {path}")
