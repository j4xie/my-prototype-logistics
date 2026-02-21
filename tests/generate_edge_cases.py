#!/usr/bin/env python3
"""
SmartBI 边界条件独立测试文件生成器。

每个边界场景生成一个独立的小型 Excel 文件，可单独上传测试。
与 generate_test_excel.py / generate_restaurant_excel.py 中嵌入的边界 sheet 不同，
这里每个文件只包含 1 个目标 sheet，便于隔离测试和定位问题。

Usage:
    python tests/generate_edge_cases.py              # 生成所有边界文件
    python tests/generate_edge_cases.py --case wide  # 只生成超宽表
"""
import argparse
import random
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Styles
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

OUTPUT_DIR = Path(__file__).parent / "test-data" / "edge-cases"


def style_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def gen_wide_table(seed=42):
    """超宽表 — 120+ 列，3级合并表头。"""
    random.seed(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "超宽数据表"

    # 12 products × 10 metrics = 120 data columns + 1 label column = 121 total
    products = [f"产品{chr(65+i)}" for i in range(12)]
    metrics = ["销量", "收入", "成本", "毛利", "毛利率", "退货量", "退货率", "库存", "周转天数", "预测销量"]
    pct_metrics = {"毛利率", "退货率"}

    # Level 1: product groups
    col = 2
    style_cell(ws, 1, 1, "区域/产品", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    for prod in products:
        end_col = col + len(metrics) - 1
        style_cell(ws, 1, col, prod, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        # Level 2: metric sub-groups
        style_cell(ws, 2, col, "销售指标", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 4)
        style_cell(ws, 2, col + 5, "退货指标", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=2, start_column=col + 5, end_row=2, end_column=col + 6)
        style_cell(ws, 2, col + 7, "库存指标", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=2, start_column=col + 7, end_row=2, end_column=col + 9)
        # Level 3: individual metrics
        for mi, m in enumerate(metrics):
            style_cell(ws, 3, col + mi, m, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        col += len(metrics)

    # Data rows: 50 regions
    regions = [f"{'华东华南华北西南华中东北西北'[i%7*2:(i%7+1)*2]}_{i+1:02d}" for i in range(50)]
    for ri, region in enumerate(regions, 4):
        style_cell(ws, ri, 1, region, border=THIN_BORDER)
        col = 2
        for _ in products:
            for mi, m in enumerate(metrics):
                if m in pct_metrics:
                    val = round(random.uniform(0.05, 0.55), 4)
                    style_cell(ws, ri, col + mi, val, border=THIN_BORDER, number_format="0.0%")
                elif m in ("周转天数",):
                    val = round(random.uniform(5, 90), 1)
                    style_cell(ws, ri, col + mi, val, border=THIN_BORDER)
                else:
                    val = round(random.uniform(100, 50000), 2)
                    style_cell(ws, ri, col + mi, val, border=THIN_BORDER, number_format="#,##0.00")
            col += len(metrics)

    filepath = OUTPUT_DIR / "Edge-wide-120col.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


def gen_mixed_types(seed=42):
    """混合类型列 — 同一列中包含数字、文字、None、布尔值、日期。"""
    random.seed(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "混合类型测试"

    headers = ["编号", "名称", "金额", "状态", "日期", "备注"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # 30 rows with intentional type mixing
    for ri in range(2, 32):
        style_cell(ws, ri, 1, ri - 1, border=THIN_BORDER)
        style_cell(ws, ri, 2, f"项目{ri-1}", border=THIN_BORDER)

        # 金额列: mostly numeric, but some strings/None
        if ri % 7 == 0:
            style_cell(ws, ri, 3, "待确认", border=THIN_BORDER)
        elif ri % 11 == 0:
            style_cell(ws, ri, 3, None, border=THIN_BORDER)
        elif ri % 13 == 0:
            style_cell(ws, ri, 3, "N/A", border=THIN_BORDER)
        else:
            style_cell(ws, ri, 3, round(random.uniform(1000, 99999), 2),
                       border=THIN_BORDER, number_format="#,##0.00")

        # 状态列: mixed boolean/string/int
        status_pool = [True, False, 1, 0, "是", "否", "已完成", "进行中", None, "待审批"]
        style_cell(ws, ri, 4, random.choice(status_pool), border=THIN_BORDER)

        # 日期列: mixed date objects and date strings
        if ri % 5 == 0:
            style_cell(ws, ri, 5, f"2025-{random.randint(1,12):02d}-{random.randint(1,28):02d}",
                       border=THIN_BORDER)
        elif ri % 8 == 0:
            style_cell(ws, ri, 5, None, border=THIN_BORDER)
        else:
            d = date(2025, 1, 1) + timedelta(days=random.randint(0, 364))
            style_cell(ws, ri, 5, d, border=THIN_BORDER, number_format="YYYY-MM-DD")

        # 备注列: very mixed
        notes = [None, None, 42, 3.14, True, "正常", "异常!", "⚠️警告", "", 0, -1, "合格率95%"]
        style_cell(ws, ri, 6, random.choice(notes), border=THIN_BORDER)

    filepath = OUTPUT_DIR / "Edge-mixed-types.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


def gen_empty_regions(seed=42):
    """空值区域 — 模拟 data_only=True 场景，大面积 None + 全空行 + 全零列。"""
    random.seed(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空值区域测试"

    headers = ["指标"] + [f"{m+1}月" for m in range(12)] + ["合计", "备注"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    metrics = ["营业收入", "营业成本", "毛利", "销售费用", "管理费用", "研发费用",
               "财务费用", "其他收入", "营业利润", "所得税", "净利润", "EBITDA",
               "", "现金流量", "应收账款"]  # row 13 = empty separator

    for ri, metric in enumerate(metrics, 2):
        style_cell(ws, ri, 1, metric if metric else None, border=THIN_BORDER)

        if metric == "":
            # Full empty row
            for ci in range(2, 16):
                style_cell(ws, ri, ci, None, border=THIN_BORDER)
            continue

        for ci in range(2, 14):  # months
            # 30% chance of None (simulating formula cells)
            if random.random() < 0.30:
                style_cell(ws, ri, ci, None, border=THIN_BORDER)
            else:
                style_cell(ws, ri, ci, round(random.uniform(-50000, 500000), 2),
                           border=THIN_BORDER, number_format="#,##0.00")

        # 合计 column: always None (formula would be here)
        style_cell(ws, ri, 14, None, border=THIN_BORDER)
        # 备注 column: always zero (全零列)
        style_cell(ws, ri, 15, 0, border=THIN_BORDER)

    filepath = OUTPUT_DIR / "Edge-empty-regions.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


def gen_formula_cells(seed=42):
    """公式测试 — SUM/AVERAGE/IF 公式，测试 data_only=True 返回 None。"""
    random.seed(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "公式测试"

    headers = ["产品"] + [f"{m+1}月" for m in range(12)] + ["合计", "平均", "达标"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    products = ["火锅底料", "复合调味料", "酱料", "速食产品", "预制菜", "休闲零食",
                "饮品", "冷冻食品"]
    for ri, prod in enumerate(products, 2):
        style_cell(ws, ri, 1, prod, border=THIN_BORDER)
        for mi in range(12):
            val = round(random.uniform(30000, 300000), 2)
            style_cell(ws, ri, mi + 2, val, border=THIN_BORDER, number_format="#,##0.00")
        # Formulas
        ws.cell(row=ri, column=14).value = f"=SUM(B{ri}:M{ri})"
        ws.cell(row=ri, column=14).border = THIN_BORDER
        ws.cell(row=ri, column=15).value = f"=AVERAGE(B{ri}:M{ri})"
        ws.cell(row=ri, column=15).border = THIN_BORDER
        ws.cell(row=ri, column=16).value = f'=IF(N{ri}>1500000,"达标","未达标")'
        ws.cell(row=ri, column=16).border = THIN_BORDER

    # Total row with formulas
    total_row = 2 + len(products)
    style_cell(ws, total_row, 1, "合计", HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 17):
        col_letter = get_column_letter(ci)
        ws.cell(row=total_row, column=ci).value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        ws.cell(row=total_row, column=ci).border = THIN_BORDER

    # Percentage row
    pct_row = total_row + 1
    style_cell(ws, pct_row, 1, "月度占比", HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 14):
        col_letter = get_column_letter(ci)
        ws.cell(row=pct_row, column=ci).value = f"={col_letter}{total_row}/N{total_row}"
        ws.cell(row=pct_row, column=ci).border = THIN_BORDER
        ws.cell(row=pct_row, column=ci).number_format = "0.0%"

    filepath = OUTPUT_DIR / "Edge-formula-cells.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


def gen_numeric_colnames(seed=42):
    """纯数字列名 — 年份作为列标题的整数值，测试 header vs data 判断。"""
    random.seed(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "纯数字列名"

    years = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
    style_cell(ws, 1, 1, "指标", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for ci, yr in enumerate(years, 2):
        # Integer year as column header — triggers 30% numeric threshold
        style_cell(ws, 1, ci, yr, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    metrics = [
        ("营业收入(万元)", False), ("营业成本(万元)", False), ("毛利(万元)", False),
        ("毛利率", True), ("净利润(万元)", False), ("净利率", True),
        ("员工数", False), ("人均产值(万元)", False),
        ("研发费用(万元)", False), ("研发占比", True),
        ("资产总额(万元)", False), ("负债率", True),
        ("ROE", True), ("ROA", True),
    ]

    for ri, (metric_name, is_pct) in enumerate(metrics, 2):
        style_cell(ws, ri, 1, metric_name, border=THIN_BORDER)
        for ci, _ in enumerate(years, 2):
            if is_pct:
                val = round(random.uniform(0.02, 0.50), 4)
                style_cell(ws, ri, ci, val, border=THIN_BORDER, number_format="0.0%")
            else:
                val = round(random.uniform(500, 50000), 1)
                style_cell(ws, ri, ci, val, border=THIN_BORDER, number_format="#,##0.0")

    filepath = OUTPUT_DIR / "Edge-numeric-colnames.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


def gen_cross_year(seed=42):
    """跨年对比 — 2024 vs 2025 数据，3级合并表头，YoY 百分比列。"""
    random.seed(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "跨年对比"

    # 3-level merge header
    style_cell(ws, 1, 1, "区域/年度", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells("A1:A3")

    col = 2
    for year in [2024, 2025]:
        style_cell(ws, 1, col, f"{year}年度", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        for qi, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
            style_cell(ws, 2, col + qi, q, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        style_cell(ws, 3, col, "收入", HEADER_FONT, border=THIN_BORDER)
        style_cell(ws, 3, col + 1, "成本", HEADER_FONT, border=THIN_BORDER)
        style_cell(ws, 3, col + 2, "毛利", HEADER_FONT, border=THIN_BORDER)
        style_cell(ws, 3, col + 3, "毛利率", HEADER_FONT, border=THIN_BORDER)
        col += 4

    # YoY columns
    style_cell(ws, 1, col, "同比变化", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col + 1)
    style_cell(ws, 3, col, "收入增长率", HEADER_FONT, border=THIN_BORDER)
    style_cell(ws, 3, col + 1, "毛利增长率", HEADER_FONT, border=THIN_BORDER)

    regions = ["华东", "华南", "华北", "西南", "华中", "东北", "西北", "海外",
               "线上直营", "线上加盟"]
    for ri, region in enumerate(regions, 4):
        style_cell(ws, ri, 1, region, border=THIN_BORDER)
        rev_2024, rev_2025 = [], []
        base_col = 2
        for year_vals, year_list in [(rev_2024, [2024]), (rev_2025, [2025])]:
            for qi in range(4):
                rev = round(random.uniform(100000, 800000), 2)
                cost = round(rev * random.uniform(0.55, 0.75), 2)
                profit = round(rev - cost, 2)
                margin = round(profit / rev, 4) if rev else 0
                year_vals.append(rev)
                style_cell(ws, ri, base_col, rev, border=THIN_BORDER, number_format="#,##0.00")
                style_cell(ws, ri, base_col + 1, cost, border=THIN_BORDER, number_format="#,##0.00")
                style_cell(ws, ri, base_col + 2, profit, border=THIN_BORDER, number_format="#,##0.00")
                style_cell(ws, ri, base_col + 3, margin, border=THIN_BORDER, number_format="0.0%")
                base_col += 4
                if qi == 3:  # reset for next year block
                    pass
            if year_list[0] == 2024:
                base_col = 6  # start of 2025 block

        # YoY
        total_24 = sum(rev_2024)
        total_25 = sum(rev_2025)
        yoy_rev = round((total_25 / total_24 - 1), 4) if total_24 else 0
        yoy_profit = round(random.uniform(-0.20, 0.40), 4)
        style_cell(ws, ri, col, yoy_rev, border=THIN_BORDER, number_format="+0.0%;-0.0%")
        style_cell(ws, ri, col + 1, yoy_profit, border=THIN_BORDER, number_format="+0.0%;-0.0%")

    filepath = OUTPUT_DIR / "Edge-cross-year-yoy.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


# Registry of all edge case generators
EDGE_CASES = {
    "wide": ("超宽表(120+列3级合并)", gen_wide_table),
    "mixed": ("混合类型列", gen_mixed_types),
    "empty": ("空值区域(None+全空行+全零列)", gen_empty_regions),
    "formula": ("公式单元格(SUM/AVERAGE/IF)", gen_formula_cells),
    "numeric-col": ("纯数字列名(年份整数)", gen_numeric_colnames),
    "cross-year": ("跨年对比(2024v2025+YoY)", gen_cross_year),
}


def main():
    parser = argparse.ArgumentParser(description="SmartBI 边界条件独立测试文件生成器")
    parser.add_argument("--case", choices=list(EDGE_CASES.keys()),
                        help="只生成指定边界场景")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.case:
        desc, func = EDGE_CASES[args.case]
        filepath = func(args.seed)
        print(f"Generated: {filepath} ({desc})")
    else:
        print(f"生成所有 {len(EDGE_CASES)} 个边界测试文件...\n")
        for key, (desc, func) in EDGE_CASES.items():
            filepath = func(args.seed)
            size_kb = filepath.stat().st_size / 1024
            print(f"  {key:15s} {desc:30s} → {filepath.name} ({size_kb:.0f} KB)")
        print(f"\n输出目录: {OUTPUT_DIR}")
        print(f"总计: {len(EDGE_CASES)} 个文件")


if __name__ == "__main__":
    main()
