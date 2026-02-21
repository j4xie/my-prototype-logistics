#!/usr/bin/env python3
"""
餐饮行业 SmartBI E2E 测试 Excel 生成器
模仿青花椒酸菜鱼的收入管理报表和订单销售明细表结构

Usage:
    python tests/generate_restaurant_excel.py                      # 默认：青花椒模板
    python tests/generate_restaurant_excel.py --template hotpot    # 火锅品牌
    python tests/generate_restaurant_excel.py --template bakery    # 烘焙连锁
    python tests/generate_restaurant_excel.py --seed 123           # 指定随机种子
    python tests/generate_restaurant_excel.py --scenario growth    # 高增长场景
    python tests/generate_restaurant_excel.py --all                # 生成全部组合
"""
import argparse
import random
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 模板定义 — 餐饮品牌
# ============================================================

TEMPLATES = {
    "fish": {
        "brand": "青花椒",
        "company": "青花椒砂锅鱼",
        "city": "上海市",
        "province": "上海市",
        "region": "华东地区",
        "stores": [
            "青花椒南方百联店", "青花椒徐汇光启城店", "青花椒成山巴春店",
            "青花椒颛桥万达店", "青花椒南桥百联店", "青花椒徐汇日月光店",
            "青花椒大丸百货店", "青花椒长寿巴春店", "青花椒川沙百联店",
            "青花椒莘庄龙之梦店", "青花椒七宝凯德店", "青花椒松江万达店",
            "青花椒金山百联店", "青花椒浦东三钢里店", "青花椒宝山龙湖店",
        ],
        "store_codes": {
            "青花椒南方百联店": "CS003001121", "青花椒徐汇光启城店": "8004161",
            "青花椒成山巴春店": "CS003001121", "青花椒颛桥万达店": "CS003025525",
        },
        "menu_items": [
            ("招牌青花椒鱼(微麻微辣)", 128, "主菜"), ("招牌青花椒鱼(中麻中辣)", 128, "主菜"),
            ("营养多C番茄鱼(两人份)", 108, "主菜"), ("鱼羊鲜双人餐", 255, "套餐"),
            ("儿童套餐A", 18, "套餐"), ("暖冬鱼火锅套餐", 168, "套餐"),
            ("峨边脆笋", 12, "配菜"), ("白灼生菜", 18, "配菜"),
            ("娃娃菜", 8, "配菜"), ("豆腐皮", 8, "配菜"),
            ("四川粉条", 10, "配菜"), ("金针菇", 10, "配菜"),
            ("年糕", 8, "配菜"), ("罗定鱼腐", 12, "配菜"),
            ("手工糍粑", 12, "甜品"), ("榴莲飞饼", 28, "甜品"),
            ("菠萝飞饼", 25, "甜品"), ("小酥肉", 22, "小吃"),
            ("米饭", 3, "主食"), ("大打包盒", 2, "其他"),
        ],
        "avg_spend_dine_in": (70, 200),
        "avg_spend_delivery": (35, 140),
        "daily_orders_per_store": (40, 120),
    },
    "hotpot": {
        "brand": "蜀大侠",
        "company": "蜀大侠火锅",
        "city": "成都市",
        "province": "四川省",
        "region": "西南地区",
        "stores": [
            "蜀大侠春熙路旗舰店", "蜀大侠太古里店", "蜀大侠建设路店",
            "蜀大侠万象城店", "蜀大侠双流万达店", "蜀大侠温江万达店",
            "蜀大侠西安赛格店", "蜀大侠武汉楚河汉街店", "蜀大侠杭州银泰店",
            "蜀大侠重庆解放碑店", "蜀大侠郑州丹尼斯店", "蜀大侠南京虹悦城店",
        ],
        "store_codes": {},
        "menu_items": [
            ("蜀大侠招牌锅底", 88, "锅底"), ("清汤锅底", 38, "锅底"),
            ("鸳鸯锅底", 68, "锅底"), ("番茄锅底", 48, "锅底"),
            ("毛肚", 58, "荤菜"), ("鲜牛肉", 48, "荤菜"),
            ("嫩牛肉", 42, "荤菜"), ("黑猪五花肉", 38, "荤菜"),
            ("虾滑", 38, "荤菜"), ("鹅肠", 48, "荤菜"),
            ("藕片", 12, "素菜"), ("土豆片", 8, "素菜"),
            ("冬瓜", 8, "素菜"), ("豆皮", 10, "素菜"),
            ("宽粉", 12, "主食"), ("红糖糍粑", 22, "甜品"),
            ("冰粉", 12, "甜品"), ("酸梅汤", 8, "饮品"),
            ("王老吉", 8, "饮品"), ("啤酒", 12, "饮品"),
        ],
        "avg_spend_dine_in": (100, 280),
        "avg_spend_delivery": (50, 160),
        "daily_orders_per_store": (30, 90),
    },
    "bakery": {
        "brand": "奈雪の茶",
        "company": "奈雪の茶",
        "city": "深圳市",
        "province": "广东省",
        "region": "华南地区",
        "stores": [
            "奈雪の茶南山海岸城店", "奈雪の茶福田COCO Park店", "奈雪の茶华强北店",
            "奈雪の茶龙华壹方城店", "奈雪の茶宝安壹方城店", "奈雪の茶罗湖万象城店",
            "奈雪の茶广州天河城店", "奈雪の茶广州正佳广场店", "奈雪の茶佛山南海万达店",
            "奈雪の茶东莞国贸城店", "奈雪の茶珠海华发商都店", "奈雪の茶惠州华贸天地店",
        ],
        "store_codes": {},
        "menu_items": [
            ("霸气芝士草莓", 32, "茶饮"), ("霸气橙子", 28, "茶饮"),
            ("鸭屎香宝藏茶", 25, "茶饮"), ("金色山脉", 22, "茶饮"),
            ("满杯红柚", 28, "茶饮"), ("杨枝甘露", 30, "茶饮"),
            ("生椰拿铁", 22, "咖啡"), ("美式咖啡", 18, "咖啡"),
            ("魔法黑森林蛋糕", 35, "烘焙"), ("奶油草莓蛋糕", 32, "烘焙"),
            ("软欧包(原味)", 18, "烘焙"), ("软欧包(芝士)", 22, "烘焙"),
            ("可颂", 15, "烘焙"), ("吐司(半条)", 25, "烘焙"),
            ("蛋挞(4个)", 20, "烘焙"), ("曲奇礼盒", 58, "礼盒"),
            ("冰淇淋", 18, "甜品"), ("水果杯", 22, "鲜果"),
        ],
        "avg_spend_dine_in": (25, 80),
        "avg_spend_delivery": (20, 65),
        "daily_orders_per_store": (80, 250),
    },
}

SCENARIOS = {
    "normal":  {"revenue_mult": 1.0, "noise": 0.10, "sparse_pct": 0.0},
    "growth":  {"revenue_mult": 1.5, "noise": 0.12, "sparse_pct": 0.0},
    "loss":    {"revenue_mult": 0.5, "noise": 0.15, "sparse_pct": 0.0},
    "sparse":  {"revenue_mult": 0.8, "noise": 0.10, "sparse_pct": 0.5},
}

# ============================================================
# 样式常量
# ============================================================

TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True)
SECTION_FONT = Font(name="微软雅黑", size=10, bold=True)
HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="008000")


def style_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell


# ============================================================
# Sheet 1: 订单销售明细表 (模仿 CSV 结构)
# ============================================================

def create_order_detail_sheet(wb, tpl, scn, year=2025):
    """订单销售明细表 — 模仿青花椒200K+行CSV结构，生成缩减版(核心30列)。"""
    ws = wb.create_sheet("订单销售明细表")
    scale = scn["revenue_mult"]

    # Title rows (mimicking CSV metadata)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    style_cell(ws, 1, 1, "订单销售明细表", TITLE_FONT, alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    style_cell(ws, 2, 1, f"门店名称:{tpl['company']}", alignment=Alignment(horizontal="left"))

    # Header row
    headers = [
        "门店名称", "营业日期", "省份", "城市", "大区", "品牌", "账单号",
        "订单状态", "订单类型", "区域", "桌位", "订单来源", "开单时间", "结单时间",
        "班次", "服务员", "收银员", "客流量", "人均消费", "营业额", "折扣率",
        "折扣额", "应收金额", "实收额", "实收(不计)", "商品结账总数",
        "商品信息", "商品折前金额", "商品折后金额", "外送费",
    ]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 3, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Seating zones per store
    zones_dine = ["A区", "B区", "C区", "大厅", "包间"]
    tables_by_zone = {"A区": 10, "B区": 12, "C区": 8, "大厅": 6, "包间": 4}
    delivery_sources = ["美团外卖", "饿了么", "京东外卖"]
    dine_sources = ["店内桌位单", "微信", "支付宝"]
    surnames = "张王李赵陈刘杨黄周吴徐孙马朱胡林郭何高罗"
    given_names = "伟芳娜秀英敏静丽强磊洋勇艳杰娟涛明超"

    stores = tpl["stores"][:8] if scn["sparse_pct"] == 0 else tpl["stores"][:4]
    menu = tpl["menu_items"]
    lo_dine, hi_dine = tpl["avg_spend_dine_in"]
    lo_dlvr, hi_dlvr = tpl["avg_spend_delivery"]
    lo_orders, hi_orders = tpl["daily_orders_per_store"]

    all_orders = []
    # Generate orders for 30 days
    n_days = 30 if scn["sparse_pct"] == 0 else 15
    for store in stores:
        for day_offset in range(n_days):
            dt = date(year, 1, 1) + timedelta(days=day_offset)
            n_orders = int(random.randint(lo_orders, hi_orders) * scale)
            for oi in range(n_orders):
                is_delivery = random.random() < 0.35
                order_type = "外卖" if is_delivery else "堂食"

                if is_delivery:
                    source = random.choice(delivery_sources)
                    zone = f"无桌位({source})"
                    table = zone
                    n_guests = 1
                    spend_range = (lo_dlvr, hi_dlvr)
                else:
                    source = random.choice(dine_sources)
                    zone = random.choice(zones_dine)
                    tbl_num = random.randint(1, tables_by_zone.get(zone, 8))
                    table = f"{zone[0]}{tbl_num}"
                    n_guests = random.randint(1, 6)
                    spend_range = (lo_dine, hi_dine)

                shift = "午市" if random.random() < 0.45 else "晚市"
                hour = random.randint(10, 13) if shift == "午市" else random.randint(17, 21)
                minute = random.randint(0, 59)
                open_time = datetime(year, dt.month, dt.day, hour, minute, random.randint(0, 59))
                duration = random.randint(20, 90) if not is_delivery else random.randint(15, 60)
                close_time = open_time + timedelta(minutes=duration)

                revenue = round(random.uniform(*spend_range) * n_guests * scale, 2)
                discount_rate = random.choice([1.0, 1.0, 1.0, 0.95, 0.93, 0.90, 0.85]) if not is_delivery else 1.0
                discount_amt = round(revenue * (1 - discount_rate), 2)
                receivable = round(revenue - discount_amt, 2)
                platform_fee = round(receivable * random.uniform(0.15, 0.25), 2) if is_delivery else 0
                actual = round(receivable - platform_fee, 2)

                # Generate menu items
                n_items = random.randint(2, 8)
                chosen_items = random.sample(menu, min(n_items, len(menu)))
                items_str = "+".join(f"{item[0]}_1份*{item[1]}" for item in chosen_items)
                items_total = sum(item[1] for item in chosen_items) * n_guests

                server = random.choice(surnames) + random.choice(given_names)
                cashier = "收银"
                bill_no = f"{dt.strftime('%Y%m%d')}{oi+1:04d}\t"

                all_orders.append({
                    "store": store, "date": dt, "province": tpl["province"],
                    "city": tpl["city"], "region": tpl["region"], "brand": tpl["brand"],
                    "bill": bill_no, "status": "已结账", "type": order_type,
                    "zone": zone, "table": table, "source": source,
                    "open_time": open_time, "close_time": close_time,
                    "shift": shift, "server": server, "cashier": cashier,
                    "guests": n_guests, "avg_spend": round(revenue / max(n_guests, 1), 2),
                    "revenue": revenue, "discount_rate": round(discount_rate, 4),
                    "discount_amt": discount_amt, "receivable": receivable,
                    "actual": actual, "actual_excl": platform_fee,
                    "item_count": round(sum(1 for _ in chosen_items) + random.uniform(0, 3), 1),
                    "items_str": items_str, "items_pre": round(items_total * random.uniform(0.9, 1.1), 2),
                    "items_post": round(items_total * discount_rate, 2),
                    "delivery_fee": round(random.uniform(3, 8), 2) if is_delivery else 0,
                })

    # Sort by date then store
    all_orders.sort(key=lambda x: (x["date"], x["store"], x["open_time"]))

    # Limit to manageable size for Excel
    max_rows = 5000 if scn["sparse_pct"] == 0 else 2000
    all_orders = all_orders[:max_rows]

    # Write data rows
    for ri, o in enumerate(all_orders, 4):
        c = 1
        for val, fmt in [
            (o["store"], None), (o["date"], "YYYY-MM-DD"), (o["province"], None),
            (o["city"], None), (o["region"], None), (o["brand"], None),
            (o["bill"], None), (o["status"], None), (o["type"], None),
            (o["zone"], None), (o["table"], None), (o["source"], None),
            (o["open_time"], "YYYY-MM-DD HH:MM:SS"), (o["close_time"], "YYYY-MM-DD HH:MM:SS"),
            (o["shift"], None), (o["server"], None), (o["cashier"], None),
            (o["guests"], "#,##0"), (o["avg_spend"], "#,##0.00"), (o["revenue"], "#,##0.00"),
            (o["discount_rate"], "0.00%"), (o["discount_amt"], "#,##0.00"),
            (o["receivable"], "#,##0.00"), (o["actual"], "#,##0.00"),
            (o["actual_excl"], "#,##0.00"), (o["item_count"], "#,##0.0"),
            (o["items_str"], None), (o["items_pre"], "#,##0.00"),
            (o["items_post"], "#,##0.00"), (o["delivery_fee"], "#,##0.00"),
        ]:
            style_cell(ws, ri, c, val, border=THIN_BORDER, number_format=fmt)
            c += 1

    # Summary row
    tr = len(all_orders) + 4
    style_cell(ws, tr, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 18, sum(o["guests"] for o in all_orders), SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    style_cell(ws, tr, 20, round(sum(o["revenue"] for o in all_orders), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    style_cell(ws, tr, 23, round(sum(o["receivable"] for o in all_orders), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    style_cell(ws, tr, 24, round(sum(o["actual"] for o in all_orders), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

    # Column widths
    widths = {1: 22, 2: 12, 7: 20, 9: 8, 10: 20, 11: 16, 12: 16, 13: 20, 14: 20, 20: 12, 27: 50}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    return all_orders


# ============================================================
# Sheet 2: 收入管理报表 (模仿收入管理报表.xlsx)
# ============================================================

def create_revenue_report_sheet(wb, tpl, scn, year=2025):
    """收入管理报表 — 4个section: 可比同比 / 环比 / 堂食外卖占比 / 客单人数分析。"""
    ws = wb.create_sheet("收入管理报表")
    stores = tpl["stores"][:8]
    scale = scn["revenue_mult"]

    # ---- Section 1: 可比同比 (YoY) ----
    ws.merge_cells("A1:N1")
    style_cell(ws, 1, 1, f"{tpl['company']}收入管理报表", TITLE_FONT, alignment=CENTER_ALIGN)

    style_cell(ws, 3, 2, "可比同比", SECTION_FONT, LIGHT_YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 3, "2025年1月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells("D3:F3")
    style_cell(ws, 3, 4, "午市晚市", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells("G3:J3")
    style_cell(ws, 3, 7, "堂食", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells("K3:N3")
    style_cell(ws, 3, 11, "外卖", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    yoy_headers = ["门店名称", "汇总实际收入", "本期", "去年同期", "同比率",
                    "实际收入", "本期", "去年同期", "同比率",
                    "实际收入", "本期", "去年同期", "同比率"]
    for ci, h in enumerate(yoy_headers, 2):
        style_cell(ws, 4, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    total_revenue = 0
    for si, store in enumerate(stores):
        row = 5 + si
        base_rev = random.uniform(200000, 800000) * scale
        yoy_growth = random.uniform(-0.15, 0.35)
        last_year = round(base_rev / (1 + yoy_growth), 2)
        dine_pct = random.uniform(0.55, 0.75)
        dine_rev = round(base_rev * dine_pct, 2)
        dlvr_rev = round(base_rev * (1 - dine_pct), 2)
        dine_ly = round(last_year * random.uniform(0.55, 0.75), 2)
        dlvr_ly = round(last_year - dine_ly, 2)

        style_cell(ws, row, 2, store, border=THIN_BORDER)
        style_cell(ws, row, 3, round(base_rev, 2), border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 4, round(base_rev, 2), border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 5, round(last_year, 2), border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 6, round(yoy_growth, 4), GREEN_FONT if yoy_growth >= 0 else RED_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 7, dine_rev, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 8, dine_rev, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 9, dine_ly, border=THIN_BORDER, number_format="#,##0.00")
        dine_yoy = round((dine_rev / dine_ly - 1), 4) if dine_ly else 0
        style_cell(ws, row, 10, dine_yoy, GREEN_FONT if dine_yoy >= 0 else RED_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 11, dlvr_rev, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 12, dlvr_rev, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 13, dlvr_ly, border=THIN_BORDER, number_format="#,##0.00")
        dlvr_yoy = round((dlvr_rev / dlvr_ly - 1), 4) if dlvr_ly else 0
        style_cell(ws, row, 14, dlvr_yoy, GREEN_FONT if dlvr_yoy >= 0 else RED_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        total_revenue += base_rev

    # Total row
    tr = 5 + len(stores)
    style_cell(ws, tr, 2, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 3, round(total_revenue, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

    # ---- Section 2: 环比 (MoM) ----
    sec2_start = tr + 3
    style_cell(ws, sec2_start, 2, "环比", SECTION_FONT, LIGHT_GREEN_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, sec2_start, 3, "2025年1月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=sec2_start, start_column=4, end_row=sec2_start, end_column=6)
    style_cell(ws, sec2_start, 4, "午市晚市", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=sec2_start, start_column=7, end_row=sec2_start, end_column=10)
    style_cell(ws, sec2_start, 7, "堂食", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=sec2_start, start_column=11, end_row=sec2_start, end_column=14)
    style_cell(ws, sec2_start, 11, "外卖", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    mom_headers = ["门店名称", "汇总实际收入", "本期", "环比", "环比率",
                    "实际收入", "本期", "环比", "环比率",
                    "实际收入", "本期", "环比", "环比率"]
    for ci, h in enumerate(mom_headers, 2):
        style_cell(ws, sec2_start + 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    for si, store in enumerate(stores):
        row = sec2_start + 2 + si
        this_month = round(random.uniform(200000, 800000) * scale, 2)
        last_month = round(this_month * random.uniform(0.8, 1.2), 2)
        mom_diff = round(this_month - last_month, 2)
        mom_rate = round((this_month / last_month - 1), 4) if last_month else 0
        dine_pct = random.uniform(0.55, 0.75)
        dine_this = round(this_month * dine_pct, 2)
        dine_last = round(last_month * random.uniform(0.55, 0.75), 2)
        dlvr_this = round(this_month - dine_this, 2)
        dlvr_last = round(last_month - dine_last, 2)

        style_cell(ws, row, 2, store, border=THIN_BORDER)
        style_cell(ws, row, 3, this_month, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 4, this_month, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 5, mom_diff, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 6, mom_rate, GREEN_FONT if mom_rate >= 0 else RED_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 7, dine_this, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 8, dine_this, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 9, round(dine_this - dine_last, 2), border=THIN_BORDER, number_format="#,##0.00")
        dine_mom = round((dine_this / dine_last - 1), 4) if dine_last else 0
        style_cell(ws, row, 10, dine_mom, GREEN_FONT if dine_mom >= 0 else RED_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 11, dlvr_this, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 12, dlvr_this, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 13, round(dlvr_this - dlvr_last, 2), border=THIN_BORDER, number_format="#,##0.00")
        dlvr_mom = round((dlvr_this / dlvr_last - 1), 4) if dlvr_last else 0
        style_cell(ws, row, 14, dlvr_mom, GREEN_FONT if dlvr_mom >= 0 else RED_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    # ---- Section 3: 堂食外卖占比 ----
    sec3_start = sec2_start + 2 + len(stores) + 2
    style_cell(ws, sec3_start, 2, "堂食外卖占比", SECTION_FONT, LIGHT_RED_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, sec3_start, 3, "2025年1月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=sec3_start, start_column=4, end_row=sec3_start, end_column=5)
    style_cell(ws, sec3_start, 4, "午市晚市", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    ratio_headers = ["门店名称", "实际收入堂食", "实际收入外卖", "收入比例", "客单量堂食", "客单量外卖", "客单比例"]
    for ci, h in enumerate(ratio_headers, 2):
        style_cell(ws, sec3_start + 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    for si, store in enumerate(stores):
        row = sec3_start + 2 + si
        dine_rev = round(random.uniform(150000, 600000) * scale, 2)
        dlvr_rev = round(random.uniform(80000, 250000) * scale, 2)
        total_rev = dine_rev + dlvr_rev
        rev_ratio = f"{round(dine_rev/total_rev*100, 1)}%:{round(dlvr_rev/total_rev*100, 1)}%"
        dine_orders = random.randint(500, 2000)
        dlvr_orders = random.randint(400, 1500)
        order_ratio = f"{round(dine_orders/(dine_orders+dlvr_orders)*100, 1)}%:{round(dlvr_orders/(dine_orders+dlvr_orders)*100, 1)}%"

        style_cell(ws, row, 2, store, border=THIN_BORDER)
        style_cell(ws, row, 3, dine_rev, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 4, dlvr_rev, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 5, rev_ratio, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws, row, 6, dine_orders, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 7, dlvr_orders, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 8, order_ratio, border=THIN_BORDER, alignment=CENTER_ALIGN)

    # ---- Section 4: 客单人数分析 ----
    sec4_start = sec3_start + 2 + len(stores) + 2
    style_cell(ws, sec4_start, 2, tpl["stores"][0], SECTION_FONT, LIGHT_YELLOW_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, sec4_start, 3, "1月1日-1月31日", HEADER_FONT, alignment=CENTER_ALIGN)

    cust_headers = ["客单人数", "客单量", "客单占比", "点单份数", "人均点单数量", "实收额", "实际人均", "营业额占比"]
    for ci, h in enumerate(cust_headers, 2):
        style_cell(ws, sec4_start + 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    total_orders = 0
    total_revenue_sec4 = 0
    data_rows = []
    for pax in range(1, 11):
        if pax <= 3:
            order_count = random.randint(50, 300)
        elif pax <= 6:
            order_count = random.randint(5, 50)
        else:
            order_count = random.randint(0, 8)
        if order_count == 0:
            continue
        items_per_person = round(random.uniform(1.2, 2.8), 1)
        total_items = round(order_count * pax * items_per_person)
        avg_per_person = round(random.uniform(60, 120) * scale, 0)
        rev = round(order_count * pax * avg_per_person, 0)
        data_rows.append((pax, order_count, total_items, items_per_person, rev, avg_per_person))
        total_orders += order_count
        total_revenue_sec4 += rev

    for di, (pax, count, items, items_pp, rev, avg_pp) in enumerate(data_rows):
        row = sec4_start + 2 + di
        style_cell(ws, row, 2, pax, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws, row, 3, count, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 4, round(count / total_orders, 3), border=THIN_BORDER, number_format="0.000")
        style_cell(ws, row, 5, items, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 6, items_pp, border=THIN_BORDER, number_format="0.0")
        style_cell(ws, row, 7, int(rev), border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 8, int(avg_pp), border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 9, round(rev / total_revenue_sec4, 3) if total_revenue_sec4 else 0, border=THIN_BORDER, number_format="0.000")

    # Total
    trow = sec4_start + 2 + len(data_rows)
    style_cell(ws, trow, 2, "总计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, trow, 3, total_orders, SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    style_cell(ws, trow, 4, 1, SECTION_FONT, border=THIN_BORDER, number_format="0.000")
    style_cell(ws, trow, 7, int(total_revenue_sec4), SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    # Note
    style_cell(ws, trow + 1, 2, "注：1、不含套餐订单；2、统计订单中不含米饭及打包盒")

    for c, w in [(2, 24), (3, 16), (4, 14), (5, 12), (6, 12), (7, 14), (8, 12), (9, 12), (10, 12), (11, 14), (12, 14), (13, 14), (14, 12)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# Sheet 3: 日营业汇总表
# ============================================================

def create_daily_summary_sheet(wb, tpl, scn, year=2025):
    """日营业汇总 — 每日×门店 的营业额/客流/人均消费/翻台率。"""
    ws = wb.create_sheet("日营业汇总")
    stores = tpl["stores"][:8]
    scale = scn["revenue_mult"]
    days_in_month = 31

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月日营业汇总表", TITLE_FONT, alignment=CENTER_ALIGN)

    # Headers
    headers = ["日期", "门店", "营业额", "客流量", "人均消费", "翻台率"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    row = 3
    for day in range(1, days_in_month + 1):
        dt = date(year, 1, day)
        is_weekend = dt.weekday() >= 5
        for store in stores:
            base = random.uniform(15000, 50000) * scale
            weekend_mult = random.uniform(1.3, 1.8) if is_weekend else 1.0
            revenue = round(base * weekend_mult, 2)
            guests = random.randint(40, 200)
            avg_spend = round(revenue / guests, 2)
            turnover_rate = round(random.uniform(1.0, 3.5), 2)

            style_cell(ws, row, 1, dt, border=THIN_BORDER, number_format="YYYY-MM-DD")
            style_cell(ws, row, 2, store, border=THIN_BORDER)
            style_cell(ws, row, 3, revenue, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, guests, border=THIN_BORDER, number_format="#,##0")
            style_cell(ws, row, 5, avg_spend, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 6, turnover_rate, border=THIN_BORDER, number_format="0.00")
            row += 1

    for c, w in [(1, 14), (2, 24), (3, 14), (4, 10), (5, 12), (6, 10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# Sheet 4: 菜品销售排行
# ============================================================

def create_menu_ranking_sheet(wb, tpl, scn, year=2025):
    """菜品销售排行 — 按门店/汇总，含销量/金额/占比/排名。"""
    ws = wb.create_sheet("菜品销售排行")
    stores = tpl["stores"][:8]
    menu = tpl["menu_items"]
    scale = scn["revenue_mult"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月菜品销售排行", TITLE_FONT, alignment=CENTER_ALIGN)

    headers = ["排名", "菜品名称", "分类", "销量(份)", "销售额(元)", "占总销售额比",
               "点单率", "退菜率", "平均单价", "毛利率"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Generate sales data per item
    items_data = []
    total_sales = 0
    for name, price, category in menu:
        popularity = random.uniform(0.3, 3.0)
        qty = int(random.uniform(50, 1500) * popularity * scale)
        actual_price = round(price * random.uniform(0.85, 1.05), 2)
        sales = round(qty * actual_price, 2)
        order_rate = round(random.uniform(5, 65), 1)
        return_rate = round(random.uniform(0.1, 3.5), 1)
        margin = round(random.uniform(40, 78), 1)
        items_data.append((name, category, qty, sales, order_rate, return_rate, actual_price, margin))
        total_sales += sales

    # Sort by sales descending
    items_data.sort(key=lambda x: -x[3])

    for rank, (name, cat, qty, sales, order_rate, ret_rate, avg_price, margin) in enumerate(items_data, 1):
        row = 2 + rank
        pct = round(sales / total_sales, 4) if total_sales else 0
        style_cell(ws, row, 1, rank, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws, row, 2, name, border=THIN_BORDER)
        style_cell(ws, row, 3, cat, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws, row, 4, qty, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 5, sales, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 6, pct, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 7, round(order_rate / 100, 4), border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 8, round(ret_rate / 100, 4), border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 9, avg_price, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 10, round(margin / 100, 4), border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    # Total row
    tr = 3 + len(items_data)
    style_cell(ws, tr, 1, "", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 2, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 4, sum(d[2] for d in items_data), SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    style_cell(ws, tr, 5, round(total_sales, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    style_cell(ws, tr, 6, 1.0, SECTION_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    for c, w in [(1, 6), (2, 30), (3, 8), (4, 10), (5, 14), (6, 12), (7, 10), (8, 10), (9, 10), (10, 10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# Sheet 5: 门店月度经营对比
# ============================================================

def create_store_monthly_comparison(wb, tpl, scn, year=2025):
    """门店月度经营对比 — 门店×月份 交叉表，含营业额/客流/人均。"""
    ws = wb.create_sheet("门店月度对比")
    stores = tpl["stores"][:8]
    scale = scn["revenue_mult"]
    months = list(range(1, 13))

    # Title
    total_cols = 2 + len(months)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    style_cell(ws, 1, 1, f"{tpl['company']}{year}年门店营业额月度对比", TITLE_FONT, alignment=CENTER_ALIGN)
    style_cell(ws, 2, 1, "单位：万元", alignment=RIGHT_ALIGN)

    # Headers
    style_cell(ws, 3, 1, "门店名称", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 2, "年累计", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for mi, m in enumerate(months):
        style_cell(ws, 3, 3 + mi, f"{m}月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    row = 4
    grand_total = [0] * 12
    for store in stores:
        base = random.uniform(20, 80) * scale  # 万元
        seasonal = [0.8, 0.7, 0.9, 1.0, 1.1, 1.2, 1.3, 1.1, 1.0, 1.1, 1.0, 1.4]
        monthly = [round(base * seasonal[m] * random.uniform(0.85, 1.15), 1) for m in range(12)]
        total = round(sum(monthly), 1)

        style_cell(ws, row, 1, store, border=THIN_BORDER)
        style_cell(ws, row, 2, total, SECTION_FONT, border=THIN_BORDER, number_format="#,##0.0")
        for mi, val in enumerate(monthly):
            style_cell(ws, row, 3 + mi, val, border=THIN_BORDER, number_format="#,##0.0")
            grand_total[mi] += val
        row += 1

    # Total row
    style_cell(ws, row, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, row, 2, round(sum(grand_total), 1), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.0")
    for mi in range(12):
        style_cell(ws, row, 3 + mi, round(grand_total[mi], 1), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.0")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    for c in range(3, 3 + 12):
        ws.column_dimensions[get_column_letter(c)].width = 10


# ============================================================
# Sheet 6: 人效分析表
# ============================================================

def create_labor_efficiency_sheet(wb, tpl, scn, year=2025):
    """门店人效分析 — 员工人数/人均产值/工时/人效比等。"""
    ws = wb.create_sheet("人效分析")
    stores = tpl["stores"][:8]
    scale = scn["revenue_mult"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月门店人效分析表", TITLE_FONT, alignment=CENTER_ALIGN)

    headers = ["门店名称", "员工人数", "营业额(万)", "人均产值(万)", "总工时(h)",
               "人均工时(h)", "每工时产值(元)", "人力成本(万)", "人效比", "人力成本率"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    for si, store in enumerate(stores):
        row = 3 + si
        n_staff = random.randint(12, 35)
        revenue = round(random.uniform(30, 80) * scale, 1)
        per_staff = round(revenue / n_staff, 2)
        total_hours = n_staff * random.randint(160, 200)
        per_hour = round(total_hours / n_staff, 0)
        per_hour_rev = round(revenue * 10000 / total_hours, 1)
        labor_cost = round(n_staff * random.uniform(0.5, 0.9), 1)
        efficiency = round(revenue / labor_cost, 2)
        cost_rate = round(labor_cost / revenue, 4)

        style_cell(ws, row, 1, store, border=THIN_BORDER)
        style_cell(ws, row, 2, n_staff, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 3, revenue, border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, row, 4, per_staff, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 5, total_hours, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 6, per_hour, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 7, per_hour_rev, border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, row, 8, labor_cost, border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, row, 9, efficiency, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 10, cost_rate, RED_FONT if cost_rate > 0.30 else None, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    for c, w in [(1, 24), (2, 10), (3, 12), (4, 12), (5, 10), (6, 10), (7, 14), (8, 12), (9, 10), (10, 12)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# Sheet 7: 食材成本分析
# ============================================================

def create_food_cost_sheet(wb, tpl, scn, year=2025):
    """食材成本分析 — 按品类 成本/用量/损耗/毛利。"""
    ws = wb.create_sheet("食材成本分析")
    scale = scn["revenue_mult"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月食材成本分析", TITLE_FONT, alignment=CENTER_ALIGN)

    headers = ["食材品类", "采购金额(元)", "使用量(kg)", "单位成本(元/kg)",
               "损耗量(kg)", "损耗率", "实际成本(元)", "对应收入(元)", "毛利率"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    categories = {
        "fish": [("鲈鱼", 35), ("草鱼", 18), ("黑鱼", 28), ("虾", 45), ("蔬菜", 5), ("豆制品", 8),
                 ("调味料", 15), ("油料", 12), ("面粉/淀粉", 6), ("水果", 10), ("包材", 3), ("饮品原料", 8)],
        "hotpot": [("牛肉", 55), ("羊肉", 48), ("猪肉", 25), ("毛肚", 38), ("鹅肠", 42), ("虾滑", 35),
                   ("蔬菜", 5), ("豆制品", 8), ("火锅底料", 20), ("调味料", 15), ("饮品", 8), ("包材", 3)],
        "bakery": [("面粉", 8), ("黄油", 35), ("奶油", 28), ("鸡蛋", 12), ("水果", 15), ("巧克力", 45),
                   ("茶叶", 80), ("鲜奶", 10), ("糖类", 6), ("包材", 5), ("咖啡豆", 60), ("其他", 10)],
    }
    template_key = [k for k in TEMPLATES if TEMPLATES[k]["brand"] == tpl["brand"]][0]
    items = categories.get(template_key, categories["fish"])

    total_purchase = 0
    total_cost = 0
    total_revenue_items = 0
    for name, unit_cost in items:
        row = 3 + items.index((name, unit_cost))
        qty = round(random.uniform(50, 3000) * scale, 1)
        purchase = round(qty * unit_cost * random.uniform(0.9, 1.1), 2)
        waste_rate_pct = round(random.uniform(1.5, 12), 1)  # keep as 0-100 for calc
        waste_qty = round(qty * waste_rate_pct / 100, 1)
        actual_cost = round(purchase * (1 + waste_rate_pct / 100), 2)
        revenue_item = round(actual_cost / random.uniform(0.25, 0.45), 2)
        margin = round(1 - actual_cost / revenue_item, 4)

        style_cell(ws, row, 1, name, border=THIN_BORDER)
        style_cell(ws, row, 2, purchase, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 3, qty, border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, row, 4, unit_cost, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 5, waste_qty, border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, row, 6, round(waste_rate_pct / 100, 4), RED_FONT if waste_rate_pct > 8 else None, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, row, 7, actual_cost, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 8, revenue_item, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 9, margin, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        total_purchase += purchase
        total_cost += actual_cost
        total_revenue_items += revenue_item

    tr = 3 + len(items)
    style_cell(ws, tr, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 2, round(total_purchase, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    style_cell(ws, tr, 7, round(total_cost, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    style_cell(ws, tr, 8, round(total_revenue_items, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    overall_margin = round(1 - total_cost / total_revenue_items, 4) if total_revenue_items else 0
    style_cell(ws, tr, 9, overall_margin, SECTION_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    for c, w in [(1, 14), (2, 14), (3, 12), (4, 14), (5, 10), (6, 10), (7, 14), (8, 14), (9, 10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# Sheet 8: 时段客流分析
# ============================================================

def create_hourly_traffic_sheet(wb, tpl, scn, year=2025):
    """时段客流分析 — 门店×时段(10:00-22:00) 热力图式数据。"""
    ws = wb.create_sheet("时段客流分析")
    stores = tpl["stores"][:8]
    scale = scn["revenue_mult"]
    hours = list(range(10, 23))  # 10:00 - 22:00

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hours) + 1)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月时段客流分析（日均）", TITLE_FONT, alignment=CENTER_ALIGN)

    style_cell(ws, 2, 1, "门店名称", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for hi, h in enumerate(hours):
        style_cell(ws, 2, 2 + hi, f"{h}:00", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Typical traffic curve: lunch peak 11-13, dinner peak 17-20
    base_curve = {
        10: 0.3, 11: 0.7, 12: 1.0, 13: 0.8, 14: 0.3, 15: 0.2,
        16: 0.3, 17: 0.6, 18: 0.9, 19: 1.0, 20: 0.7, 21: 0.4, 22: 0.2,
    }

    for si, store in enumerate(stores):
        row = 3 + si
        base_traffic = random.uniform(5, 25) * scale
        style_cell(ws, row, 1, store, border=THIN_BORDER)
        for hi, h in enumerate(hours):
            curve = base_curve.get(h, 0.3)
            traffic = round(base_traffic * curve * random.uniform(0.7, 1.3))
            fill = None
            if traffic > base_traffic * 0.8:
                fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif traffic > base_traffic * 0.5:
                fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
            elif traffic > base_traffic * 0.3:
                fill = LIGHT_YELLOW_FILL
            style_cell(ws, row, 2 + hi, traffic, fill=fill, border=THIN_BORDER, number_format="#,##0", alignment=CENTER_ALIGN)

    ws.column_dimensions["A"].width = 24
    for c in range(2, 2 + len(hours)):
        ws.column_dimensions[get_column_letter(c)].width = 8


# ============================================================
# Sheet 9: 外卖平台对比
# ============================================================

def create_delivery_platform_sheet(wb, tpl, scn, year=2025):
    """外卖平台对比 — 美团/饿了么/京东 按门店的订单量/金额/扣点/实收对比。"""
    ws = wb.create_sheet("外卖平台对比")
    stores = tpl["stores"][:8]
    scale = scn["revenue_mult"]
    platforms = ["美团外卖", "饿了么", "京东外卖"]

    # Title
    total_cols = 1 + len(platforms) * 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月外卖平台经营对比", TITLE_FONT, alignment=CENTER_ALIGN)

    # R2: platform headers (merged 5 cols each)
    style_cell(ws, 2, 1, "门店名称", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for pi, platform in enumerate(platforms):
        col = 2 + pi * 5
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 4)
        style_cell(ws, 2, col, platform, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R3: sub-headers
    sub_headers = ["订单量", "营业额", "平台扣点率", "平台扣点额", "实收额"]
    for pi in range(len(platforms)):
        col = 2 + pi * 5
        for si, sh in enumerate(sub_headers):
            style_cell(ws, 3, col + si, sh, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    for si, store in enumerate(stores):
        row = 4 + si
        style_cell(ws, row, 1, store, border=THIN_BORDER)
        for pi, platform in enumerate(platforms):
            col = 2 + pi * 5
            # Meituan > Eleme > JD in typical volume
            platform_weight = [0.5, 0.35, 0.15][pi]
            orders = int(random.uniform(200, 1000) * platform_weight * scale)
            revenue = round(orders * random.uniform(50, 120), 2)
            rate = round(random.uniform(0.15, 0.25), 4)
            deduction = round(revenue * rate, 2)
            actual = round(revenue - deduction, 2)

            style_cell(ws, row, col, orders, border=THIN_BORDER, number_format="#,##0")
            style_cell(ws, row, col + 1, revenue, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, col + 2, rate, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
            style_cell(ws, row, col + 3, deduction, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, col + 4, actual, border=THIN_BORDER, number_format="#,##0.00")

    ws.column_dimensions["A"].width = 24
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12


# ============================================================
# Sheet 10: 会员消费分析
# ============================================================

def create_membership_sheet(wb, tpl, scn, year=2025):
    """会员消费分析 — 会员等级/消费频次/客单价/储值余额分布。"""
    ws = wb.create_sheet("会员消费分析")
    scale = scn["revenue_mult"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    style_cell(ws, 1, 1, f"{tpl['company']}2025年1月会员消费分析", TITLE_FONT, alignment=CENTER_ALIGN)

    headers = ["会员等级", "会员数", "消费人次", "消费金额(元)", "人均消费(元)",
               "月均消费频次", "储值余额(元)", "储值消费占比"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    levels = [
        ("普通会员", 0.5, 1.2, 80),
        ("银卡会员", 0.25, 2.5, 120),
        ("金卡会员", 0.15, 4.0, 160),
        ("钻石会员", 0.08, 6.0, 220),
        ("黑卡会员", 0.02, 8.0, 350),
    ]
    total_members = int(random.uniform(5000, 20000) * scale)
    total_revenue_mem = 0

    for li, (level, pct, freq, avg) in enumerate(levels):
        row = 3 + li
        n_members = int(total_members * pct)
        visits = int(n_members * freq * random.uniform(0.8, 1.2))
        avg_spend = round(avg * random.uniform(0.85, 1.15) * scale, 2)
        revenue = round(visits * avg_spend, 2)
        stored_balance = round(n_members * random.uniform(50, 500), 2)
        stored_pct = round(random.uniform(0.15, 0.55), 4)

        style_cell(ws, row, 1, level, border=THIN_BORDER)
        style_cell(ws, row, 2, n_members, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 3, visits, border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, 4, revenue, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 5, avg_spend, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 6, round(freq * random.uniform(0.8, 1.2), 1), border=THIN_BORDER, number_format="0.0")
        style_cell(ws, row, 7, stored_balance, border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, row, 8, stored_pct, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        total_revenue_mem += revenue

    tr = 3 + len(levels)
    style_cell(ws, tr, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 2, total_members, SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    style_cell(ws, tr, 4, round(total_revenue_mem, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

    for c, w in [(1, 14), (2, 10), (3, 10), (4, 14), (5, 12), (6, 12), (7, 14), (8, 12)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# 边界测试 sheet
# ============================================================

def create_wide_menu_sheet(wb, tpl, scn):
    """超宽菜品矩阵 — 100+ 列, 门店×菜品交叉表, 3-level merge headers。"""
    ws = wb.create_sheet("超宽菜品矩阵")
    scale = scn["revenue_mult"]

    stores = tpl["stores"]
    # Group menu_items by category
    from collections import OrderedDict
    cat_items = OrderedDict()
    for name, price, category in tpl["menu_items"]:
        cat_items.setdefault(category, []).append(name)

    # Build 100+ columns: categories → items → metrics
    col = 2
    col_map = []  # [(col_idx, category, item, metric)]
    for cat_name, items in cat_items.items():
        cat_start = col
        for item in items:
            item_start = col
            for metric in ["销量", "金额", "退单率"]:
                col_map.append((col, cat_name, item, metric))
                col += 1
            # Merge item name across 3 metric cols
            ws.merge_cells(start_row=2, start_column=item_start, end_row=2, end_column=col - 1)
            style_cell(ws, 2, item_start, item,
                       Font(bold=True, size=9), PatternFill("solid", fgColor="E8EAF6"),
                       Alignment(horizontal="center"), Border(
                           left=Side("thin"), right=Side("thin"),
                           top=Side("thin"), bottom=Side("thin")))
        # Merge category across all items
        ws.merge_cells(start_row=1, start_column=cat_start, end_row=1, end_column=col - 1)
        style_cell(ws, 1, cat_start, cat_name,
                   Font(bold=True, size=10, color="FFFFFF"),
                   PatternFill("solid", fgColor="3F51B5"),
                   Alignment(horizontal="center"),
                   Border(left=Side("thin"), right=Side("thin"),
                          top=Side("thin"), bottom=Side("thin")))

    total_cols = col - 1
    # Row 3: metric headers
    style_cell(ws, 1, 1, "门店", Font(bold=True), PatternFill("solid", fgColor="FFC107"),
               Alignment(horizontal="center"),
               Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")))
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    for ci, (col_idx, cat, item, metric) in enumerate(col_map):
        style_cell(ws, 3, col_idx, metric, Font(size=9),
                   PatternFill("solid", fgColor="C5CAE9"),
                   Alignment(horizontal="center"),
                   Border(left=Side("thin"), right=Side("thin"),
                          top=Side("thin"), bottom=Side("thin")))

    # Data rows — one per store
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    for si, store in enumerate(stores):
        row = 4 + si
        style_cell(ws, row, 1, store, border=border)
        for col_idx, cat, item, metric in col_map:
            if metric == "销量":
                val = int(random.uniform(20, 500) * scale)
            elif metric == "金额":
                val = round(random.uniform(500, 15000) * scale, 2)
            else:  # 退单率
                val = round(random.uniform(0.001, 0.08), 4)
            fmt = "#,##0" if metric == "销量" else "#,##0.00" if metric == "金额" else "0.00%"
            style_cell(ws, row, col_idx, val, border=border, number_format=fmt)

    ws.column_dimensions["A"].width = 16
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10


def create_mixed_type_menu_sheet(wb, tpl, scn):
    """混合类型数据 — 同一列混合数字、文字、日期。"""
    ws = wb.create_sheet("混合类型测试")
    scale = scn["revenue_mult"]
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    headers = ["菜品名称", "本月销量", "上月销量", "环比变化", "备注", "更新日期"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, Font(bold=True), PatternFill("solid", fgColor="FFE0B2"),
                   Alignment(horizontal="center"), border)

    all_items = [name for name, price, cat in tpl["menu_items"]]

    for ri, item in enumerate(all_items):
        row = 2 + ri
        this_month = int(random.uniform(50, 800) * scale)
        last_month = int(random.uniform(50, 800) * scale)

        style_cell(ws, row, 1, item, border=border)

        # Mixed type: most rows numeric, some rows text
        if ri == 3:
            style_cell(ws, row, 2, "停售", border=border)  # text in numeric column
            style_cell(ws, row, 3, "停售", border=border)
        elif ri == 7:
            style_cell(ws, row, 2, "新品上架", border=border)
            style_cell(ws, row, 3, None, border=border)
        else:
            style_cell(ws, row, 2, this_month, border=border, number_format="#,##0")
            style_cell(ws, row, 3, last_month, border=border, number_format="#,##0")

        # 环比变化 — mix of float percentage and text
        if ri == 3:
            style_cell(ws, row, 4, "N/A", border=border)
        elif ri == 7:
            style_cell(ws, row, 4, "新品", border=border)
        else:
            change = round((this_month / last_month - 1) if last_month else 0, 4)
            style_cell(ws, row, 4, change, border=border, number_format="+0.0%;-0.0%")

        # 备注 — text annotations
        notes = ["", "热销", "需推广", None, "季节性下降", "涨价后", "新配方", "限时特价"]
        style_cell(ws, row, 5, random.choice(notes), border=border)

        # 更新日期 — mix of date objects and date strings
        if ri % 5 == 0:
            style_cell(ws, row, 6, f"2025-{random.randint(1,12):02d}-{random.randint(1,28):02d}", border=border)
        else:
            d = date(2025, random.randint(1, 12), random.randint(1, 28))
            style_cell(ws, row, 6, d, border=border, number_format="YYYY-MM-DD")

    for c, w in [(1, 16), (2, 12), (3, 12), (4, 12), (5, 14), (6, 14)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_empty_regions_restaurant_sheet(wb, tpl, scn):
    """空值区域测试 — 模拟公式返回 None + 全空行分隔 + 全零列。"""
    ws = wb.create_sheet("空值区域测试")
    scale = scn["revenue_mult"]
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    months = [f"{m}月" for m in range(1, 13)]
    headers = ["项目"] + months + ["合计", "全零列"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, Font(bold=True), PatternFill("solid", fgColor="B2DFDB"),
                   Alignment(horizontal="center"), border)

    items = ["堂食收入", "外卖收入", "团购收入", "",  # empty row separator
             "食材成本", "人工成本", "租金", "",  # empty row separator
             "=SUM模拟(None)", "=VLOOKUP模拟(None)",  # formula simulation
             "营业利润", "净利润"]

    for ri, item in enumerate(items):
        row = 2 + ri
        style_cell(ws, row, 1, item if item else None,
                   Font(bold=True) if "利润" in item else None, border=border)

        if item == "":
            continue
        elif "模拟(None)" in item:
            for ci in range(2, 14):
                style_cell(ws, row, ci, None, border=border)
            style_cell(ws, row, 14, None, border=border)
        else:
            for ci in range(2, 14):
                val = round(random.uniform(50000, 800000) * scale, 2)
                if "成本" in item or "租金" in item:
                    val = -abs(val) * random.uniform(0.2, 0.6)
                style_cell(ws, row, ci, round(val, 2), border=border, number_format="#,##0.00")
            style_cell(ws, row, 14, round(random.uniform(200000, 5000000) * scale, 2),
                       border=border, number_format="#,##0.00")

        # 全零列
        style_cell(ws, row, 15, 0, border=border, number_format="#,##0.00")

    for c, w in [(1, 18)] + [(i, 11) for i in range(2, 16)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_cross_year_restaurant_sheet(wb, tpl, scn):
    """跨年对比数据 — 2024 vs 2025 门店营业额 for YoY。"""
    ws = wb.create_sheet("跨年对比(2024-2025)")
    scale = scn["revenue_mult"]
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    center = Alignment(horizontal="center")

    # 3-level header: year → half → metrics
    ws.merge_cells("A1:A3")
    style_cell(ws, 1, 1, "门店", Font(bold=True), PatternFill("solid", fgColor="FFC107"),
               Alignment(horizontal="center", vertical="center"), border)

    col = 2
    for year in [2024, 2025]:
        end_col = col + 7
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        style_cell(ws, 1, col, f"{year}年", Font(bold=True, color="FFFFFF"),
                   PatternFill("solid", fgColor="3F51B5"), center, border)

        for qi, half in enumerate(["上半年", "下半年"]):
            qstart = col + qi * 4
            ws.merge_cells(start_row=2, start_column=qstart, end_row=2, end_column=qstart + 3)
            style_cell(ws, 2, qstart, half, Font(bold=True),
                       PatternFill("solid", fgColor="7986CB"), center, border)

        for mi in range(8):
            metric = ["营业额", "客流量", "客单价", "毛利率"][mi % 4]
            style_cell(ws, 3, col + mi, metric, Font(bold=True, size=9),
                       PatternFill("solid", fgColor="C5CAE9"), center, border)
        col = end_col + 1

    # YoY columns
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col + 1)
    style_cell(ws, 1, col, "同比变化", Font(bold=True), PatternFill("solid", fgColor="FF7043"),
               center, border)
    style_cell(ws, 3, col, "营业额增长", Font(bold=True, size=9),
               PatternFill("solid", fgColor="FFCCBC"), center, border)
    style_cell(ws, 3, col + 1, "客流增长", Font(bold=True, size=9),
               PatternFill("solid", fgColor="FFCCBC"), center, border)

    # Data rows
    stores = tpl["stores"]
    for si, store in enumerate(stores):
        row = 4 + si
        style_cell(ws, row, 1, store, border=border)

        rev_2024 = [round(random.uniform(80000, 500000) * scale, 2) for _ in range(8)]
        rev_2025 = [round(v * random.uniform(0.85, 1.25), 2) for v in rev_2024]

        for yi, year_data in enumerate([rev_2024, rev_2025]):
            base_col = 2 + yi * 8
            for mi in range(8):
                metric_type = mi % 4
                if metric_type == 0:  # 营业额
                    val = year_data[mi]
                    fmt = "#,##0.00"
                elif metric_type == 1:  # 客流量
                    val = int(random.uniform(3000, 15000) * scale)
                    fmt = "#,##0"
                elif metric_type == 2:  # 客单价
                    val = round(random.uniform(45, 120) * scale, 2)
                    fmt = "#,##0.00"
                else:  # 毛利率
                    val = round(random.uniform(0.35, 0.65), 4)
                    fmt = "0.0%"
                style_cell(ws, row, base_col + mi, val, border=border, number_format=fmt)

        # YoY
        total_2024 = sum(rev_2024[i] for i in range(0, 8, 4))
        total_2025 = sum(rev_2025[i] for i in range(0, 8, 4))
        yoy_rev = round((total_2025 / total_2024 - 1), 4) if total_2024 else 0
        yoy_traffic = round(random.uniform(-0.15, 0.25), 4)
        style_cell(ws, row, col, yoy_rev, border=border, number_format="+0.0%;-0.0%")
        style_cell(ws, row, col + 1, yoy_traffic, border=border, number_format="+0.0%;-0.0%")

    ws.column_dimensions["A"].width = 16
    for c in range(2, col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 11


def create_formula_restaurant_sheet(wb, tpl, scn):
    """公式测试 — SUM/AVERAGE/IF 公式，模拟 data_only=True 返回 None。"""
    ws = wb.create_sheet(title="公式测试")

    # Title
    style_cell(ws, 1, 1, "门店月度营业额汇总（含公式）",
               font=TITLE_FONT, fill=HEADER_FILL)
    ws.merge_cells("A1:N1")

    headers = ["门店"] + [f"{m+1}月" for m in range(12)] + ["年度合计"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, font=HEADER_FONT, fill=HEADER_FILL,
                   border=THIN_BORDER, alignment=CENTER_ALIGN)

    stores = tpl["stores"][:8]
    for ri, store in enumerate(stores, 3):
        style_cell(ws, ri, 1, store, border=THIN_BORDER)
        for mi in range(12):
            val = round(random.uniform(80000, 500000) * scn["revenue_mult"], 2)
            style_cell(ws, ri, mi + 2, val, border=THIN_BORDER, number_format="#,##0.00")
        # SUM formula
        ws.cell(row=ri, column=14).value = f"=SUM(B{ri}:M{ri})"
        ws.cell(row=ri, column=14).border = THIN_BORDER
        ws.cell(row=ri, column=14).number_format = "#,##0.00"

    sum_row = 3 + len(stores)
    style_cell(ws, sum_row, 1, "月度合计", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 15):
        col_letter = get_column_letter(ci)
        ws.cell(row=sum_row, column=ci).value = f"=SUM({col_letter}3:{col_letter}{sum_row - 1})"
        ws.cell(row=sum_row, column=ci).border = THIN_BORDER
        ws.cell(row=sum_row, column=ci).number_format = "#,##0.00"

    avg_row = sum_row + 1
    style_cell(ws, avg_row, 1, "门店平均", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 15):
        col_letter = get_column_letter(ci)
        ws.cell(row=avg_row, column=ci).value = f"=AVERAGE({col_letter}3:{col_letter}{sum_row - 1})"
        ws.cell(row=avg_row, column=ci).border = THIN_BORDER
        ws.cell(row=avg_row, column=ci).number_format = "#,##0.00"

    # IF formula row — achievement check
    cond_row = avg_row + 1
    style_cell(ws, cond_row, 1, "达标判定", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 14):
        col_letter = get_column_letter(ci)
        ws.cell(row=cond_row, column=ci).value = f'=IF({col_letter}{sum_row}>2000000,"达标","未达标")'
        ws.cell(row=cond_row, column=ci).border = THIN_BORDER

    # Percentage formula
    pct_row = cond_row + 1
    style_cell(ws, pct_row, 1, "月度占比", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 14):
        col_letter = get_column_letter(ci)
        ws.cell(row=pct_row, column=ci).value = f"={col_letter}{sum_row}/N{sum_row}"
        ws.cell(row=pct_row, column=ci).border = THIN_BORDER
        ws.cell(row=pct_row, column=ci).number_format = "0.0%"


def create_numeric_colname_restaurant_sheet(wb, tpl, scn):
    """纯数字列名测试 — 年份/季度作为列标题的整数值。"""
    ws = wb.create_sheet(title="纯数字列名")

    years = [2020, 2021, 2022, 2023, 2024, 2025]
    style_cell(ws, 1, 1, "经营指标", font=HEADER_FONT, fill=HEADER_FILL,
               border=THIN_BORDER, alignment=CENTER_ALIGN)
    for ci, yr in enumerate(years, 2):
        style_cell(ws, 1, ci, yr, font=HEADER_FONT, fill=HEADER_FILL,
                   border=THIN_BORDER, alignment=CENTER_ALIGN)

    metrics = [
        ("总营业额(万元)", [round(random.uniform(3000, 8000) * scn["revenue_mult"], 1) for _ in years]),
        ("堂食收入(万元)", [round(random.uniform(1500, 5000) * scn["revenue_mult"], 1) for _ in years]),
        ("外卖收入(万元)", [round(random.uniform(800, 3000) * scn["revenue_mult"], 1) for _ in years]),
        ("食材成本率", [round(random.uniform(0.30, 0.42), 4) for _ in years]),
        ("人工成本率", [round(random.uniform(0.20, 0.30), 4) for _ in years]),
        ("毛利率", [round(random.uniform(0.50, 0.65), 4) for _ in years]),
        ("净利率", [round(random.uniform(0.03, 0.12), 4) for _ in years]),
        ("门店数", [random.randint(5, 30) for _ in years]),
        ("日均客流", [random.randint(200, 1500) for _ in years]),
        ("客单价(元)", [round(random.uniform(50, 120) * scn["revenue_mult"], 1) for _ in years]),
        ("翻台率", [round(random.uniform(2.0, 5.0), 1) for _ in years]),
        ("好评率", [round(random.uniform(0.85, 0.98), 4) for _ in years]),
    ]

    pct_rows = {"食材成本率", "人工成本率", "毛利率", "净利率", "好评率"}

    for ri, (metric_name, values) in enumerate(metrics, 2):
        is_pct = metric_name in pct_rows
        style_cell(ws, ri, 1, metric_name, border=THIN_BORDER)
        for ci, val in enumerate(values, 2):
            fmt = "0.0%" if is_pct else "#,##0.0"
            style_cell(ws, ri, ci, val, border=THIN_BORDER, number_format=fmt)


# ============================================================
# 主函数
# ============================================================

def generate(template_name="fish", scenario="normal", seed=42, output_dir=None):
    """生成餐饮行业测试 Excel 文件。"""
    random.seed(seed)
    tpl = TEMPLATES[template_name]
    scn = SCENARIOS[scenario]

    wb = openpyxl.Workbook()

    # Remove default sheet
    ws_default = wb.active
    ws_default.title = "目录"

    # Sheet names for index
    sheet_names = ["订单销售明细表", "收入管理报表", "日营业汇总", "菜品销售排行",
                   "门店月度对比", "人效分析", "食材成本分析", "时段客流分析",
                   "外卖平台对比", "会员消费分析",
                   "超宽菜品矩阵", "混合类型测试", "空值区域测试", "跨年对比(2024-2025)",
                   "公式测试", "纯数字列名"]

    # Build index sheet
    ws_default.merge_cells("A1:C1")
    style_cell(ws_default, 1, 1, f"{tpl['company']}经营分析报表", TITLE_FONT, alignment=CENTER_ALIGN)
    style_cell(ws_default, 2, 1, "序号", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws_default, 2, 2, "报表名称", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws_default, 2, 3, "说明", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    descriptions = [
        "订单级别交易流水", "YoY/MoM/堂食外卖占比/客单分析", "每日门店营业汇总",
        "菜品销量金额排行", "门店12个月营业额趋势", "门店人效/工时/成本",
        "食材品类成本毛利", "10-22点客流热力数据", "美团/饿了么/京东对比",
        "会员等级消费频次储值",
        "100+列门店×菜品交叉表(3级合并表头)",
        "同列混合数字/文字/日期",
        "None空值+全空行+全零列",
        "2024vs2025门店营业额YoY",
        "SUM/AVERAGE/IF公式(data_only=True测试)",
        "年份整数列名(30%阈值测试)",
    ]
    for i, (name, desc) in enumerate(zip(sheet_names, descriptions)):
        style_cell(ws_default, 3 + i, 1, i + 1, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws_default, 3 + i, 2, name, border=THIN_BORDER)
        style_cell(ws_default, 3 + i, 3, desc, border=THIN_BORDER)
    ws_default.column_dimensions["A"].width = 6
    ws_default.column_dimensions["B"].width = 20
    ws_default.column_dimensions["C"].width = 35

    # Generate all sheets
    create_order_detail_sheet(wb, tpl, scn)
    create_revenue_report_sheet(wb, tpl, scn)
    create_daily_summary_sheet(wb, tpl, scn)
    create_menu_ranking_sheet(wb, tpl, scn)
    create_store_monthly_comparison(wb, tpl, scn)
    create_labor_efficiency_sheet(wb, tpl, scn)
    create_food_cost_sheet(wb, tpl, scn)
    create_hourly_traffic_sheet(wb, tpl, scn)
    create_delivery_platform_sheet(wb, tpl, scn)
    create_membership_sheet(wb, tpl, scn)

    # Edge-case sheets
    create_wide_menu_sheet(wb, tpl, scn)
    create_mixed_type_menu_sheet(wb, tpl, scn)
    create_empty_regions_restaurant_sheet(wb, tpl, scn)
    create_cross_year_restaurant_sheet(wb, tpl, scn)
    create_formula_restaurant_sheet(wb, tpl, scn)
    create_numeric_colname_restaurant_sheet(wb, tpl, scn)

    # Save
    if output_dir is None:
        output_dir = Path(__file__).parent / "test-data" / "restaurant"
    else:
        output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"Restaurant-{template_name}-{scenario}-s{seed}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)

    # Print summary
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    print(f"Generated: {filepath}")
    print(f"  Template: {template_name} ({tpl['company']})")
    print(f"  Scenario: {scenario}")
    print(f"  Seed: {seed}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        merges = len(ws.merged_cells.ranges)
        print(f"    {sname}: {ws.max_row}r x {ws.max_column}c, merges={merges}")
    wb.close()
    return filepath


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="餐饮行业 SmartBI E2E 测试 Excel 生成器")
    parser.add_argument("--template", choices=list(TEMPLATES.keys()), default="fish")
    parser.add_argument("--scenario", choices=list(SCENARIOS.keys()), default="normal")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--all", action="store_true", help="生成全部 template×scenario 组合")
    args = parser.parse_args()

    if args.all:
        count = 0
        for t in TEMPLATES:
            for s in SCENARIOS:
                generate(t, s, args.seed, args.output)
                count += 1
        print(f"\nGenerated {count} files.")
    else:
        generate(args.template, args.scenario, args.seed, args.output)
