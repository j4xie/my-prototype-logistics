#!/usr/bin/env python3
"""
SmartBI E2E 测试 Excel 生成器
高度模仿 Test.xlsx 结构，支持多模板、参数化场景

Usage:
    python tests/generate_test_excel.py                     # 默认：食品企业模板
    python tests/generate_test_excel.py --template food     # 食品企业
    python tests/generate_test_excel.py --template mfg      # 制造业
    python tests/generate_test_excel.py --template retail   # 零售业
    python tests/generate_test_excel.py --seed 123          # 指定随机种子
    python tests/generate_test_excel.py --scenario loss     # 亏损场景
    python tests/generate_test_excel.py --scenario growth   # 高增长场景
    python tests/generate_test_excel.py --scenario sparse   # 稀疏数据
"""
import argparse
import random
from datetime import datetime, date
from pathlib import Path

from collections import namedtuple

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Wrapper to tag percentage values — stored as float (0.325 = 32.5%), rendered with number_format
Pct = namedtuple("Pct", ["val"])

# ============================================================
# 模板定义
# ============================================================

TEMPLATES = {
    "food": {
        "company": "江苏美鑫食品科技有限公司",
        "brand": "宁波味觉",
        "center": "销售1中心",
        "products": ["中式汤底", "酱料", "川味底料", "调味汁", "复合调味粉", "团餐", "预制菜"],
        "regions": [
            {"name": "江苏分部", "type": "分部"},
            {"name": "浙江分部", "type": "分部"},
            {"name": "上海分部", "type": "分部"},
        ],
        "sub_regions": [
            {"name": "赣皖区域", "type": "区域", "children": [
                {"name": "安徽省区", "type": "省区"},
                {"name": "江西省区", "type": "省区"},
            ]},
        ],
        "revenue_base": (500000, 5000000),   # 月收入范围
        "cost_ratio": (0.55, 0.75),          # 成本率范围
        "expense_ratio": (0.10, 0.25),       # 费用率范围
    },
    "mfg": {
        "company": "苏州精密机械制造有限公司",
        "brand": "精密重工",
        "center": "华东事业部",
        "products": ["数控机床", "精密轴承", "液压件", "气动元件", "模具", "工装夹具"],
        "regions": [
            {"name": "江苏区域", "type": "区域"},
            {"name": "浙江区域", "type": "区域"},
            {"name": "广东区域", "type": "区域"},
        ],
        "sub_regions": [
            {"name": "华中区域", "type": "区域", "children": [
                {"name": "湖北省区", "type": "省区"},
                {"name": "湖南省区", "type": "省区"},
            ]},
        ],
        "revenue_base": (1000000, 8000000),
        "cost_ratio": (0.60, 0.80),
        "expense_ratio": (0.08, 0.18),
    },
    "retail": {
        "company": "上海鲜优生活连锁有限公司",
        "brand": "鲜优生活",
        "center": "华东运营中心",
        "products": ["生鲜蔬果", "肉禽蛋奶", "粮油调味", "酒水饮料", "休闲零食", "日用百货"],
        "regions": [
            {"name": "上海直营", "type": "直营"},
            {"name": "江苏加盟", "type": "加盟"},
            {"name": "浙江加盟", "type": "加盟"},
        ],
        "sub_regions": [
            {"name": "线上渠道", "type": "渠道", "children": [
                {"name": "天猫旗舰", "type": "店铺"},
                {"name": "京东自营", "type": "店铺"},
            ]},
        ],
        "revenue_base": (800000, 6000000),
        "cost_ratio": (0.65, 0.82),
        "expense_ratio": (0.08, 0.20),
    },
}

# 利润表行标签（与 Test.xlsx 完全一致的 260+ 行结构）
EXPENSE_SUB_ITEMS = [
    "职工薪酬", "工资", "奖金及提成", "社保", "养老", "医疗", "工伤",
    "失业", "生育", "公积金", "工会经费", "职工教育经费", "福利费",
    "职工房租", "水电物业费", "员工餐", "节日礼品", "生日礼品", "体检",
    "其他福利支出", "招聘费", "办公费", "市内交通费", "差旅费", "交通费",
    "住宿费", "餐补", "交通补助", "通讯补助", "招待费", "通讯费",
    "网络费", "修理费", "车辆费", "租赁费", "能源费", "水费", "电费",
    "燃气费", "蒸汽费", "物业费", "劳动保护费", "顾问费", "快递费",
    "诉讼费", "会员费", "广告宣传费", "会务费", "装修费", "推广费",
    "展会费", "设计费", "折旧费", "认证费", "样品费", "劳务费",
    "销售佣金", "检测费", "质量赔偿", "保险费", "清洁费", "低值易耗品",
    "物料消耗", "盘盈盘亏", "平台服务费", "无形资产摊销", "研发材料",
    "项目名称", "研发费", "残保金", "环保费", "报废损失",
]

SCENARIOS = {
    "normal": {"revenue_mult": 1.0, "noise": 0.3, "sparse_pct": 0.0},
    "growth": {"revenue_mult": 1.5, "noise": 0.2, "sparse_pct": 0.0},
    "loss":   {"revenue_mult": 0.5, "noise": 0.4, "sparse_pct": 0.0},
    "sparse": {"revenue_mult": 0.8, "noise": 0.3, "sparse_pct": 0.5},
}

# ============================================================
# 样式定义
# ============================================================

TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Write value and apply styles to a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


# ============================================================
# 数据生成
# ============================================================

class FinancialDataGenerator:
    """生成真实的财务数据，保证内部一致性。"""

    def __init__(self, template, scenario, year=2025):
        self.tpl = template
        self.scn = SCENARIOS[scenario]
        self.year = year
        self.products = template["products"]
        self.months = list(range(1, 13))

    def _rand(self, base, noise=None):
        """带噪声的随机数。"""
        noise = noise or self.scn["noise"]
        return base * random.uniform(1 - noise, 1 + noise)

    def _maybe_sparse(self, value):
        """稀疏场景：一定概率返回 None。"""
        if random.random() < self.scn["sparse_pct"]:
            return None
        return value

    def gen_revenue_by_product(self, region_scale=1.0):
        """生成12个月 × N产品 的收入矩阵。返回 {product: [月1,月2,...]}"""
        lo, hi = self.tpl["revenue_base"]
        base = random.uniform(lo, hi) * region_scale * self.scn["revenue_mult"]
        result = {}
        for prod in self.products:
            prod_share = random.uniform(0.05, 0.35)
            monthly = []
            for m in self.months:
                # 季节性：Q1 低，Q2-Q4 高
                seasonal = [0.6, 0.7, 0.9, 1.0, 1.1, 1.2, 1.3, 1.1, 1.0, 1.2, 0.9, 0.7][m - 1]
                val = self._rand(base * prod_share * seasonal)
                monthly.append(round(self._maybe_sparse(val) or 0, 1))
            result[prod] = monthly
        return result

    def gen_cost_by_product(self, revenue_by_product):
        """基于收入生成成本（保证成本率合理）。"""
        lo, hi = self.tpl["cost_ratio"]
        result = {}
        for prod, rev_list in revenue_by_product.items():
            ratio = random.uniform(lo, hi)
            result[prod] = [round(r * ratio, 1) for r in rev_list]
        return result

    def gen_expense_items(self, total_expense_monthly):
        """将总费用分配到 70+ 细项。返回 {item_name: [月1,...]}"""
        # 主要项分配比例
        major_items = {
            "工资": 0.30, "奖金及提成": 0.10, "社保": 0.08, "差旅费": 0.06,
            "招待费": 0.04, "办公费": 0.03, "运费": 0.05, "折旧费": 0.04,
            "租赁费": 0.05, "广告宣传费": 0.03, "样品费": 0.02,
        }
        result = {}
        remaining_pct = 1.0 - sum(major_items.values())
        minor_count = len(EXPENSE_SUB_ITEMS) - len(major_items)

        for item in EXPENSE_SUB_ITEMS:
            if item in major_items:
                share = major_items[item]
            elif item in ("职工薪酬", "福利费", "能源费"):
                # 汇总行，后面会计算
                share = 0
            else:
                share = remaining_pct / max(minor_count, 1) * random.uniform(0, 2)
            monthly = [round(t * share * random.uniform(0.7, 1.3), 1) for t in total_expense_monthly]
            result[item] = monthly

        # 汇总：福利费 = 职工房租+水电物业费+员工餐+节日礼品+生日礼品+体检+其他福利支出
        welfare_items = ["职工房租", "水电物业费", "员工餐", "节日礼品", "生日礼品", "体检", "其他福利支出"]
        result["福利费"] = [
            round(sum(result.get(si, [0]*12)[m] for si in welfare_items if si in result), 1)
            for m in range(12)
        ]

        # 汇总：能源费 = 水费+电费+燃气费+蒸汽费
        energy_items = ["水费", "电费", "燃气费", "蒸汽费"]
        result["能源费"] = [
            round(sum(result.get(si, [0]*12)[m] for si in energy_items if si in result), 1)
            for m in range(12)
        ]

        # 汇总：职工薪酬 = 工资+奖金+社保+公积金+工会+教育+福利费
        salary_items = ["工资", "奖金及提成", "社保", "公积金", "工会经费", "职工教育经费", "福利费"]
        result["职工薪酬"] = [
            round(sum(result.get(si, [0]*12)[m] for si in salary_items if si in result), 1)
            for m in range(12)
        ]
        return result

    def gen_budget(self, actual_list, accuracy=0.15):
        """基于实际值生成预算值（预算通常更平滑）。"""
        avg = sum(actual_list) / len(actual_list) if actual_list else 0
        return [round(avg * random.uniform(1 - accuracy, 1 + accuracy), 1) for _ in actual_list]

    def gen_rebate_detail(self, regions, salespeople_per_region=3, rows=50):
        """生成返利明细表数据。"""
        customer_templates = [
            "{}区{}冷冻食品批发部", "{}市{}食品商行", "{}供应链管理有限公司",
            "{}实业有限公司", "{}食品有限公司", "{}商贸有限公司",
        ]
        surnames = "张王李赵陈刘杨黄周吴徐孙马朱胡林郭何高罗"
        given_names = "伟芳娜秀英敏静丽强磊洋勇艳杰娟涛明超秀兰霞"

        rows_data = []
        for _ in range(rows):
            month = random.randint(1, 12)
            day = random.randint(1, 28)
            dt = date(self.year - 1, month, day)
            region = random.choice(regions)
            name = random.choice(surnames) + random.choice(given_names) + random.choice(given_names[:5])
            customer = random.choice(customer_templates).format(
                random.choice(["钟楼", "武进", "姑苏", "吴中", "海宁", "衢州", "余杭"]),
                random.choice(["京功", "大匠", "优鲜", "百味", "食佳", "淼沃"]),
            )
            amount = round(-random.uniform(500, 15000), 2)
            if random.random() < 0.15:  # 15% 冲减
                amount = abs(amount)
                customer = "冲减计提" + customer

            rows_data.append({
                "date": dt,
                "month": dt.strftime("%Y-%m"),
                "summary": f"{customer}经销商返点实现{abs(int(amount))}",
                "amount": amount,
                "region": region,
                "salesperson": name,
            })

        rows_data.sort(key=lambda x: x["date"])
        return rows_data


# ============================================================
# Sheet 生成器
# ============================================================

def create_index_sheet(wb, tpl, sheet_names):
    """创建索引页。"""
    ws = wb.active
    ws.title = "索引"

    ws.merge_cells("A1:C1")
    style_cell(ws, 1, 1, f"{tpl['center']}经营口径报表索引", TITLE_FONT, alignment=CENTER_ALIGN)

    headers = ["序号", "索引", "报表编制说明"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    descriptions = {
        "收入及净利简表": "1、分部及区域的预算净利及实际净利，数据来源为本文件中分部利润表",
        "利润表": "1、为中心、分部、区域的合计；\n2、所得税为扣除二级BU利润后计提",
    }

    for ri, name in enumerate(sheet_names[1:], 1):
        style_cell(ws, ri + 2, 1, ri, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws, ri + 2, 2, name, border=THIN_BORDER)
        desc = descriptions.get(name, f"编制规则同{sheet_names[2]}") if "利润表" in name else descriptions.get(name, "")
        style_cell(ws, ri + 2, 3, desc, border=THIN_BORDER, alignment=LEFT_ALIGN)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60


def create_summary_sheet(wb, tpl, gen, all_region_data):
    """创建收入及净利简表（多层表头 + 合并单元格）。"""
    ws = wb.create_sheet("收入及净利简表")
    months = list(range(1, 13))

    # R1: Title (merged)
    last_col = 1 + len(months) * 3 + 1  # region + 12months*3cols + total
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(last_col, 40))
    style_cell(ws, 1, 1,
               f"{tpl['center']}{gen.year}年各分部及区域收入&净利简表",
               TITLE_FONT, alignment=CENTER_ALIGN)

    # === 收入 section (R2-R9) ===
    # R2: month headers (merged 3 cols each)
    style_cell(ws, 2, 1, "区域", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for mi, m in enumerate(months):
        col_start = 2 + mi * 3
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + 2)
        style_cell(ws, 2, col_start, f"{m}月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R3: sub-headers
    style_cell(ws, 3, 1, None, border=THIN_BORDER)
    for mi in range(len(months)):
        col_start = 2 + mi * 3
        for ci, sub in enumerate(["预算收入", "24年同期实际", "实际收入"]):
            style_cell(ws, 3, col_start + ci, sub, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R4-R8: region data
    region_names = list(all_region_data.keys())
    for ri, rname in enumerate(region_names, 4):
        data = all_region_data[rname]
        style_cell(ws, ri, 1, rname, SECTION_FONT, border=THIN_BORDER)
        for mi in range(len(months)):
            col = 2 + mi * 3
            rev = data["revenue_total"][mi]
            budget = data["budget_revenue"][mi]
            prev_year = round(rev * random.uniform(0.6, 1.1), 1)
            style_cell(ws, ri, col, budget, border=THIN_BORDER, number_format="#,##0.0")
            style_cell(ws, ri, col + 1, prev_year, border=THIN_BORDER, number_format="#,##0.0")
            style_cell(ws, ri, col + 2, rev, border=THIN_BORDER, number_format="#,##0.0")

    # 合计行
    total_row = 4 + len(region_names)
    style_cell(ws, total_row, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    for mi in range(len(months)):
        col = 2 + mi * 3
        for offset in range(3):
            total = sum(
                ws.cell(row=r, column=col + offset).value or 0
                for r in range(4, total_row)
            )
            style_cell(ws, total_row, col + offset, round(total, 1), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.0")

    # === 净利 section (with gap row) ===
    gap_row = total_row + 1
    net_header_row = gap_row + 1

    # R11: month headers (same merge pattern)
    style_cell(ws, net_header_row, 1, "区域", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for mi, m in enumerate(months):
        col_start = 2 + mi * 3
        ws.merge_cells(start_row=net_header_row, start_column=col_start,
                       end_row=net_header_row, end_column=col_start + 2)
        style_cell(ws, net_header_row, col_start, f"{m}月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R12: sub-headers
    sub_row = net_header_row + 1
    style_cell(ws, sub_row, 1, None, border=THIN_BORDER)
    for mi in range(len(months)):
        col_start = 2 + mi * 3
        for ci, sub in enumerate(["预算净利", "24年同期实际", "实际净利"]):
            style_cell(ws, sub_row, col_start + ci, sub, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R13+: net profit data
    data_start = sub_row + 1
    for ri, rname in enumerate(region_names, data_start):
        data = all_region_data[rname]
        style_cell(ws, ri, 1, rname, SECTION_FONT, border=THIN_BORDER)
        for mi in range(len(months)):
            col = 2 + mi * 3
            net = data["net_profit"][mi]
            budget_net = data["budget_net"][mi]
            prev_net = round(net * random.uniform(0.5, 1.5), 1)
            style_cell(ws, ri, col, budget_net, border=THIN_BORDER, number_format="#,##0.0")
            style_cell(ws, ri, col + 1, prev_net, border=THIN_BORDER, number_format="#,##0.0")
            style_cell(ws, ri, col + 2, net, border=THIN_BORDER, number_format="#,##0.0")

    # 合计
    net_total_row = data_start + len(region_names)
    style_cell(ws, net_total_row, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    for mi in range(len(months)):
        col = 2 + mi * 3
        for offset in range(3):
            total = sum(
                ws.cell(row=r, column=col + offset).value or 0
                for r in range(data_start, net_total_row)
            )
            style_cell(ws, net_total_row, col + offset, round(total, 1), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.0")

    # Notes
    note_row = net_total_row + 1
    style_cell(ws, note_row, 1, "备注：")
    style_cell(ws, note_row + 1, 1, "本表收入为扣除返利后收入，利润为分红前利润")
    style_cell(ws, note_row + 2, 1, "编制说明：")

    ws.column_dimensions["A"].width = 16


def create_profit_sheet(ws, tpl, gen, region_name, region_data, sheet_type="分部"):
    """创建利润表 sheet（与 Test.xlsx 的 268 行结构完全一致）。"""
    products = gen.products
    months = list(range(1, 13))
    year = gen.year

    # ---- Header rows (R1-R5) ----
    title = f"{region_name}利润表 "
    if sheet_type == "center":
        title = f"{tpl['brand']}{tpl['center']}利润表 "
    elif sheet_type == "center_internal":
        title = f"{tpl['center'][:-2] if tpl['center'].endswith('中心') else tpl['center']}利润表 "

    # Total columns: label(1) + 行次(1) + 12 months × 2 + 年合计(2) = 28
    total_cols = 2 + len(months) * 2 + 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    style_cell(ws, 1, 1, title, Font(bold=True, size=16), alignment=CENTER_ALIGN)
    ws.row_dimensions[1].height = 28

    style_cell(ws, 2, 1, datetime(year, 12, 30))
    style_cell(ws, 2, total_cols, "会企01表")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    style_cell(ws, 3, 1, f"编制单位：{tpl['company']}")
    ws.row_dimensions[3].height = 22

    # R4: month headers (merged 2 cols each: 预算数 + 本月实际)
    style_cell(ws, 4, 1, "项　　　  　目　", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 4, 2, "行次", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)

    for mi, m in enumerate(months):
        col = 3 + mi * 2
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        style_cell(ws, 4, col, datetime(year, m, 1), HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        style_cell(ws, 4, col + 1, None, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # 年合计 header (merged 2 cols)
    annual_col = 3 + len(months) * 2
    ws.merge_cells(start_row=4, start_column=annual_col, end_row=4, end_column=annual_col + 1)
    style_cell(ws, 4, annual_col, f"{year}年合计", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R5: sub-headers
    for mi in range(len(months)):
        col = 3 + mi * 2
        style_cell(ws, 5, col, "预算数", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        style_cell(ws, 5, col + 1, "本月实际", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 5, annual_col, "预算数", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 5, annual_col + 1, "本年实际", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24

    # ---- Data generation ----
    rev_by_prod = region_data["revenue_by_product"]
    cost_by_prod = region_data["cost_by_product"]
    budget_rev = region_data["budget_revenue"]
    budget_net = region_data["budget_net"]

    def write_row(row_num, label, line_no, monthly_budget, monthly_actual, is_section=False, is_rate=False):
        """Write a data row with budget + actual for each month + annual totals."""
        font = SECTION_FONT if is_section else None
        fmt = "#,##0.0" if not is_rate else "0.0"
        style_cell(ws, row_num, 1, label, font, border=THIN_BORDER)
        if line_no:
            style_cell(ws, row_num, 2, line_no, border=THIN_BORDER, alignment=CENTER_ALIGN)
        budget_total = 0.0
        actual_total = 0.0
        for mi in range(len(months)):
            col = 3 + mi * 2
            bval = monthly_budget[mi] if monthly_budget else None
            aval = monthly_actual[mi] if monthly_actual else None
            if bval is not None:
                style_cell(ws, row_num, col, bval, border=THIN_BORDER, number_format=fmt)
                budget_total += bval
            if aval is not None:
                style_cell(ws, row_num, col + 1, aval, border=THIN_BORDER, number_format=fmt)
                actual_total += aval
        # Annual totals (skip for ratio/rate rows — average, not sum)
        if not is_rate:
            acol = 3 + len(months) * 2
            if monthly_budget:
                style_cell(ws, row_num, acol, round(budget_total, 1), font, border=THIN_BORDER, number_format=fmt)
            if monthly_actual:
                style_cell(ws, row_num, acol + 1, round(actual_total, 1), font, border=THIN_BORDER, number_format=fmt)
        return row_num + 1

    row = 6

    # 一、营业收入 (line 1)
    total_rev = [sum(rev_by_prod[p][m] for p in products) for m in range(12)]
    returns = [round(r * random.uniform(0.01, 0.03), 1) for r in total_rev]  # 返利
    other_rev = [round(random.uniform(0, total_rev[m] * 0.02), 1) for m in range(12)]
    net_rev = [round(total_rev[m] - returns[m] + other_rev[m], 1) for m in range(12)]
    row = write_row(row, "一、营业收入", "1", budget_rev, net_rev, True)

    # Product revenue rows
    for prod in products:
        row = write_row(row, prod, None, None, rev_by_prod[prod])
    row = write_row(row, "返利", None, None, [-r for r in returns])
    row = write_row(row, "其他", None, None, other_rev)

    # 减：营业成本 (line 2)
    total_cost = [sum(cost_by_prod[p][m] for p in products) for m in range(12)]
    freight = [round(random.uniform(0.02, 0.05) * total_rev[m], 1) for m in range(12)]
    full_cost = [round(total_cost[m] + freight[m], 1) for m in range(12)]
    budget_cost = gen.gen_budget(full_cost)
    row = write_row(row, "减：营业成本", "2", budget_cost, full_cost, True)

    for prod in products:
        row = write_row(row, prod, None, None, cost_by_prod[prod])
    other_cost = [round(random.uniform(0, total_cost[m] * 0.02), 1) for m in range(12)]
    row = write_row(row, "其他", None, None, other_cost)
    row = write_row(row, "运费", None, None, freight)

    # 毛利 (line 3)
    gross_profit = [round(net_rev[m] - full_cost[m], 1) for m in range(12)]
    budget_gp = [round(budget_rev[m] - budget_cost[m], 1) for m in range(12)]
    row = write_row(row, "毛利", "3", budget_gp, gross_profit, True)

    # 毛利率 (line 4)
    gp_rate = [round(gross_profit[m] / net_rev[m] * 100, 1) if net_rev[m] else 0 for m in range(12)]
    budget_gp_rate = [round(budget_gp[m] / budget_rev[m] * 100, 1) if budget_rev[m] else 0 for m in range(12)]
    row = write_row(row, "毛利率", "4", budget_gp_rate, gp_rate, is_rate=True)

    # 税金及附加 (line 5)
    tax = [round(net_rev[m] * random.uniform(0.005, 0.015), 1) for m in range(12)]
    row = write_row(row, "税金及附加", "5", gen.gen_budget(tax), tax, True)

    # 销售费用 (line 6), 管理费用 (line 7), 研发费用 (line 8)
    for line_no, expense_name in [("6", "销售费用"), ("7", "管理费用"), ("8", "研发费用")]:
        if expense_name == "销售费用":
            total_exp = [round(net_rev[m] * random.uniform(*tpl["expense_ratio"]), 1) for m in range(12)]
        elif expense_name == "管理费用":
            total_exp = [round(net_rev[m] * random.uniform(0.03, 0.08), 1) for m in range(12)]
        else:
            total_exp = [round(net_rev[m] * random.uniform(0.01, 0.04), 1) for m in range(12)]

        budget_exp = gen.gen_budget(total_exp)
        row = write_row(row, expense_name, line_no, budget_exp, total_exp, True)

        # Expense sub-items
        sub_data = gen.gen_expense_items(total_exp)
        for item_name in EXPENSE_SUB_ITEMS:
            if item_name in sub_data:
                row = write_row(row, item_name, None, None, sub_data[item_name])

    # 财务费用~其他收益 (lines 9-16)
    fin_expense = [round(random.uniform(-5000, 20000), 1) for _ in range(12)]
    row = write_row(row, "财务费用", "9", gen.gen_budget(fin_expense), fin_expense, True)
    row = write_row(row, "资产减值损失", "10", [0]*12, [0]*12)
    row = write_row(row, "信用减值损失", "11", [0]*12, [round(random.uniform(-5000, 5000), 1) for _ in range(12)])
    row = write_row(row, "加：公允价值变动收益（损失以\"-\"号填列）", "12", [0]*12, [0]*12)
    row = write_row(row, "投资收益（损失以\"-\"号填列）", "13", [0]*12, [0]*12)
    row = write_row(row, "其中：对联营企业和合营企业的投资收益", "14", [0]*12, [0]*12)
    row = write_row(row, "资产处置收益（损失以\"-\"号填列）", "15", [0]*12, [0]*12)
    other_income = [round(random.uniform(0, 30000), 1) for _ in range(12)]
    row = write_row(row, "其他收益", "16", gen.gen_budget(other_income), other_income)

    # 二、营业利润 (line 17) — calculated
    # Simplified: gross_profit - expenses - tax + other
    op_profit = [round(gross_profit[m] * random.uniform(0.2, 0.5), 1) for m in range(12)]
    row = write_row(row, "二、营业利润（亏损以\"-\"号填列）", "17", gen.gen_budget(op_profit), op_profit, True)

    row = write_row(row, "加：营业外收入", "18", [0]*12, [round(random.uniform(0, 5000), 1) for _ in range(12)])
    row = write_row(row, "减：营业外支出", "19", [0]*12, [round(random.uniform(0, 3000), 1) for _ in range(12)])

    # 三、利润总额
    total_profit = [round(op_profit[m] * random.uniform(0.9, 1.1), 1) for m in range(12)]
    row = write_row(row, "三、利润总额（亏损总额以\"-\"号填列）", "20", gen.gen_budget(total_profit), total_profit, True)

    income_tax = [round(max(0, total_profit[m] * 0.25), 1) for m in range(12)]
    row = write_row(row, "减：所得税费用", "21", gen.gen_budget(income_tax), income_tax)

    net_profit = [round(total_profit[m] - income_tax[m], 1) for m in range(12)]
    row = write_row(row, "四、净利润（净亏损以\"-\"号填列）", "22", budget_net, net_profit, True)

    dividend = [round(max(0, net_profit[m] * random.uniform(0, 0.3)), 1) for m in range(12)]
    row = write_row(row, "五、分红", "23", [0]*12, dividend)

    after_div = [round(net_profit[m] - dividend[m], 1) for m in range(12)]
    row = write_row(row, "六、分红后净利润（净亏损以\"-\"号填列）", "24", budget_net, after_div, True)

    # Ratios
    net_rate = [round(net_profit[m] / net_rev[m] * 100, 1) if net_rev[m] else 0 for m in range(12)]
    row = write_row(row, "净利率", None, None, net_rate, is_rate=True)

    sales_achieve = [round(net_rev[m] / budget_rev[m] * 100, 1) if budget_rev[m] else 0 for m in range(12)]
    row = write_row(row, "销售达成率", None, None, sales_achieve, is_rate=True)

    net_achieve = [round(net_profit[m] / budget_net[m] * 100, 1) if budget_net[m] else 0 for m in range(12)]
    row = write_row(row, "净利达成率", None, None, net_achieve, is_rate=True)

    # Footer note (merged)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=total_cols)
    style_cell(ws, row, 1, "编制说明：")
    style_cell(ws, row, 2, f"1、营业收入的取数来源为{year}年1-12月销售明细表\n2、数据基于ERP系统导出")

    # Column widths
    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["B"].width = 6
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    # Store net_profit for summary
    return net_profit


def create_rebate_sheet(wb, gen, regions):
    """创建返利明细表。"""
    ws = wb.create_sheet("24年返利明细")
    region_names = [r["name"] for r in regions]
    data = gen.gen_rebate_detail(region_names)

    headers = ["日期", "月份", "摘要", "贷方金额", "区域", "业务员"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    for ri, row_data in enumerate(data, 2):
        style_cell(ws, ri, 1, row_data["date"], border=THIN_BORDER, number_format="YYYY-MM-DD")
        style_cell(ws, ri, 2, row_data["month"], border=THIN_BORDER)
        style_cell(ws, ri, 3, row_data["summary"], border=THIN_BORDER)
        style_cell(ws, ri, 4, row_data["amount"], border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, ri, 5, row_data["region"], border=THIN_BORDER)
        style_cell(ws, ri, 6, row_data["salesperson"], border=THIN_BORDER)

    # 合计行
    total_row = len(data) + 2
    style_cell(ws, total_row, 3, "合计", SECTION_FONT, border=THIN_BORDER)
    total_amount = round(sum(d["amount"] for d in data), 2)
    style_cell(ws, total_row, 4, total_amount, SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10


# ============================================================
# 新增复杂 Sheet 生成器
# ============================================================

def create_balance_sheet(wb, tpl, gen):
    """资产负债表 — 3级合并表头，资产/负债/权益 三大板块。"""
    ws = wb.create_sheet("资产负债表")
    year = gen.year

    # Title (3-level merged header)
    ws.merge_cells("A1:F1")
    style_cell(ws, 1, 1, f"{tpl['company']}资产负债表", TITLE_FONT, alignment=CENTER_ALIGN)
    ws.merge_cells("A2:E2")
    style_cell(ws, 2, 1, f"编制日期：{year}年12月31日", alignment=Alignment(horizontal="center"))
    style_cell(ws, 2, 8, "单位：元")

    # R3: 8-col layout: A-D = 资产 side, E-H = 负债 side
    # Write headers avoiding merged cells
    style_cell(ws, 3, 1, "资产", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 2, "行次", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 3, "期末余额", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 4, "年初余额", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 5, "负债和所有者权益", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 6, "行次", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 7, "期末余额", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 8, "年初余额", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Asset items (left side)
    asset_items = [
        ("流动资产：", None, True),
        ("货币资金", 1, False),
        ("交易性金融资产", 2, False),
        ("应收票据", 3, False),
        ("应收账款", 4, False),
        ("预付款项", 5, False),
        ("其他应收款", 6, False),
        ("存货", 7, False),
        ("其中：原材料", None, False),
        ("在制品", None, False),
        ("库存商品", None, False),
        ("合同资产", 8, False),
        ("一年内到期的非流动资产", 9, False),
        ("其他流动资产", 10, False),
        ("流动资产合计", 11, True),
        ("非流动资产：", None, True),
        ("长期股权投资", 12, False),
        ("固定资产", 13, False),
        ("在建工程", 14, False),
        ("无形资产", 15, False),
        ("长期待摊费用", 16, False),
        ("递延所得税资产", 17, False),
        ("其他非流动资产", 18, False),
        ("非流动资产合计", 19, True),
        ("资产总计", 20, True),
    ]

    # Liability items (right side)
    liability_items = [
        ("流动负债：", None, True),
        ("短期借款", 21, False),
        ("应付票据", 22, False),
        ("应付账款", 23, False),
        ("预收款项", 24, False),
        ("合同负债", 25, False),
        ("应付职工薪酬", 26, False),
        ("应交税费", 27, False),
        ("其他应付款", 28, False),
        ("一年内到期的非流动负债", 29, False),
        ("其他流动负债", 30, False),
        ("流动负债合计", 31, True),
        ("非流动负债：", None, True),
        ("长期借款", 32, False),
        ("长期应付款", 33, False),
        ("递延收益", 34, False),
        ("递延所得税负债", 35, False),
        ("非流动负债合计", 36, True),
        ("负债合计", 37, True),
        ("所有者权益：", None, True),
        ("实收资本", 38, False),
        ("资本公积", 39, False),
        ("盈余公积", 40, False),
        ("未分配利润", 41, False),
        ("所有者权益合计", 42, True),
        ("负债和所有者权益总计", 43, True),
    ]

    # Generate data — scale by scenario
    scale = gen.scn["revenue_mult"]
    total_assets_end = random.uniform(50_000_000, 200_000_000) * scale
    total_assets_begin = total_assets_end * random.uniform(0.85, 1.15)

    def gen_bs_values(total, count):
        """Distribute total among count items with realistic proportions."""
        weights = [random.uniform(0.5, 5) for _ in range(count)]
        total_w = sum(weights)
        return [round(total * w / total_w, 2) for w in weights]

    # Asset values
    current_items = [i for i in asset_items if not i[2] and i[1] is not None and i[1] <= 10]
    noncurrent_items = [i for i in asset_items if not i[2] and i[1] is not None and i[1] > 11 and i[1] < 20]
    current_end = gen_bs_values(total_assets_end * 0.6, len(current_items))
    current_begin = gen_bs_values(total_assets_begin * 0.6, len(current_items))
    noncurrent_end = gen_bs_values(total_assets_end * 0.4, len(noncurrent_items))
    noncurrent_begin = gen_bs_values(total_assets_begin * 0.4, len(noncurrent_items))

    # Liability values
    total_liab_end = total_assets_end * random.uniform(0.4, 0.7)
    total_liab_begin = total_assets_begin * random.uniform(0.4, 0.7)
    equity_end = total_assets_end - total_liab_end
    equity_begin = total_assets_begin - total_liab_begin

    cl_items = [i for i in liability_items if not i[2] and i[1] is not None and i[1] <= 30]
    ncl_items = [i for i in liability_items if not i[2] and i[1] is not None and 32 <= i[1] <= 35]
    eq_items = [i for i in liability_items if not i[2] and i[1] is not None and i[1] >= 38]
    cl_end = gen_bs_values(total_liab_end * 0.7, len(cl_items))
    cl_begin = gen_bs_values(total_liab_begin * 0.7, len(cl_items))
    ncl_end = gen_bs_values(total_liab_end * 0.3, len(ncl_items))
    ncl_begin = gen_bs_values(total_liab_begin * 0.3, len(ncl_items))
    eq_end = gen_bs_values(equity_end, len(eq_items))
    eq_begin = gen_bs_values(equity_begin, len(eq_items))

    # Write asset side (cols A-D)
    row = 4
    ci_cur = 0
    ci_nc = 0
    sum_cur_end = sum_cur_begin = 0
    sum_nc_end = sum_nc_begin = 0
    for label, line_no, is_sect in asset_items:
        font = SECTION_FONT if is_sect else None
        style_cell(ws, row, 1, label, font, border=THIN_BORDER)
        if line_no:
            style_cell(ws, row, 2, line_no, border=THIN_BORDER, alignment=CENTER_ALIGN)
        if label == "流动资产合计":
            style_cell(ws, row, 3, round(sum(current_end), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, round(sum(current_begin), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif label == "非流动资产合计":
            style_cell(ws, row, 3, round(sum(noncurrent_end), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, round(sum(noncurrent_begin), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif label == "资产总计":
            style_cell(ws, row, 3, round(total_assets_end, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, round(total_assets_begin, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif not is_sect and line_no and line_no <= 10:
            style_cell(ws, row, 3, current_end[ci_cur], border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, current_begin[ci_cur], border=THIN_BORDER, number_format="#,##0.00")
            ci_cur += 1
        elif not is_sect and line_no and 12 <= line_no <= 18:
            style_cell(ws, row, 3, noncurrent_end[ci_nc], border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, noncurrent_begin[ci_nc], border=THIN_BORDER, number_format="#,##0.00")
            ci_nc += 1
        elif label.startswith("其中：") or label in ("在制品", "库存商品"):
            parent_val = current_end[ci_cur - 1] if ci_cur > 0 else 0
            style_cell(ws, row, 3, round(parent_val * random.uniform(0.1, 0.5), 2), border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 4, round(parent_val * random.uniform(0.1, 0.5), 2), border=THIN_BORDER, number_format="#,##0.00")
        row += 1

    # Write liability side (cols E-H), starting at row 4
    row = 4
    ci_cl = ci_ncl = ci_eq = 0
    for label, line_no, is_sect in liability_items:
        font = SECTION_FONT if is_sect else None
        style_cell(ws, row, 5, label, font, border=THIN_BORDER)
        if line_no:
            style_cell(ws, row, 6, line_no, border=THIN_BORDER, alignment=CENTER_ALIGN)
        if label == "流动负债合计":
            style_cell(ws, row, 7, round(sum(cl_end), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 8, round(sum(cl_begin), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif label == "非流动负债合计":
            style_cell(ws, row, 7, round(sum(ncl_end), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 8, round(sum(ncl_begin), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif label == "负债合计":
            style_cell(ws, row, 7, round(total_liab_end, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 8, round(total_liab_begin, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif label == "所有者权益合计":
            style_cell(ws, row, 7, round(equity_end, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 8, round(equity_begin, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif label == "负债和所有者权益总计":
            style_cell(ws, row, 7, round(total_assets_end, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 8, round(total_assets_begin, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
        elif not is_sect and line_no:
            if line_no <= 30:
                style_cell(ws, row, 7, cl_end[ci_cl], border=THIN_BORDER, number_format="#,##0.00")
                style_cell(ws, row, 8, cl_begin[ci_cl], border=THIN_BORDER, number_format="#,##0.00")
                ci_cl += 1
            elif 32 <= line_no <= 35:
                style_cell(ws, row, 7, ncl_end[ci_ncl], border=THIN_BORDER, number_format="#,##0.00")
                style_cell(ws, row, 8, ncl_begin[ci_ncl], border=THIN_BORDER, number_format="#,##0.00")
                ci_ncl += 1
            elif line_no >= 38:
                style_cell(ws, row, 7, eq_end[ci_eq], border=THIN_BORDER, number_format="#,##0.00")
                style_cell(ws, row, 8, eq_begin[ci_eq], border=THIN_BORDER, number_format="#,##0.00")
                ci_eq += 1
        row += 1

    # Column widths
    for c, w in [(1, 22), (2, 6), (3, 16), (4, 16), (5, 22), (6, 6), (7, 16), (8, 16)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_sales_detail(wb, tpl, gen):
    """销售明细表 — 300+ 行交易流水，10 列，包含日期/客户/产品/数量/单价/金额。"""
    ws = wb.create_sheet("销售明细")
    year = gen.year
    products = tpl["products"]
    regions = [r["name"] for r in tpl["regions"]] + [sr["name"] for sr in tpl["sub_regions"]]

    # Title
    ws.merge_cells("A1:J1")
    style_cell(ws, 1, 1, f"{tpl['company']}{year}年销售明细表", TITLE_FONT, alignment=CENTER_ALIGN)

    headers = ["日期", "订单号", "客户名称", "产品", "规格型号", "数量", "单价(元)", "金额(元)", "区域", "业务员"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    surnames = "张王李赵陈刘杨黄周吴徐孙马朱胡林郭何高罗"
    given = "伟芳娜秀英敏静丽强磊洋勇艳杰娟涛明超秀兰"
    specs = ["A级", "B级", "C级", "特级", "标准", "精装", "简装", "500g", "1kg", "2.5kg", "5kg", "10kg", "25kg"]
    cust_tpls = ["{}食品有限公司", "{}商贸有限公司", "{}超市", "{}餐饮管理有限公司",
                 "{}供应链有限公司", "{}酒店管理集团", "{}电商平台"]
    cities = ["南京", "苏州", "无锡", "杭州", "宁波", "上海", "合肥", "南昌", "温州", "常州", "嘉兴", "镇江"]

    rows_data = []
    n_rows = random.randint(280, 350) if gen.scn["sparse_pct"] == 0 else random.randint(100, 180)
    order_counter = 10001
    for _ in range(n_rows):
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        dt = date(year, month, day)
        prod = random.choice(products)
        qty = random.randint(1, 500) * 10
        unit_price = round(random.uniform(5, 200) * gen.scn["revenue_mult"], 2)
        amount = round(qty * unit_price, 2)
        city = random.choice(cities)
        cust = random.choice(cust_tpls).format(city + random.choice(["京功", "大匠", "优鲜", "百味", "食佳", "淼沃"]))
        salesperson = random.choice(surnames) + random.choice(given) + random.choice(given[:6])
        rows_data.append({
            "date": dt, "order": f"SO{year}{order_counter}",
            "customer": cust, "product": prod,
            "spec": random.choice(specs), "qty": qty,
            "unit_price": unit_price, "amount": amount,
            "region": random.choice(regions), "salesperson": salesperson,
        })
        order_counter += 1

    rows_data.sort(key=lambda x: x["date"])
    for ri, d in enumerate(rows_data, 3):
        style_cell(ws, ri, 1, d["date"], border=THIN_BORDER, number_format="YYYY-MM-DD")
        style_cell(ws, ri, 2, d["order"], border=THIN_BORDER)
        style_cell(ws, ri, 3, d["customer"], border=THIN_BORDER)
        style_cell(ws, ri, 4, d["product"], border=THIN_BORDER)
        style_cell(ws, ri, 5, d["spec"], border=THIN_BORDER)
        style_cell(ws, ri, 6, d["qty"], border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, ri, 7, d["unit_price"], border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, ri, 8, d["amount"], border=THIN_BORDER, number_format="#,##0.00")
        style_cell(ws, ri, 9, d["region"], border=THIN_BORDER)
        style_cell(ws, ri, 10, d["salesperson"], border=THIN_BORDER)

    # Summary row
    tr = len(rows_data) + 3
    style_cell(ws, tr, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, tr, 6, sum(d["qty"] for d in rows_data), SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    style_cell(ws, tr, 8, round(sum(d["amount"] for d in rows_data), 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

    for c, w in [(1, 12), (2, 18), (3, 28), (4, 14), (5, 10), (6, 8), (7, 10), (8, 14), (9, 12), (10, 10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_ar_aging_sheet(wb, tpl, gen):
    """应收账款账龄分析表 — 客户×账龄交叉表 + 合并表头。"""
    ws = wb.create_sheet("应收账款账龄")
    year = gen.year

    ws.merge_cells("A1:H1")
    style_cell(ws, 1, 1, f"{tpl['company']}应收账款账龄分析表", TITLE_FONT, alignment=CENTER_ALIGN)
    style_cell(ws, 2, 1, f"截止日期：{year}年12月31日")
    style_cell(ws, 2, 8, "单位：元")

    # 2-level header with merges
    # R3: 客户 | 应收余额 | 账龄分布(merged 5 cols) | 信用额度
    style_cell(ws, 3, 1, "客户名称", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 2, "应收余额", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells("C3:G3")
    style_cell(ws, 3, 3, "账龄分布", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, 8, "信用额度", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:B4")
    ws.merge_cells("H3:H4")

    # R4: age bucket sub-headers
    age_buckets = ["0-30天", "31-60天", "61-90天", "91-180天", "180天以上"]
    for ci, bucket in enumerate(age_buckets, 3):
        style_cell(ws, 4, ci, bucket, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Customer data
    cities = ["南京", "苏州", "无锡", "杭州", "宁波", "上海", "合肥", "深圳", "广州", "成都",
              "武汉", "长沙", "济南", "青岛", "厦门", "福州", "昆明", "西安", "重庆", "天津"]
    suffixes = ["食品有限公司", "商贸有限公司", "餐饮集团", "连锁超市", "供应链有限公司",
                "实业有限公司", "进出口有限公司", "电子商务有限公司"]

    n_customers = random.randint(25, 40)
    customers = [f"{random.choice(cities)}{random.choice(['京功','大匠','优鲜','百味','鑫达','汇通','荣泰','永利'])}{random.choice(suffixes)}"
                 for _ in range(n_customers)]

    row = 5
    total_by_bucket = [0] * 5
    total_balance = 0
    total_credit = 0
    for cust in customers:
        balance = round(random.uniform(10000, 2000000) * gen.scn["revenue_mult"], 2)
        # Distribute across age buckets (younger = more typical)
        weights = [random.uniform(3, 8), random.uniform(1, 4), random.uniform(0, 2), random.uniform(0, 1.5), random.uniform(0, 0.8)]
        total_w = sum(weights)
        buckets = [round(balance * w / total_w, 2) for w in weights]
        # Adjust to match total
        buckets[-1] = round(balance - sum(buckets[:-1]), 2)
        credit = round(balance * random.uniform(1.0, 2.5), 2)

        style_cell(ws, row, 1, cust, border=THIN_BORDER)
        style_cell(ws, row, 2, balance, border=THIN_BORDER, number_format="#,##0.00")
        for bi, bv in enumerate(buckets):
            style_cell(ws, row, 3 + bi, bv, border=THIN_BORDER, number_format="#,##0.00")
            total_by_bucket[bi] += bv
        style_cell(ws, row, 8, credit, border=THIN_BORDER, number_format="#,##0.00")
        total_balance += balance
        total_credit += credit
        row += 1

    # Total row
    style_cell(ws, row, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, row, 2, round(total_balance, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    for bi in range(5):
        style_cell(ws, row, 3 + bi, round(total_by_bucket[bi], 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")
    style_cell(ws, row, 8, round(total_credit, 2), SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

    # Percentage row
    row += 1
    style_cell(ws, row, 1, "占比", SECTION_FONT, border=THIN_BORDER)
    style_cell(ws, row, 2, 1.0, SECTION_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
    for bi in range(5):
        pct = round(total_by_bucket[bi] / total_balance, 4) if total_balance else 0
        style_cell(ws, row, 3 + bi, pct, SECTION_FONT, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    for c, w in [(1, 30), (2, 16), (3, 14), (4, 14), (5, 14), (6, 14), (7, 14), (8, 16)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_inventory_sheet(wb, tpl, gen):
    """库存台账 — 产品×仓库 交叉表 + 期初/入库/出库/期末 多层表头。"""
    ws = wb.create_sheet("库存台账")
    year = gen.year
    products = tpl["products"]
    warehouses = ["成品仓", "原材料仓", "半成品仓", "冷链仓"]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(warehouses) * 4)
    style_cell(ws, 1, 1, f"{tpl['company']}{year}年12月库存台账", TITLE_FONT, alignment=CENTER_ALIGN)

    # R2: warehouse headers (merged 4 cols each)
    style_cell(ws, 2, 1, "产品名称", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 2, 2, "规格", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    for wi, wh in enumerate(warehouses):
        col = 3 + wi * 4
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        style_cell(ws, 2, col, wh, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R3: sub-headers per warehouse
    sub_headers = ["期初", "入库", "出库", "期末"]
    for wi in range(len(warehouses)):
        col = 3 + wi * 4
        for si, sh in enumerate(sub_headers):
            style_cell(ws, 3, col + si, sh, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # 合计 columns
    total_col = 3 + len(warehouses) * 4
    ws.merge_cells(start_row=2, start_column=total_col, end_row=2, end_column=total_col + 3)
    style_cell(ws, 2, total_col, "合计", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for si, sh in enumerate(sub_headers):
        style_cell(ws, 3, total_col + si, sh, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Data rows
    specs = ["500g/袋", "1kg/箱", "2.5kg/箱", "5kg/箱", "10kg/桶", "25kg/袋"]
    row = 4
    totals = {wh: {"begin": 0, "in": 0, "out": 0, "end": 0} for wh in warehouses + ["合计"]}

    for prod in products:
        spec = random.choice(specs)
        style_cell(ws, row, 1, prod, border=THIN_BORDER)
        style_cell(ws, row, 2, spec, border=THIN_BORDER)

        prod_total = {"begin": 0, "in": 0, "out": 0, "end": 0}
        for wi, wh in enumerate(warehouses):
            col = 3 + wi * 4
            begin = random.randint(50, 5000)
            in_qty = random.randint(100, 8000)
            out_qty = random.randint(80, min(begin + in_qty, 7000))
            end = begin + in_qty - out_qty
            style_cell(ws, row, col, begin, border=THIN_BORDER, number_format="#,##0")
            style_cell(ws, row, col + 1, in_qty, border=THIN_BORDER, number_format="#,##0")
            style_cell(ws, row, col + 2, out_qty, border=THIN_BORDER, number_format="#,##0")
            style_cell(ws, row, col + 3, end, border=THIN_BORDER, number_format="#,##0")
            totals[wh]["begin"] += begin
            totals[wh]["in"] += in_qty
            totals[wh]["out"] += out_qty
            totals[wh]["end"] += end
            prod_total["begin"] += begin
            prod_total["in"] += in_qty
            prod_total["out"] += out_qty
            prod_total["end"] += end

        # Product total
        style_cell(ws, row, total_col, prod_total["begin"], border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, total_col + 1, prod_total["in"], border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, total_col + 2, prod_total["out"], border=THIN_BORDER, number_format="#,##0")
        style_cell(ws, row, total_col + 3, prod_total["end"], border=THIN_BORDER, number_format="#,##0")
        totals["合计"]["begin"] += prod_total["begin"]
        totals["合计"]["in"] += prod_total["in"]
        totals["合计"]["out"] += prod_total["out"]
        totals["合计"]["end"] += prod_total["end"]
        row += 1

    # Total row
    style_cell(ws, row, 1, "合计", SECTION_FONT, border=THIN_BORDER)
    for wi, wh in enumerate(warehouses):
        col = 3 + wi * 4
        for si, key in enumerate(["begin", "in", "out", "end"]):
            style_cell(ws, row, col + si, totals[wh][key], SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
    for si, key in enumerate(["begin", "in", "out", "end"]):
        style_cell(ws, row, total_col + si, totals["合计"][key], SECTION_FONT, border=THIN_BORDER, number_format="#,##0")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    for c in range(3, total_col + 5):
        ws.column_dimensions[get_column_letter(c)].width = 10


def create_monthly_kpi_sheet(wb, tpl, gen, all_region_data):
    """月度经营KPI — 多区块布局: KPI指标区 + 产品销量排名 + 区域对比。"""
    ws = wb.create_sheet("月度经营分析")
    year = gen.year

    # === Section 1: KPI Summary (R1-R8) ===
    ws.merge_cells("A1:H1")
    style_cell(ws, 1, 1, f"{tpl['company']}{year}年度经营KPI", TITLE_FONT, alignment=CENTER_ALIGN)

    kpi_headers = ["指标", "年度目标", "年度实际", "达成率", "去年同期", "同比增长"]
    for ci, h in enumerate(kpi_headers, 1):
        style_cell(ws, 2, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Compute totals from region data
    total_rev = sum(
        sum(rd["revenue_total"])
        for rn, rd in all_region_data.items()
        if rd.get("type") not in ("expense", "center")
    )
    total_budget = sum(
        sum(rd["budget_revenue"])
        for rn, rd in all_region_data.items()
        if rd.get("type") not in ("expense", "center")
    )
    prev_year = total_rev * random.uniform(0.75, 0.95)
    net_profit = total_rev * random.uniform(0.03, 0.10)

    kpi_rows = [
        ("营业收入(万元)", round(total_budget / 10000, 1), round(total_rev / 10000, 1),
         Pct(round(total_rev / total_budget, 4)), round(prev_year / 10000, 1),
         Pct(round((total_rev - prev_year) / prev_year, 4))),
        ("净利润(万元)", round(net_profit * 1.2 / 10000, 1), round(net_profit / 10000, 1),
         Pct(round(net_profit / (net_profit * 1.2), 4)), round(net_profit * 0.8 / 10000, 1),
         Pct(round((net_profit - net_profit * 0.8) / (net_profit * 0.8), 4))),
        ("毛利率", Pct(0.30), Pct(round(random.uniform(0.25, 0.35), 4)), "-",
         Pct(round(random.uniform(0.25, 0.35), 4)), "-"),
        ("净利率", Pct(0.05), Pct(round(net_profit / total_rev, 4)), "-",
         Pct(round(random.uniform(0.03, 0.08), 4)), "-"),
        ("客户数(家)", random.randint(300, 500), random.randint(250, 500), "-",
         random.randint(200, 400), "-"),
    ]
    for ri, kpi in enumerate(kpi_rows, 3):
        for ci, v in enumerate(kpi, 1):
            font = SECTION_FONT if ci == 1 else None
            if isinstance(v, Pct):
                style_cell(ws, ri, ci, v.val, font, border=THIN_BORDER,
                           alignment=CENTER_ALIGN, number_format="0.0%")
            else:
                style_cell(ws, ri, ci, v, font, border=THIN_BORDER,
                           alignment=CENTER_ALIGN if ci > 1 else None)

    # === Section 2: Product ranking (R10+) ===
    ws.merge_cells("A10:F10")
    style_cell(ws, 10, 1, "产品销售排名", SECTION_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    rank_headers = ["排名", "产品", "销售额(万元)", "占比", "同比", "毛利率"]
    for ci, h in enumerate(rank_headers, 1):
        style_cell(ws, 11, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    prod_sales = {}
    for rn, rd in all_region_data.items():
        if rd.get("type") in ("expense", "center"):
            continue
        for prod in tpl["products"]:
            if prod not in prod_sales:
                prod_sales[prod] = 0
            prod_sales[prod] += sum(rd["revenue_by_product"].get(prod, [0] * 12))
    sorted_prods = sorted(prod_sales.items(), key=lambda x: -x[1])
    total_prod_rev = sum(v for _, v in sorted_prods)

    for ri, (prod, sales) in enumerate(sorted_prods, 12):
        rank = ri - 11
        pct = round(sales / total_prod_rev, 4)
        yoy = round(random.uniform(-0.15, 0.30), 4)
        gm = round(random.uniform(0.20, 0.45), 4)
        style_cell(ws, ri, 1, rank, border=THIN_BORDER, alignment=CENTER_ALIGN)
        style_cell(ws, ri, 2, prod, border=THIN_BORDER)
        style_cell(ws, ri, 3, round(sales / 10000, 1), border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, ri, 4, pct, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, ri, 5, yoy, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="+0.0%;-0.0%")
        style_cell(ws, ri, 6, gm, border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")

    # === Section 3: Regional comparison (offset to right side) ===
    region_start_row = 10
    region_start_col = 8
    ws.merge_cells(start_row=region_start_row, start_column=region_start_col,
                   end_row=region_start_row, end_column=region_start_col + 4)
    style_cell(ws, region_start_row, region_start_col, "区域经营对比", SECTION_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    reg_headers = ["区域", "收入(万元)", "净利(万元)", "净利率", "达成率"]
    for ci, h in enumerate(reg_headers):
        style_cell(ws, region_start_row + 1, region_start_col + ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    region_row = region_start_row + 2
    for rn, rd in all_region_data.items():
        if rd.get("type") in ("expense", "center"):
            continue
        rev = sum(rd["revenue_total"])
        net = sum(rd["net_profit"])
        budget = sum(rd["budget_revenue"])
        style_cell(ws, region_row, region_start_col, rn, border=THIN_BORDER)
        style_cell(ws, region_row, region_start_col + 1, round(rev / 10000, 1), border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, region_row, region_start_col + 2, round(net / 10000, 1), border=THIN_BORDER, number_format="#,##0.0")
        style_cell(ws, region_row, region_start_col + 3, round(net / rev, 4), border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        style_cell(ws, region_row, region_start_col + 4, round(rev / budget, 4), border=THIN_BORDER, alignment=CENTER_ALIGN, number_format="0.0%")
        region_row += 1

    ws.column_dimensions["A"].width = 16
    for c in range(2, 13):
        ws.column_dimensions[get_column_letter(c)].width = 14


def create_budget_comparison_sheet(wb, tpl, gen):
    """费用预算执行对比 — 宽表: 部门×月度 预算/实际/差异 三联列。"""
    ws = wb.create_sheet("费用预算执行")
    year = gen.year
    months = list(range(1, 13))
    departments = ["销售部", "市场部", "研发部", "生产部", "采购部", "人力资源部", "财务部", "行政部"]

    # Title
    total_cols = 2 + len(months) * 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 38))
    style_cell(ws, 1, 1, f"{tpl['company']}{year}年费用预算执行表", TITLE_FONT, alignment=CENTER_ALIGN)

    # R2: month headers (merged 3 cols each)
    style_cell(ws, 2, 1, "部门", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 2, 2, "费用类别", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    for mi, m in enumerate(months):
        col = 3 + mi * 3
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        style_cell(ws, 2, col, f"{m}月", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # R3: sub-headers
    for mi in range(len(months)):
        col = 3 + mi * 3
        style_cell(ws, 3, col, "预算", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        style_cell(ws, 3, col + 1, "实际", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        style_cell(ws, 3, col + 2, "差异", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    RED_FONT = Font(color="FF0000")
    expense_types = ["人工费用", "差旅费", "业务招待费", "办公费", "折旧费"]

    row = 4
    for dept in departments:
        dept_start_row = row
        dept_base = random.uniform(50000, 300000) * gen.scn["revenue_mult"]

        for exp_type in expense_types:
            style_cell(ws, row, 2, exp_type, border=THIN_BORDER)
            exp_share = {"人工费用": 0.45, "差旅费": 0.15, "业务招待费": 0.12, "办公费": 0.10, "折旧费": 0.08}.get(exp_type, 0.10)
            for mi in range(12):
                col = 3 + mi * 3
                seasonal = [0.8, 0.7, 0.9, 1.0, 1.0, 1.1, 1.2, 1.0, 0.9, 1.1, 1.0, 1.3][mi]
                budget = round(dept_base * exp_share * seasonal * random.uniform(0.9, 1.1), 0)
                actual = round(budget * random.uniform(0.7, 1.3), 0)
                diff = round(actual - budget, 0)
                style_cell(ws, row, col, budget, border=THIN_BORDER, number_format="#,##0")
                style_cell(ws, row, col + 1, actual, border=THIN_BORDER, number_format="#,##0")
                diff_font = RED_FONT if diff > 0 else None
                style_cell(ws, row, col + 2, diff, diff_font, border=THIN_BORDER, number_format="#,##0")
            row += 1

        # Department subtotal
        style_cell(ws, row, 2, "小计", SECTION_FONT, border=THIN_BORDER)
        for mi in range(12):
            col = 3 + mi * 3
            for offset in range(3):
                total = sum(ws.cell(r, col + offset).value or 0 for r in range(dept_start_row, row))
                style_cell(ws, row, col + offset, round(total, 0), SECTION_FONT, border=THIN_BORDER, number_format="#,##0")
        row += 1

        # Merge department name across its rows
        ws.merge_cells(start_row=dept_start_row, start_column=1, end_row=row - 1, end_column=1)
        style_cell(ws, dept_start_row, 1, dept, SECTION_FONT, border=THIN_BORDER,
                   alignment=Alignment(horizontal="center", vertical="center", text_rotation=90))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    for c in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10


def create_cashflow_sheet(wb, tpl, gen):
    """现金流量表 — 经营/投资/筹资 三大活动 + 分季度对比。"""
    ws = wb.create_sheet("现金流量表")
    year = gen.year

    # Title with merges
    ws.merge_cells("A1:F1")
    style_cell(ws, 1, 1, f"{tpl['company']}现金流量表", TITLE_FONT, alignment=CENTER_ALIGN)
    ws.merge_cells("A2:E2")
    style_cell(ws, 2, 1, f"{year}年度", alignment=Alignment(horizontal="center"))
    style_cell(ws, 2, 7, "单位：元")

    # Headers
    headers = ["项目", "行次", "Q1", "Q2", "Q3", "Q4", "全年合计"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 3, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    cf_items = [
        # Operating activities
        ("一、经营活动产生的现金流量", None, True),
        ("销售商品、提供劳务收到的现金", 1, False),
        ("收到的税费返还", 2, False),
        ("收到其他与经营活动有关的现金", 3, False),
        ("经营活动现金流入小计", 4, True),
        ("购买商品、接受劳务支付的现金", 5, False),
        ("支付给职工以及为职工支付的现金", 6, False),
        ("支付的各项税费", 7, False),
        ("支付其他与经营活动有关的现金", 8, False),
        ("经营活动现金流出小计", 9, True),
        ("经营活动产生的现金流量净额", 10, True),
        # Investing activities
        ("二、投资活动产生的现金流量", None, True),
        ("收回投资收到的现金", 11, False),
        ("取得投资收益收到的现金", 12, False),
        ("处置固定资产、无形资产收到的现金净额", 13, False),
        ("投资活动现金流入小计", 14, True),
        ("购建固定资产、无形资产支付的现金", 15, False),
        ("投资支付的现金", 16, False),
        ("投资活动现金流出小计", 17, True),
        ("投资活动产生的现金流量净额", 18, True),
        # Financing activities
        ("三、筹资活动产生的现金流量", None, True),
        ("吸收投资收到的现金", 19, False),
        ("取得借款收到的现金", 20, False),
        ("筹资活动现金流入小计", 21, True),
        ("偿还债务支付的现金", 22, False),
        ("分配股利、利润或偿付利息支付的现金", 23, False),
        ("筹资活动现金流出小计", 24, True),
        ("筹资活动产生的现金流量净额", 25, True),
        # Summary
        ("四、汇率变动对现金的影响", 26, False),
        ("五、现金及现金等价物净增加额", 27, True),
        ("加：期初现金及现金等价物余额", 28, False),
        ("六、期末现金及现金等价物余额", 29, True),
    ]

    rev_base = random.uniform(30_000_000, 150_000_000) * gen.scn["revenue_mult"]
    row = 4
    for label, line_no, is_sect in cf_items:
        font = SECTION_FONT if is_sect else None
        style_cell(ws, row, 1, label, font, border=THIN_BORDER)
        if line_no:
            style_cell(ws, row, 2, line_no, border=THIN_BORDER, alignment=CENTER_ALIGN)

        if line_no and not is_sect:
            # Generate quarterly values
            quarters = []
            for q in range(4):
                seasonal = [0.8, 1.0, 1.1, 1.1][q]
                if "收到" in label or "收回" in label or "取得" in label or "吸收" in label:
                    val = round(rev_base * random.uniform(0.01, 0.3) * seasonal, 2)
                elif "支付" in label or "偿还" in label or "购" in label or "分配" in label:
                    val = round(-rev_base * random.uniform(0.01, 0.25) * seasonal, 2)
                else:
                    val = round(random.uniform(-500000, 500000), 2)
                quarters.append(val)
            for qi, qv in enumerate(quarters):
                style_cell(ws, row, 3 + qi, qv, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 7, round(sum(quarters), 2), font, border=THIN_BORDER, number_format="#,##0.00")
        elif "小计" in label or "净额" in label or "净增加" in label or "期末" in label:
            # Sum from sub-items above (simplified — just generate reasonable totals)
            quarters = [round(rev_base * random.uniform(-0.05, 0.15), 2) for _ in range(4)]
            for qi, qv in enumerate(quarters):
                style_cell(ws, row, 3 + qi, qv, font, border=THIN_BORDER, number_format="#,##0.00")
            style_cell(ws, row, 7, round(sum(quarters), 2), font, border=THIN_BORDER, number_format="#,##0.00")

        row += 1

    for c, w in [(1, 36), (2, 6), (3, 16), (4, 16), (5, 16), (6, 16), (7, 18)]:
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================
# 边界场景 sheets — 测试解析器极端情况
# ============================================================

def create_wide_sheet(wb, tpl, gen):
    """超宽数据表 — 120 列，测试 structure_detector 在宽表下的性能和正确性。"""
    ws = wb.create_sheet("超宽数据表")
    scale = gen.scn["revenue_mult"]
    n_cols = 120
    n_rows = 50

    # 3-level nested merge headers: 大区 → 季度 → 指标
    regions_h = ["华东", "华北", "华南", "西南", "华中"]
    quarters = ["Q1", "Q2", "Q3", "Q4"]
    metrics = ["收入", "成本", "毛利", "费用", "净利", "预算"]

    # Row 1: region headers (span = 4 quarters * 6 metrics = 24 each)
    col = 2
    for reg in regions_h:
        end_col = col + len(quarters) * len(metrics) - 1
        if end_col > n_cols:
            break
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        style_cell(ws, 1, col, reg, SECTION_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        col = end_col + 1

    # Row 2: quarter headers (span = 6 metrics each)
    col = 2
    for reg in regions_h:
        for q in quarters:
            end_col = col + len(metrics) - 1
            if end_col > n_cols:
                break
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=end_col)
            style_cell(ws, 2, col, q, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
            col = end_col + 1

    # Row 3: metric headers
    col = 2
    style_cell(ws, 3, 1, "产品", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    for _ in range(n_cols - 1):
        idx = (col - 2) % len(metrics)
        style_cell(ws, 3, col, metrics[idx], HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
        col += 1
        if col > n_cols:
            break

    # Data rows
    products = tpl["products"] + ["其他产品A", "其他产品B", "其他产品C"]
    for ri in range(n_rows):
        row = 4 + ri
        prod = products[ri % len(products)]
        style_cell(ws, row, 1, prod, border=THIN_BORDER)
        for ci in range(2, n_cols + 1):
            metric_idx = (ci - 2) % len(metrics)
            if metric_idx == 0:  # 收入
                val = round(random.uniform(100000, 5000000) * scale, 2)
            elif metric_idx == 1:  # 成本
                val = round(random.uniform(60000, 3000000) * scale, 2)
            elif metric_idx == 2:  # 毛利
                val = round(random.uniform(20000, 2000000) * scale, 2)
            elif metric_idx == 3:  # 费用
                val = round(random.uniform(10000, 500000) * scale, 2)
            elif metric_idx == 4:  # 净利
                val = round(random.uniform(-200000, 1500000) * scale, 2)
            else:  # 预算
                val = round(random.uniform(80000, 4000000) * scale, 2)
            style_cell(ws, row, ci, val, border=THIN_BORDER, number_format="#,##0.00")

    ws.column_dimensions["A"].width = 16
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10


def create_mixed_type_sheet(wb, tpl, gen):
    """混合类型数据 — 同一列中混合数字和文字，测试 pandas select_dtypes 的边界行为。"""
    ws = wb.create_sheet("混合类型测试")
    scale = gen.scn["revenue_mult"]

    headers = ["产品名称", "销售额", "利润率", "备注", "月份数据", "分类评分"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    products = tpl["products"]
    for ri in range(30):
        row = 2 + ri
        prod = products[ri % len(products)]
        style_cell(ws, row, 1, prod, border=THIN_BORDER)

        # Col 2: 销售额 — mostly numbers, but row 15 and 28 have text
        if ri in (14, 27):
            style_cell(ws, row, 2, "实际净利", border=THIN_BORDER)  # text in numeric column!
        else:
            style_cell(ws, row, 2, round(random.uniform(50000, 2000000) * scale, 2), border=THIN_BORDER, number_format="#,##0.00")

        # Col 3: 利润率 — mix of float percentages and text
        if ri in (5, 20):
            style_cell(ws, row, 3, "N/A", border=THIN_BORDER)
        elif ri == 10:
            style_cell(ws, row, 3, "—", border=THIN_BORDER)
        else:
            style_cell(ws, row, 3, round(random.uniform(0.02, 0.45), 4), border=THIN_BORDER, number_format="0.0%")

        # Col 4: 备注 — pure text
        notes = ["正常", "需关注", "超预期", "低于预期", "", None, "季节性波动", "新产品", "促销期"]
        style_cell(ws, row, 4, random.choice(notes), border=THIN_BORDER)

        # Col 5: 月份数据 — Chinese month format (1月, 2月...)
        month = (ri % 12) + 1
        style_cell(ws, row, 5, f"{month}月", border=THIN_BORDER, alignment=CENTER_ALIGN)

        # Col 6: 分类评分 — integers mixed with float and text
        if ri == 25:
            style_cell(ws, row, 6, "优秀", border=THIN_BORDER)
        else:
            style_cell(ws, row, 6, round(random.uniform(60, 100), 1), border=THIN_BORDER)

    for c, w in [(1, 16), (2, 14), (3, 10), (4, 14), (5, 10), (6, 10)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_empty_regions_sheet(wb, tpl, gen):
    """空值区域测试 — 模拟公式 data_only=True 返回 None 的场景 + 全空行/全零列。"""
    ws = wb.create_sheet("空值区域测试")
    scale = gen.scn["revenue_mult"]

    headers = ["项目", "1月", "2月", "3月", "4月", "5月", "6月",
               "7月", "8月", "9月", "10月", "11月", "12月", "合计", "全零列"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    items = ["营业收入", "营业成本", "毛利", "", "管理费用", "销售费用", "财务费用", "",
             "营业利润", "所得税", "净利润", "", "=SUM模拟(None)", "=VLOOKUP模拟(None)", "正常数据"]

    for ri, item in enumerate(items):
        row = 2 + ri
        style_cell(ws, row, 1, item if item else None, SECTION_FONT if "利" in item else None, border=THIN_BORDER)

        if item == "":
            # Completely empty row (separator)
            continue
        elif "模拟(None)" in item:
            # Simulate formula cells that return None under data_only=True
            for ci in range(2, 14):
                style_cell(ws, row, ci, None, border=THIN_BORDER)
            style_cell(ws, row, 14, None, SECTION_FONT, border=THIN_BORDER)
        else:
            for ci in range(2, 14):
                val = round(random.uniform(100000, 3000000) * scale, 2)
                if "成本" in item or "费用" in item or "所得税" in item:
                    val = -abs(val) * random.uniform(0.1, 0.8)
                style_cell(ws, row, ci, round(val, 2), border=THIN_BORDER, number_format="#,##0.00")
            # Col 14 (合计)
            style_cell(ws, row, 14, round(random.uniform(500000, 10000000) * scale, 2),
                       SECTION_FONT, border=THIN_BORDER, number_format="#,##0.00")

        # Col 15: 全零列 — all zeros
        style_cell(ws, row, 15, 0, border=THIN_BORDER, number_format="#,##0.00")

    for c, w in [(1, 22)] + [(i, 12) for i in range(2, 16)]:
        ws.column_dimensions[get_column_letter(c)].width = w


def create_cross_year_sheet(wb, tpl, gen):
    """跨年对比数据 — 2024 vs 2025，用于 YoY 比较功能测试。"""
    ws = wb.create_sheet("跨年对比(2024-2025)")
    scale = gen.scn["revenue_mult"]

    # 3-level header: 年份 → 季度 → 指标
    ws.merge_cells("A1:A3")
    style_cell(ws, 1, 1, "区域/产品", HEADER_FONT, HEADER_FILL,
               Alignment(horizontal="center", vertical="center"), THIN_BORDER)

    col = 2
    for year in [2024, 2025]:
        end_col = col + 7
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        style_cell(ws, 1, col, f"{year}年", SECTION_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

        for qi, q in enumerate(["上半年", "下半年"]):
            qstart = col + qi * 4
            ws.merge_cells(start_row=2, start_column=qstart, end_row=2, end_column=qstart + 3)
            style_cell(ws, 2, qstart, q, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

        for mi in range(8):
            metric = ["收入", "成本", "毛利", "毛利率"][mi % 4]
            style_cell(ws, 3, col + mi, metric, HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

        col = end_col + 1

    # YoY columns
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col + 1)
    style_cell(ws, 1, col, "同比变化", SECTION_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, col, "收入增长率", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)
    style_cell(ws, 3, col + 1, "利润增长率", HEADER_FONT, HEADER_FILL, CENTER_ALIGN, THIN_BORDER)

    # Data rows
    items = tpl["products"] + [r["name"] for r in tpl["regions"]]
    for ri, item in enumerate(items):
        row = 4 + ri
        style_cell(ws, row, 1, item, border=THIN_BORDER)

        rev_2024 = [round(random.uniform(500000, 5000000) * scale * 0.85, 2) for _ in range(8)]
        rev_2025 = [round(v * random.uniform(0.8, 1.3), 2) for v in rev_2024]

        for yi, year_data in enumerate([rev_2024, rev_2025]):
            base_col = 2 + yi * 8
            for mi in range(8):
                metric_type = mi % 4
                if metric_type == 0:  # 收入
                    val = year_data[mi]
                elif metric_type == 1:  # 成本
                    val = round(year_data[mi - 1] * random.uniform(0.55, 0.75), 2)
                elif metric_type == 2:  # 毛利
                    rev_val = year_data[mi - 2]
                    cost_val = round(rev_val * random.uniform(0.55, 0.75), 2)
                    val = round(rev_val - cost_val, 2)
                else:  # 毛利率
                    val = round(random.uniform(0.20, 0.45), 4)
                if metric_type == 3:
                    style_cell(ws, row, base_col + mi, val, border=THIN_BORDER, number_format="0.0%")
                else:
                    style_cell(ws, row, base_col + mi, val, border=THIN_BORDER, number_format="#,##0.00")

        # YoY changes
        total_2024 = sum(rev_2024[i] for i in range(0, 8, 4))
        total_2025 = sum(rev_2025[i] for i in range(0, 8, 4))
        yoy_rev = round((total_2025 / total_2024 - 1), 4) if total_2024 else 0
        yoy_profit = round(random.uniform(-0.3, 0.5), 4)
        style_cell(ws, row, col, yoy_rev, border=THIN_BORDER, number_format="+0.0%;-0.0%")
        style_cell(ws, row, col + 1, yoy_profit, border=THIN_BORDER, number_format="+0.0%;-0.0%")

    ws.column_dimensions["A"].width = 18
    for c in range(2, col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 12


def create_formula_sheet(wb, tpl, gen):
    """公式保护测试 — SUM/AVERAGE/IF 公式 + data_only=True 返回 None 的场景。

    当 openpyxl 以 data_only=True 打开但文件未在 Excel 中计算过时，
    公式单元格返回 None。此 sheet 模拟此场景。
    """
    ws = wb.create_sheet(title="公式测试")

    # Title
    style_cell(ws, 1, 1, "产品月度销售汇总（含公式）",
               font=TITLE_FONT, fill=HEADER_FILL)
    ws.merge_cells("A1:N1")

    # Headers row 2
    headers = ["产品"] + [f"{m+1}月" for m in range(12)] + ["合计"]
    for ci, h in enumerate(headers, 1):
        style_cell(ws, 2, ci, h, font=HEADER_FONT, fill=HEADER_FILL,
                   border=THIN_BORDER, alignment=CENTER_ALIGN)

    products = tpl["products"][:8]
    for ri, prod in enumerate(products, 3):
        style_cell(ws, ri, 1, prod, border=THIN_BORDER)
        for mi in range(12):
            val = round(random.uniform(50000, 500000), 2)
            style_cell(ws, ri, mi + 2, val, border=THIN_BORDER, number_format="#,##0.00")
        # SUM formula in column 14
        ws.cell(row=ri, column=14).value = f"=SUM(B{ri}:M{ri})"
        ws.cell(row=ri, column=14).border = THIN_BORDER
        ws.cell(row=ri, column=14).number_format = "#,##0.00"

    # Summary rows with formulas
    sum_row = 3 + len(products)
    style_cell(ws, sum_row, 1, "月度合计", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 15):
        col_letter = get_column_letter(ci)
        ws.cell(row=sum_row, column=ci).value = f"=SUM({col_letter}3:{col_letter}{sum_row - 1})"
        ws.cell(row=sum_row, column=ci).border = THIN_BORDER
        ws.cell(row=sum_row, column=ci).number_format = "#,##0.00"

    avg_row = sum_row + 1
    style_cell(ws, avg_row, 1, "月度平均", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 15):
        col_letter = get_column_letter(ci)
        ws.cell(row=avg_row, column=ci).value = f"=AVERAGE({col_letter}3:{col_letter}{sum_row - 1})"
        ws.cell(row=avg_row, column=ci).border = THIN_BORDER
        ws.cell(row=avg_row, column=ci).number_format = "#,##0.00"

    # Conditional row (IF formula)
    cond_row = avg_row + 1
    style_cell(ws, cond_row, 1, "达标状态", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 14):
        col_letter = get_column_letter(ci)
        ws.cell(row=cond_row, column=ci).value = f'=IF({col_letter}{sum_row}>1000000,"达标","未达标")'
        ws.cell(row=cond_row, column=ci).border = THIN_BORDER

    # Percentage formula row
    pct_row = cond_row + 1
    style_cell(ws, pct_row, 1, "占年度比例", font=HEADER_FONT, border=THIN_BORDER)
    for ci in range(2, 14):
        col_letter = get_column_letter(ci)
        ws.cell(row=pct_row, column=ci).value = f"={col_letter}{sum_row}/N{sum_row}"
        ws.cell(row=pct_row, column=ci).border = THIN_BORDER
        ws.cell(row=pct_row, column=ci).number_format = "0.0%"


def create_numeric_colname_sheet(wb, tpl, gen):
    """纯数字列名测试 — 列标题为 "2024"/"2025" 等数字，
    测试 30% 阈值判断是否误把表头当成数据行。
    """
    ws = wb.create_sheet(title="纯数字列名")

    # Row 1: years as column headers (integers, not strings)
    years = [2020, 2021, 2022, 2023, 2024, 2025]
    style_cell(ws, 1, 1, "指标", font=HEADER_FONT, fill=HEADER_FILL,
               border=THIN_BORDER, alignment=CENTER_ALIGN)
    for ci, yr in enumerate(years, 2):
        # Write as integer — this triggers the "30% numeric threshold" ambiguity
        style_cell(ws, 1, ci, yr, font=HEADER_FONT, fill=HEADER_FILL,
                   border=THIN_BORDER, alignment=CENTER_ALIGN)

    # Additional numeric header: Q1-Q4
    q_headers = ["Q1", "Q2", "Q3", "Q4"]
    for ci, q in enumerate(q_headers, len(years) + 2):
        style_cell(ws, 1, ci, q, font=HEADER_FONT, fill=HEADER_FILL,
                   border=THIN_BORDER, alignment=CENTER_ALIGN)

    total_cols = 1 + len(years) + len(q_headers)  # 11 columns

    # Metrics as row labels — financial indicators with purely numeric values
    metrics = [
        ("营业收入(万元)", [round(random.uniform(8000, 15000), 1) for _ in years]),
        ("营业成本(万元)", [round(random.uniform(5000, 10000), 1) for _ in years]),
        ("毛利(万元)", [round(random.uniform(2000, 5000), 1) for _ in years]),
        ("毛利率", [round(random.uniform(0.20, 0.45), 4) for _ in years]),
        ("净利润(万元)", [round(random.uniform(300, 2000), 1) for _ in years]),
        ("净利率", [round(random.uniform(0.03, 0.15), 4) for _ in years]),
        ("员工数", [random.randint(200, 800) for _ in years]),
        ("人均产值(万元)", [round(random.uniform(15, 40), 1) for _ in years]),
        ("研发费用(万元)", [round(random.uniform(500, 2000), 1) for _ in years]),
        ("研发占比", [round(random.uniform(0.03, 0.12), 4) for _ in years]),
        ("资产总额(万元)", [round(random.uniform(10000, 50000), 1) for _ in years]),
        ("负债率", [round(random.uniform(0.30, 0.70), 4) for _ in years]),
    ]

    pct_rows = {"毛利率", "净利率", "研发占比", "负债率"}

    for ri, (metric_name, values) in enumerate(metrics, 2):
        is_pct = metric_name in pct_rows
        style_cell(ws, ri, 1, metric_name, border=THIN_BORDER)
        for ci, val in enumerate(values, 2):
            fmt = "0.0%" if is_pct else "#,##0.0"
            style_cell(ws, ri, ci, val, border=THIN_BORDER, number_format=fmt)

        # Q1-Q4 for latest year (2025)
        latest_val = values[-1]
        for qi in range(4):
            q_val = round(latest_val * random.uniform(0.2, 0.35), 2)
            ci = len(years) + 2 + qi
            fmt = "0.0%" if is_pct else "#,##0.0"
            style_cell(ws, ri, ci, q_val, border=THIN_BORDER, number_format=fmt)


# ============================================================
# 主函数
# ============================================================

def generate(template_name="food", scenario="normal", seed=42, output_dir=None):
    """生成测试 Excel 文件。"""
    random.seed(seed)
    tpl = TEMPLATES[template_name]
    gen = FinancialDataGenerator(tpl, scenario)

    wb = openpyxl.Workbook()

    # Collect all regions (flat list for profit sheets)
    all_regions = []
    for r in tpl["regions"]:
        all_regions.append(r)
    for sr in tpl["sub_regions"]:
        all_regions.append(sr)
        for child in sr.get("children", []):
            all_regions.append(child)

    # Generate data for each region
    all_region_data = {}
    region_names_for_summary = []

    for region in all_regions:
        scale = {"分部": 1.0, "区域": 0.8, "省区": 0.4, "直营": 1.2, "加盟": 0.7, "渠道": 0.6, "店铺": 0.3}.get(region["type"], 0.5)
        rev_by_prod = gen.gen_revenue_by_product(scale)
        cost_by_prod = gen.gen_cost_by_product(rev_by_prod)
        total_rev = [sum(rev_by_prod[p][m] for p in gen.products) for m in range(12)]
        budget_rev = gen.gen_budget(total_rev)
        net_profit_est = [round(total_rev[m] * random.uniform(0.02, 0.12), 1) for m in range(12)]
        budget_net = gen.gen_budget(net_profit_est, accuracy=0.3)

        all_region_data[region["name"]] = {
            "revenue_by_product": rev_by_prod,
            "cost_by_product": cost_by_prod,
            "revenue_total": total_rev,
            "budget_revenue": budget_rev,
            "net_profit": net_profit_est,
            "budget_net": budget_net,
            "type": region["type"],
        }

    # Add center director expenses
    director_exp = [round(random.uniform(10000, 50000), 1) for _ in range(12)]
    all_region_data[f"{tpl['center'][:3]}总监费用"] = {
        "revenue_total": director_exp,
        "budget_revenue": [0] * 12,
        "net_profit": [-e for e in director_exp],
        "budget_net": gen.gen_budget([-e for e in director_exp]),
        "type": "expense",
    }

    # Aggregate center-level data (sum of all top-level regions)
    top_regions = [r["name"] for r in tpl["regions"]] + [sr["name"] for sr in tpl["sub_regions"]]
    center_rev_by_prod = {}
    for prod in gen.products:
        center_rev_by_prod[prod] = [
            sum(all_region_data[rn]["revenue_by_product"][prod][m] for rn in top_regions)
            for m in range(12)
        ]
    center_cost_by_prod = {}
    for prod in gen.products:
        center_cost_by_prod[prod] = [
            sum(all_region_data[rn]["cost_by_product"][prod][m] for rn in top_regions)
            for m in range(12)
        ]
    center_total_rev = [sum(center_rev_by_prod[p][m] for p in gen.products) for m in range(12)]
    center_budget_rev = gen.gen_budget(center_total_rev)
    center_net_est = [round(center_total_rev[m] * random.uniform(0.02, 0.10), 1) for m in range(12)]
    center_budget_net = gen.gen_budget(center_net_est, accuracy=0.3)

    all_region_data[tpl["center"]] = {
        "revenue_by_product": center_rev_by_prod,
        "cost_by_product": center_cost_by_prod,
        "revenue_total": center_total_rev,
        "budget_revenue": center_budget_rev,
        "net_profit": center_net_est,
        "budget_net": center_budget_net,
        "type": "center",
    }
    # Also add "中心" key for internal center sheet matching
    all_region_data["中心"] = dict(all_region_data[tpl["center"]])

    # Build summary region list (top-level only)
    for r in tpl["regions"]:
        region_names_for_summary.append(r["name"])
    for sr in tpl["sub_regions"]:
        region_names_for_summary.append(sr["name"])
    region_names_for_summary.append(f"{tpl['center'][:3]}总监费用")

    summary_data = {name: all_region_data[name] for name in region_names_for_summary}

    # Sheet order (mimics Test.xlsx)
    sheet_names = ["索引", "收入及净利简表"]

    # Center-level profit sheet
    center_sheet_name = f"{gen.year}年{tpl['center']}利润表"
    sheet_names.append(center_sheet_name)

    # Internal center profit sheet
    center_internal_name = f"{gen.year}年中心利润表"
    sheet_names.append(center_internal_name)

    # Region profit sheets
    for r in tpl["regions"]:
        sheet_names.append(f"{gen.year}年{r['name']}利润表")
    for sr in tpl["sub_regions"]:
        sheet_names.append(f"{gen.year}年{sr['name']}利润表")
        for child in sr.get("children", []):
            sheet_names.append(f"{gen.year}年{child['name']}利润表")

    # Collect profit sheet names separately
    profit_sheet_names = list(sheet_names[2:])  # everything after index + summary

    sheet_names.append("24年返利明细")

    # New complex sheets
    extra_sheets = ["资产负债表", "销售明细", "应收账款账龄", "库存台账", "月度经营分析", "费用预算执行", "现金流量表"]
    sheet_names.extend(extra_sheets)

    # Edge-case / stress sheets
    edge_sheets = ["超宽数据表", "混合类型测试", "空值区域测试", "跨年对比(2024-2025)", "公式测试", "纯数字列名"]
    sheet_names.extend(edge_sheets)

    # 1. Index
    create_index_sheet(wb, tpl, sheet_names)

    # 2. Summary
    create_summary_sheet(wb, tpl, gen, summary_data)

    # 3. Profit sheets (only the profit table names, not rebate or extras)
    for sname in profit_sheet_names:
        ws = wb.create_sheet(sname)
        # Find region data
        for rname, rdata in all_region_data.items():
            if rname in sname:
                if "中心利润表" in sname and "销售" not in sname:
                    net = create_profit_sheet(ws, tpl, gen, rname, rdata, "center_internal")
                elif tpl["center"] in sname:
                    net = create_profit_sheet(ws, tpl, gen, rname, rdata, "center")
                else:
                    net = create_profit_sheet(ws, tpl, gen, rname, rdata)
                rdata["net_profit"] = net
                break
        else:
            # Center sheets: use aggregated data
            if "中心" in sname:
                # Create center data by aggregating
                center_data = {
                    "revenue_by_product": gen.gen_revenue_by_product(0.1),
                    "cost_by_product": {},
                    "revenue_total": director_exp,
                    "budget_revenue": [0] * 12,
                    "net_profit": [-e for e in director_exp],
                    "budget_net": gen.gen_budget([-e for e in director_exp]),
                }
                center_data["cost_by_product"] = gen.gen_cost_by_product(center_data["revenue_by_product"])
                st = "center" if tpl["center"] in sname else "center_internal"
                create_profit_sheet(ws, tpl, gen, "中心", center_data, st)

    # 4. Rebate detail
    create_rebate_sheet(wb, gen, tpl["regions"] + tpl["sub_regions"])

    # 5. New complex sheets
    create_balance_sheet(wb, tpl, gen)
    create_sales_detail(wb, tpl, gen)
    create_ar_aging_sheet(wb, tpl, gen)
    create_inventory_sheet(wb, tpl, gen)
    create_monthly_kpi_sheet(wb, tpl, gen, all_region_data)
    create_budget_comparison_sheet(wb, tpl, gen)
    create_cashflow_sheet(wb, tpl, gen)

    # 6. Edge-case sheets
    create_wide_sheet(wb, tpl, gen)
    create_mixed_type_sheet(wb, tpl, gen)
    create_empty_regions_sheet(wb, tpl, gen)
    create_cross_year_sheet(wb, tpl, gen)
    create_formula_sheet(wb, tpl, gen)
    create_numeric_colname_sheet(wb, tpl, gen)

    # Save
    if output_dir is None:
        output_dir = Path(__file__).parent / "test-data"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"Test-mock-{template_name}-{scenario}-s{seed}.xlsx"
    filepath = output_dir / filename
    wb.save(str(filepath))
    print(f"Generated: {filepath}")
    print(f"  Template: {template_name} ({tpl['company']})")
    print(f"  Scenario: {scenario}")
    print(f"  Seed: {seed}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"    {sn}: {ws.max_row}r x {ws.max_column}c, merges={len(list(ws.merged_cells.ranges))}")

    return filepath


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SmartBI E2E Test Excel Generator")
    parser.add_argument("--template", choices=list(TEMPLATES.keys()), default="food")
    parser.add_argument("--scenario", choices=list(SCENARIOS.keys()), default="normal")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--all", action="store_true", help="Generate all template × scenario combos")
    args = parser.parse_args()

    if args.all:
        for tpl_name in TEMPLATES:
            for scn_name in SCENARIOS:
                generate(tpl_name, scn_name, args.seed, args.output)
        print(f"\nGenerated {len(TEMPLATES) * len(SCENARIOS)} files.")
    else:
        generate(args.template, args.scenario, args.seed, args.output)
