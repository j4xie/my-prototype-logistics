from __future__ import annotations
"""
Data Exporter Service

Multi-format export for parsed Excel data.
Supports JSON (internal), Markdown (LLM), and CSV (export/debug) formats.

Based on research showing Markdown key:value format has ~60.7% accuracy for LLMs,
while CSV has only ~44.3%. This module provides the optimal format for each use case.

Architecture:
    Excel File → StructuredData (internal) → JSON / Markdown / CSV
"""
import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

import openpyxl
import pandas as pd
import numpy as np

from services.structure_detector import StructureDetector, StructureDetectionResult, MergedCellInfo
from services.context_extractor import ContextExtractor, ContextInfo

logger = logging.getLogger(__name__)


@dataclass
class ColumnDef:
    """Column definition with metadata"""
    name: str                      # Flattened column name (e.g., "01月_预算数")
    original_name: str             # Original column name from Excel
    data_type: str = "text"        # Data type: numeric, date, text, percentage
    sub_type: Optional[str] = None # Sub-type: amount, rate, quantity, category
    index: int = 0                 # Column index (0-based)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "original_name": self.original_name,
            "data_type": self.data_type,
            "sub_type": self.sub_type,
            "index": self.index
        }


@dataclass
class StructuredData:
    """
    Unified internal data representation.

    This is the central data structure that all export formats derive from.
    """
    # Metadata (Layer 1 of Three-Layer Model)
    metadata: Dict[str, Any] = field(default_factory=dict)

    # Column definitions
    columns: List[ColumnDef] = field(default_factory=list)

    # Data rows (each row is a dict with column names as keys)
    rows: List[Dict[str, Any]] = field(default_factory=list)

    # Context (Layer 3 of Three-Layer Model)
    context: Optional[ContextInfo] = None

    # Processing info
    row_count: int = 0
    column_count: int = 0
    source_sheet: str = ""
    processing_notes: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization"""
        return {
            "metadata": self.metadata,
            "columns": [c.to_dict() for c in self.columns],
            "rows": self.rows,
            "context": self.context.to_dict() if self.context else None,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "source_sheet": self.source_sheet,
            "processing_notes": self.processing_notes
        }

    def to_dataframe(self) -> pd.DataFrame:
        """Convert to pandas DataFrame"""
        if not self.rows:
            return pd.DataFrame()
        return pd.DataFrame(self.rows)


class HeaderFlattener:
    """
    Multi-level header flattening.

    Uses rule-based approach to combine hierarchical Excel headers
    into flat column names suitable for data processing.

    Example:
        Row 1: |  2025年1月  |  2025年2月  |
        Row 2: | 预算 | 实际 | 预算 | 实际 |

        Result: ["01月_预算", "01月_实际", "02月_预算", "02月_实际"]
    """

    # Patterns to skip (title rows, unit rows, metadata rows, etc.)
    SKIP_PATTERNS = [
        re.compile(r'利润表|资产负债表|现金流量表|报表|汇总表|明细表|统计表|分析表|对比表', re.IGNORECASE),
        re.compile(r'单位[:：]|Unit[:：]', re.IGNORECASE),
        re.compile(r'编制单位|编制日期', re.IGNORECASE),
        re.compile(r'统计期间|报告期间|数据期间', re.IGNORECASE),
    ]

    # Month normalization patterns
    MONTH_PATTERNS = [
        (re.compile(r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})?'), r'\2月'),  # 2025-01-01 → 01月
        (re.compile(r'(\d{1,2})月'), r'\1月'),  # Already month format
    ]

    def __init__(self):
        pass

    def flatten(
        self,
        header_rows: List[List[Any]],
        merge_map: Optional[Dict[tuple, str]] = None,
        max_parts: int = 3
    ) -> List[ColumnDef]:
        """
        Flatten multi-level headers into single-level column definitions.

        Args:
            header_rows: List of header row values
            merge_map: Optional mapping of (row, col) -> merged cell value
            max_parts: Maximum number of parts to combine in column name

        Returns:
            List of ColumnDef with flattened names
        """
        if not header_rows:
            return []

        merge_map = merge_map or {}
        num_cols = max(len(row) for row in header_rows)

        # Filter out title/unit/date rows, keep column header rows
        meaningful_rows = []
        for row_idx, row in enumerate(header_rows):
            row_text = " ".join(str(v) for v in row if v)

            # Skip if matches skip patterns (title rows like "利润表")
            if any(p.search(row_text) for p in self.SKIP_PATTERNS):
                logger.debug(f"Skipping row {row_idx}: matches skip pattern")
                continue

            # Skip if mostly empty (less than 20% non-empty)
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if len(non_empty) < num_cols * 0.2:
                logger.debug(f"Skipping row {row_idx}: too few non-empty cells ({len(non_empty)}/{num_cols})")
                continue

            # Skip if mostly numeric (likely data row leaked into headers)
            numeric_count = sum(1 for v in row if self._is_numeric(v))
            if numeric_count > len(non_empty) * 0.5 and len(non_empty) > 5:
                logger.debug(f"Skipping row {row_idx}: too many numeric values")
                continue

            meaningful_rows.append((row_idx, row))

        if not meaningful_rows:
            # Fallback: use the last row as column names
            meaningful_rows = [(len(header_rows) - 1, header_rows[-1])]

        logger.debug(f"Meaningful header rows: {[r[0] for r in meaningful_rows]}")

        # Build column definitions with proper multi-level merging
        columns = []
        used_names = {}  # Track name usage for deduplication

        for col_idx in range(num_cols):
            parts = []
            seen_normalized = set()

            for row_idx, row in meaningful_rows:
                value = None

                # Check merge map first (1-indexed for openpyxl)
                if merge_map and (row_idx + 1, col_idx + 1) in merge_map:
                    value = merge_map[(row_idx + 1, col_idx + 1)]
                elif col_idx < len(row):
                    value = row[col_idx]

                if value is None:
                    continue

                value_str = str(value).strip()
                if not value_str:
                    continue

                # Skip pure numeric values
                if self._is_numeric(value_str):
                    continue

                # Normalize month/date patterns
                normalized = self._normalize_month(value_str)

                # Skip datetime timestamps (keep only date part)
                if "00:00:00" in normalized:
                    date_part = normalized.split(" ")[0]
                    normalized = self._normalize_month(date_part)

                # Skip if we already have this normalized value
                if normalized in seen_normalized:
                    continue

                parts.append(normalized)
                seen_normalized.add(normalized)

            # Generate final column name
            if parts:
                # Limit to max_parts (keep last parts which are more specific)
                if len(parts) > max_parts:
                    parts = parts[-max_parts:]
                name = "_".join(parts)
            else:
                name = f"Column_{col_idx + 1}"

            # Handle duplicate names by adding suffix
            base_name = name
            if name in used_names:
                used_names[name] += 1
                # Try to find a distinguishing suffix from the row data
                # For budget/actual columns, check the last meaningful row
                suffix = None
                for row_idx, row in reversed(meaningful_rows):
                    if col_idx < len(row) and row[col_idx]:
                        cell_value = str(row[col_idx]).strip()
                        if cell_value and not self._is_numeric(cell_value):
                            # Check for common budget/actual keywords
                            if any(kw in cell_value for kw in ['预算', '本月实际', '实际', '同期']):
                                suffix = cell_value
                                break

                if suffix:
                    name = f"{base_name}_{suffix}"
                else:
                    name = f"{base_name}_{used_names[base_name]}"
            else:
                used_names[name] = 1

            # Get original name from the last meaningful row
            original_name = name
            if meaningful_rows:
                last_row = meaningful_rows[-1][1]
                if col_idx < len(last_row) and last_row[col_idx]:
                    original_name = str(last_row[col_idx])

            columns.append(ColumnDef(
                name=name,
                original_name=original_name,
                index=col_idx
            ))

        return columns

    def _is_numeric(self, value: Any) -> bool:
        """Check if value is numeric"""
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        try:
            cleaned = str(value).replace(',', '').replace('¥', '').replace('%', '').strip()
            if not cleaned:
                return False
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False

    def _normalize_month(self, value: str) -> str:
        """Normalize date/month values to standard format"""
        for pattern, replacement in self.MONTH_PATTERNS:
            match = pattern.search(value)
            if match:
                # Extract month number and format
                try:
                    if '年' in value or '-' in value or '/' in value:
                        # Extract month from date pattern
                        parts = re.split(r'[-/年月日]', value)
                        for part in parts:
                            if part.isdigit() and 1 <= int(part) <= 12:
                                return f"{int(part):02d}月"
                except (ValueError, IndexError):
                    pass
        return value


class DataExporter:
    """
    Multi-format data exporter.

    Exports StructuredData to:
    - JSON: Internal processing, structured storage
    - Markdown: LLM prompts (highest accuracy ~60.7%)
    - CSV: Export, debugging, Excel-compatible
    """

    SMARTBI_VERSION = "1.0"

    def __init__(self):
        self._structure_detector = StructureDetector()
        self._context_extractor = ContextExtractor()
        self._header_flattener = HeaderFlattener()

    async def from_excel(
        self,
        file_bytes: bytes,
        sheet_index: int = 0,
        max_rows: Optional[int] = None,
        skip_empty_rows: bool = True
    ) -> StructuredData:
        """
        Parse Excel file into StructuredData.

        Args:
            file_bytes: Raw Excel file bytes
            sheet_index: Sheet index to parse (0-based)
            max_rows: Maximum rows to extract (None for all)
            skip_empty_rows: Skip rows with all empty values

        Returns:
            StructuredData with parsed content
        """
        result = StructuredData()

        try:
            # Load workbook
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            if sheet_index >= len(wb.sheetnames):
                sheet_index = 0

            ws = wb[wb.sheetnames[sheet_index]]
            result.source_sheet = ws.title

            # Detect structure
            structure = await self._structure_detector.detect(
                file_bytes,
                sheet_index=sheet_index,
                max_header_rows=10
            )

            if not structure.success:
                result.processing_notes.append(f"Structure detection failed: {structure.error}")
                wb.close()
                return result

            # Build metadata
            result.metadata = {
                "title": None,
                "unit": None,
                "period": None,
                "company": None,
                "sheet": ws.title,
                "header_rows": structure.header_row_count,
                "data_start_row": structure.data_start_row
            }

            # Extract metadata from context
            context_info = self._context_extractor.extract(
                ws=ws,
                data_end_row=ws.max_row or 1,
                total_rows=ws.max_row or 1,
                header_rows=structure.header_row_count
            )

            if context_info:
                result.context = context_info
                if context_info.title:
                    result.metadata["title"] = context_info.title
                if context_info.unit:
                    result.metadata["unit"] = context_info.unit
                if context_info.period:
                    result.metadata["period"] = context_info.period
                if context_info.compiled_by:
                    result.metadata["company"] = context_info.compiled_by

            # Build merge map for header flattening
            merge_map = {}
            for m in structure.merged_cells:
                cell_value = ws.cell(row=m.min_row, column=m.min_col).value
                value_str = str(cell_value) if cell_value else ""
                for r in range(m.min_row, m.max_row + 1):
                    for c in range(m.min_col, m.max_col + 1):
                        merge_map[(r, c)] = value_str

            # Extract header rows for flattening
            header_rows = []
            header_end_row = structure.header_row_count

            # Check if the "data start row" is actually a sub-header row
            # This happens when merged cells create a multi-level header structure
            potential_subheader_row = structure.data_start_row
            if potential_subheader_row <= (ws.max_row or 1):
                subheader_values = []
                numeric_count = 0
                non_empty_count = 0

                for col_idx in range(1, min((ws.max_column or 1) + 1, 30)):  # Check first 30 cols
                    cell = ws.cell(row=potential_subheader_row, column=col_idx)
                    val = cell.value
                    if val is not None and str(val).strip():
                        non_empty_count += 1
                        subheader_values.append(str(val))
                        # Check if numeric
                        try:
                            cleaned = str(val).replace(',', '').replace('¥', '').strip()
                            if cleaned:
                                float(cleaned)
                                numeric_count += 1
                        except (ValueError, TypeError):
                            pass

                # If row has mostly text values with sub-header keywords, it's a sub-header row
                is_subheader = False
                if non_empty_count > 0 and numeric_count < non_empty_count * 0.3:
                    # Check for common sub-header patterns
                    subheader_keywords = ['预算', '实际', '本月', '同期', '合计', '累计',
                                          'Budget', 'Actual', 'YTD', 'MTD']
                    text_lower = ' '.join(subheader_values).lower()
                    if any(kw.lower() in text_lower for kw in subheader_keywords):
                        is_subheader = True
                        header_end_row = potential_subheader_row
                        logger.debug(f"Row {potential_subheader_row} detected as sub-header row")

            for row_idx in range(1, header_end_row + 1):
                row_values = []
                for col_idx in range(1, (ws.max_column or 1) + 1):
                    if (row_idx, col_idx) in merge_map:
                        row_values.append(merge_map[(row_idx, col_idx)])
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        row_values.append(cell.value)
                header_rows.append(row_values)

            # Update data start row if we included an extra header row
            actual_data_start = header_end_row + 1

            # Flatten headers
            result.columns = self._header_flattener.flatten(header_rows, merge_map)

            # Detect data types from first few data rows
            data_start = actual_data_start  # 1-indexed for openpyxl
            sample_rows = min(10, (ws.max_row or data_start) - data_start + 1)

            for col_idx, col_def in enumerate(result.columns):
                values = []
                for row_idx in range(data_start, data_start + sample_rows):
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    if cell.value is not None:
                        values.append(cell.value)

                col_def.data_type = self._detect_data_type(values)
                col_def.sub_type = self._detect_sub_type(col_def.name, values)

            # Extract data rows
            rows = []
            max_row = ws.max_row or data_start
            if max_rows:
                max_row = min(max_row, data_start + max_rows - 1)

            for row_idx in range(data_start, max_row + 1):
                row_data = {}
                has_data = False

                for col_idx, col_def in enumerate(result.columns):
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    value = cell.value

                    if value is not None:
                        has_data = True
                        # Clean value based on data type
                        value = self._clean_value(value, col_def.data_type)

                    row_data[col_def.name] = value

                if not has_data and skip_empty_rows:
                    continue

                rows.append(row_data)

            result.rows = rows
            result.row_count = len(rows)
            result.column_count = len(result.columns)

            result.metadata["row_count"] = result.row_count
            result.metadata["column_count"] = result.column_count

            wb.close()

            result.processing_notes.append(
                f"Parsed {result.row_count} rows, {result.column_count} columns from sheet '{ws.title}'"
            )

        except Exception as e:
            logger.error(f"Excel parsing failed: {e}", exc_info=True)
            result.processing_notes.append(f"Error: {str(e)}")

        return result

    def to_json(
        self,
        data: StructuredData,
        include_metadata: bool = True,
        indent: int = 2
    ) -> str:
        """
        Export to JSON format.

        Best for: Internal processing, storage, API responses.

        Args:
            data: StructuredData to export
            include_metadata: Include metadata in output
            indent: JSON indentation level

        Returns:
            JSON string
        """
        output = {}

        if include_metadata:
            output["metadata"] = data.metadata
            output["columns"] = [c.to_dict() for c in data.columns]

        output["rows"] = data.rows
        output["row_count"] = data.row_count
        output["column_count"] = data.column_count

        if data.context and include_metadata:
            output["context"] = data.context.to_dict()

        return json.dumps(output, ensure_ascii=False, indent=indent, default=str)

    def to_markdown(
        self,
        data: StructuredData,
        max_rows: int = 50,
        include_metadata: bool = True,
        format_numbers: bool = True
    ) -> str:
        """
        Export to Markdown format.

        Best for: LLM prompts (highest accuracy ~60.7%).

        Args:
            data: StructuredData to export
            max_rows: Maximum rows to include (for LLM context limits)
            include_metadata: Include metadata header
            format_numbers: Format numbers with thousands separator

        Returns:
            Markdown string
        """
        lines = []

        # Metadata header
        if include_metadata and data.metadata:
            title = data.metadata.get("title", "Data Table")
            if title:
                lines.append(f"## {title}")
                lines.append("")

            meta_parts = []
            if data.metadata.get("unit"):
                meta_parts.append(f"单位：{data.metadata['unit']}")
            if data.metadata.get("period"):
                meta_parts.append(f"期间：{data.metadata['period']}")
            if data.metadata.get("company"):
                meta_parts.append(f"编制单位：{data.metadata['company']}")

            if meta_parts:
                lines.append(" | ".join(meta_parts))
                lines.append("")

        # Table header
        col_names = [c.name for c in data.columns]
        lines.append("| " + " | ".join(col_names) + " |")
        lines.append("|" + "|".join(["---" for _ in col_names]) + "|")

        # Table rows
        display_rows = data.rows[:max_rows]
        for row in display_rows:
            values = []
            for col in data.columns:
                value = row.get(col.name)
                if value is None:
                    values.append("")
                elif format_numbers and isinstance(value, (int, float)):
                    if col.data_type == "percentage":
                        values.append(f"{value * 100:.2f}%")
                    else:
                        values.append(f"{value:,.2f}" if isinstance(value, float) else f"{value:,}")
                else:
                    values.append(str(value))
            lines.append("| " + " | ".join(values) + " |")

        # Row count note
        if len(data.rows) > max_rows:
            lines.append("")
            lines.append(f"*（显示前 {max_rows} 行，共 {len(data.rows)} 行）*")

        # Context notes
        if include_metadata and data.context and data.context.has_content():
            lines.append("")
            lines.append("### 备注")
            for note in data.context.notes:
                lines.append(f"- {note}")
            for explanation in data.context.explanations:
                lines.append(f"- {explanation}")

        return "\n".join(lines)

    def to_csv(
        self,
        data: StructuredData,
        include_header_comments: bool = True,
        encoding: str = "utf-8-sig"
    ) -> str:
        """
        Export to CSV format.

        Best for: Export, debugging, Excel compatibility.

        Args:
            data: StructuredData to export
            include_header_comments: Include metadata as header comments
            encoding: Output encoding (utf-8-sig for Excel compatibility)

        Returns:
            CSV string
        """
        output = io.StringIO()

        # Header comments (metadata)
        if include_header_comments:
            output.write(f"# smartbi_version: {self.SMARTBI_VERSION}\n")

            if data.metadata:
                if data.metadata.get("title"):
                    output.write(f"# title: {data.metadata['title']}\n")
                if data.metadata.get("unit"):
                    output.write(f"# unit: {data.metadata['unit']}\n")
                if data.metadata.get("company"):
                    output.write(f"# company: {data.metadata['company']}\n")
                if data.metadata.get("period"):
                    output.write(f"# period: {data.metadata['period']}\n")
                if data.metadata.get("sheet"):
                    output.write(f"# sheet: {data.metadata['sheet']}\n")

            output.write(f"# row_count: {data.row_count}\n")
            output.write(f"# column_count: {data.column_count}\n")

        # CSV content
        col_names = [c.name for c in data.columns]
        writer = csv.DictWriter(output, fieldnames=col_names, extrasaction='ignore')
        writer.writeheader()

        for row in data.rows:
            # Convert values to strings, handling None
            clean_row = {}
            for col_name in col_names:
                value = row.get(col_name)
                if value is None:
                    clean_row[col_name] = ""
                elif isinstance(value, float):
                    # Keep full precision for numbers
                    clean_row[col_name] = value
                else:
                    clean_row[col_name] = value
            writer.writerow(clean_row)

        return output.getvalue()

    def _detect_data_type(self, values: List[Any]) -> str:
        """Detect column data type from sample values"""
        if not values:
            return "text"

        numeric_count = 0
        percentage_count = 0
        date_count = 0

        for v in values:
            if v is None:
                continue

            v_str = str(v)

            # Check percentage
            if '%' in v_str:
                percentage_count += 1
                continue

            # Check date
            if isinstance(v, datetime):
                date_count += 1
                continue

            # Check numeric
            try:
                cleaned = v_str.replace(',', '').replace('¥', '').replace('$', '')
                float(cleaned)
                numeric_count += 1
            except (ValueError, TypeError):
                pass

        total = len([v for v in values if v is not None])
        if total == 0:
            return "text"

        if percentage_count > total * 0.5:
            return "percentage"
        if date_count > total * 0.5:
            return "date"
        if numeric_count > total * 0.5:
            return "numeric"

        return "text"

    def _detect_sub_type(self, col_name: str, values: List[Any]) -> Optional[str]:
        """Detect column sub-type from name and values"""
        name_lower = col_name.lower()

        # Amount indicators
        if any(kw in name_lower for kw in ['金额', '收入', '支出', '成本', 'amount', 'revenue', 'cost']):
            return "amount"

        # Rate indicators
        if any(kw in name_lower for kw in ['率', '比例', 'rate', 'ratio', '%']):
            return "rate"

        # Quantity indicators
        if any(kw in name_lower for kw in ['数量', '件数', 'count', 'quantity', 'num']):
            return "quantity"

        # Budget indicators
        if any(kw in name_lower for kw in ['预算', 'budget']):
            return "budget"

        # Actual indicators
        if any(kw in name_lower for kw in ['实际', 'actual', '本月', '本期']):
            return "actual"

        return None

    def _clean_value(self, value: Any, data_type: str) -> Any:
        """Clean value based on data type"""
        if value is None:
            return None

        if data_type == "numeric":
            if isinstance(value, (int, float)):
                return float(value)
            try:
                cleaned = str(value).replace(',', '').replace('¥', '').replace('$', '').strip()
                if not cleaned or cleaned.lower() in ('nan', 'none', '-', '—'):
                    return None
                return float(cleaned)
            except (ValueError, TypeError):
                # Keep text value if it can't be converted to number
                # This preserves notes/remarks in numeric columns
                text_val = str(value).strip()
                if text_val:
                    return text_val
                return None

        elif data_type == "percentage":
            if isinstance(value, (int, float)):
                # Assume already decimal if between -1 and 1
                if -1 <= value <= 1:
                    return float(value)
                return float(value) / 100

            try:
                cleaned = str(value).replace('%', '').replace(',', '').strip()
                if not cleaned:
                    return None
                num = float(cleaned)
                # Convert to decimal if > 1
                if num > 1 or num < -1:
                    return num / 100
                return num
            except (ValueError, TypeError):
                return None

        elif data_type == "date":
            if isinstance(value, datetime):
                return value.isoformat()
            return str(value) if value else None

        else:
            # Text
            if value is None:
                return None
            s = str(value).strip()
            if not s or s.lower() in ('nan', 'none'):
                return None
            return s


# Convenience functions for direct use
async def export_excel_to_json(
    file_bytes: bytes,
    sheet_index: int = 0,
    include_metadata: bool = True
) -> str:
    """Export Excel file directly to JSON"""
    exporter = DataExporter()
    data = await exporter.from_excel(file_bytes, sheet_index)
    return exporter.to_json(data, include_metadata=include_metadata)


async def export_excel_to_markdown(
    file_bytes: bytes,
    sheet_index: int = 0,
    max_rows: int = 50
) -> str:
    """Export Excel file directly to Markdown (best for LLM)"""
    exporter = DataExporter()
    data = await exporter.from_excel(file_bytes, sheet_index)
    return exporter.to_markdown(data, max_rows=max_rows)


async def export_excel_to_csv(
    file_bytes: bytes,
    sheet_index: int = 0,
    include_header_comments: bool = True
) -> str:
    """Export Excel file directly to CSV"""
    exporter = DataExporter()
    data = await exporter.from_excel(file_bytes, sheet_index)
    return exporter.to_csv(data, include_header_comments=include_header_comments)


@dataclass
class SheetExportResult:
    """Result of exporting a single sheet"""
    index: int
    name: str
    safe_name: str  # Filename-safe version
    row_count: int
    column_count: int
    data: StructuredData
    markdown: str
    csv: str


@dataclass
class BatchExportResult:
    """Result of batch exporting all sheets"""
    source_file: str
    sheet_count: int
    sheets: List[SheetExportResult]
    combined_json: str

    def get_file_structure(self) -> Dict[str, str]:
        """Return dict of filename -> content for all export files"""
        files = {}

        # Combined JSON
        files["all_sheets.json"] = self.combined_json

        # Individual MD and CSV files
        for sheet in self.sheets:
            files[f"md/{sheet.safe_name}.md"] = sheet.markdown
            files[f"csv/{sheet.safe_name}.csv"] = sheet.csv

        return files


class BatchExporter:
    """
    Batch export all sheets from an Excel file.

    Output structure:
    - all_sheets.json: Combined JSON with all sheets
    - md/{sheet_name}.md: Individual Markdown files
    - csv/{sheet_name}.csv: Individual CSV files
    """

    def __init__(self):
        self._exporter = DataExporter()

    async def export_all_sheets(
        self,
        file_bytes: bytes,
        source_filename: str = "workbook",
        max_rows_per_md: int = 500
    ) -> BatchExportResult:
        """
        Export all sheets from Excel file.

        Args:
            file_bytes: Excel file bytes
            source_filename: Original filename (for metadata)
            max_rows_per_md: Max rows per Markdown file

        Returns:
            BatchExportResult with all exported content
        """
        # Get sheet names
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        sheets = []
        all_sheets_data = []

        for idx, sheet_name in enumerate(sheet_names):
            logger.info(f"Exporting sheet {idx}: {sheet_name}")

            # Parse sheet
            data = await self._exporter.from_excel(file_bytes, sheet_index=idx)

            # Generate safe filename
            safe_name = self._make_safe_filename(idx, sheet_name)

            # Generate Markdown
            markdown = self._exporter.to_markdown(
                data,
                max_rows=max_rows_per_md,
                include_metadata=True
            )

            # Generate CSV
            csv_content = self._exporter.to_csv(data, include_header_comments=True)

            # Store sheet result
            sheet_result = SheetExportResult(
                index=idx,
                name=sheet_name,
                safe_name=safe_name,
                row_count=data.row_count,
                column_count=data.column_count,
                data=data,
                markdown=markdown,
                csv=csv_content
            )
            sheets.append(sheet_result)

            # Prepare for combined JSON
            all_sheets_data.append({
                "index": idx,
                "name": sheet_name,
                "metadata": data.metadata,
                "columns": [c.to_dict() for c in data.columns],
                "rows": data.rows,
                "row_count": data.row_count,
                "column_count": data.column_count
            })

        # Generate combined JSON
        combined = {
            "source_file": source_filename,
            "sheet_count": len(sheets),
            "exported_at": datetime.now().isoformat(),
            "sheets": all_sheets_data
        }
        combined_json = json.dumps(combined, ensure_ascii=False, indent=2, default=str)

        return BatchExportResult(
            source_file=source_filename,
            sheet_count=len(sheets),
            sheets=sheets,
            combined_json=combined_json
        )

    def _make_safe_filename(self, index: int, name: str) -> str:
        """Convert sheet name to safe filename"""
        # Remove/replace unsafe characters
        safe = re.sub(r'[<>:"/\\|?*]', '_', name)
        safe = safe.strip()
        # Truncate if too long
        if len(safe) > 50:
            safe = safe[:50]
        return f"{index:02d}_{safe}"

    async def save_to_directory(
        self,
        file_bytes: bytes,
        output_dir: str,
        source_filename: str = "workbook"
    ) -> Dict[str, str]:
        """
        Export all sheets and save to directory.

        Args:
            file_bytes: Excel file bytes
            output_dir: Output directory path
            source_filename: Original filename

        Returns:
            Dict of saved file paths
        """
        import os

        result = await self.export_all_sheets(file_bytes, source_filename)

        # Create directories
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(os.path.join(output_dir, "md"), exist_ok=True)
        os.makedirs(os.path.join(output_dir, "csv"), exist_ok=True)

        saved_files = {}

        # Save combined JSON
        json_path = os.path.join(output_dir, "all_sheets.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            f.write(result.combined_json)
        saved_files["all_sheets.json"] = json_path

        # Save individual files
        for sheet in result.sheets:
            # Markdown
            md_path = os.path.join(output_dir, "md", f"{sheet.safe_name}.md")
            with open(md_path, 'w', encoding='utf-8') as f:
                f.write(sheet.markdown)
            saved_files[f"md/{sheet.safe_name}.md"] = md_path

            # CSV
            csv_path = os.path.join(output_dir, "csv", f"{sheet.safe_name}.csv")
            with open(csv_path, 'w', encoding='utf-8-sig') as f:
                f.write(sheet.csv)
            saved_files[f"csv/{sheet.safe_name}.csv"] = csv_path

        logger.info(f"Saved {len(saved_files)} files to {output_dir}")
        return saved_files
