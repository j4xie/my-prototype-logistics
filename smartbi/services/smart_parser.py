"""
Smart Excel Parser - 动态规则 + LLM 结合

架构：
1. 规则优先：已知格式用配置规则快速解析
2. LLM回退：未知/复杂格式用LLM辅助理解
3. 自动学习：LLM分析结果可保存为新规则

使用场景：
- 财务报表（利润表、资产负债表、现金流量表）
- 销售数据（按区域、产品、时间维度）
- 自定义业务表格
"""

import json
import logging
import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)


# ============================================================
# 规则定义
# ============================================================

@dataclass
class HeaderRule:
    """表头识别规则"""
    # 表头行数检测
    header_keywords: List[str] = field(default_factory=lambda: [
        "项目", "科目", "行次", "预算", "实际", "合计"
    ])
    # 跳过行的关键词（标题、单位等）
    skip_keywords: List[str] = field(default_factory=lambda: [
        "利润表", "资产负债表", "现金流量表", "报表",
        "编制单位", "单位：", "期间："
    ])
    # 数据行开始的标志
    data_start_keywords: List[str] = field(default_factory=lambda: [
        "一、", "二、", "营业收入", "资产总计"
    ])


@dataclass
class ColumnRule:
    """列映射规则"""
    # 时间列模式
    time_patterns: List[str] = field(default_factory=lambda: [
        r"\d{4}[-/年]\d{1,2}[-/月]",  # 2025-01, 2025年1月
        r"\d{1,2}月",                  # 1月, 01月
        r"Q[1-4]",                     # Q1, Q2
    ])
    # 预算/实际子列
    subcolumn_keywords: Dict[str, str] = field(default_factory=lambda: {
        "预算": "budget",
        "实际": "actual",
        "本月": "actual",
        "同期": "prior",
        "累计": "ytd",
    })


@dataclass
class MetadataRule:
    """元信息提取规则"""
    # 标题提取位置（行号）
    title_rows: List[int] = field(default_factory=lambda: [1, 2])
    # 单位提取关键词
    unit_keywords: List[str] = field(default_factory=lambda: [
        "单位：", "单位:", "Unit:"
    ])
    # 公司名提取关键词
    company_keywords: List[str] = field(default_factory=lambda: [
        "编制单位：", "编制单位:", "公司："
    ])
    # 期间提取关键词
    period_keywords: List[str] = field(default_factory=lambda: [
        "期间：", "期间:", "日期：", "Period:"
    ])


@dataclass
class ExcelParseRule:
    """完整的Excel解析规则"""
    name: str                                    # 规则名称
    description: str = ""                        # 描述
    match_conditions: Dict[str, Any] = field(default_factory=dict)  # 匹配条件
    header_rule: HeaderRule = field(default_factory=HeaderRule)
    column_rule: ColumnRule = field(default_factory=ColumnRule)
    metadata_rule: MetadataRule = field(default_factory=MetadataRule)

    # 特殊处理
    skip_empty_rows: bool = True
    detect_subheader: bool = True               # 检测子表头（预算/实际）
    flatten_merged_cells: bool = True           # 展开合并单元格

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "match_conditions": self.match_conditions,
            "header_rule": {
                "header_keywords": self.header_rule.header_keywords,
                "skip_keywords": self.header_rule.skip_keywords,
                "data_start_keywords": self.header_rule.data_start_keywords,
            },
            "column_rule": {
                "time_patterns": self.column_rule.time_patterns,
                "subcolumn_keywords": self.column_rule.subcolumn_keywords,
            },
            "metadata_rule": {
                "title_rows": self.metadata_rule.title_rows,
                "unit_keywords": self.metadata_rule.unit_keywords,
                "company_keywords": self.metadata_rule.company_keywords,
                "period_keywords": self.metadata_rule.period_keywords,
            },
            "skip_empty_rows": self.skip_empty_rows,
            "detect_subheader": self.detect_subheader,
            "flatten_merged_cells": self.flatten_merged_cells,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ExcelParseRule":
        rule = cls(
            name=data.get("name", "unnamed"),
            description=data.get("description", ""),
            match_conditions=data.get("match_conditions", {}),
            skip_empty_rows=data.get("skip_empty_rows", True),
            detect_subheader=data.get("detect_subheader", True),
            flatten_merged_cells=data.get("flatten_merged_cells", True),
        )

        if "header_rule" in data:
            hr = data["header_rule"]
            rule.header_rule = HeaderRule(
                header_keywords=hr.get("header_keywords", []),
                skip_keywords=hr.get("skip_keywords", []),
                data_start_keywords=hr.get("data_start_keywords", []),
            )

        if "column_rule" in data:
            cr = data["column_rule"]
            rule.column_rule = ColumnRule(
                time_patterns=cr.get("time_patterns", []),
                subcolumn_keywords=cr.get("subcolumn_keywords", {}),
            )

        if "metadata_rule" in data:
            mr = data["metadata_rule"]
            rule.metadata_rule = MetadataRule(
                title_rows=mr.get("title_rows", [1, 2]),
                unit_keywords=mr.get("unit_keywords", []),
                company_keywords=mr.get("company_keywords", []),
                period_keywords=mr.get("period_keywords", []),
            )

        return rule


# ============================================================
# 预定义规则
# ============================================================

# 财务利润表规则
PROFIT_STATEMENT_RULE = ExcelParseRule(
    name="profit_statement",
    description="财务利润表（月度/年度）",
    match_conditions={
        "title_contains": ["利润表", "损益表", "Income Statement"],
        "has_keywords": ["营业收入", "营业成本", "净利润"],
    },
    header_rule=HeaderRule(
        header_keywords=["项目", "行次", "预算", "实际", "本月", "累计"],
        skip_keywords=["利润表", "编制单位", "单位："],
        data_start_keywords=["一、营业收入", "营业收入"],
    ),
    column_rule=ColumnRule(
        time_patterns=[r"\d{4}[-/年]\d{1,2}", r"\d{1,2}月"],
        subcolumn_keywords={"预算": "budget", "实际": "actual", "本月实际": "actual"},
    ),
)

# 销售数据规则
SALES_DATA_RULE = ExcelParseRule(
    name="sales_data",
    description="销售数据表",
    match_conditions={
        "title_contains": ["销售", "收入", "Sales"],
        "has_keywords": ["产品", "区域", "金额", "数量"],
    },
    header_rule=HeaderRule(
        header_keywords=["产品", "区域", "客户", "金额", "数量", "日期"],
        skip_keywords=["销售报表", "统计表"],
        data_start_keywords=[],  # 通常第一行就是数据
    ),
)

# 通用表格规则（兜底）
GENERIC_TABLE_RULE = ExcelParseRule(
    name="generic_table",
    description="通用表格（自动检测）",
    match_conditions={},  # 总是匹配
)

# 预定义规则库
BUILTIN_RULES = [
    PROFIT_STATEMENT_RULE,
    SALES_DATA_RULE,
    GENERIC_TABLE_RULE,
]


# ============================================================
# 规则引擎
# ============================================================

class RuleEngine:
    """规则引擎 - 管理和匹配解析规则"""

    def __init__(self, rules_dir: Optional[str] = None):
        self.rules: List[ExcelParseRule] = list(BUILTIN_RULES)
        self.rules_dir = rules_dir

        # 加载自定义规则
        if rules_dir and os.path.exists(rules_dir):
            self._load_custom_rules(rules_dir)

    def _load_custom_rules(self, rules_dir: str):
        """从目录加载自定义规则"""
        for filename in os.listdir(rules_dir):
            if filename.endswith(".json"):
                filepath = os.path.join(rules_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    rule = ExcelParseRule.from_dict(data)
                    # 插入到预定义规则之前（优先级更高）
                    self.rules.insert(0, rule)
                    logger.info(f"Loaded custom rule: {rule.name}")
                except Exception as e:
                    logger.warning(f"Failed to load rule {filename}: {e}")

    def match_rule(self, excel_info: Dict[str, Any]) -> ExcelParseRule:
        """
        根据Excel信息匹配最佳规则

        Args:
            excel_info: 包含 title, keywords, sheet_names 等信息

        Returns:
            匹配的规则（如果无匹配则返回通用规则）
        """
        title = excel_info.get("title", "").lower()
        keywords = [k.lower() for k in excel_info.get("keywords", [])]

        for rule in self.rules:
            conditions = rule.match_conditions

            if not conditions:
                # 无条件的规则（通用规则）放最后
                continue

            matched = True

            # 检查标题包含
            if "title_contains" in conditions:
                title_match = any(
                    t.lower() in title
                    for t in conditions["title_contains"]
                )
                if not title_match:
                    matched = False

            # 检查关键词
            if matched and "has_keywords" in conditions:
                required_keywords = [k.lower() for k in conditions["has_keywords"]]
                keyword_match = any(
                    rk in " ".join(keywords)
                    for rk in required_keywords
                )
                if not keyword_match:
                    matched = False

            if matched:
                logger.info(f"Matched rule: {rule.name}")
                return rule

        # 返回通用规则
        logger.info("No specific rule matched, using generic rule")
        return GENERIC_TABLE_RULE

    def save_rule(self, rule: ExcelParseRule, filename: Optional[str] = None):
        """保存规则到文件"""
        if not self.rules_dir:
            raise ValueError("rules_dir not configured")

        os.makedirs(self.rules_dir, exist_ok=True)

        if not filename:
            filename = f"{rule.name}.json"

        filepath = os.path.join(self.rules_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(rule.to_dict(), f, ensure_ascii=False, indent=2)

        logger.info(f"Saved rule to {filepath}")
        return filepath


# ============================================================
# LLM 集成
# ============================================================

class LLMParser:
    """LLM辅助解析器 - 处理复杂/未知格式"""

    # LLM分析提示词
    ANALYZE_PROMPT = """分析以下Excel表格结构，返回JSON格式的解析规则。

表格信息：
- Sheet名: {sheet_name}
- 前10行内容:
{preview_rows}

请分析并返回JSON：
{{
  "title": "表格标题",
  "table_type": "profit_statement/sales_data/inventory/other",
  "header_rows": 表头行数(数字),
  "data_start_row": 数据开始行(数字),
  "has_subheader": true/false (是否有预算/实际等子表头),
  "columns": [
    {{"name": "列名", "type": "text/numeric/date", "purpose": "用途说明"}}
  ],
  "metadata": {{
    "unit": "单位(如有)",
    "period": "期间(如有)",
    "company": "公司名(如有)"
  }},
  "notes": "其他说明"
}}

只返回JSON，不要其他内容。"""

    COLUMN_MAPPING_PROMPT = """将以下Excel列名映射为标准化名称。

原始列名：
{original_columns}

要求：
1. 日期列统一为 "YYYY-MM" 或 "MM月" 格式
2. 如有预算/实际子列，格式为 "MM月_预算" / "MM月_实际"
3. 保持中文，简洁明了

返回JSON数组：
[{{"original": "原始名", "mapped": "映射后名称", "type": "numeric/text/date"}}]

只返回JSON数组。"""

    def __init__(self, llm_client=None):
        """
        Args:
            llm_client: LLM客户端（需实现 async chat(prompt) -> str）
        """
        self.llm_client = llm_client

    async def analyze_structure(
        self,
        sheet_name: str,
        preview_rows: List[List[Any]]
    ) -> Dict[str, Any]:
        """
        使用LLM分析表格结构

        Args:
            sheet_name: Sheet名称
            preview_rows: 前N行数据预览

        Returns:
            结构分析结果
        """
        if not self.llm_client:
            logger.warning("LLM client not configured, using fallback")
            return self._fallback_analyze(preview_rows)

        # 格式化预览行
        preview_text = "\n".join(
            f"Row {i+1}: {row}"
            for i, row in enumerate(preview_rows[:10])
        )

        prompt = self.ANALYZE_PROMPT.format(
            sheet_name=sheet_name,
            preview_rows=preview_text
        )

        try:
            response = await self.llm_client.chat(prompt)
            # 提取JSON
            result = self._extract_json(response)
            return result
        except Exception as e:
            logger.error(f"LLM analysis failed: {e}")
            return self._fallback_analyze(preview_rows)

    async def suggest_column_mapping(
        self,
        original_columns: List[str]
    ) -> List[Dict[str, str]]:
        """
        使用LLM建议列名映射

        Args:
            original_columns: 原始列名列表

        Returns:
            映射建议列表
        """
        if not self.llm_client:
            return [{"original": c, "mapped": c, "type": "text"} for c in original_columns]

        prompt = self.COLUMN_MAPPING_PROMPT.format(
            original_columns="\n".join(f"- {c}" for c in original_columns)
        )

        try:
            response = await self.llm_client.chat(prompt)
            result = self._extract_json(response)
            return result if isinstance(result, list) else []
        except Exception as e:
            logger.error(f"LLM column mapping failed: {e}")
            return [{"original": c, "mapped": c, "type": "text"} for c in original_columns]

    def _extract_json(self, text: str) -> Any:
        """从LLM响应中提取JSON"""
        # 尝试找到JSON块
        json_match = re.search(r'```json\s*([\s\S]*?)\s*```', text)
        if json_match:
            text = json_match.group(1)
        else:
            # 尝试找到 { } 或 [ ]
            brace_match = re.search(r'[\[{][\s\S]*[\]}]', text)
            if brace_match:
                text = brace_match.group(0)

        return json.loads(text)

    def _fallback_analyze(self, preview_rows: List[List[Any]]) -> Dict[str, Any]:
        """无LLM时的回退分析"""
        return {
            "title": None,
            "table_type": "other",
            "header_rows": 1,
            "data_start_row": 2,
            "has_subheader": False,
            "columns": [],
            "metadata": {},
            "notes": "Fallback analysis (no LLM)"
        }


# ============================================================
# 智能解析器（整合规则+LLM）
# ============================================================

class SmartExcelParser:
    """
    智能Excel解析器

    工作流程：
    1. 快速扫描Excel，提取特征
    2. 尝试匹配已有规则
    3. 如无匹配，调用LLM分析
    4. 可选：将LLM分析结果保存为新规则
    """

    def __init__(
        self,
        rules_dir: Optional[str] = None,
        llm_client=None,
        auto_save_rules: bool = False
    ):
        self.rule_engine = RuleEngine(rules_dir)
        self.llm_parser = LLMParser(llm_client)
        self.auto_save_rules = auto_save_rules

    async def parse(
        self,
        file_bytes: bytes,
        sheet_index: int = 0,
        force_llm: bool = False
    ) -> Tuple[Dict[str, Any], ExcelParseRule]:
        """
        智能解析Excel

        Args:
            file_bytes: Excel文件字节
            sheet_index: Sheet索引
            force_llm: 强制使用LLM（跳过规则匹配）

        Returns:
            (解析结果, 使用的规则)
        """
        import io
        import openpyxl

        # 1. 加载Excel
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        if sheet_index >= len(wb.sheetnames):
            sheet_index = 0
        ws = wb[wb.sheetnames[sheet_index]]

        # 2. 快速扫描，提取特征
        excel_info = self._quick_scan(ws)

        # 3. 匹配规则
        if not force_llm:
            rule = self.rule_engine.match_rule(excel_info)
            if rule.name != "generic_table":
                # 找到特定规则，使用规则解析
                logger.info(f"Using rule: {rule.name}")
                result = self._parse_with_rule(ws, rule)
                wb.close()
                return result, rule

        # 4. 无特定规则或强制LLM，使用LLM分析
        logger.info("Using LLM analysis")
        llm_result = await self.llm_parser.analyze_structure(
            ws.title,
            excel_info["preview_rows"]
        )

        # 5. 根据LLM结果生成临时规则
        temp_rule = self._create_rule_from_llm(llm_result)

        # 6. 使用临时规则解析
        result = self._parse_with_rule(ws, temp_rule)

        # 7. 可选：保存新规则
        if self.auto_save_rules and llm_result.get("table_type") != "other":
            try:
                self.rule_engine.save_rule(temp_rule)
            except Exception as e:
                logger.warning(f"Failed to save rule: {e}")

        wb.close()
        return result, temp_rule

    def _quick_scan(self, ws) -> Dict[str, Any]:
        """快速扫描Sheet，提取特征"""
        preview_rows = []
        keywords = []
        title = ""

        for row_idx in range(1, min(15, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    val_str = str(val)
                    row_values.append(val_str)
                    # 提取关键词
                    if len(val_str) < 20:
                        keywords.append(val_str)
                else:
                    row_values.append("")
            preview_rows.append(row_values)

            # 第一行可能是标题
            if row_idx == 1 and row_values[0]:
                title = row_values[0]

        return {
            "title": title,
            "keywords": keywords,
            "preview_rows": preview_rows,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }

    def _parse_with_rule(self, ws, rule: ExcelParseRule) -> Dict[str, Any]:
        """使用规则解析Sheet"""
        # 这里调用 DataExporter 的逻辑
        # 简化版实现
        return {
            "rule_used": rule.name,
            "sheet": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }

    def _create_rule_from_llm(self, llm_result: Dict[str, Any]) -> ExcelParseRule:
        """从LLM分析结果创建规则"""
        table_type = llm_result.get("table_type", "other")

        rule = ExcelParseRule(
            name=f"llm_generated_{table_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}",
            description=f"LLM generated rule for {table_type}",
            match_conditions={
                "title_contains": [llm_result.get("title", "")] if llm_result.get("title") else [],
            },
        )

        # 设置表头行数
        if llm_result.get("header_rows"):
            rule.header_rule.header_keywords = []  # LLM已确定行数

        rule.detect_subheader = llm_result.get("has_subheader", False)

        return rule
