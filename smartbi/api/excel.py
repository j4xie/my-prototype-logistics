from __future__ import annotations
"""
Excel Parsing API

Endpoints for Excel file parsing and processing.
Includes Zero-Code auto-parse functionality.
"""
import logging
import time
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, File, UploadFile, Form, HTTPException, Query
from pydantic import BaseModel

from services.excel_parser import ExcelParser
from services.field_detector import FieldDetector
from services.data_feature_analyzer import DataFeatureAnalyzer
from services.field_mapping import FieldMappingService

# Zero-Code services
from services.structure_detector import StructureDetector
from services.semantic_mapper import SemanticMapper
from services.fixed_executor import FixedExecutor
from services.schema_cache import get_schema_cache

logger = logging.getLogger(__name__)
router = APIRouter()

# Initialize services
excel_parser = ExcelParser()
field_detector = FieldDetector()
data_feature_analyzer = DataFeatureAnalyzer()
field_mapping_service = FieldMappingService()

# Zero-Code services (lazy initialized)
_structure_detector: Optional[StructureDetector] = None
_semantic_mapper: Optional[SemanticMapper] = None
_fixed_executor: Optional[FixedExecutor] = None


def get_structure_detector() -> StructureDetector:
    global _structure_detector
    if _structure_detector is None:
        _structure_detector = StructureDetector()
    return _structure_detector


def get_semantic_mapper() -> SemanticMapper:
    global _semantic_mapper
    if _semantic_mapper is None:
        _semantic_mapper = SemanticMapper()
    return _semantic_mapper


def get_fixed_executor() -> FixedExecutor:
    global _fixed_executor
    if _fixed_executor is None:
        _fixed_executor = FixedExecutor()
    return _fixed_executor


class ParseMetadata(BaseModel):
    """Parse metadata model - Java compatible"""
    sheetName: Optional[str] = None
    originalColumnCount: Optional[int] = None
    sampledRowCount: Optional[int] = None
    parseTimeMs: Optional[int] = None
    hasMultiHeader: Optional[bool] = None
    headerRowCount: Optional[int] = None
    originalHeaders: Optional[List[List[str]]] = None
    detectedOrientation: Optional[str] = None
    totalRowCount: Optional[int] = None


class ParseResponse(BaseModel):
    """Excel parse response model - compatible with Java ExcelParseResponse"""
    success: bool
    errorMessage: Optional[str] = None
    uploadId: Optional[int] = None
    headers: List[str] = []
    rowCount: int = 0
    columnCount: int = 0
    fieldMappings: Optional[List[Dict]] = None
    dataFeatures: Optional[List[Dict]] = None
    previewData: Optional[List[Dict]] = None
    missingRequiredFields: Optional[List[str]] = None
    status: Optional[str] = None
    metadata: Optional[ParseMetadata] = None
    # Backward compatibility fields
    rows: List[list] = []
    direction: Optional[str] = None
    sheetNames: Optional[List[str]] = None
    currentSheet: Optional[str] = None


class SheetInfo(BaseModel):
    """Detailed information about a sheet"""
    name: str
    index: int
    rowCount: Optional[int] = None
    columnCount: Optional[int] = None
    hasData: bool = True
    previewHeaders: Optional[List[str]] = None


class SheetNamesResponse(BaseModel):
    """Sheet names response model"""
    success: bool
    sheetNames: List[str] = []
    error: Optional[str] = None


class ListSheetsResponse(BaseModel):
    """Detailed sheet listing response model"""
    success: bool
    totalSheets: int = 0
    sheets: List[SheetInfo] = []
    errorMessage: Optional[str] = None


class MultiHeaderDetectionResponse(BaseModel):
    """Multi-header detection response model"""
    success: bool
    isMultiHeader: bool = False
    recommendedHeaderRows: int = 1
    previewRows: Optional[List[list]] = None
    error: Optional[str] = None


def _perform_analysis(
    headers: List[str],
    rows: List[List[Any]],
    factory_id: Optional[str] = None
) -> tuple[List[Dict], List[Dict], List[str]]:
    """
    Perform field detection, data feature analysis, and field mapping.

    Args:
        headers: Column headers
        rows: Data rows
        factory_id: Optional factory ID for factory-specific mappings

    Returns:
        Tuple of (field_mappings, data_features, missing_required_fields)
    """
    # Analyze data features for each column
    data_features = []
    for i, header in enumerate(headers):
        column_values = [row[i] if i < len(row) else None for row in rows]
        feature = data_feature_analyzer.analyze_column(header, column_values, i)
        data_features.append(feature)

    # Map fields using the mapping service
    field_mappings_results = field_mapping_service.map_fields(
        headers,
        data_features,
        factory_id
    )

    # Convert to dictionaries for JSON response
    field_mappings = [fm.to_dict() for fm in field_mappings_results]
    data_features_dict = [df.to_dict() for df in data_features]

    # Find missing required fields
    mapped_fields = {
        fm.standard_field for fm in field_mappings_results
        if fm.standard_field is not None
    }
    missing_required = field_mapping_service.dictionary.get_missing_recommended_fields(mapped_fields)

    return field_mappings, data_features_dict, missing_required


@router.post("/parse", response_model=ParseResponse)
async def parse_excel(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),  # Alias for Java client compatibility
    header_rows: int = Form(1),
    skip_rows: int = Form(0),
    transpose: bool = Form(False),
    analyze: bool = Query(False, description="Include field analysis in response")
):
    """
    Parse Excel file and return structured data

    - **file**: Excel file (.xlsx, .xls) or CSV file
    - **sheet_name**: Sheet name to parse (default: first sheet)
    - **sheet_index**: Sheet index to parse (0-based, alternative to sheet_name)
    - **sheetIndex**: Alias for sheet_index (Java client compatibility)
    - **header_rows**: Number of header rows (1 for single, 2 for multi-level)
    - **skip_rows**: Number of rows to skip at the top
    - **transpose**: Whether to transpose the data
    - **analyze**: Include field analysis (field mappings, data features) in response
    """
    start_time = time.time()

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in excel_parser.supported_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Supported: {excel_parser.supported_extensions}"
            )

        # Read file content
        content = await file.read()

        # Parse based on file type
        if ext == ".csv":
            result = excel_parser.parse_csv(content)
        else:
            # Determine sheet to parse: sheet_name takes priority, then sheet_index/sheetIndex
            # Use sheetIndex as alias for sheet_index (Java client compatibility)
            effective_index = sheet_index if sheet_index is not None else sheetIndex

            if sheet_name:
                sheet = sheet_name
            elif effective_index is not None:
                sheet = int(effective_index)  # Ensure it's an integer
            else:
                sheet = 0  # Default to first sheet

            logger.info(f"Parsing Excel sheet: {sheet} (sheet_name={sheet_name}, sheet_index={sheet_index}, sheetIndex={sheetIndex})")

            result = excel_parser.parse(
                content,
                sheet_name=sheet,
                header_rows=header_rows,
                skip_rows=skip_rows,
                transpose=transpose
            )

        # Convert rows to previewData (List[Dict]) for Java compatibility
        headers = result.get("headers", [])
        rows = result.get("rows", [])
        preview_data = []
        for row in rows[:100]:  # Limit to 100 rows for preview
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            preview_data.append(row_dict)

        # Calculate parse time
        parse_time_ms = int((time.time() - start_time) * 1000)

        # Build metadata
        metadata = ParseMetadata(
            sheetName=result.get("currentSheet"),
            originalColumnCount=result.get("columnCount"),
            sampledRowCount=min(len(rows), 100),
            parseTimeMs=parse_time_ms,
            hasMultiHeader=header_rows > 1,
            headerRowCount=header_rows,
            detectedOrientation=result.get("direction"),
            totalRowCount=result.get("rowCount", len(rows))
        )

        # Optionally perform analysis
        field_mappings = None
        data_features = None
        missing_required_fields = None

        if analyze and headers and rows:
            field_mappings, data_features, missing_required_fields = _perform_analysis(
                headers, rows, None
            )

        return ParseResponse(
            success=result.get("success", True),
            errorMessage=result.get("error"),
            headers=headers,
            rowCount=result.get("rowCount", len(rows)),
            columnCount=result.get("columnCount", len(headers)),
            previewData=preview_data,
            fieldMappings=field_mappings,
            dataFeatures=data_features,
            missingRequiredFields=missing_required_fields,
            metadata=metadata,
            status="PARSED" if not analyze else "ANALYZED",
            # Keep backward compatibility
            rows=rows,
            direction=result.get("direction"),
            sheetNames=result.get("sheetNames"),
            currentSheet=result.get("currentSheet")
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel parse error: {e}", exc_info=True)
        return ParseResponse(success=False, errorMessage=str(e))


@router.post("/parse-analyze", response_model=ParseResponse)
async def parse_and_analyze(
    file: UploadFile = File(...),
    factory_id: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),  # Alias for Java client compatibility
    header_rows: int = Form(1),
    skip_rows: int = Form(0),
    auto_detect_multi_header: bool = Form(True),
    auto_detect_orientation: bool = Form(True),
    auto_detect_title_rows: bool = Form(True)  # NEW: Auto-detect and skip title rows
):
    """
    Parse Excel file with full field analysis.

    This endpoint performs:
    1. Basic parsing (like /parse)
    2. Data feature analysis (type detection, statistics)
    3. Field mapping (maps columns to standard fields)
    4. Missing field detection

    - **file**: Excel file (.xlsx, .xls) or CSV file
    - **factory_id**: Factory ID for factory-specific field mappings
    - **sheet_name**: Sheet name to parse (default: first sheet)
    - **sheet_index**: Sheet index to parse (0-based)
    - **sheetIndex**: Alias for sheet_index (Java client compatibility)
    - **header_rows**: Number of header rows (default: 1)
    - **skip_rows**: Number of rows to skip at the top
    - **auto_detect_multi_header**: Automatically detect multi-level headers
    - **auto_detect_orientation**: Automatically detect data orientation (row vs column)
    - **auto_detect_title_rows**: Automatically detect and skip title rows (NEW)
    """
    start_time = time.time()

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in excel_parser.supported_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Supported: {excel_parser.supported_extensions}"
            )

        # Read file content
        content = await file.read()

        # Determine sheet to parse
        effective_index = sheet_index if sheet_index is not None else sheetIndex

        if sheet_name:
            sheet = sheet_name
        elif effective_index is not None:
            sheet = int(effective_index)
        else:
            sheet = 0

        logger.info(f"Parse-analyze: sheet={sheet}, factory_id={factory_id}, auto_detect_multi_header={auto_detect_multi_header}, auto_detect_title_rows={auto_detect_title_rows}")

        # NEW: Use StructureDetector for comprehensive header detection
        detected_header_rows = header_rows
        detected_skip_rows = skip_rows
        structure_result = None

        if auto_detect_title_rows and ext != ".csv":
            try:
                detector = get_structure_detector()
                structure_result = await detector.detect(
                    content,
                    sheet_index=sheet if isinstance(sheet, int) else 0,
                    max_header_rows=10
                )

                if structure_result.success and structure_result.confidence >= 0.5:
                    # data_start_row tells us where actual data begins (0-indexed)
                    # header_row_count tells us how many header rows there are
                    # skip_rows = data_start_row - header_row_count

                    # Calculate rows to skip (title rows before headers)
                    title_rows_count = structure_result.data_start_row - structure_result.header_row_count
                    if title_rows_count > 0:
                        detected_skip_rows = title_rows_count
                        logger.info(f"StructureDetector: skipping {title_rows_count} title rows")

                    detected_header_rows = structure_result.header_row_count
                    logger.info(f"StructureDetector: detected header_rows={detected_header_rows}, data_start={structure_result.data_start_row}, confidence={structure_result.confidence:.2f}")
            except Exception as e:
                logger.warning(f"StructureDetector failed, falling back to simple detection: {e}")

        # Fallback: Auto-detect multi-header using simple method if StructureDetector didn't run
        if structure_result is None and auto_detect_multi_header and ext != ".csv":
            detection_result = excel_parser.detect_multi_header(content, sheet)
            if detection_result.get("success") and detection_result.get("isMultiHeader"):
                detected_header_rows = detection_result.get("recommendedHeaderRows", header_rows)
                logger.info(f"Simple detection: multi-header with {detected_header_rows} rows")

        # Parse the file with detected parameters
        if ext == ".csv":
            result = excel_parser.parse_csv(content)
        else:
            result = excel_parser.parse(
                content,
                sheet_name=sheet,
                header_rows=detected_header_rows,
                skip_rows=detected_skip_rows,
                transpose=False  # Will handle orientation separately
            )

        if not result.get("success"):
            return ParseResponse(
                success=False,
                errorMessage=result.get("error", "Parse failed"),
                status="PARSE_ERROR"
            )

        headers = result.get("headers", [])
        rows = result.get("rows", [])

        # Auto-detect orientation if enabled
        detected_orientation = result.get("direction", "row")
        if auto_detect_orientation and hasattr(excel_parser, 'detect_orientation'):
            try:
                orientation_result = excel_parser.detect_orientation(content, sheet)
                if orientation_result.get("success"):
                    detected_orientation = orientation_result.get("direction", "row")
            except Exception as e:
                logger.warning(f"Orientation detection failed: {e}")

        # Perform full analysis
        field_mappings, data_features, missing_required_fields = _perform_analysis(
            headers, rows, factory_id
        )

        # Convert rows to previewData
        preview_data = []
        for row in rows[:100]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            preview_data.append(row_dict)

        # Calculate parse time
        parse_time_ms = int((time.time() - start_time) * 1000)

        # Build metadata
        metadata = ParseMetadata(
            sheetName=result.get("currentSheet"),
            originalColumnCount=len(headers),
            sampledRowCount=min(len(rows), 100),
            parseTimeMs=parse_time_ms,
            hasMultiHeader=detected_header_rows > 1,
            headerRowCount=detected_header_rows,
            detectedOrientation=detected_orientation,
            totalRowCount=result.get("rowCount", len(rows))
        )

        return ParseResponse(
            success=True,
            headers=headers,
            rowCount=result.get("rowCount", len(rows)),
            columnCount=len(headers),
            previewData=preview_data,
            fieldMappings=field_mappings,
            dataFeatures=data_features,
            missingRequiredFields=missing_required_fields,
            metadata=metadata,
            status="ANALYZED",
            # Backward compatibility
            rows=rows,
            direction=detected_orientation,
            sheetNames=result.get("sheetNames"),
            currentSheet=result.get("currentSheet")
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Parse-analyze error: {e}", exc_info=True)
        return ParseResponse(
            success=False,
            errorMessage=str(e),
            status="ERROR"
        )


@router.post("/list-sheets", response_model=ListSheetsResponse)
async def list_sheets_detailed(file: UploadFile = File(...)):
    """
    Get detailed information about all sheets in an Excel file.

    Returns sheet names, indices, dimensions, and preview headers.

    - **file**: Excel file (.xlsx, .xls)
    """
    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="This endpoint only supports Excel files (.xlsx, .xls)"
            )

        content = await file.read()
        sheet_names = excel_parser.get_sheet_names(content)

        sheets: List[SheetInfo] = []
        for i, name in enumerate(sheet_names):
            try:
                # Parse each sheet to get dimensions
                result = excel_parser.parse(content, sheet_name=i, header_rows=1)

                sheet_info = SheetInfo(
                    name=name,
                    index=i,
                    rowCount=result.get("rowCount", 0),
                    columnCount=result.get("columnCount", 0),
                    hasData=result.get("rowCount", 0) > 0,
                    previewHeaders=result.get("headers", [])[:10]  # First 10 headers
                )
            except Exception as e:
                logger.warning(f"Failed to get info for sheet '{name}': {e}")
                sheet_info = SheetInfo(
                    name=name,
                    index=i,
                    hasData=False
                )

            sheets.append(sheet_info)

        return ListSheetsResponse(
            success=True,
            totalSheets=len(sheets),
            sheets=sheets
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"List sheets error: {e}", exc_info=True)
        return ListSheetsResponse(
            success=False,
            errorMessage=str(e)
        )


@router.post("/sheets", response_model=SheetNamesResponse)
async def get_sheet_names(file: UploadFile = File(...)):
    """
    Get all sheet names from an Excel file

    - **file**: Excel file (.xlsx, .xls)
    """
    try:
        content = await file.read()
        sheet_names = excel_parser.get_sheet_names(content)

        return SheetNamesResponse(
            success=True,
            sheetNames=sheet_names
        )

    except Exception as e:
        logger.error(f"Get sheet names error: {e}", exc_info=True)
        return SheetNamesResponse(success=False, error=str(e))


@router.post("/detect-header", response_model=MultiHeaderDetectionResponse)
async def detect_multi_header(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None)
):
    """
    Detect if Excel file has multi-level headers

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_name**: Sheet name to analyze (default: first sheet)
    - **sheet_index**: Sheet index to analyze (0-based)
    - **sheetIndex**: Alias for sheet_index
    """
    try:
        content = await file.read()

        # Determine sheet: sheet_name takes priority
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if sheet_name:
            sheet = sheet_name
        elif effective_index is not None:
            sheet = int(effective_index)
        else:
            sheet = 0

        result = excel_parser.detect_multi_header(content, sheet)

        return MultiHeaderDetectionResponse(**result)

    except Exception as e:
        logger.error(f"Multi-header detection error: {e}", exc_info=True)
        return MultiHeaderDetectionResponse(success=False, error=str(e))


@router.post("/preview")
async def preview_excel(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    rows: int = Form(10)
):
    """
    Preview first N rows of Excel file

    - **file**: Excel file
    - **sheet_name**: Sheet name to preview
    - **sheet_index**: Sheet index to preview (0-based)
    - **sheetIndex**: Alias for sheet_index
    - **rows**: Number of rows to preview (default: 10)
    """
    try:
        content = await file.read()

        # Determine sheet: sheet_name takes priority
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if sheet_name:
            sheet = sheet_name
        elif effective_index is not None:
            sheet = int(effective_index)
        else:
            sheet = 0

        result = excel_parser.parse(content, sheet_name=sheet)

        if result.get("success"):
            preview_rows = result["rows"][:rows]
            return {
                "success": True,
                "headers": result["headers"],
                "rows": preview_rows,
                "totalRows": result["rowCount"],
                "previewedRows": len(preview_rows)
            }

        return result

    except Exception as e:
        logger.error(f"Excel preview error: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


class SheetAnalysisResponse(BaseModel):
    """Sheet analysis response model"""
    success: bool
    sheets: Optional[List[Dict]] = None
    recommended: Optional[List[int]] = None
    summary: Optional[str] = None
    method: Optional[str] = None
    error: Optional[str] = None


@router.post("/analyze-sheets", response_model=SheetAnalysisResponse)
async def analyze_sheets(file: UploadFile = File(...)):
    """
    Analyze Excel sheets using LLM to identify which contain valid analyzable data.

    Returns:
    - sheets: List of sheet analysis results
    - recommended: List of recommended sheet indices to analyze
    - summary: Analysis summary

    - **file**: Excel file (.xlsx, .xls)
    """
    from services.llm_mapper import LLMMapper

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="This endpoint only supports Excel files (.xlsx, .xls)"
            )

        content = await file.read()

        # Get sheet information
        sheet_names = excel_parser.get_sheet_names(content)
        sheets_info = []

        for i, name in enumerate(sheet_names):
            try:
                result = excel_parser.parse(content, sheet_name=i, header_rows=1)
                sheet_info = {
                    "index": i,
                    "name": name,
                    "rowCount": result.get("rowCount", 0),
                    "columnCount": result.get("columnCount", 0),
                    "previewHeaders": result.get("headers", [])[:10]
                }
            except Exception as e:
                logger.warning(f"Failed to get info for sheet '{name}': {e}")
                sheet_info = {
                    "index": i,
                    "name": name,
                    "rowCount": 0,
                    "columnCount": 0,
                    "previewHeaders": []
                }
            sheets_info.append(sheet_info)

        # Analyze sheets using LLM
        llm_mapper = LLMMapper()
        try:
            analysis_result = await llm_mapper.analyze_sheets(sheets_info)
            return SheetAnalysisResponse(**analysis_result)
        finally:
            await llm_mapper.close()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Sheet analysis error: {e}", exc_info=True)
        return SheetAnalysisResponse(success=False, error=str(e))


# =============================================================================
# Zero-Code Auto-Parse Endpoint
# =============================================================================


class ContextResponse(BaseModel):
    """Context information from Excel sheet (Three-Layer Model - Layer 3)"""
    notes: List[str] = []
    explanations: List[str] = []
    definitions: Dict[str, str] = {}
    source_rows: List[int] = []
    metadata: Optional[Dict[str, Any]] = None


class AutoParseResponse(BaseModel):
    """
    Response model for auto-parse endpoint.
    Uses camelCase field names for Java compatibility.
    """
    # Core fields (matching Java ExcelParseResponse)
    success: bool
    errorMessage: Optional[str] = None  # Java uses errorMessage, not error
    headers: List[str] = []
    rowCount: int = 0  # camelCase for Java
    columnCount: int = 0  # camelCase for Java
    fieldMappings: Optional[List[Dict[str, Any]]] = None  # camelCase for Java
    dataFeatures: Optional[List[Dict[str, Any]]] = None  # Java expects this
    previewData: Optional[List[Dict[str, Any]]] = None  # camelCase for Java
    missingRequiredFields: Optional[List[str]] = None  # Java expects this
    status: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None  # ParseMetadata

    # Backward compatibility (keep snake_case aliases)
    rows: List[list] = []  # Raw rows for backward compatibility
    direction: Optional[str] = None
    sheetNames: Optional[List[str]] = None
    currentSheet: Optional[str] = None

    # Auto-parse specific fields
    autoDetected: bool = True
    tableType: Optional[str] = None
    structure: Optional[Dict[str, Any]] = None
    unmappedFields: Optional[List[str]] = None
    originalHeaders: List[str] = []
    dataTypes: Optional[Dict[str, str]] = None
    statistics: Optional[Dict[str, Dict[str, Any]]] = None
    context: Optional[ContextResponse] = None
    detectionMethod: Optional[str] = None
    confidence: float = 0.0
    cacheKey: Optional[str] = None
    processingTimeMs: int = 0
    notes: List[str] = []
    recommendedCharts: Optional[List[str]] = None

    class Config:
        # Allow population by field name (Python) or alias (camelCase)
        populate_by_name = True


@router.post("/auto-parse", response_model=AutoParseResponse)
async def auto_parse_excel(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    factory_id: Optional[str] = Form(None),
    use_cache: bool = Form(True),
    max_rows: int = Form(10000),
    skip_empty_rows: bool = Form(True),
    calculate_stats: bool = Form(True)
):
    """
    Zero-Code Excel Auto-Parse

    Automatically detects Excel structure and maps fields without user configuration.

    Process:
    1. Check cache for previously parsed files
    2. Auto-detect header structure (title rows, merged cells, data start)
    3. Semantically map columns to standard business fields
    4. Extract and standardize data

    Features:
    - Fully automatic - no headerRow or field mapping configuration needed
    - Multi-layer detection: rules → LLM-fast → LLM-VL → multi-model
    - Result caching for cost optimization
    - Learns from user corrections

    - **file**: Excel file (.xlsx, .xls) or CSV
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    - **factory_id**: Optional factory ID for custom mappings
    - **use_cache**: Use cached results if available (default: true)
    - **max_rows**: Maximum rows to extract (default: 10000)
    - **skip_empty_rows**: Skip empty rows (default: true)
    - **calculate_stats**: Calculate column statistics (default: true)
    """
    start_time = time.time()

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls", ".csv"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls, .csv"
            )

        # Read file content
        content = await file.read()

        # Determine sheet index
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        logger.info(f"Auto-parse starting: file={filename}, sheet={effective_index}")

        # Initialize services
        detector = get_structure_detector()
        mapper = get_semantic_mapper()
        executor = get_fixed_executor()
        cache = get_schema_cache()

        # Step 1: Check cache
        cache_key = None
        structure_result = None
        mapping_result = None

        if use_cache:
            cached = cache.get(content, effective_index)
            if cached:
                structure_result, mapping_result = cached
                logger.info("Using cached schema")

        # Step 2: Detect structure if not cached
        if structure_result is None:
            structure_result = await detector.detect(
                content,
                sheet_index=effective_index,
                max_header_rows=10
            )
            if not structure_result.success:
                return AutoParseResponse(
                    success=False,
                    auto_detected=True,
                    error=structure_result.error or "Structure detection failed"
                )

            logger.info(
                f"Structure detected: method={structure_result.method}, "
                f"header_rows={structure_result.header_row_count}, "
                f"confidence={structure_result.confidence:.2f}"
            )

        # Step 3: Map fields if not cached
        if mapping_result is None:
            columns = [col.name for col in structure_result.columns]
            sample_data = structure_result.preview_rows[structure_result.data_start_row:][:5] if structure_result.preview_rows else None

            mapping_result = await mapper.map_fields(
                columns=columns,
                sample_data=sample_data,
                factory_id=factory_id,
                table_context=structure_result.sheet_name
            )

            logger.info(
                f"Fields mapped: mapped={len(mapping_result.field_mappings)}, "
                f"unmapped={len(mapping_result.unmapped_fields)}, "
                f"confidence={mapping_result.confidence:.2f}"
            )

        # Step 4: Cache results
        if use_cache and structure_result and mapping_result:
            cache_key = cache.set(
                content, effective_index,
                structure_result, mapping_result,
                metadata={"filename": filename}
            )

        # Step 5: Extract data
        extracted = executor.execute_with_pandas(
            content,
            structure_result,
            mapping_result,
            options={
                "max_rows": max_rows,
                "skip_empty_rows": skip_empty_rows,
                "calculate_stats": calculate_stats
            }
        )

        if not extracted.success:
            return AutoParseResponse(
                success=False,
                autoDetected=True,
                errorMessage=extracted.error or "Data extraction failed",
                status="EXTRACT_ERROR"
            )

        # Build response
        processing_time = int((time.time() - start_time) * 1000)

        # Generate recommended charts based on data types
        recommended_charts = _recommend_charts(
            mapping_result.field_mappings,
            mapping_result.table_type
        )

        # Collect notes
        notes = extracted.processing_notes.copy()
        if structure_result.note:
            notes.append(structure_result.note)
        if mapping_result.note:
            notes.append(mapping_result.note)

        # Convert field_mappings to Java-compatible format
        field_mappings_java = []
        for fm in mapping_result.field_mappings:
            fm_dict = fm.to_dict()
            # Ensure camelCase keys
            field_mappings_java.append({
                "originalColumn": fm_dict.get("original") or fm_dict.get("originalColumn"),
                "standardField": fm_dict.get("standard") or fm_dict.get("standardField"),
                "dataType": fm_dict.get("data_type") or fm_dict.get("dataType"),
                "confidence": fm_dict.get("confidence", 0.0),
                "source": fm_dict.get("source", "auto")
            })

        # Build metadata in Java ParseMetadata format
        metadata = {
            "sheetName": structure_result.sheet_name,
            "originalColumnCount": structure_result.total_cols,
            "sampledRowCount": min(extracted.row_count, 100),
            "parseTimeMs": processing_time,
            "hasMultiHeader": structure_result.header_row_count > 1,
            "headerRowCount": structure_result.header_row_count,
            "detectedOrientation": "row",  # Default
            "totalRowCount": extracted.row_count
        }

        # Convert rows to previewData (List[Dict]) format
        preview_data = []
        for row in extracted.rows[:100]:
            if isinstance(row, dict):
                preview_data.append(row)
            elif isinstance(row, (list, tuple)):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(extracted.headers):
                        row_dict[extracted.headers[i]] = value
                preview_data.append(row_dict)

        return AutoParseResponse(
            # Core Java-compatible fields
            success=True,
            errorMessage=None,
            headers=extracted.headers,
            rowCount=extracted.row_count,
            columnCount=extracted.column_count,
            fieldMappings=field_mappings_java,
            dataFeatures=None,  # Will be populated if needed
            previewData=preview_data,
            missingRequiredFields=mapping_result.unmapped_fields,
            status="ANALYZED",
            metadata=metadata,

            # Backward compatibility
            rows=[],  # Don't send raw rows to save bandwidth
            direction="row",
            sheetNames=None,
            currentSheet=structure_result.sheet_name,

            # Auto-parse specific (camelCase)
            autoDetected=True,
            tableType=mapping_result.table_type,
            structure={
                "headerRows": structure_result.header_row_count,
                "dataStartRow": structure_result.data_start_row,
                "totalRows": structure_result.total_rows,
                "totalCols": structure_result.total_cols,
                "mergedCellsCount": len(structure_result.merged_cells)
            },
            unmappedFields=mapping_result.unmapped_fields,
            originalHeaders=extracted.original_headers,
            dataTypes=extracted.data_types,
            statistics=extracted.statistics if calculate_stats else None,
            context=ContextResponse(
                notes=extracted.context.notes if extracted.context else [],
                explanations=extracted.context.explanations if extracted.context else [],
                definitions=extracted.context.definitions if extracted.context else {},
                source_rows=extracted.context.source_rows if extracted.context else [],
                metadata=extracted.context.to_dict().get("metadata") if extracted.context else None
            ) if extracted.context and extracted.context.has_content() else None,
            detectionMethod=structure_result.method,
            confidence=min(structure_result.confidence, mapping_result.confidence),
            cacheKey=cache_key,
            processingTimeMs=processing_time,
            notes=notes,
            recommendedCharts=recommended_charts
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Auto-parse error: {e}", exc_info=True)
        return AutoParseResponse(
            success=False,
            autoDetected=True,
            errorMessage=str(e),
            status="ERROR",
            processingTimeMs=int((time.time() - start_time) * 1000)
        )


def _recommend_charts(
    field_mappings: List[Any],
    table_type: Optional[str]
) -> List[str]:
    """Recommend chart types based on detected fields and table type"""
    charts = []

    # Get mapped standard fields
    standard_fields = {
        fm.standard for fm in field_mappings
        if fm.standard is not None
    }

    # Based on table type
    if table_type == "profit_statement":
        charts.extend(["waterfall", "bar_comparison", "trend_line"])
    elif table_type == "sales_report":
        charts.extend(["bar_chart", "pie_chart", "trend_line"])
    elif table_type == "budget_report":
        charts.extend(["bar_comparison", "gauge", "waterfall"])

    # Based on available fields
    if "budget_amount" in standard_fields and "actual_amount" in standard_fields:
        if "bar_comparison" not in charts:
            charts.append("bar_comparison")
        if "variance" in standard_fields and "waterfall" not in charts:
            charts.append("waterfall")

    if "achievement_rate" in standard_fields:
        if "gauge" not in charts:
            charts.append("gauge")

    if "yoy_rate" in standard_fields or "mom_rate" in standard_fields:
        if "trend_line" not in charts:
            charts.append("trend_line")

    # Default charts if none recommended
    if not charts:
        charts = ["bar_chart", "table"]

    return charts[:5]  # Limit to 5 recommendations


@router.post("/auto-parse/feedback")
async def submit_auto_parse_feedback(
    cache_key: str = Form(...),
    correction_type: str = Form(...),
    original_value: str = Form(...),
    correct_value: str = Form(...)
):
    """
    Submit feedback for auto-parse results to improve future detection.

    - **cache_key**: The cache key from auto-parse response
    - **correction_type**: Type of correction (mapping, structure, data_type)
    - **original_value**: The original detected value
    - **correct_value**: The user's corrected value
    """
    try:
        cache = get_schema_cache()

        correction = {
            "type": correction_type,
            "original": original_value,
            "correct": correct_value
        }

        success = cache.add_user_correction(cache_key, correction)

        if success:
            logger.info(f"Feedback recorded: {correction_type} correction for {cache_key}")
            return {"success": True, "message": "Feedback recorded"}
        else:
            return {"success": False, "error": "Cache key not found"}

    except Exception as e:
        logger.error(f"Feedback submission error: {e}")
        return {"success": False, "error": str(e)}


# =============================================================================
# Context Extraction Endpoint
# =============================================================================

class ExtractContextResponse(BaseModel):
    """Response model for context extraction endpoint"""
    success: bool
    error: Optional[str] = None
    context: Optional[ContextResponse] = None
    has_content: bool = False


@router.post("/extract-context", response_model=ExtractContextResponse)
async def extract_context(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    data_end_row: Optional[int] = Form(None)
):
    """
    Extract context information from Excel file.

    Extracts:
    - Notes (备注)
    - Compilation explanations (编制说明)
    - Term definitions
    - Metadata (title, unit, period, department)

    This is part of the Three-Layer Data Model:
    - Layer 1: Metadata (title, unit, period)
    - Layer 2: Data (business records)
    - Layer 3: Context (notes, explanations, definitions)

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    - **data_end_row**: Optional row where data ends (auto-detected if not provided)
    """
    from services.context_extractor import ContextExtractor

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls"
            )

        content = await file.read()

        # Determine sheet index
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        # Extract context
        extractor = ContextExtractor()
        context_info = extractor.extract_from_bytes(
            file_bytes=content,
            sheet_index=effective_index,
            data_end_row=data_end_row
        )

        return ExtractContextResponse(
            success=True,
            context=ContextResponse(
                notes=context_info.notes,
                explanations=context_info.explanations,
                definitions=context_info.definitions,
                source_rows=context_info.source_rows,
                metadata=context_info.to_dict().get("metadata")
            ) if context_info.has_content() else None,
            has_content=context_info.has_content()
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Context extraction error: {e}", exc_info=True)
        return ExtractContextResponse(
            success=False,
            error=str(e)
        )


# =============================================================================
# Multi-Format Export Endpoints
# =============================================================================


class ExportResponse(BaseModel):
    """Response model for export endpoints"""
    success: bool
    format: str
    content: str
    metadata: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    processing_time_ms: int = 0


class ExportMetadataResponse(BaseModel):
    """Metadata-only response for export preview"""
    success: bool
    metadata: Optional[Dict[str, Any]] = None
    columns: Optional[List[Dict[str, Any]]] = None
    row_count: int = 0
    column_count: int = 0
    context: Optional[ContextResponse] = None
    error: Optional[str] = None


@router.post("/export", response_model=ExportResponse)
async def export_excel(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    format: str = Form("json"),
    include_metadata: bool = Form(True),
    max_rows: Optional[int] = Form(None)
):
    """
    Export Excel file to multiple formats.

    Supported formats:
    - **json**: Structured JSON (best for internal processing, storage)
    - **markdown**: Markdown table (best for LLM prompts, ~60.7% accuracy)
    - **csv**: CSV with header comments (best for export, debugging)

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    - **format**: Output format (json, markdown, csv)
    - **include_metadata**: Include metadata in output (default: true)
    - **max_rows**: Maximum rows to export (default: all for json/csv, 50 for markdown)
    """
    from services.data_exporter import DataExporter

    start_time = time.time()

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls"
            )

        # Validate format
        format_lower = format.lower()
        if format_lower not in ["json", "markdown", "csv"]:
            raise HTTPException(
                status_code=400,
                detail="Supported formats: json, markdown, csv"
            )

        content = await file.read()

        # Determine sheet index
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        # Parse Excel
        exporter = DataExporter()
        data = await exporter.from_excel(
            content,
            sheet_index=effective_index,
            max_rows=max_rows
        )

        # Export to requested format
        if format_lower == "json":
            output = exporter.to_json(data, include_metadata=include_metadata)
        elif format_lower == "markdown":
            md_max_rows = max_rows if max_rows else 50
            output = exporter.to_markdown(data, max_rows=md_max_rows, include_metadata=include_metadata)
        else:  # csv
            output = exporter.to_csv(data, include_header_comments=include_metadata)

        processing_time = int((time.time() - start_time) * 1000)

        return ExportResponse(
            success=True,
            format=format_lower,
            content=output,
            metadata=data.metadata if include_metadata else None,
            processing_time_ms=processing_time
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export error: {e}", exc_info=True)
        return ExportResponse(
            success=False,
            format=format,
            content="",
            error=str(e),
            processing_time_ms=int((time.time() - start_time) * 1000)
        )


@router.post("/export/json", response_model=ExportResponse)
async def export_to_json(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    include_metadata: bool = Form(True),
    max_rows: Optional[int] = Form(None)
):
    """
    Export Excel file to JSON format.

    Best for: Internal processing, storage, API responses.

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    - **include_metadata**: Include metadata in output
    - **max_rows**: Maximum rows to export (default: all)
    """
    return await export_excel(
        file=file,
        sheet_index=sheet_index,
        sheetIndex=sheetIndex,
        format="json",
        include_metadata=include_metadata,
        max_rows=max_rows
    )


@router.post("/export/markdown", response_model=ExportResponse)
async def export_to_markdown(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    include_metadata: bool = Form(True),
    max_rows: int = Form(50)
):
    """
    Export Excel file to Markdown format.

    Best for: LLM prompts (highest accuracy ~60.7%).

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    - **include_metadata**: Include metadata header
    - **max_rows**: Maximum rows to include (default: 50 for LLM context limits)
    """
    return await export_excel(
        file=file,
        sheet_index=sheet_index,
        sheetIndex=sheetIndex,
        format="markdown",
        include_metadata=include_metadata,
        max_rows=max_rows
    )


@router.post("/export/csv", response_model=ExportResponse)
async def export_to_csv(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None),
    include_metadata: bool = Form(True),
    max_rows: Optional[int] = Form(None)
):
    """
    Export Excel file to CSV format.

    Best for: Export, debugging, Excel compatibility.

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    - **include_metadata**: Include metadata as header comments
    - **max_rows**: Maximum rows to export (default: all)
    """
    return await export_excel(
        file=file,
        sheet_index=sheet_index,
        sheetIndex=sheetIndex,
        format="csv",
        include_metadata=include_metadata,
        max_rows=max_rows
    )


@router.post("/export/metadata", response_model=ExportMetadataResponse)
async def get_export_metadata(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(None),
    sheetIndex: Optional[int] = Form(None)
):
    """
    Get metadata and column information without full export.

    Useful for previewing data structure before choosing export format.

    - **file**: Excel file (.xlsx, .xls)
    - **sheet_index**: Sheet index (0-based, default: 0)
    - **sheetIndex**: Alias for sheet_index
    """
    from services.data_exporter import DataExporter

    try:
        # Validate file type
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls"
            )

        content = await file.read()

        # Determine sheet index
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        # Parse Excel (limited rows for metadata only)
        exporter = DataExporter()
        data = await exporter.from_excel(
            content,
            sheet_index=effective_index,
            max_rows=10  # Only need a few rows for metadata
        )

        return ExportMetadataResponse(
            success=True,
            metadata=data.metadata,
            columns=[c.to_dict() for c in data.columns],
            row_count=data.row_count,
            column_count=data.column_count,
            context=ContextResponse(
                notes=data.context.notes if data.context else [],
                explanations=data.context.explanations if data.context else [],
                definitions=data.context.definitions if data.context else {},
                source_rows=data.context.source_rows if data.context else [],
                metadata=data.context.to_dict().get("metadata") if data.context else None
            ) if data.context and data.context.has_content() else None
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export metadata error: {e}", exc_info=True)
        return ExportMetadataResponse(
            success=False,
            error=str(e)
        )


# ============================================================
# Batch Export API - 批量导出所有Sheets
# ============================================================

class BatchExportSheetInfo(BaseModel):
    """单个Sheet的导出信息"""
    index: int
    name: str
    safe_name: str
    row_count: int
    column_count: int


class BatchExportResponse(BaseModel):
    """批量导出响应"""
    success: bool
    source_file: Optional[str] = None
    sheet_count: Optional[int] = None
    sheets: Optional[List[BatchExportSheetInfo]] = None
    files: Optional[Dict[str, str]] = None  # filename -> content (for small files)
    error: Optional[str] = None
    message: Optional[str] = None


@router.post("/export/batch", response_model=BatchExportResponse)
async def batch_export_all_sheets(
    file: UploadFile = File(...),
    max_rows_per_md: int = Form(500),
    return_content: bool = Form(False)
):
    """
    批量导出Excel所有Sheets。

    输出结构：
    - all_sheets.json: 合并的JSON文件（包含所有sheets）
    - md/{sheet_name}.md: 每个sheet的Markdown文件
    - csv/{sheet_name}.csv: 每个sheet的CSV文件

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **max_rows_per_md**: Markdown文件最大行数 (默认500)
    - **return_content**: 是否在响应中返回文件内容 (默认false，仅返回结构信息)
    """
    from services.data_exporter import BatchExporter

    try:
        filename = file.filename or "workbook.xlsx"
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls"
            )

        content = await file.read()
        exporter = BatchExporter()

        result = await exporter.export_all_sheets(
            content,
            source_filename=filename,
            max_rows_per_md=max_rows_per_md
        )

        # 构建响应
        sheets_info = [
            BatchExportSheetInfo(
                index=s.index,
                name=s.name,
                safe_name=s.safe_name,
                row_count=s.row_count,
                column_count=s.column_count
            )
            for s in result.sheets
        ]

        response = BatchExportResponse(
            success=True,
            source_file=filename,
            sheet_count=result.sheet_count,
            sheets=sheets_info,
            message=f"Successfully exported {result.sheet_count} sheets"
        )

        # 如果请求返回内容
        if return_content:
            response.files = result.get_file_structure()

        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Batch export error: {e}", exc_info=True)
        return BatchExportResponse(
            success=False,
            error=str(e)
        )


@router.post("/export/batch/save")
async def batch_export_and_save(
    file: UploadFile = File(...),
    output_dir: str = Form(...),
    max_rows_per_md: int = Form(500)
):
    """
    批量导出并保存到指定目录。

    参数:
    - **file**: Excel文件
    - **output_dir**: 输出目录路径
    - **max_rows_per_md**: Markdown文件最大行数
    """
    from services.data_exporter import BatchExporter

    try:
        filename = file.filename or "workbook.xlsx"
        content = await file.read()

        exporter = BatchExporter()
        saved_files = await exporter.save_to_directory(
            content,
            output_dir=output_dir,
            source_filename=filename
        )

        return {
            "success": True,
            "output_dir": output_dir,
            "files_saved": len(saved_files),
            "files": list(saved_files.keys())
        }

    except Exception as e:
        logger.error(f"Batch export save error: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e)
        }


# ============================================================
# Smart Parse API - 智能解析（规则+LLM）
# ============================================================

class SmartParseResponse(BaseModel):
    """智能解析响应"""
    success: bool
    rule_used: Optional[str] = None
    rule_type: Optional[str] = None  # builtin / custom / llm_generated
    sheet_name: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None
    columns: Optional[List[Dict[str, Any]]] = None
    row_count: Optional[int] = None
    column_count: Optional[int] = None
    preview_rows: Optional[List[Dict[str, Any]]] = None
    error: Optional[str] = None


class RuleMatchResponse(BaseModel):
    """规则匹配响应"""
    success: bool
    matched_rule: Optional[str] = None
    rule_description: Optional[str] = None
    confidence: Optional[float] = None
    available_rules: Optional[List[str]] = None
    error: Optional[str] = None


@router.post("/smart-parse/match-rule", response_model=RuleMatchResponse)
async def match_parse_rule(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0)
):
    """
    匹配最佳解析规则（不执行解析）。

    用于预览将使用哪个规则，以及可用的规则列表。

    参数:
    - **file**: Excel文件
    - **sheet_index**: Sheet索引
    """
    from services.smart_parser import SmartExcelParser, RuleEngine
    import io
    import openpyxl

    try:
        content = await file.read()

        # 快速扫描
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        if sheet_index >= len(wb.sheetnames):
            sheet_index = 0
        ws = wb[wb.sheetnames[sheet_index]]

        # 提取特征
        preview_rows = []
        keywords = []
        title = ""

        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    val_str = str(val)
                    row_values.append(val_str)
                    if len(val_str) < 20:
                        keywords.append(val_str)
            if row_idx == 1 and row_values:
                title = row_values[0] if row_values else ""

        wb.close()

        excel_info = {
            "title": title,
            "keywords": keywords,
        }

        # 匹配规则
        engine = RuleEngine()
        rule = engine.match_rule(excel_info)

        return RuleMatchResponse(
            success=True,
            matched_rule=rule.name,
            rule_description=rule.description,
            confidence=1.0 if rule.name != "generic_table" else 0.5,
            available_rules=[r.name for r in engine.rules]
        )

    except Exception as e:
        logger.error(f"Rule match error: {e}", exc_info=True)
        return RuleMatchResponse(
            success=False,
            error=str(e)
        )


@router.post("/smart-parse", response_model=SmartParseResponse)
async def smart_parse_excel(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0),
    force_llm: bool = Form(False),
    preview_rows_count: int = Form(10)
):
    """
    智能解析Excel - 规则优先，LLM回退。

    工作流程:
    1. 快速扫描Excel，提取特征
    2. 尝试匹配已有规则（builtin或custom）
    3. 如无匹配，调用LLM分析结构
    4. 返回解析结果和使用的规则

    参数:
    - **file**: Excel文件
    - **sheet_index**: Sheet索引
    - **force_llm**: 强制使用LLM（跳过规则匹配）
    - **preview_rows_count**: 预览行数
    """
    from services.data_exporter import DataExporter
    from services.smart_parser import RuleEngine

    try:
        filename = file.filename or "workbook.xlsx"
        content = await file.read()

        # 1. 使用现有 DataExporter 解析
        exporter = DataExporter()
        data = await exporter.from_excel(content, sheet_index=sheet_index)

        # 2. 匹配规则（仅用于标记）
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[sheet_index]] if sheet_index < len(wb.sheetnames) else wb[wb.sheetnames[0]]

        keywords = []
        title = ""
        for row_idx in range(1, min(5, ws.max_row + 1)):
            for col_idx in range(1, min(5, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    val_str = str(val)
                    if len(val_str) < 30:
                        keywords.append(val_str)
                    if row_idx == 1 and col_idx == 1:
                        title = val_str
        wb.close()

        engine = RuleEngine()
        rule = engine.match_rule({"title": title, "keywords": keywords})

        # 3. 构建响应
        preview = data.rows[:preview_rows_count] if data.rows else []

        return SmartParseResponse(
            success=True,
            rule_used=rule.name,
            rule_type="builtin" if rule.name in ["profit_statement", "sales_data", "generic_table"] else "custom",
            sheet_name=data.source_sheet,
            metadata=data.metadata,
            columns=[c.to_dict() for c in data.columns],
            row_count=data.row_count,
            column_count=data.column_count,
            preview_rows=preview
        )

    except Exception as e:
        logger.error(f"Smart parse error: {e}", exc_info=True)
        return SmartParseResponse(
            success=False,
            error=str(e)
        )


# ============================================================
# Export with Validation & Fix API - LLM 二次核对 + 自动修复
# ============================================================

class ValidationIssueModel(BaseModel):
    """验证问题模型"""
    level: str  # "error" | "warning" | "info"
    category: str
    message: str
    details: Optional[Dict[str, Any]] = None


class ValidationResultModel(BaseModel):
    """验证结果模型"""
    success: bool
    sheet_name: str
    has_errors: bool
    issue_count: int
    issues: List[ValidationIssueModel]
    summary: Dict[str, Any]
    llm_review: Optional[str] = None


class ExportWithValidationResponse(BaseModel):
    """导出+验证响应"""
    success: bool
    # 导出数据
    metadata: Optional[Dict[str, Any]] = None
    columns: Optional[List[Dict[str, Any]]] = None
    row_count: Optional[int] = None
    column_count: Optional[int] = None
    preview_rows: Optional[List[Dict[str, Any]]] = None
    # 验证结果
    validation: Optional[ValidationResultModel] = None
    # 修复记录（如果启用）
    fix_log: Optional[List[str]] = None
    # 错误信息
    error: Optional[str] = None
    processing_time_ms: int = 0


@router.post("/export/validate", response_model=ExportWithValidationResponse)
async def export_with_validation(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0),
    sheetIndex: Optional[int] = Form(None),
    use_llm: bool = Form(True),
    preview_rows_count: int = Form(20)
):
    """
    导出Excel并进行LLM二次核对。

    流程:
    1. 解析Excel并导出数据
    2. 规则验证（行数、列数、元信息、备注）
    3. LLM二次核对（如果启用）
    4. 返回导出数据和验证结果

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **sheet_index**: Sheet索引 (默认0)
    - **sheetIndex**: sheet_index的别名
    - **use_llm**: 是否使用LLM二次核对 (默认true)
    - **preview_rows_count**: 响应中预览的行数 (默认20)
    """
    from services.export_validator import ExportValidator, SimpleLLMClient
    from services.data_exporter import DataExporter
    from config import get_settings
    settings = get_settings()

    start_time = time.time()

    try:
        # 确定sheet索引
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        content = await file.read()

        # 1. 导出数据
        exporter = DataExporter()
        data = await exporter.from_excel(content, sheet_index=effective_index)

        exported = {
            "metadata": data.metadata,
            "columns": [c.to_dict() for c in data.columns],
            "rows": data.rows,
            "row_count": data.row_count,
            "column_count": data.column_count
        }

        # 2. 创建LLM客户端（如果需要）
        llm_client = None
        if use_llm:
            try:
                api_key = settings.llm_api_key
                base_url = settings.llm_base_url
                model = settings.llm_model

                if api_key:
                    llm_client = SimpleLLMClient(
                        api_key=api_key,
                        base_url=base_url,
                        model=model
                    )
            except Exception as e:
                logger.warning(f"LLM client init failed, skipping LLM validation: {e}")

        # 3. 验证
        validator = ExportValidator(llm_client=llm_client)
        validation_result = await validator.validate_export(
            content,
            exported,
            sheet_index=effective_index,
            use_llm=llm_client is not None
        )

        processing_time = int((time.time() - start_time) * 1000)

        # 4. 构建响应
        validation_model = ValidationResultModel(
            success=validation_result.success,
            sheet_name=validation_result.sheet_name,
            has_errors=validation_result.has_errors(),
            issue_count=len(validation_result.issues),
            issues=[
                ValidationIssueModel(
                    level=i.level,
                    category=i.category,
                    message=i.message,
                    details=i.details
                )
                for i in validation_result.issues
            ],
            summary=validation_result.summary,
            llm_review=validation_result.llm_review
        )

        preview = data.rows[:preview_rows_count] if data.rows else []

        return ExportWithValidationResponse(
            success=True,
            metadata=data.metadata,
            columns=[c.to_dict() for c in data.columns],
            row_count=data.row_count,
            column_count=data.column_count,
            preview_rows=preview,
            validation=validation_model,
            processing_time_ms=processing_time
        )

    except Exception as e:
        logger.error(f"Export with validation error: {e}", exc_info=True)
        return ExportWithValidationResponse(
            success=False,
            error=str(e),
            processing_time_ms=int((time.time() - start_time) * 1000)
        )


@router.post("/export/validate-fix", response_model=ExportWithValidationResponse)
async def export_with_validation_and_fix(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0),
    sheetIndex: Optional[int] = Form(None),
    auto_fix: bool = Form(True),
    preview_rows_count: int = Form(20)
):
    """
    导出Excel + LLM核对 + 自动修复差异。

    流程:
    1. 解析Excel并导出数据
    2. 规则验证 + LLM二次核对
    3. 如果发现差异且启用auto_fix，调用LLM修复
    4. 返回修复后的数据和修复记录

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **sheet_index**: Sheet索引 (默认0)
    - **sheetIndex**: sheet_index的别名
    - **auto_fix**: 发现差异时自动修复 (默认true)
    - **preview_rows_count**: 响应中预览的行数 (默认20)
    """
    from services.export_validator import (
        ExportValidator, ExportFixer, SimpleLLMClient
    )
    from services.data_exporter import DataExporter
    from config import get_settings
    settings = get_settings()

    start_time = time.time()

    try:
        # 确定sheet索引
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        content = await file.read()

        # 1. 导出数据
        exporter = DataExporter()
        data = await exporter.from_excel(content, sheet_index=effective_index)

        exported = {
            "metadata": data.metadata,
            "columns": [c.to_dict() for c in data.columns],
            "rows": data.rows,
            "row_count": data.row_count,
            "column_count": data.column_count
        }

        # 2. 创建LLM客户端
        llm_client = None
        try:
            api_key = settings.llm_api_key
            base_url = settings.llm_base_url
            model = settings.llm_model

            if api_key:
                llm_client = SimpleLLMClient(
                    api_key=api_key,
                    base_url=base_url,
                    model=model
                )
        except Exception as e:
            logger.warning(f"LLM client init failed: {e}")

        # 3. 验证
        validator = ExportValidator(llm_client=llm_client)
        validation_result = await validator.validate_export(
            content,
            exported,
            sheet_index=effective_index,
            use_llm=llm_client is not None
        )

        fix_log = []

        # 4. 如果有问题且启用自动修复
        if auto_fix and llm_client and validation_result.issues:
            fixable_issues = [i for i in validation_result.issues if i.level in ["error", "warning"]]

            if fixable_issues:
                fixer = ExportFixer(llm_client)
                exported, fix_log = await fixer.fix_export(
                    content,
                    exported,
                    fixable_issues,
                    sheet_index=effective_index
                )

                # 重新验证（用规则验证确认修复效果）
                validation_result = await validator.validate_export(
                    content,
                    exported,
                    sheet_index=effective_index,
                    use_llm=False
                )

        processing_time = int((time.time() - start_time) * 1000)

        # 5. 构建响应
        validation_model = ValidationResultModel(
            success=validation_result.success,
            sheet_name=validation_result.sheet_name,
            has_errors=validation_result.has_errors(),
            issue_count=len(validation_result.issues),
            issues=[
                ValidationIssueModel(
                    level=i.level,
                    category=i.category,
                    message=i.message,
                    details=i.details
                )
                for i in validation_result.issues
            ],
            summary=validation_result.summary,
            llm_review=validation_result.llm_review
        )

        # 获取预览行
        rows = exported.get("rows", [])
        preview = rows[:preview_rows_count] if rows else []

        return ExportWithValidationResponse(
            success=True,
            metadata=exported.get("metadata"),
            columns=exported.get("columns"),
            row_count=exported.get("row_count", len(rows)),
            column_count=exported.get("column_count", 0),
            preview_rows=preview,
            validation=validation_model,
            fix_log=fix_log if fix_log else None,
            processing_time_ms=processing_time
        )

    except Exception as e:
        logger.error(f"Export with validation and fix error: {e}", exc_info=True)
        return ExportWithValidationResponse(
            success=False,
            error=str(e),
            processing_time_ms=int((time.time() - start_time) * 1000)
        )


@router.post("/export/batch/validate")
async def batch_export_with_validation(
    file: UploadFile = File(...),
    use_llm: bool = Form(True),
    max_rows_per_md: int = Form(500)
):
    """
    批量导出所有Sheets并验证。

    对每个Sheet进行:
    1. 导出数据
    2. 规则验证 + LLM二次核对

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **use_llm**: 是否使用LLM核对 (默认true)
    - **max_rows_per_md**: Markdown最大行数
    """
    from services.export_validator import ExportValidator, SimpleLLMClient
    from services.data_exporter import BatchExporter
    from config import get_settings
    settings = get_settings()

    start_time = time.time()

    try:
        content = await file.read()
        filename = file.filename or "workbook.xlsx"

        # 1. 批量导出
        batch_exporter = BatchExporter()
        export_result = await batch_exporter.export_all_sheets(
            content,
            source_filename=filename,
            max_rows_per_md=max_rows_per_md
        )

        # 2. 创建LLM客户端
        llm_client = None
        if use_llm:
            try:
                api_key = settings.llm_api_key
                base_url = settings.llm_base_url
                model = settings.llm_model

                if api_key:
                    llm_client = SimpleLLMClient(
                        api_key=api_key,
                        base_url=base_url,
                        model=model
                    )
            except Exception as e:
                logger.warning(f"LLM client init failed: {e}")

        # 3. 验证每个Sheet
        validator = ExportValidator(llm_client=llm_client)
        validation_results = []
        sheets_with_errors = 0
        sheets_with_warnings = 0

        for idx, sheet_result in enumerate(export_result.sheets):
            sheet_data = {
                "metadata": sheet_result.data.metadata,
                "columns": [c.to_dict() for c in sheet_result.data.columns],
                "rows": sheet_result.data.rows,
                "row_count": sheet_result.data.row_count,
                "column_count": sheet_result.data.column_count
            }

            validation = await validator.validate_export(
                content,
                sheet_data,
                sheet_index=idx,
                use_llm=llm_client is not None
            )

            validation_results.append({
                "sheet_index": idx,
                "sheet_name": sheet_result.name,
                "success": validation.success,
                "has_errors": validation.has_errors(),
                "issue_count": len(validation.issues),
                "issues": [
                    {
                        "level": i.level,
                        "category": i.category,
                        "message": i.message
                    }
                    for i in validation.issues
                ],
                "llm_review": validation.llm_review
            })

            if validation.has_errors():
                sheets_with_errors += 1
            elif any(i.level == "warning" for i in validation.issues):
                sheets_with_warnings += 1

        processing_time = int((time.time() - start_time) * 1000)

        return {
            "success": sheets_with_errors == 0,
            "source_file": filename,
            "total_sheets": export_result.sheet_count,
            "sheets_validated": len(validation_results),
            "sheets_with_errors": sheets_with_errors,
            "sheets_with_warnings": sheets_with_warnings,
            "validation_results": validation_results,
            "processing_time_ms": processing_time
        }

    except Exception as e:
        logger.error(f"Batch export with validation error: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "processing_time_ms": int((time.time() - start_time) * 1000)
        }


# ============================================================
# Raw Export API - 原始数据导出（保真模式）
# ============================================================

class RawExportResponse(BaseModel):
    """原始导出响应"""
    success: bool
    format: str
    content: str
    sheet_name: Optional[str] = None
    total_rows: Optional[int] = None
    total_cols: Optional[int] = None
    merged_cells_count: Optional[int] = None
    stats: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    processing_time_ms: int = 0


@router.post("/raw-export", response_model=RawExportResponse)
async def raw_export_excel(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0),
    sheetIndex: Optional[int] = Form(None),
    format: str = Form("json"),
    max_rows: Optional[int] = Form(None),
    truncate: bool = Form(True),
    include_metadata: bool = Form(True)
):
    """
    原始数据导出 - 100%保真模式

    不做任何智能处理，把Excel数据原样导出，保留：
    - 所有行（不猜测表头）
    - 合并单元格信息
    - 原始行号、列号
    - 单元格数据类型

    支持格式:
    - **json**: 结构化JSON（完整信息）
    - **json_simple**: 简化JSON（紧凑格式）
    - **markdown**: Markdown表格（适合LLM处理）
    - **csv**: CSV格式（带可选元数据注释）

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **sheet_index**: Sheet索引 (0-based, 默认0)
    - **sheetIndex**: sheet_index的别名
    - **format**: 输出格式 (json/json_simple/markdown/csv)
    - **max_rows**: 最大行数限制（None表示全部）
    - **truncate**: Markdown格式是否截断长单元格（默认True）
    - **include_metadata**: CSV格式是否包含元数据注释（默认True）
    """
    from services.raw_exporter import RawExporter

    start_time = time.time()

    try:
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls"
            )

        content = await file.read()

        # 确定sheet索引
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        # 导出原始数据
        exporter = RawExporter()
        raw_data = exporter.export_sheet(
            content,
            sheet_index=effective_index,
            max_rows=max_rows
        )

        # 转换为指定格式
        format_lower = format.lower()
        if format_lower == "json":
            output = exporter.to_json(raw_data, simple=False)
        elif format_lower == "json_simple":
            output = exporter.to_json(raw_data, simple=True)
        elif format_lower == "markdown":
            output = exporter.to_markdown(
                raw_data,
                max_rows=max_rows or 50,
                truncate=truncate
            )
        elif format_lower == "csv":
            output = exporter.to_csv(
                raw_data,
                include_row_number=True,
                include_metadata=include_metadata
            )
        else:
            raise HTTPException(
                status_code=400,
                detail="Supported formats: json, json_simple, markdown, csv"
            )

        processing_time = int((time.time() - start_time) * 1000)

        return RawExportResponse(
            success=True,
            format=format_lower,
            content=output,
            sheet_name=raw_data.sheet_name,
            total_rows=raw_data.total_rows,
            total_cols=raw_data.total_cols,
            merged_cells_count=len(raw_data.merged_cells),
            stats=raw_data.stats,
            processing_time_ms=processing_time
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Raw export error: {e}", exc_info=True)
        return RawExportResponse(
            success=False,
            format=format,
            content="",
            error=str(e),
            processing_time_ms=int((time.time() - start_time) * 1000)
        )


@router.post("/raw-export/all-sheets")
async def raw_export_all_sheets(
    file: UploadFile = File(...),
    format: str = Form("json"),
    max_rows_per_sheet: Optional[int] = Form(None)
):
    """
    导出所有Sheets的原始数据

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **format**: 输出格式 (json/json_simple)
    - **max_rows_per_sheet**: 每个Sheet的最大行数
    """
    from services.raw_exporter import RawExporter

    start_time = time.time()

    try:
        content = await file.read()

        exporter = RawExporter()
        all_data = exporter.export_all_sheets(content, max_rows_per_sheet)

        sheets_output = []
        for raw_data in all_data:
            if format.lower() == "json_simple":
                sheet_json = raw_data.to_simple_dict()
            else:
                sheet_json = raw_data.to_dict()
            sheets_output.append(sheet_json)

        processing_time = int((time.time() - start_time) * 1000)

        return {
            "success": True,
            "sheet_count": len(sheets_output),
            "sheets": sheets_output,
            "processing_time_ms": processing_time
        }

    except Exception as e:
        logger.error(f"Raw export all sheets error: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "processing_time_ms": int((time.time() - start_time) * 1000)
        }


# ============================================================
# LLM Structure Analysis API - LLM结构分析
# ============================================================

class ColumnInfoModel(BaseModel):
    """列信息模型"""
    col_letter: str
    col_index: int
    name: str
    data_type: str
    meaning: str
    role: str
    is_key: bool = False


class StructureAnalysisModel(BaseModel):
    """结构分析模型"""
    success: bool
    sheet_name: str
    total_rows: int
    total_cols: int
    title_rows: List[int]
    header_rows: List[int]
    data_start_row: int
    data_end_row: Optional[int] = None
    columns: List[ColumnInfoModel]
    table_type: str
    table_type_confidence: float
    merged_cells_meaning: Dict[str, str] = {}
    notes: List[str] = []
    method: str = "llm"
    error: Optional[str] = None


class AnalysisRecommendationModel(BaseModel):
    """分析推荐模型"""
    analysis_type: str
    description: str
    priority: int
    chart_types: List[str]
    required_columns: List[str]


class LLMAnalysisResponse(BaseModel):
    """LLM分析响应"""
    success: bool
    structure: Optional[StructureAnalysisModel] = None
    recommendations: List[AnalysisRecommendationModel] = []
    insights: List[str] = []
    warnings: List[str] = []
    error: Optional[str] = None
    processing_time_ms: int = 0


@router.post("/analyze-structure", response_model=LLMAnalysisResponse)
async def analyze_excel_structure(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0),
    sheetIndex: Optional[int] = Form(None),
    include_recommendations: bool = Form(True)
):
    """
    使用LLM分析Excel结构

    流程:
    1. 使用RawExporter导出原始数据（保真）
    2. 将数据转为Markdown格式
    3. 调用LLM识别表头、数据起始行、列含义
    4. 返回结构化分析结果和分析推荐

    输出:
    - structure: 结构分析（表头行、数据起始行、列信息、表格类型）
    - recommendations: 推荐的分析类型和图表
    - insights: 初步洞察
    - warnings: 数据质量警告

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **sheet_index**: Sheet索引 (0-based, 默认0)
    - **sheetIndex**: sheet_index的别名
    - **include_recommendations**: 是否包含分析推荐（默认True）
    """
    from services.llm_structure_analyzer import LLMStructureAnalyzer

    start_time = time.time()

    try:
        filename = file.filename or ""
        ext = "." + filename.split(".")[-1].lower() if "." in filename else ""

        if ext not in [".xlsx", ".xls"]:
            raise HTTPException(
                status_code=400,
                detail="Supported file types: .xlsx, .xls"
            )

        content = await file.read()

        # 确定sheet索引
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        # 分析结构
        analyzer = LLMStructureAnalyzer()
        try:
            result = await analyzer.analyze_from_bytes(
                content,
                sheet_index=effective_index,
                include_recommendations=include_recommendations
            )
        finally:
            await analyzer.close()

        processing_time = int((time.time() - start_time) * 1000)

        # 转换为响应模型
        structure_model = StructureAnalysisModel(
            success=result.structure.success,
            sheet_name=result.structure.sheet_name,
            total_rows=result.structure.total_rows,
            total_cols=result.structure.total_cols,
            title_rows=result.structure.title_rows,
            header_rows=result.structure.header_rows,
            data_start_row=result.structure.data_start_row,
            data_end_row=result.structure.data_end_row,
            columns=[
                ColumnInfoModel(
                    col_letter=c.col_letter,
                    col_index=c.col_index,
                    name=c.name,
                    data_type=c.data_type,
                    meaning=c.meaning,
                    role=c.role,
                    is_key=c.is_key
                )
                for c in result.structure.columns
            ],
            table_type=result.structure.table_type,
            table_type_confidence=result.structure.table_type_confidence,
            merged_cells_meaning=result.structure.merged_cells_meaning,
            notes=result.structure.notes,
            method=result.structure.method,
            error=result.structure.error
        )

        recommendations_models = [
            AnalysisRecommendationModel(
                analysis_type=r.analysis_type,
                description=r.description,
                priority=r.priority,
                chart_types=r.chart_types,
                required_columns=r.required_columns
            )
            for r in result.recommendations
        ]

        return LLMAnalysisResponse(
            success=True,
            structure=structure_model,
            recommendations=recommendations_models,
            insights=result.insights,
            warnings=result.warnings,
            processing_time_ms=processing_time
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"LLM structure analysis error: {e}", exc_info=True)
        return LLMAnalysisResponse(
            success=False,
            error=str(e),
            processing_time_ms=int((time.time() - start_time) * 1000)
        )


@router.post("/smart-analyze")
async def smart_analyze_excel(
    file: UploadFile = File(...),
    sheet_index: Optional[int] = Form(0),
    sheetIndex: Optional[int] = Form(None),
    max_analyses: int = Form(5)
):
    """
    智能分析Excel - 完整流程

    流程:
    1. 原始导出（保真）
    2. LLM结构分析
    3. 基于结构提取数据
    4. 执行推荐的分析
    5. 生成洞察和建议

    参数:
    - **file**: Excel文件 (.xlsx, .xls)
    - **sheet_index**: Sheet索引 (0-based, 默认0)
    - **sheetIndex**: sheet_index的别名
    - **max_analyses**: 最多执行的分析数量（默认5）
    """
    from services.raw_exporter import RawExporter
    from services.llm_structure_analyzer import LLMStructureAnalyzer
    from services.smart_analyzer import SmartAnalyzer

    start_time = time.time()

    try:
        content = await file.read()

        # 确定sheet索引
        effective_index = sheet_index if sheet_index is not None else sheetIndex
        if effective_index is None:
            effective_index = 0

        # 1. 原始导出
        raw_exporter = RawExporter()
        raw_data = raw_exporter.export_sheet(content, sheet_index=effective_index)

        # 2. LLM结构分析
        llm_analyzer = LLMStructureAnalyzer()
        try:
            structure_result = await llm_analyzer.analyze(raw_data, include_recommendations=True)
        finally:
            await llm_analyzer.close()

        # 3. 基于结构提取数据并分析
        # 将raw_data转为smart_analyzer需要的格式
        data_start = structure_result.structure.data_start_row
        columns = structure_result.structure.columns

        # 提取列名
        headers = [c.name for c in columns]

        # 提取数据行
        rows = []
        for row in raw_data.rows[data_start:]:
            row_dict = {}
            for i, cell in enumerate(row.cells):
                col_name = columns[i].name if i < len(columns) else f"Col{i}"
                row_dict[col_name] = cell.value
            rows.append(row_dict)

        # 构建exported_data格式
        exported_data = {
            "metadata": {
                "title": raw_data.sheet_name,
                "table_type": structure_result.structure.table_type
            },
            "columns": [
                {
                    "name": c.name,
                    "data_type": c.data_type,
                    "role": c.role
                }
                for c in columns
            ],
            "rows": rows
        }

        # 4. 执行智能分析
        smart_analyzer = SmartAnalyzer()
        analysis_result = await smart_analyzer.analyze(exported_data, max_analyses=max_analyses)

        processing_time = int((time.time() - start_time) * 1000)

        return {
            "success": True,
            # 原始数据统计
            "raw_stats": {
                "sheet_name": raw_data.sheet_name,
                "total_rows": raw_data.total_rows,
                "total_cols": raw_data.total_cols,
                "merged_cells": len(raw_data.merged_cells)
            },
            # 结构分析
            "structure": {
                "table_type": structure_result.structure.table_type,
                "table_type_confidence": structure_result.structure.table_type_confidence,
                "header_rows": structure_result.structure.header_rows,
                "data_start_row": structure_result.structure.data_start_row,
                "columns": [
                    {
                        "letter": c.col_letter,
                        "name": c.name,
                        "type": c.data_type,
                        "role": c.role
                    }
                    for c in columns
                ],
                "method": structure_result.structure.method
            },
            # 分析结果
            "scenario": {
                "type": analysis_result.scenario.scenario.value,
                "confidence": analysis_result.scenario.confidence,
                "evidence": analysis_result.scenario.evidence
            },
            "analyses": [
                {
                    "type": a.analysis_type,
                    "title": a.title,
                    "data": a.data,
                    "insights": a.insights,
                    "warnings": a.warnings,
                    "chart_type": a.chart_config.get("type") if a.chart_config else None
                }
                for a in analysis_result.analyses
            ],
            "recommendations": [
                {
                    "type": r.analysis_type,
                    "description": r.description,
                    "chart_type": r.chart_type
                }
                for r in analysis_result.recommendations
            ],
            "summary": analysis_result.summary,
            "processing_notes": analysis_result.processing_notes,
            # LLM洞察
            "llm_insights": structure_result.insights,
            "llm_warnings": structure_result.warnings,
            "processing_time_ms": processing_time
        }

    except Exception as e:
        logger.error(f"Smart analyze error: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "processing_time_ms": int((time.time() - start_time) * 1000)
        }
