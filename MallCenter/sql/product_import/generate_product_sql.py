#!/usr/bin/env python3
"""
generate_product_sql.py - 从 Excel 生成商品导入 SQL

用法:
    cd /Users/jietaoxie/my-prototype-logistics/MallCenter
    source .venv/bin/activate
    python sql/product_import/generate_product_sql.py
"""

import openpyxl
import uuid
import re
from datetime import datetime

# 分类映射
CATEGORY_MAP = {
    '丸滑产品': 'CAT001',
    '家禽蛋副': 'CAT002',
    '小吃点心': 'CAT003',
    '水发产品': 'CAT004',
    '海鲜水产': 'CAT005',
    '牛羊肉类': 'CAT006',
    '猪肉猪副': 'CAT007',
    '米面制品': 'CAT008',
    '肉肠罐头': 'CAT009',
    '蔬菜菌菇': 'CAT010',
    '蘸料底料': 'CAT011',
    '调理肉类': 'CAT012',
    '豆制品类': 'CAT013',
    '饮料甜品': 'CAT014',
}

def escape_sql(value):
    """转义 SQL 字符串"""
    if value is None:
        return ''
    value = str(value)
    return value.replace("'", "''").replace("\\", "\\\\")

def generate_id():
    """生成32位UUID"""
    return uuid.uuid4().hex

def parse_price(value):
    """解析价格"""
    if value is None:
        return 0.00
    try:
        return float(value)
    except:
        return 0.00

def generate_spu_code(index, shelf):
    """生成商品编码"""
    prefix = "SPU" if shelf == '1' else "SPU_OFF"
    return f"{prefix}{index:04d}"

def main():
    # 读取 Excel
    wb = openpyxl.load_workbook('小程序上线产品.xlsx')

    sql_lines = []
    sql_lines.append("-- =====================================================")
    sql_lines.append("-- 03_products.sql - 导入商品数据")
    sql_lines.append(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sql_lines.append("-- =====================================================")
    sql_lines.append("")
    sql_lines.append("-- 插入商品 SPU")
    sql_lines.append("INSERT INTO goods_spu (id, spu_code, name, sell_point, description, category_first, category_second, pic_urls, shelf, sort, sales_price, market_price, cost_price, stock, sale_num, create_time, del_flag)")
    sql_lines.append("VALUES")

    values = []
    index = 1

    # 处理在售商品
    sheet = wb['在售']
    for row in range(2, sheet.max_row + 1):
        category = sheet.cell(row=row, column=2).value
        name = sheet.cell(row=row, column=3).value
        storage = sheet.cell(row=row, column=4).value  # 储存方式 -> sell_point
        spec = sheet.cell(row=row, column=5).value     # 包装规格 -> description
        cost_price = parse_price(sheet.cell(row=row, column=7).value)
        sales_price = parse_price(sheet.cell(row=row, column=10).value)
        market_price = parse_price(sheet.cell(row=row, column=11).value)
        supplier = sheet.cell(row=row, column=13).value

        if not name:
            continue

        spu_id = generate_id()
        spu_code = generate_spu_code(index, '1')
        category_id = CATEGORY_MAP.get(category, 'CAT001')

        # 组合 sell_point (储存方式 + 供应商)
        sell_point = f"{storage or ''}"
        if supplier:
            sell_point += f" | 供应商: {supplier}"

        # 默认图片 (稍后替换)
        pic_urls = '/uploads/products/default.png'

        value = f"('{spu_id}', '{spu_code}', '{escape_sql(name)}', '{escape_sql(sell_point)}', '{escape_sql(spec or '')}', 'ROOT', '{category_id}', '{pic_urls}', '1', {index}, {sales_price:.2f}, {market_price:.2f}, {cost_price:.2f}, 999, 0, NOW(), '0')"
        values.append(value)
        index += 1

    # 处理下架商品
    sheet = wb['下架']
    for row in range(2, sheet.max_row + 1):
        category = sheet.cell(row=row, column=2).value
        name = sheet.cell(row=row, column=3).value
        storage = sheet.cell(row=row, column=4).value
        spec = sheet.cell(row=row, column=5).value
        cost_price = parse_price(sheet.cell(row=row, column=7).value)
        sales_price = parse_price(sheet.cell(row=row, column=10).value)
        market_price = parse_price(sheet.cell(row=row, column=11).value)
        supplier = sheet.cell(row=row, column=13).value

        if not name:
            continue

        spu_id = generate_id()
        spu_code = generate_spu_code(index, '0')
        category_id = CATEGORY_MAP.get(category, 'CAT001')

        sell_point = f"{storage or ''}"
        if supplier:
            sell_point += f" | 供应商: {supplier}"

        pic_urls = '/uploads/products/default.png'

        value = f"('{spu_id}', '{spu_code}', '{escape_sql(name)}', '{escape_sql(sell_point)}', '{escape_sql(spec or '')}', 'ROOT', '{category_id}', '{pic_urls}', '0', {index}, {sales_price:.2f}, {market_price:.2f}, {cost_price:.2f}, 999, 0, NOW(), '0')"
        values.append(value)
        index += 1

    # 组合 SQL
    sql_lines.append(",\n".join(values) + ";")
    sql_lines.append("")
    sql_lines.append(f"-- 共导入 {len(values)} 条商品")
    sql_lines.append("SELECT COUNT(*) AS spu_count, '商品导入完成' AS status FROM goods_spu;")

    # 写入文件
    output_path = 'sql/product_import/03_products.sql'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(sql_lines))

    print(f"生成完成: {output_path}")
    print(f"共 {len(values)} 条商品 (在售: {index - 17}, 下架: 16)")

if __name__ == '__main__':
    main()
